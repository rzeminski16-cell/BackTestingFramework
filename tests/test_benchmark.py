"""
Tests for benchmark/index collection and report comparison.

Covers:
- transform_index_data parsing of an INDEX_DATA-style response
- compute_benchmark_comparison math (known beta/correlation/alpha)
- BenchmarkLoader file resolution + registry alias resolution
- The Excel report's Benchmark sheet (valid data and graceful unavailable case)
"""

import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace

import numpy as np
import pandas as pd
from openpyxl import Workbook

from Classes.DataCollection.benchmark_collector import (
    transform_index_data,
    resolve_benchmark,
    load_benchmark_registry,
)
from Classes.Analysis.benchmark import (
    BenchmarkLoader,
    compute_benchmark_comparison,
    write_comparison_sheet,
)
from Classes.Analysis.excel_report_generator import ExcelReportGenerator
from Classes.Analysis.portfolio_report_generator import PortfolioReportGenerator
from Classes.Analysis.enhanced_portfolio_report import EnhancedPortfolioReportGenerator

REGISTRY_PATH = Path("config/benchmarks.json")


def _known_beta_frames(beta_true=1.5, n=252, seed=0):
    np.random.seed(seed)
    dates = pd.bdate_range("2022-01-03", periods=n)
    r_b = np.random.normal(0.0004, 0.01, n)
    r_s = beta_true * r_b
    bench_close = 100 * np.cumprod(1 + r_b)
    strat_equity = 100000 * np.cumprod(1 + r_s)
    ec = pd.DataFrame({"date": dates, "equity": strat_equity})
    bp = pd.DataFrame({"date": dates, "close": bench_close})
    return ec, bp


class TestTransformIndexData(unittest.TestCase):
    def test_parses_and_sorts(self):
        resp = {
            "Meta Data": {"2. Symbol": "SPX"},
            "Time Series (Daily)": {
                "2024-01-03": {"1. open": "470", "2. high": "472", "3. low": "468", "4. close": "471"},
                "2024-01-02": {"1. open": "469", "2. high": "471", "3. low": "467", "4. close": "470"},
                "2024-01-04": {"1. open": "471", "2. high": "474", "3. low": "470", "4. close": "473"},
            },
        }
        df = transform_index_data(resp)
        self.assertEqual(list(df.columns)[:5], ["date", "open", "high", "low", "close"])
        self.assertTrue(pd.api.types.is_numeric_dtype(df["close"]))
        self.assertEqual(list(df["date"].astype(str)), ["2024-01-02", "2024-01-03", "2024-01-04"])

    def test_empty(self):
        self.assertTrue(transform_index_data({}).empty)
        self.assertTrue(transform_index_data({"Meta Data": {}}).empty)


class TestComparisonMath(unittest.TestCase):
    def test_known_beta(self):
        ec, bp = _known_beta_frames(beta_true=1.5)
        cmp = compute_benchmark_comparison(
            ec, bp, benchmark_name="S&P 500", benchmark_symbol="SPX", risk_free_rate=0.0
        )
        self.assertTrue(cmp.is_valid)
        self.assertAlmostEqual(cmp.beta, 1.5, places=4)
        self.assertAlmostEqual(cmp.correlation, 1.0, places=4)
        self.assertAlmostEqual(cmp.alpha, 0.0, places=2)
        # leveraged 1.5x series outperforms in a net-up window
        self.assertGreater(cmp.strategy_total_return_pct, cmp.benchmark_total_return_pct)
        self.assertAlmostEqual(
            cmp.excess_return_pct,
            cmp.strategy_total_return_pct - cmp.benchmark_total_return_pct, places=6
        )

    def test_invalid_inputs(self):
        ec, bp = _known_beta_frames()
        self.assertFalse(compute_benchmark_comparison(pd.DataFrame(), bp).is_valid)
        self.assertFalse(compute_benchmark_comparison(ec, pd.DataFrame()).is_valid)

    def test_no_overlap(self):
        ec = pd.DataFrame({"date": pd.bdate_range("2022-01-03", periods=10), "equity": range(10)})
        bp = pd.DataFrame({"date": pd.bdate_range("2010-01-04", periods=10), "close": range(10)})
        # benchmark entirely before strategy -> merge_asof backward leaves coverage,
        # but strategy starts after benchmark ends so closes are carried; ensure no crash
        result = compute_benchmark_comparison(ec, bp)
        self.assertIsNotNone(result)


class TestBenchmarkLoader(unittest.TestCase):
    def test_registry_resolution(self):
        reg = load_benchmark_registry(REGISTRY_PATH)
        self.assertEqual(resolve_benchmark("S&P 500", reg)[1]["symbol"], "SPX")
        self.assertEqual(resolve_benchmark("SPX", reg)[1]["symbol"], "SPX")
        self.assertEqual(resolve_benchmark("^GSPC", reg)[1]["symbol"], "SPX")  # alias
        self.assertIsNone(resolve_benchmark("NOPE", reg))

    def test_load_and_compare(self):
        ec, bp = _known_beta_frames()
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            bp.to_csv(tmp / "SPX_daily.csv", index=False)
            loader = BenchmarkLoader(benchmarks_dir=tmp, registry_path=REGISTRY_PATH)
            self.assertFalse(loader.load_series("S&P 500").empty)
            cmp = loader.compare(ec, "S&P 500", risk_free_rate=0.0)
            self.assertTrue(cmp.is_valid)
            self.assertAlmostEqual(cmp.beta, 1.5, places=4)

    def test_missing_data_is_graceful(self):
        ec, _ = _known_beta_frames()
        with tempfile.TemporaryDirectory() as tmp:
            loader = BenchmarkLoader(benchmarks_dir=Path(tmp), registry_path=REGISTRY_PATH)
            cmp = loader.compare(ec, "S&P 500")
            self.assertFalse(cmp.is_valid)
            self.assertIn("no stored data", cmp.reason)


class TestExcelBenchmarkSheet(unittest.TestCase):
    def _cells(self, ws):
        return [c.value for row in ws.iter_rows() for c in row if c.value is not None]

    def test_sheet_with_valid_data(self):
        ec, bp = _known_beta_frames()
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            bp.to_csv(tmp / "SPX_daily.csv", index=False)
            loader = BenchmarkLoader(benchmarks_dir=tmp, registry_path=REGISTRY_PATH)
            gen = ExcelReportGenerator(
                output_directory=tmp, benchmark_name="S&P 500",
                risk_free_rate=0.0, benchmark_loader=loader,
            )
            wb = Workbook()
            gen._create_benchmark_analysis(wb, SimpleNamespace(equity_curve=ec), {})
            self.assertIn("Benchmark", wb.sheetnames)
            cells = self._cells(wb["Benchmark"])
            self.assertIn("Beta", cells)
            self.assertIn("Alpha (annualized)", cells)
            self.assertTrue(any("S&P 500" in str(c) for c in cells))

    def test_sheet_unavailable_is_graceful(self):
        ec, _ = _known_beta_frames()
        with tempfile.TemporaryDirectory() as tmp:
            loader = BenchmarkLoader(benchmarks_dir=Path(tmp), registry_path=REGISTRY_PATH)
            gen = ExcelReportGenerator(
                output_directory=Path(tmp), benchmark_name="S&P 500",
                benchmark_loader=loader,
            )
            wb = Workbook()
            gen._create_benchmark_analysis(wb, SimpleNamespace(equity_curve=ec), {})
            cells = self._cells(wb["Benchmark"])
            self.assertTrue(any("unavailable" in str(c).lower() for c in cells))


class TestSharedRenderer(unittest.TestCase):
    def _cells(self, ws):
        return [c.value for row in ws.iter_rows() for c in row if c.value is not None]

    def test_valid(self):
        ec, bp = _known_beta_frames()
        cmp = compute_benchmark_comparison(ec, bp, benchmark_name="S&P 500", benchmark_symbol="SPX")
        wb = Workbook()
        ws = wb.active
        write_comparison_sheet(ws, cmp, title="PORTFOLIO VS BENCHMARK")
        cells = self._cells(ws)
        self.assertIn("PORTFOLIO VS BENCHMARK", cells)
        self.assertIn("Beta", cells)

    def test_invalid(self):
        wb = Workbook()
        ws = wb.active
        write_comparison_sheet(ws, None)
        self.assertTrue(any("unavailable" in str(c).lower() for c in self._cells(ws)))


class TestPortfolioWiring(unittest.TestCase):
    """The portfolio benchmark sheet renders without raising (data optional)."""

    def test_basic_generator_sheet(self):
        ec, _ = _known_beta_frames()
        with tempfile.TemporaryDirectory() as tmp:
            gen = PortfolioReportGenerator(Path(tmp), use_enhanced=False)
            wb = Workbook()
            gen._create_benchmark_sheet(wb, SimpleNamespace(portfolio_equity_curve=ec))
            self.assertIn("Benchmark", wb.sheetnames)
            self.assertTrue(len([c for r in wb["Benchmark"].iter_rows() for c in r if c.value]) > 0)

    def test_enhanced_generator_sheet(self):
        ec, _ = _known_beta_frames()
        with tempfile.TemporaryDirectory() as tmp:
            gen = EnhancedPortfolioReportGenerator(Path(tmp))
            wb = Workbook()
            gen._create_benchmark_analysis(wb, SimpleNamespace(portfolio_equity_curve=ec))
            self.assertIn("Benchmark", wb.sheetnames)


class _FakeResp:
    def __init__(self, success, data=None, error_message=None):
        self.success = success
        self.data = data
        self.error_message = error_message


def _daily_payload():
    return {"Meta Data": {"2. Symbol": "SPY"},
            "Time Series (Daily)": {
                "2024-01-02": {"1. open": "470", "2. high": "472", "3. low": "468",
                               "4. close": "471", "5. volume": "1000"},
                "2024-01-03": {"1. open": "471", "2. high": "474", "3. low": "470",
                               "4. close": "473", "5. volume": "1100"}}}


class TestBenchmarkCollector(unittest.TestCase):
    def _registry(self):
        from Classes.DataCollection.benchmark_collector import load_benchmark_registry
        return load_benchmark_registry(REGISTRY_PATH)

    def test_index_data_success(self):
        from Classes.DataCollection.benchmark_collector import BenchmarkCollector
        idx = {"Meta Data": {}, "Time Series (Daily)": {
            "2024-01-02": {"1. open": "1", "2. high": "1", "3. low": "1", "4. close": "100"}}}

        class C:
            def get_index_data(self, *a, **k): return _FakeResp(True, idx)
            def get_daily_prices(self, *a, **k): return _FakeResp(False, error_message="should not be called")
        res = BenchmarkCollector(C(), registry=self._registry()).collect("SPX")
        self.assertFalse(res.empty)
        self.assertEqual(res.source, "index_data")

    def test_falls_back_to_etf_proxy(self):
        from Classes.DataCollection.benchmark_collector import BenchmarkCollector

        class C:
            def get_index_data(self, *a, **k):
                return _FakeResp(False, error_message="Invalid API call / premium endpoint")
            def get_daily_prices(self, symbol, adjusted=True, outputsize="full"):
                assert symbol == "SPY"  # resolved from registry etf
                return _FakeResp(True, _daily_payload())
        res = BenchmarkCollector(C(), registry=self._registry()).collect("SPX")
        self.assertFalse(res.empty)
        self.assertEqual(res.source, "etf_proxy")
        self.assertIn("close", res.df.columns)
        self.assertEqual(res.df["symbol"].iloc[0], "SPX")  # stored under the index name

    def test_both_fail_surfaces_real_error(self):
        from Classes.DataCollection.benchmark_collector import BenchmarkCollector

        class C:
            def get_index_data(self, *a, **k):
                return _FakeResp(False, error_message="premium endpoint")
            def get_daily_prices(self, *a, **k):
                return _FakeResp(False, error_message="rate limit")
        res = BenchmarkCollector(C(), registry=self._registry()).collect("SPX")
        self.assertTrue(res.empty)
        self.assertIn("premium endpoint", res.error)
        self.assertIn("rate limit", res.error)


if __name__ == "__main__":
    unittest.main()
