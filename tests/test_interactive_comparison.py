"""
Tests for the AUTO baseline twin and the discretion-vs-rules comparison.

Deterministic scenario: two symbols, the user rejects one entry that the
rules-only baseline takes. The comparison must show the baseline trade,
the per-decision delta, and matching aggregate deltas.
"""
import json

import pandas as pd
import pytest
from openpyxl import load_workbook

from Classes.Config.config import (
    CommissionConfig, CommissionMode, PortfolioConfig,
)
from Classes.Engine.portfolio_engine import PortfolioEngine
from Classes.Interactive.baseline import (
    extract_trades_and_equity,
    run_auto_baseline,
    write_baseline_outputs,
)
from Classes.Interactive.comparison import (
    compare_runs,
    write_comparison_json,
    write_comparison_workbook,
)
from Classes.Interactive.decision_provider import ScriptedDecisionProvider
from Classes.Interactive.models import (
    BacktestRunManifest,
    DecisionAction,
    DecisionResponse,
    PromptRecord,
)
from Classes.Interactive.session import InteractiveSession
from Classes.Interactive.store import InteractiveRunStore
from Classes.Models.signal import Signal
from Classes.Models.trade_direction import TradeDirection
from Classes.Strategy.base_strategy import BaseStrategy


def _config():
    return PortfolioConfig(
        initial_capital=100_000.0,
        commission=CommissionConfig(mode=CommissionMode.PERCENTAGE, value=0.0),
        slippage_percent=0.0)


class TwoSymbolStrategy(BaseStrategy):
    """AAA enters bar 2 exits bar 8; BBB enters bar 3 exits bar 9."""

    _validate_on_init = False

    ENTRY_BARS = {"AAA": 2, "BBB": 3}
    EXIT_BARS = {"AAA": 8, "BBB": 9}

    def __init__(self, **params):
        super().__init__(**params)

    @property
    def trade_direction(self):
        return TradeDirection.LONG

    def required_columns(self):
        return ["date", "close"]

    def generate_entry_signal(self, context):
        if context.current_index == self.ENTRY_BARS.get(context.symbol, -1):
            return Signal.buy(size=1.0,
                              stop_loss=context.current_price * 0.5,
                              direction=TradeDirection.LONG, reason="entry")
        return None

    def calculate_initial_stop_loss(self, context):
        return context.current_price * 0.5

    def generate_exit_signal(self, context):
        if context.current_index == self.EXIT_BARS.get(context.symbol, -1):
            return Signal.sell(reason="exit")
        return None

    def position_size(self, context, signal):
        return (context.total_equity * 0.3) / context.current_price


def _data_dict():
    n = 12
    dates = pd.date_range("2024-01-01", periods=n, freq="D")

    def make(base):
        prices = [base + i for i in range(n)]  # rising: trades are winners
        return pd.DataFrame({
            "date": dates, "close": prices, "open": prices,
            "high": prices, "low": prices, "volume": [1000.0] * n,
        })

    return {"AAA": make(100.0), "BBB": make(50.0)}


@pytest.fixture
def paired_runs(tmp_path):
    """Interactive run (BBB entry rejected) + AUTO baseline twin."""
    store = InteractiveRunStore(tmp_path / "run")
    manifest = BacktestRunManifest(
        run_id="cmp", backtest_name="cmp", mode="interactive",
        engine_type="portfolio")
    store.write_manifest(manifest)

    def scripted(request):
        if (request.event.symbol == "BBB"
                and request.event.signal_type == "BUY"):
            return DecisionResponse(action=DecisionAction.REJECT,
                                    rationale="valuation stretched")
        return DecisionResponse(action=DecisionAction.ACCEPT)

    session = InteractiveSession(store, ScriptedDecisionProvider(scripted),
                                 manifest)
    interactive = PortfolioEngine(_config()).run(
        _data_dict(), TwoSymbolStrategy(), decision_session=session)
    session.finalize("completed")

    baseline = run_auto_baseline(
        "portfolio", _config(), _data_dict(), TwoSymbolStrategy)
    return interactive, baseline, store, tmp_path


class TestBaseline:
    def test_baseline_takes_the_rejected_trade(self, paired_runs):
        interactive, baseline, _, _ = paired_runs
        interactive_trades, _ = extract_trades_and_equity(interactive)
        baseline_trades, _ = extract_trades_and_equity(baseline)
        assert {t.symbol for t in interactive_trades} == {"AAA"}
        assert {t.symbol for t in baseline_trades} == {"AAA", "BBB"}
        # AAA is identical in both (rejection happened after AAA's entry,
        # but AAA sizing depends only on equity at bar 2, still equal).
        aaa_int = [t for t in interactive_trades if t.symbol == "AAA"][0]
        aaa_base = [t for t in baseline_trades if t.symbol == "AAA"][0]
        assert aaa_int.quantity == aaa_base.quantity
        assert aaa_int.pl == aaa_base.pl

    def test_write_baseline_outputs(self, paired_runs):
        _, baseline, store, tmp_path = paired_runs
        baseline_dir = write_baseline_outputs(baseline, store.run_dir, "cmp")
        assert (baseline_dir / "baseline_trades.csv").exists()
        assert (baseline_dir / "baseline_equity.csv").exists()
        manifest = json.loads(
            (baseline_dir / "baseline_manifest.json").read_text())
        assert manifest['mode'] == "auto_baseline"
        assert manifest['interactive_run_id'] == "cmp"
        assert manifest['num_trades'] == 2


class TestComparison:
    def test_compare_runs(self, paired_runs):
        interactive, baseline, store, _ = paired_runs
        comparison = compare_runs(
            interactive, baseline,
            events=store.load_events(),
            decisions=store.load_decisions(),
            outcomes=store.load_outcomes())

        # The rejected BBB entry is the one divergent decision.
        assert comparison.headline['num_divergent_decisions'] == 1
        row = comparison.decision_rows[0]
        assert row['symbol'] == "BBB"
        assert row['action'] == "reject"
        assert row['rationale'] == "valuation stretched"
        assert row['executed'] is False
        assert row['attribution_quality'] == "exact"
        # BBB rose from 53 to 61 in the baseline: a winning trade the
        # discretionary run skipped.
        assert row['baseline_trade_pl'] is not None
        assert row['baseline_trade_pl'] > 0
        assert row['interactive_trade_pl'] is None
        assert row['pl_delta'] == pytest.approx(-row['baseline_trade_pl'])

        # Aggregate delta equals the skipped trade's P&L (only divergence).
        assert comparison.headline['discretion_delta'] == pytest.approx(
            -row['baseline_trade_pl'])
        # Equity curves aligned and end at the final equities.
        curves = comparison.equity_curves
        assert curves['interactive'].iloc[-1] == pytest.approx(
            interactive.final_equity)
        assert curves['baseline'].iloc[-1] == pytest.approx(
            baseline.final_equity)

    def test_workbook_and_json(self, paired_runs, tmp_path):
        interactive, baseline, store, _ = paired_runs
        comparison = compare_runs(
            interactive, baseline,
            events=store.load_events(),
            decisions=store.load_decisions(),
            outcomes=store.load_outcomes())
        prompts = [PromptRecord(prompt_id=1, event_id=1, run_id="cmp",
                                prompt_text="Assume today is 2024-01-04...",
                                response_summary="looks rich")]
        path = write_comparison_workbook(comparison, tmp_path / "reports",
                                         prompts=prompts)
        assert path.exists()
        wb = load_workbook(path)
        for sheet in ("Summary", "Equity Curves", "Decisions", "Prompts",
                      "Methodology"):
            assert sheet in wb.sheetnames
        decisions_ws = wb["Decisions"]
        values = [c.value for c in decisions_ws[2]]
        assert "BBB" in values
        assert "valuation stretched" in values

        json_path = write_comparison_json(
            comparison, tmp_path / "exports" / "comparison.json")
        payload = json.loads(json_path.read_text())
        assert payload['headline']['num_divergent_decisions'] == 1

    def test_export_flat_includes_outcomes(self, paired_runs):
        _, _, store, _ = paired_runs
        paths = store.export_flat()
        assert any(p.suffix == ".csv" for p in paths)
        df = store.build_flat_table()
        rejected = df[df['action'] == 'reject']
        assert len(rejected) == 1
        assert rejected.iloc[0]['symbol'] == "BBB"
        assert bool(rejected.iloc[0]['executed']) is False
