"""
Tests for Classes/Analysis/excel_charts.py (native Excel chart helpers).

Note: openpyxl does NOT parse charts back when loading a workbook, so we assert
on the in-memory workbook (ws._charts) for chart presence, and inspect the saved
.xlsx zip parts for file-level verification (chart XML, tables, conditional
formatting, hidden data sheet).
"""

import io
import unittest
import zipfile

from openpyxl import Workbook, load_workbook

from Classes.Analysis import excel_charts as ec


def _save_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


class TestSeriesBlock(unittest.TestCase):
    def test_write_block_values_and_cursor(self):
        wb = Workbook()
        block = ec.write_series_block(
            wb, ["Date", "Equity"],
            [["2020-01-01", 100.0], ["2020-01-02", 101.5], ["2020-01-03", 99.0]],
        )
        self.assertIsNotNone(block)
        ws = ec.get_chart_data_sheet(wb)
        self.assertEqual(ws.sheet_state, "hidden")
        # header + values written
        self.assertEqual(ws.cell(row=1, column=1).value, "Date")
        self.assertEqual(ws.cell(row=1, column=2).value, "Equity")
        self.assertEqual(ws.cell(row=2, column=2).value, 100.0)
        self.assertEqual(ws.cell(row=4, column=2).value, 99.0)
        self.assertEqual(block["max_row"], 4)
        # a second block lands in later columns (cursor advanced past a gap)
        block2 = ec.write_series_block(wb, ["X", "Y"], [[1, 2], [3, 4]])
        self.assertGreater(block2["start_col"], block["start_col"] + 1)

    def test_empty_block(self):
        wb = Workbook()
        self.assertIsNone(ec.write_series_block(wb, [], []))
        self.assertIsNone(ec.write_series_block(wb, ["A", "B"], []))


class TestChartBuilders(unittest.TestCase):
    def setUp(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.block = ec.write_series_block(
            self.wb, ["Cat", "S1", "S2"],
            [["a", 1, 4], ["b", 2, 5], ["c", 3, 6]],
        )

    def test_line_bar_area_pie(self):
        ec.add_line_chart(self.ws, "E2", self.block, title="L")
        ec.add_bar_chart(self.ws, "E20", self.block, title="B")
        ec.add_stacked_area_chart(self.ws, "E40", self.block, title="A")
        ec.add_pie_chart(self.ws, "E60", self.block, title="P")
        self.assertEqual(len(self.ws._charts), 4)

    def test_scatter(self):
        ec.add_scatter_chart(self.ws, "E2", self.block, title="S", x_title="Cat", y_title="val")
        self.assertEqual(len(self.ws._charts), 1)

    def test_color_scale_and_table(self):
        for r, vals in enumerate([("h1", "h2"), (1, 2), (3, 4)], start=1):
            for c, v in enumerate(vals, start=1):
                self.ws.cell(row=r, column=c, value=v)
        ec.apply_color_scale(self.ws, "A2:B3")
        self.assertTrue(len(self.ws.conditional_formatting) >= 1)
        ec.make_table(self.ws, "A1:B3", "MyTable")
        self.assertIn("MyTable", self.ws.tables)

    def test_safe_chart_swallows_errors(self):
        # bad block -> helper returns None, no raise, no chart added
        result = ec.add_line_chart(self.ws, "E2", {"data": None, "cats": None}, title="x")
        self.assertIsNone(result)
        self.assertEqual(len(self.ws._charts), 0)


class TestSavedFileParts(unittest.TestCase):
    def test_xlsx_contains_chart_table_and_hidden_sheet(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        block = ec.write_series_block(wb, ["Cat", "Val"], [["a", 1], ["b", 2], ["c", 3]])
        ec.add_line_chart(ws, "E2", block, title="L")
        for r, vals in enumerate([("h1", "h2"), (1, 2), (3, 4)], start=1):
            for c, v in enumerate(vals, start=1):
                ws.cell(row=r, column=c, value=v)
        ec.apply_color_scale(ws, "A2:B3")
        ec.make_table(ws, "A1:B3", "T1")

        data = _save_bytes(wb)
        names = zipfile.ZipFile(io.BytesIO(data)).namelist()
        self.assertTrue(any(n.startswith("xl/charts/chart") for n in names), names)
        self.assertTrue(any(n.startswith("xl/tables/table") for n in names), names)

        # cell values survive round-trip; _ChartData hidden
        wb2 = load_workbook(io.BytesIO(data))
        self.assertIn(ec.CHART_DATA_SHEET, wb2.sheetnames)
        self.assertEqual(wb2[ec.CHART_DATA_SHEET].sheet_state, "hidden")
        self.assertEqual(wb2[ec.CHART_DATA_SHEET].cell(row=2, column=2).value, 1)

    def test_sample_series_display_only(self):
        vals = list(range(2000))
        dates = [f"d{i}" for i in range(2000)]
        sd, sv = ec.sample_series(vals, dates, max_points=500)
        self.assertLessEqual(len(sv), 502)
        self.assertEqual(sv[0], 0)
        self.assertEqual(sv[-1], 1999)  # endpoints preserved


if __name__ == "__main__":
    unittest.main()
