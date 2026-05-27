"""
Integration tests for native interactive Excel charts in the report generators.

openpyxl does not parse charts back on load, so chart presence is verified by
inspecting the saved .xlsx zip parts (xl/charts/chartN.xml). Value fidelity is
checked by reading the hidden _ChartData cells (which DO round-trip) and
comparing to the source series.
"""

import io
import tempfile
import unittest
import zipfile
from pathlib import Path
from types import SimpleNamespace

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook

from Classes.Models.trade import Trade
from Classes.Engine.backtest_result import BacktestResult
from Classes.Analysis.excel_report_generator import ExcelReportGenerator
from Classes.Analysis.enhanced_portfolio_report import EnhancedPortfolioReportGenerator


def _chart_part_count(path_or_bytes):
    if isinstance(path_or_bytes, (str, Path)):
        zf = zipfile.ZipFile(open(path_or_bytes, "rb"))
    else:
        zf = zipfile.ZipFile(io.BytesIO(path_or_bytes))
    return sum(1 for n in zf.namelist() if n.startswith("xl/charts/chart"))


def _make_trades(n=60, seed=11):
    rng = np.random.default_rng(seed)
    base = pd.Timestamp("2020-01-02")
    trades = []
    for i in range(n):
        entry = base + pd.Timedelta(days=int(i * 10))
        ex = entry + pd.Timedelta(days=int(rng.integers(2, 20)))
        ep = 100 + rng.normal(0, 5)
        plpct = rng.normal(1.0, 6.0)
        xp = ep * (1 + plpct / 100)
        trades.append(Trade(symbol="TEST", entry_date=entry, exit_date=ex, entry_price=ep,
                            exit_price=xp, quantity=100, side="LONG", pl=(xp - ep) * 100,
                            pl_pct=plpct, duration_days=(ex - entry).days,
                            initial_stop_loss=ep * 0.92))
    return trades


def _equity_df(n=620, seed=11):
    rng = np.random.default_rng(seed + 1)
    dates = pd.bdate_range("2020-01-02", periods=n)
    eq = 100000 * np.cumprod(1 + rng.normal(0.0006, 0.011, n))
    cap = eq * 0.5
    return pd.DataFrame({"date": dates, "equity": eq, "capital": cap,
                         "position_value": eq - cap, "num_positions": 1})


class TestSingleBacktestReport(unittest.TestCase):
    def _result(self):
        edf = _equity_df()
        return BacktestResult(symbol="TEST", strategy_name="Demo", trades=_make_trades(),
                              equity_curve=edf, final_equity=float(edf["equity"].iloc[-1]),
                              total_return=float(edf["equity"].iloc[-1] - 100000),
                              total_return_pct=float((edf["equity"].iloc[-1] / 100000 - 1) * 100))

    def test_native_report_has_charts(self):
        with tempfile.TemporaryDirectory() as tmp:
            gen = ExcelReportGenerator(output_directory=Path(tmp), prefer_native_charts=True)
            path = gen.generate_report(self._result(), "report.xlsx")
            wb = load_workbook(path)
            self.assertIn("Enhanced Charts", wb.sheetnames)
            self.assertIn("_ChartData", wb.sheetnames)
            self.assertEqual(wb["_ChartData"].sheet_state, "hidden")
            self.assertGreaterEqual(_chart_part_count(path), 8)

    def test_value_fidelity(self):
        from Classes.Analysis import report_data as rd
        result = self._result()
        with tempfile.TemporaryDirectory() as tmp:
            gen = ExcelReportGenerator(output_directory=Path(tmp), prefer_native_charts=True)
            path = gen.generate_report(result, "report.xlsx")
            wb = load_workbook(path)
            # The first _ChartData block is the equity series; first value must equal source.
            expected = rd.compute_equity_drawdown(result.equity_curve)["equity"][0]
            self.assertAlmostEqual(wb["_ChartData"].cell(row=2, column=2).value, expected, places=4)

    def test_png_path_still_works(self):
        # prefer_native_charts=False, use_enhanced=False -> no enhanced sheet, no crash.
        with tempfile.TemporaryDirectory() as tmp:
            gen = ExcelReportGenerator(output_directory=Path(tmp), prefer_native_charts=False,
                                       use_enhanced=False)
            path = gen.generate_report(self._result(), "report.xlsx")
            wb = load_workbook(path)
            self.assertNotIn("Enhanced Charts", wb.sheetnames)


class TestPortfolioDashboard(unittest.TestCase):
    def test_dashboard_native_charts(self):
        edf = _equity_df()
        trades = _make_trades()
        sym_results = {"AAA": SimpleNamespace(total_return=5000.0, trades=trades[:30]),
                       "BBB": SimpleNamespace(total_return=-1500.0, trades=trades[30:])}
        result = SimpleNamespace(portfolio_equity_curve=edf, symbol_results=sym_results,
                                 signal_rejections=[], vulnerability_swaps=[], vulnerability_history=[])
        with tempfile.TemporaryDirectory() as tmp:
            gen = EnhancedPortfolioReportGenerator(Path(tmp), prefer_native_charts=True)
            wb = Workbook()
            gen._create_dashboard_visualizations(wb, result, {"all_trades": trades})
            self.assertGreaterEqual(len(wb["Dashboard"]._charts), 5)


class TestOptimizationWFChart(unittest.TestCase):
    def test_wf_equity_native(self):
        try:
            from Classes.Optimization.enhanced_optimization_report import EnhancedOptimizationReportGenerator
        except Exception:
            self.skipTest("optimization deps unavailable")

        def win(i, ret):
            return SimpleNamespace(window_id=i,
                                   test_start=pd.Timestamp("2020-01-01") + pd.Timedelta(days=90 * i),
                                   test_end=pd.Timestamp("2020-01-01") + pd.Timedelta(days=90 * (i + 1)),
                                   out_sample_total_return_pct=ret, out_sample_sortino=1.0)
        mr = SimpleNamespace(individual_results={
            "AAA": SimpleNamespace(windows=[win(0, 3.0), win(1, -2.0), win(2, 5.0)]),
        })
        gen = EnhancedOptimizationReportGenerator(config={})
        wb = Workbook()
        gen._native_wf_equity_chart(wb, wb.active, "A1", mr)
        self.assertEqual(len(wb.active._charts), 1)
        # fidelity: AAA window0 normalized equity = 100 * 1.03 = 103
        self.assertAlmostEqual(wb["_ChartData"].cell(row=2, column=2).value, 103.0, places=6)


if __name__ == "__main__":
    unittest.main()
