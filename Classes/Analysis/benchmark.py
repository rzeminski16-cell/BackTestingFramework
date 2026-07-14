"""
Benchmark comparison for generated reports.

Loads a stored benchmark/index price series (collected via INDEX_DATA into
``raw_data/benchmarks/``) and compares it against a strategy equity curve,
producing the metrics reports need: benchmark equity overlay, total return and
CAGR vs benchmark, excess return, alpha, beta, correlation, tracking error,
information ratio, and up/down capture.

This is the single entry point used by every report generator so the
comparison is computed consistently everywhere.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import numpy as np
import pandas as pd

from ..Core.performance_metrics import DEFAULT_RISK_FREE_RATE, TRADING_DAYS_PER_YEAR
from ..DataCollection.benchmark_collector import (
    DEFAULT_BENCHMARK_DIR,
    DEFAULT_REGISTRY_PATH,
    load_benchmark_registry,
    resolve_benchmark,
)

logger = logging.getLogger(__name__)


@dataclass
class BenchmarkComparison:
    """Result of comparing a strategy equity curve against a benchmark."""

    is_valid: bool
    benchmark_name: str = ""
    benchmark_symbol: str = ""
    reason: str = ""

    start_date: Optional[pd.Timestamp] = None
    end_date: Optional[pd.Timestamp] = None

    strategy_total_return_pct: float = 0.0
    benchmark_total_return_pct: float = 0.0
    excess_return_pct: float = 0.0

    strategy_cagr: float = 0.0
    benchmark_cagr: float = 0.0

    alpha: float = 0.0            # annualized, percent
    beta: float = 0.0
    correlation: float = 0.0
    tracking_error: float = 0.0   # annualized, percent
    information_ratio: float = 0.0
    up_capture: float = 0.0       # percent
    down_capture: float = 0.0     # percent
    # Max drawdown of the strategy/benchmark relative line: the worst
    # peak-to-trough episode of UNDERPERFORMANCE vs the benchmark (percent).
    relative_max_drawdown_pct: float = 0.0

    # date / equity DataFrame for an overlay chart (benchmark rebased to the
    # strategy's starting equity).
    benchmark_equity: Optional[pd.DataFrame] = field(default=None, repr=False)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "is_valid": self.is_valid,
            "benchmark_name": self.benchmark_name,
            "benchmark_symbol": self.benchmark_symbol,
            "strategy_total_return_pct": self.strategy_total_return_pct,
            "benchmark_total_return_pct": self.benchmark_total_return_pct,
            "excess_return_pct": self.excess_return_pct,
            "strategy_cagr": self.strategy_cagr,
            "benchmark_cagr": self.benchmark_cagr,
            "alpha": self.alpha,
            "beta": self.beta,
            "correlation": self.correlation,
            "tracking_error": self.tracking_error,
            "information_ratio": self.information_ratio,
            "up_capture": self.up_capture,
            "down_capture": self.down_capture,
            "relative_max_drawdown_pct": self.relative_max_drawdown_pct,
        }

    def summary_rows(self) -> List[Tuple[str, Any]]:
        """(label, formatted value) rows for rendering into a report table."""
        if not self.is_valid:
            return [("Benchmark comparison", f"Unavailable: {self.reason}")]
        return [
            ("Benchmark", f"{self.benchmark_name} ({self.benchmark_symbol})"),
            ("Comparison Window",
             f"{self.start_date:%Y-%m-%d} to {self.end_date:%Y-%m-%d}"),
            ("Strategy Total Return", f"{self.strategy_total_return_pct:.2f}%"),
            ("Benchmark Total Return", f"{self.benchmark_total_return_pct:.2f}%"),
            ("Excess Return", f"{self.excess_return_pct:.2f}%"),
            ("Strategy CAGR", f"{self.strategy_cagr:.2f}%"),
            ("Benchmark CAGR", f"{self.benchmark_cagr:.2f}%"),
            ("Alpha (annualized)", f"{self.alpha:.2f}%"),
            ("Beta", f"{self.beta:.2f}"),
            ("Correlation", f"{self.correlation:.2f}"),
            ("Tracking Error (annualized)", f"{self.tracking_error:.2f}%"),
            ("Information Ratio", f"{self.information_ratio:.2f}"),
            ("Up Capture", f"{self.up_capture:.1f}%"),
            ("Down Capture", f"{self.down_capture:.1f}%"),
            ("Relative Max Drawdown", f"{self.relative_max_drawdown_pct:.2f}%"),
        ]


def _invalid(reason: str, name: str = "", symbol: str = "") -> BenchmarkComparison:
    return BenchmarkComparison(is_valid=False, benchmark_name=name, benchmark_symbol=symbol, reason=reason)


def _cagr(start_value: float, end_value: float, years: float) -> float:
    if years <= 0 or start_value <= 0 or end_value <= 0:
        return 0.0
    return (pow(end_value / start_value, 1.0 / years) - 1.0) * 100.0


def compute_benchmark_comparison(
    equity_curve: pd.DataFrame,
    benchmark_prices: pd.DataFrame,
    *,
    benchmark_name: str = "",
    benchmark_symbol: str = "",
    risk_free_rate: Optional[float] = None,
) -> BenchmarkComparison:
    """
    Compare a strategy equity curve against a benchmark price series.

    Args:
        equity_curve: DataFrame with 'date' and 'equity' columns.
        benchmark_prices: DataFrame with 'date' and 'close' columns.
        benchmark_name / benchmark_symbol: labels for the report.
        risk_free_rate: annual risk-free rate (defaults to framework default).

    Returns:
        BenchmarkComparison (is_valid=False with a reason if it cannot be
        computed, e.g. missing columns or no overlapping dates).
    """
    if risk_free_rate is None:
        risk_free_rate = DEFAULT_RISK_FREE_RATE

    if equity_curve is None or equity_curve.empty or "equity" not in equity_curve.columns or "date" not in equity_curve.columns:
        return _invalid("strategy equity curve missing date/equity", benchmark_name, benchmark_symbol)
    if benchmark_prices is None or benchmark_prices.empty or "close" not in benchmark_prices.columns or "date" not in benchmark_prices.columns:
        return _invalid("benchmark price series unavailable", benchmark_name, benchmark_symbol)

    ec = equity_curve[["date", "equity"]].copy()
    ec["date"] = pd.to_datetime(ec["date"], errors="coerce")
    ec = ec.dropna(subset=["date"]).drop_duplicates("date").sort_values("date").reset_index(drop=True)

    bp = benchmark_prices[["date", "close"]].copy()
    bp["date"] = pd.to_datetime(bp["date"], errors="coerce")
    bp["close"] = pd.to_numeric(bp["close"], errors="coerce")
    bp = bp.dropna(subset=["date", "close"]).drop_duplicates("date").sort_values("date").reset_index(drop=True)

    if len(ec) < 2 or len(bp) < 2:
        return _invalid("not enough data points", benchmark_name, benchmark_symbol)

    # Align the most recent benchmark close to each strategy date (handles a
    # benchmark sampled less frequently than the daily equity curve).
    merged = pd.merge_asof(ec, bp, on="date", direction="backward")
    merged = merged.dropna(subset=["close"]).reset_index(drop=True)
    if len(merged) < 2:
        return _invalid("no overlapping dates between strategy and benchmark", benchmark_name, benchmark_symbol)

    start_equity = float(merged["equity"].iloc[0])
    start_close = float(merged["close"].iloc[0])
    if start_equity <= 0 or start_close <= 0:
        return _invalid("non-positive starting values", benchmark_name, benchmark_symbol)

    start_date = merged["date"].iloc[0]
    end_date = merged["date"].iloc[-1]
    years = (end_date - start_date).days / 365.25

    # Benchmark rebased to the strategy's starting equity (for an overlay).
    benchmark_equity = pd.DataFrame({
        "date": merged["date"],
        "equity": start_equity * (merged["close"] / start_close),
    })

    strat_total = (float(merged["equity"].iloc[-1]) / start_equity - 1.0) * 100.0
    bench_total = (float(merged["close"].iloc[-1]) / start_close - 1.0) * 100.0

    r_s = merged["equity"].pct_change()
    r_b = merged["close"].pct_change()
    rets = pd.DataFrame({"s": r_s, "b": r_b}).replace([np.inf, -np.inf], np.nan).dropna()

    beta = correlation = alpha = tracking_error = information_ratio = 0.0
    up_capture = down_capture = 0.0

    if len(rets) >= 2:
        var_b = float(rets["b"].var())
        if var_b > 0:
            beta = float(rets["s"].cov(rets["b"]) / var_b)
        if rets["s"].std() > 0 and rets["b"].std() > 0:
            correlation = float(rets["s"].corr(rets["b"]))

        daily_rf = (1 + risk_free_rate) ** (1 / TRADING_DAYS_PER_YEAR) - 1
        alpha_daily = (rets["s"] - daily_rf).mean() - beta * (rets["b"] - daily_rf).mean()
        alpha = (pow(1 + alpha_daily, TRADING_DAYS_PER_YEAR) - 1) * 100.0

        active = rets["s"] - rets["b"]
        active_std = float(active.std())
        if active_std > 0:
            tracking_error = active_std * np.sqrt(TRADING_DAYS_PER_YEAR) * 100.0
            information_ratio = float(active.mean() / active_std * np.sqrt(TRADING_DAYS_PER_YEAR))

        up = rets[rets["b"] > 0]
        down = rets[rets["b"] < 0]
        if len(up) > 0:
            bench_up = (1 + up["b"]).prod() - 1
            strat_up = (1 + up["s"]).prod() - 1
            if bench_up != 0:
                up_capture = float(strat_up / bench_up * 100.0)
        if len(down) > 0:
            bench_dn = (1 + down["b"]).prod() - 1
            strat_dn = (1 + down["s"]).prod() - 1
            if bench_dn != 0:
                down_capture = float(strat_dn / bench_dn * 100.0)

    # Benchmark-relative max drawdown: worst peak-to-trough decline of the
    # strategy/benchmark ratio — the deepest sustained underperformance
    # episode, invisible in the absolute drawdown when both fall together.
    relative_max_drawdown = 0.0
    relative_line = (merged["equity"] / start_equity) / (merged["close"] / start_close)
    running_peak = relative_line.cummax()
    rel_dd = (running_peak - relative_line) / running_peak * 100.0
    if len(rel_dd):
        relative_max_drawdown = float(rel_dd.max())

    return BenchmarkComparison(
        is_valid=True,
        benchmark_name=benchmark_name,
        benchmark_symbol=benchmark_symbol,
        start_date=start_date,
        end_date=end_date,
        strategy_total_return_pct=strat_total,
        benchmark_total_return_pct=bench_total,
        excess_return_pct=strat_total - bench_total,
        strategy_cagr=_cagr(start_equity, float(merged["equity"].iloc[-1]), years),
        benchmark_cagr=_cagr(start_close, float(merged["close"].iloc[-1]), years),
        alpha=alpha,
        beta=beta,
        correlation=correlation,
        tracking_error=tracking_error,
        information_ratio=information_ratio,
        up_capture=up_capture,
        down_capture=down_capture,
        relative_max_drawdown_pct=relative_max_drawdown,
        benchmark_equity=benchmark_equity,
    )


def write_comparison_sheet(
    ws,
    comparisons: Union[BenchmarkComparison, Dict[str, BenchmarkComparison], None],
    *,
    equity_curve: Optional[pd.DataFrame] = None,
    title: str = "BENCHMARK COMPARISON",
    collect_hint: str = "python scripts/collect_benchmarks.py",
) -> None:
    """
    Render benchmark comparisons onto an openpyxl worksheet.

    Accepts either a single BenchmarkComparison (backward-compatible) or a
    dict of {benchmark_name: BenchmarkComparison} for multi-benchmark display.
    Produces a side-by-side metrics table plus a multi-series return-over-time
    chart. ``equity_curve`` (DataFrame with 'date'/'equity' columns) adds the
    strategy line to the chart.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # Normalise: accept single comparison or dict
    if comparisons is None:
        comp_dict: Dict[str, BenchmarkComparison] = {}
    elif isinstance(comparisons, BenchmarkComparison):
        comp_dict = {comparisons.benchmark_name or "Benchmark": comparisons}
    else:
        comp_dict = dict(comparisons)

    valid_comps: Dict[str, BenchmarkComparison] = {
        name: c for name, c in comp_dict.items() if c is not None and c.is_valid
    }

    # Style helpers
    DARK_BLUE   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    MED_BLUE    = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    ALT_FILL    = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
    WHITE_FILL  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    WHITE_FONT  = Font(color="FFFFFF", bold=True, size=11)
    BOLD_FONT   = Font(bold=True)
    CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT        = Alignment(horizontal="left", vertical="center")

    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)

    HEADER_ROW     = 3
    METRIC_COL     = 1
    STRATEGY_COL   = 2
    bench_names    = list(valid_comps.keys())
    n_bench        = len(bench_names)

    # ── No valid comparisons ─────────────────────────────────────────────────
    if not valid_comps:
        reasons = [f"{name}: {c.reason}" for name, c in comp_dict.items()
                   if c is not None and not c.is_valid]
        if not reasons:
            reasons = ["benchmark module unavailable"]
        r = 3
        ws.cell(row=r, column=1, value="Benchmark comparison unavailable").font = BOLD_FONT
        for i, reason in enumerate(reasons):
            ws.cell(row=r + 1 + i, column=1, value=f"  {reason}")
        ws.cell(row=r + len(reasons) + 1, column=1,
                value=f"Collect index data with: {collect_hint}")
        ws.column_dimensions['A'].width = 60
        return

    # ── Metric rows definition ───────────────────────────────────────────────
    # (label, strategy_fn | None, benchmark_fn, highlight_fn | None)
    # strategy_fn receives the first valid comp (strategy values are identical)
    # benchmark_fn receives the per-benchmark comp
    # highlight_fn(comp) -> "green"|"red"|None
    first = next(iter(valid_comps.values()))

    METRICS = [
        ("Comparison Window",
         None,
         lambda c: f"{c.start_date:%Y-%m-%d} → {c.end_date:%Y-%m-%d}",
         None),
        ("Total Return (%)",
         lambda c: f"{c.strategy_total_return_pct:.2f}%",
         lambda c: f"{c.benchmark_total_return_pct:.2f}%",
         None),
        ("CAGR (%)",
         lambda c: f"{c.strategy_cagr:.2f}%",
         lambda c: f"{c.benchmark_cagr:.2f}%",
         None),
        ("Excess Return (%)",
         None,
         lambda c: f"{c.excess_return_pct:+.2f}%",
         lambda c: "green" if c.excess_return_pct > 0 else ("red" if c.excess_return_pct < 0 else None)),
        ("Alpha (annualized, %)",
         None,
         lambda c: f"{c.alpha:.2f}%",
         lambda c: "green" if c.alpha > 0 else ("red" if c.alpha < 0 else None)),
        ("Beta",
         None,
         lambda c: f"{c.beta:.2f}",
         None),
        ("Correlation",
         None,
         lambda c: f"{c.correlation:.2f}",
         None),
        ("Tracking Error (ann., %)",
         None,
         lambda c: f"{c.tracking_error:.2f}%",
         None),
        ("Information Ratio",
         None,
         lambda c: f"{c.information_ratio:.2f}",
         lambda c: "green" if c.information_ratio > 0.5 else ("red" if c.information_ratio < 0 else None)),
        ("Up Capture (%)",
         None,
         lambda c: f"{c.up_capture:.1f}%",
         lambda c: "green" if c.up_capture > 100 else None),
        ("Down Capture (%)",
         None,
         lambda c: f"{c.down_capture:.1f}%",
         lambda c: "red" if c.down_capture > 100 else ("green" if c.down_capture < 80 else None)),
        ("Relative Max Drawdown (%)",
         None,
         lambda c: f"{c.relative_max_drawdown_pct:.2f}%",
         lambda c: "red" if c.relative_max_drawdown_pct > 30 else None),
    ]

    # ── Header row ───────────────────────────────────────────────────────────
    hdr = ws.cell(row=HEADER_ROW, column=METRIC_COL, value="Metric")
    hdr.font, hdr.fill, hdr.alignment = WHITE_FONT, DARK_BLUE, LEFT

    shdr = ws.cell(row=HEADER_ROW, column=STRATEGY_COL, value="Strategy")
    shdr.font, shdr.fill, shdr.alignment = WHITE_FONT, DARK_BLUE, CENTER

    for j, bname in enumerate(bench_names):
        c = ws.cell(row=HEADER_ROW, column=STRATEGY_COL + 1 + j, value=bname)
        c.font, c.fill, c.alignment = WHITE_FONT, MED_BLUE, CENTER

    # ── Data rows ────────────────────────────────────────────────────────────
    data_start = HEADER_ROW + 1
    for i, (label, strat_fn, bench_fn, hi_fn) in enumerate(METRICS):
        row = data_start + i
        row_fill = ALT_FILL if i % 2 == 0 else WHITE_FILL

        lbl_cell = ws.cell(row=row, column=METRIC_COL, value=label)
        lbl_cell.font, lbl_cell.fill, lbl_cell.alignment = BOLD_FONT, row_fill, LEFT

        strat_val = strat_fn(first) if strat_fn else "—"
        sc = ws.cell(row=row, column=STRATEGY_COL, value=strat_val)
        sc.fill, sc.alignment = row_fill, CENTER

        for j, (bname, comp) in enumerate(valid_comps.items()):
            col = STRATEGY_COL + 1 + j
            val = bench_fn(comp)
            bc = ws.cell(row=row, column=col, value=val)
            bc.alignment = CENTER
            if hi_fn:
                colour = hi_fn(comp)
                bc.fill = (GREEN_FILL if colour == "green" else
                           RED_FILL if colour == "red" else row_fill)
            else:
                bc.fill = row_fill

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions[get_column_letter(STRATEGY_COL)].width = 16
    for j in range(n_bench):
        ws.column_dimensions[get_column_letter(STRATEGY_COL + 1 + j)].width = 18
    ws.row_dimensions[HEADER_ROW].height = 28

    # ── Return-over-time chart ───────────────────────────────────────────────
    try:
        from .excel_charts import write_series_block, add_line_chart

        # Build a date->equity map for the strategy (optional)
        strat_map: Dict = {}
        if equity_curve is not None and not equity_curve.empty:
            ec_tmp = equity_curve[["date", "equity"]].copy()
            ec_tmp["date"] = pd.to_datetime(ec_tmp["date"], errors="coerce")
            ec_tmp = ec_tmp.dropna(subset=["date"])
            strat_map = dict(zip(ec_tmp["date"], ec_tmp["equity"].astype(float)))

        # Date->equity maps per benchmark (already rebased to strategy start equity)
        bench_maps: Dict[str, Dict] = {}
        for bname, comp in valid_comps.items():
            bdf = comp.benchmark_equity.copy()
            bdf["date"] = pd.to_datetime(bdf["date"], errors="coerce")
            bench_maps[bname] = dict(zip(bdf["date"], bdf["equity"].astype(float)))

        # Union of all dates, sorted
        all_dates: List[pd.Timestamp] = sorted(set().union(
            strat_map.keys(),
            *(m.keys() for m in bench_maps.values())
        ))

        include_strategy = bool(strat_map)
        headers = ["Date"] + (["Strategy"] if include_strategy else []) + bench_names

        rows_data = []
        for d in all_dates:
            row_vals: List[Any] = [d.strftime("%Y-%m-%d")]
            if include_strategy:
                row_vals.append(strat_map.get(d))
            for bname in bench_names:
                row_vals.append(bench_maps[bname].get(d))
            rows_data.append(row_vals)

        # Down-sample for chart rendering
        max_pts = 500
        if len(rows_data) > max_pts:
            step = max(1, len(rows_data) // max_pts)
            idxs = list(range(0, len(rows_data), step))
            if idxs[-1] != len(rows_data) - 1:
                idxs.append(len(rows_data) - 1)
            rows_data = [rows_data[i] for i in idxs]

        block = write_series_block(ws.parent, headers, rows_data)
        if block:
            chart_anchor_row = data_start + len(METRICS) + 2
            add_line_chart(
                ws,
                f"A{chart_anchor_row}",
                block,
                title="Return Over Time: Strategy vs Benchmarks",
                x_title="Date",
                y_title="Portfolio Value ($)",
                width=26,
                height=14,
                smooth=True,
            )
    except Exception as exc:
        logger.warning("Benchmark return-over-time chart failed: %s", exc)


class BenchmarkLoader:
    """Resolves and loads stored benchmark series, and runs comparisons."""

    def __init__(
        self,
        benchmarks_dir: Optional[Path] = None,
        registry_path: Optional[Path] = None,
    ):
        self.benchmarks_dir = Path(benchmarks_dir) if benchmarks_dir else DEFAULT_BENCHMARK_DIR
        self.registry = load_benchmark_registry(registry_path or DEFAULT_REGISTRY_PATH)

    def default_name(self) -> Optional[str]:
        return self.registry.get("default")

    def _file_candidates(self, symbol: str, interval: str) -> List[Path]:
        return [
            self.benchmarks_dir / f"{symbol}_{interval}.csv",
            self.benchmarks_dir / f"{symbol}.csv",
            self.benchmarks_dir / f"{symbol}_daily.csv",
        ]

    def load_series(self, name_or_symbol: str) -> pd.DataFrame:
        """Load a benchmark's [date, close] series. Empty if unavailable."""
        resolved = resolve_benchmark(name_or_symbol, self.registry)
        if resolved:
            _, entry = resolved
            symbol = entry.get("symbol", name_or_symbol)
            interval = entry.get("interval", "daily")
        else:
            symbol, interval = str(name_or_symbol), "daily"

        for path in self._file_candidates(symbol, interval):
            if path.exists():
                try:
                    df = pd.read_csv(path)
                    df.columns = [str(c).lower().strip() for c in df.columns]
                    if "date" in df.columns and "close" in df.columns:
                        return df[["date", "close"]]
                except Exception as exc:  # pragma: no cover - defensive
                    logger.warning("Failed reading benchmark %s: %s", path, exc)
        return pd.DataFrame()

    def resolved_label(self, name_or_symbol: str) -> Tuple[str, str]:
        """Return (friendly_name, symbol) for display."""
        resolved = resolve_benchmark(name_or_symbol, self.registry)
        if resolved:
            name, entry = resolved
            return name, entry.get("symbol", "")
        return str(name_or_symbol), str(name_or_symbol)

    def compare(
        self,
        equity_curve: pd.DataFrame,
        benchmark_name: Optional[str] = None,
        risk_free_rate: Optional[float] = None,
    ) -> BenchmarkComparison:
        """Load the configured benchmark and compare it to an equity curve."""
        name = benchmark_name or self.default_name()
        if not name:
            return _invalid("no benchmark configured")

        friendly, symbol = self.resolved_label(name)
        series = self.load_series(name)
        if series.empty:
            return _invalid(
                f"no stored data for '{friendly}' (collect it first)",
                friendly, symbol,
            )
        return compute_benchmark_comparison(
            equity_curve, series,
            benchmark_name=friendly, benchmark_symbol=symbol,
            risk_free_rate=risk_free_rate,
        )

    def compare_all(
        self,
        equity_curve: pd.DataFrame,
        risk_free_rate: Optional[float] = None,
    ) -> Dict[str, BenchmarkComparison]:
        """Compare equity_curve against every benchmark in the registry.

        Returns a dict of {benchmark_name: BenchmarkComparison}, preserving
        registry order. Invalid comparisons (missing data) are included so
        callers can surface informative messages.
        """
        results: Dict[str, BenchmarkComparison] = {}
        for name in self.registry.get("benchmarks", {}):
            results[name] = self.compare(equity_curve, benchmark_name=name,
                                         risk_free_rate=risk_free_rate)
        return results
