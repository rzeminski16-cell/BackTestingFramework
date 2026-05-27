"""
Shared helpers for building NATIVE, interactive Excel charts with openpyxl.

Reports historically embedded matplotlib figures as static PNG images. This
module centralizes the "write the plotted series to worksheet cells, then add a
native openpyxl chart that references those cells" pattern, plus other native
Excel features (conditional-formatting heatmaps, Excel Tables, freeze panes).

Design notes / reliability:
- Chart source data is written to a single hidden ``_ChartData`` sheet per
  workbook (charts that reference a hidden sheet still render in Excel). This
  keeps the visible sheets clean.
- Callers write *already-computed* Python values into the cells. We never use
  Excel formulas for metrics, so what Excel displays equals exactly what the
  framework computed.
- Every public helper is wrapped so a charting failure logs and returns None
  rather than aborting report generation.
"""

from __future__ import annotations

import functools
import logging
import re
from typing import Any, Dict, List, Optional, Sequence

from openpyxl.chart import (
    AreaChart,
    BarChart,
    LineChart,
    PieChart,
    Reference,
    ScatterChart,
    Series,
)
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo

logger = logging.getLogger(__name__)

CHART_DATA_SHEET = "_ChartData"


def safe_chart(fn):
    """Decorator: never let a charting helper abort report generation."""
    @functools.wraps(fn)
    def wrapper(*args, **kwargs):
        try:
            return fn(*args, **kwargs)
        except Exception as exc:  # pragma: no cover - defensive
            logger.warning("excel_charts.%s failed: %s", fn.__name__, exc)
            return None
    return wrapper


def get_chart_data_sheet(wb):
    """Return (creating if needed) the hidden per-workbook chart-data sheet."""
    if CHART_DATA_SHEET in wb.sheetnames:
        return wb[CHART_DATA_SHEET]
    ws = wb.create_sheet(CHART_DATA_SHEET)
    ws.sheet_state = "hidden"
    return ws


def write_series_block(wb, headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> Optional[Dict[str, Any]]:
    """
    Write a block of chart source data to the hidden _ChartData sheet.

    Args:
        wb: the openpyxl Workbook.
        headers: column headers; headers[0] is the category/x label, the rest
            are series names (used as chart legend titles).
        rows: each row is (category, series1, series2, ...).

    Returns a dict with Reference objects:
        - 'cats': categories (x), excluding the header row
        - 'data': series columns including the header row (titles_from_data)
        - plus 'ws', 'start_col', 'n_cols', 'max_row' for advanced callers.
    Returns None if there is no data.
    """
    if not headers or not rows:
        return None

    ws = get_chart_data_sheet(wb)
    start_col = int(getattr(wb, "_chartdata_next_col", 1))
    n_cols = len(headers)

    for j, header in enumerate(headers):
        ws.cell(row=1, column=start_col + j, value=header)
    for i, row in enumerate(rows, start=2):
        for j, value in enumerate(row):
            ws.cell(row=i, column=start_col + j, value=value)

    max_row = 1 + len(rows)
    cats = Reference(ws, min_col=start_col, min_row=2, max_row=max_row)
    data = Reference(
        ws,
        min_col=start_col + 1,
        max_col=start_col + n_cols - 1,
        min_row=1,
        max_row=max_row,
    )

    # Advance the cursor, leaving a one-column gap between blocks.
    wb._chartdata_next_col = start_col + n_cols + 1

    return {
        "ws": ws,
        "cats": cats,
        "data": data,
        "start_col": start_col,
        "n_cols": n_cols,
        "max_row": max_row,
    }


def _finalize(ws, chart, anchor, title, x_title, y_title, width, height, style):
    if title:
        chart.title = title
    if x_title:
        chart.x_axis.title = x_title
    if y_title:
        chart.y_axis.title = y_title
    # openpyxl sometimes hides axes unless explicitly kept.
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    if width:
        chart.width = width
    if height:
        chart.height = height
    if style:
        chart.style = style
    ws.add_chart(chart, anchor)
    return chart


@safe_chart
def add_line_chart(ws, anchor, block, title=None, x_title=None, y_title=None,
                   width=18, height=10, style=12, smooth=False):
    chart = LineChart()
    chart.add_data(block["data"], titles_from_data=True)
    chart.set_categories(block["cats"])
    for series in chart.series:
        series.smooth = smooth
    return _finalize(ws, chart, anchor, title, x_title, y_title, width, height, style)


@safe_chart
def add_bar_chart(ws, anchor, block, title=None, x_title=None, y_title=None,
                  width=18, height=10, style=10, bar_type="col"):
    chart = BarChart()
    chart.type = bar_type
    chart.add_data(block["data"], titles_from_data=True)
    chart.set_categories(block["cats"])
    return _finalize(ws, chart, anchor, title, x_title, y_title, width, height, style)


@safe_chart
def add_stacked_area_chart(ws, anchor, block, title=None, x_title=None, y_title=None,
                           width=18, height=10, style=12):
    chart = AreaChart()
    chart.grouping = "stacked"
    chart.add_data(block["data"], titles_from_data=True)
    chart.set_categories(block["cats"])
    return _finalize(ws, chart, anchor, title, x_title, y_title, width, height, style)


@safe_chart
def add_pie_chart(ws, anchor, block, title=None, width=14, height=10, style=10):
    chart = PieChart()
    chart.add_data(block["data"], titles_from_data=True)
    chart.set_categories(block["cats"])
    if title:
        chart.title = title
    if width:
        chart.width = width
    if height:
        chart.height = height
    if style:
        chart.style = style
    ws.add_chart(chart, anchor)
    return chart


@safe_chart
def add_scatter_chart(ws, anchor, block, title=None, x_title=None, y_title=None,
                      width=18, height=10, style=13):
    """Scatter chart: block's first column is X, remaining columns are Y series."""
    chart = ScatterChart()
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    data_ws = block["ws"]
    start_col = block["start_col"]
    max_row = block["max_row"]
    xref = Reference(data_ws, min_col=start_col, min_row=2, max_row=max_row)

    for col in range(start_col + 1, start_col + block["n_cols"]):
        yref = Reference(data_ws, min_col=col, min_row=1, max_row=max_row)
        series = Series(yref, xref, title_from_data=True)
        series.marker.symbol = "circle"
        series.marker.size = 5
        series.graphicalProperties.line.noFill = True  # markers only
        chart.series.append(series)

    if title:
        chart.title = title
    if width:
        chart.width = width
    if height:
        chart.height = height
    if style:
        chart.style = style
    ws.add_chart(chart, anchor)
    return chart


@safe_chart
def apply_color_scale(ws, cell_range, low="F8696B", mid="FFEB84", high="63BE7B"):
    """Apply a red->yellow->green 3-color scale (e.g. for returns/correlation heatmaps)."""
    rule = ColorScaleRule(
        start_type="min", start_color=low,
        mid_type="percentile", mid_value=50, mid_color=mid,
        end_type="max", end_color=high,
    )
    ws.conditional_formatting.add(cell_range, rule)
    return rule


@safe_chart
def make_table(ws, cell_range, name, style="TableStyleMedium2"):
    """Turn a header+data range into an Excel Table (sortable/filterable)."""
    safe = re.sub(r"[^A-Za-z0-9_]", "_", str(name))
    if not safe or not (safe[0].isalpha() or safe[0] == "_"):
        safe = "t_" + safe

    # Excel table displayNames must be unique workbook-wide.
    existing = set()
    wb = ws.parent
    for sheet in getattr(wb, "worksheets", []):
        existing.update(getattr(sheet, "tables", {}).keys())
    base, k = safe, 1
    while safe in existing:
        safe = f"{base}_{k}"
        k += 1

    table = Table(displayName=safe, ref=cell_range)
    table.tableStyleInfo = TableStyleInfo(
        name=style, showRowStripes=True, showColumnStripes=False,
        showFirstColumn=False, showLastColumn=False,
    )
    ws.add_table(table)
    return table


def freeze(ws, cell="A2"):
    """Freeze panes above/left of ``cell`` (e.g. 'A2' freezes the header row)."""
    ws.freeze_panes = cell


def sample_series(values: List[Any], dates: Optional[List[Any]] = None, max_points: int = 500):
    """
    Down-sample a long series for DISPLAY only (charts). Returns (dates, values).

    Never use the result for metric computation - metrics must use the full
    series. Keeps first and last points; evenly strides the middle.
    """
    n = len(values)
    if n <= max_points:
        return dates, values
    step = max(1, n // max_points)
    idx = list(range(0, n, step))
    if idx[-1] != n - 1:
        idx.append(n - 1)
    sampled_vals = [values[i] for i in idx]
    sampled_dates = [dates[i] for i in idx] if dates is not None else None
    return sampled_dates, sampled_vals
