"""
Enhanced Portfolio Report Generator for comprehensive portfolio analysis.

Generates advanced Excel reports with:
- Executive Summary Dashboard with KPIs and traffic-light indicators
- Table of Contents with hyperlinks
- Dashboard sheets with consolidated visualizations
- Deep trade analysis (MAE/MFE, clustering, streaks)
- Rolling performance metrics
- Statistical significance testing
- Portfolio contribution analysis
- Capital utilization tracking
- Comprehensive visualizations using matplotlib
"""
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from io import BytesIO
import warnings

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import LineChart, BarChart, Reference, PieChart, AreaChart, ScatterChart
    from openpyxl.chart.label import DataLabelList
    from openpyxl.drawing.image import Image
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.hyperlink import Hyperlink
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from ..Analysis.performance_metrics import PerformanceMetrics, DEFAULT_RISK_FREE_RATE
from ..Models.trade import Trade

# Try to import enhanced visualizations
try:
    from ..Analysis.enhanced_visualizations import EnhancedVisualizations, MATPLOTLIB_AVAILABLE
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    EnhancedVisualizations = None

# Try to import scipy for statistical tests
try:
    from scipy import stats
    SCIPY_AVAILABLE = True
except ImportError:
    SCIPY_AVAILABLE = False


class EnhancedPortfolioReportGenerator:
    """
    Generates comprehensive Excel reports for portfolio backtests with advanced analytics.
    """

    # Style definitions
    COLORS = {
        'header_dark': '1F4E79',
        'header_medium': '2E75B6',
        'header_light': '5B9BD5',
        'positive': 'C6EFCE',
        'positive_dark': '28A745',
        'negative': 'FFC7CE',
        'negative_dark': 'DC3545',
        'neutral': 'FFEB9C',
        'neutral_dark': 'FFC107',
        'white': 'FFFFFF',
        'light_gray': 'F5F5F5',
        'medium_gray': 'E0E0E0',
        'dark_gray': '666666',
    }

    def __init__(self, output_dir: Path, include_matplotlib_charts: bool = True):
        """
        Initialize enhanced report generator.

        Args:
            output_dir: Directory to save reports
            include_matplotlib_charts: Whether to include matplotlib visualizations
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for report generation")

        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.include_matplotlib_charts = include_matplotlib_charts and MATPLOTLIB_AVAILABLE

        # Initialize visualizations module if available
        if self.include_matplotlib_charts:
            try:
                self.viz = EnhancedVisualizations(dpi=150)
            except Exception:
                self.viz = None
                self.include_matplotlib_charts = False
        else:
            self.viz = None

        self._init_styles()

    def _init_styles(self):
        """Initialize reusable styles."""
        self.header_fill = PatternFill(start_color=self.COLORS['header_dark'],
                                        end_color=self.COLORS['header_dark'], fill_type="solid")
        self.header_font = Font(color=self.COLORS['white'], bold=True, size=11)
        self.subheader_fill = PatternFill(start_color=self.COLORS['header_medium'],
                                           end_color=self.COLORS['header_medium'], fill_type="solid")
        self.positive_fill = PatternFill(start_color=self.COLORS['positive'],
                                          end_color=self.COLORS['positive'], fill_type="solid")
        self.negative_fill = PatternFill(start_color=self.COLORS['negative'],
                                          end_color=self.COLORS['negative'], fill_type="solid")
        self.neutral_fill = PatternFill(start_color=self.COLORS['neutral'],
                                         end_color=self.COLORS['neutral'], fill_type="solid")
        self.light_fill = PatternFill(start_color=self.COLORS['light_gray'],
                                       end_color=self.COLORS['light_gray'], fill_type="solid")

        self.title_font = Font(bold=True, size=16, color=self.COLORS['header_dark'])
        self.section_font = Font(bold=True, size=14, color=self.COLORS['header_dark'])
        self.subsection_font = Font(bold=True, size=12, color=self.COLORS['header_medium'])

        self.thin_border = Border(
            left=Side(style='thin', color=self.COLORS['medium_gray']),
            right=Side(style='thin', color=self.COLORS['medium_gray']),
            top=Side(style='thin', color=self.COLORS['medium_gray']),
            bottom=Side(style='thin', color=self.COLORS['medium_gray'])
        )

    def generate_portfolio_report(self, result, report_name: Optional[str] = None) -> Path:
        """
        Generate comprehensive enhanced portfolio report.

        Args:
            result: PortfolioBacktestResult object
            report_name: Optional custom report name

        Returns:
            Path to generated report
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = report_name or f"enhanced_portfolio_report_{timestamp}.xlsx"
        filepath = self.output_dir / filename

        # Calculate comprehensive metrics
        metrics = self._calculate_comprehensive_metrics(result)

        # Create workbook
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # Create sheets in order
        self._create_table_of_contents(wb, result)
        self._create_executive_summary(wb, result, metrics)
        self._create_dashboard_visualizations(wb, result, metrics)
        self._create_performance_metrics(wb, result, metrics)
        self._create_per_security_analysis(wb, result, metrics)
        self._create_trade_analysis(wb, result, metrics)
        self._create_trade_log(wb, result)
        self._create_rolling_metrics(wb, result, metrics)
        self._create_statistical_analysis(wb, result, metrics)
        self._create_capital_allocation(wb, result, metrics)

        if result.signal_rejections:
            self._create_signal_rejections(wb, result)

        if result.vulnerability_swaps or result.vulnerability_history:
            self._create_vulnerability_analysis(wb, result)

        self._create_costs_analysis(wb, result, metrics)
        self._create_monthly_heatmap(wb, result, metrics)

        # Update Table of Contents with hyperlinks
        self._update_toc_hyperlinks(wb)

        wb.save(filepath)
        print(f"Enhanced portfolio report saved to {filepath}")
        return filepath

    def _calculate_comprehensive_metrics(self, result) -> Dict[str, Any]:
        """Calculate all metrics needed for the report."""
        metrics = {}

        # Collect all trades
        all_trades = []
        for sym_result in result.symbol_results.values():
            all_trades.extend(sym_result.trades)
        all_trades.sort(key=lambda t: t.entry_date)

        metrics['all_trades'] = all_trades
        metrics['num_trades'] = len(all_trades)

        # Basic metrics
        metrics['initial_capital'] = result.config.initial_capital
        metrics['final_equity'] = result.final_equity
        metrics['total_return'] = result.total_return
        metrics['total_return_pct'] = result.total_return_pct

        if len(all_trades) == 0:
            return self._get_empty_metrics(metrics)

        # Calculate from equity curve
        equity_df = result.portfolio_equity_curve

        if len(equity_df) > 0:
            metrics['start_date'] = pd.Timestamp(equity_df['date'].iloc[0])
            metrics['end_date'] = pd.Timestamp(equity_df['date'].iloc[-1])
            metrics['total_days'] = (metrics['end_date'] - metrics['start_date']).days
            metrics['years'] = metrics['total_days'] / 365.25
        else:
            metrics['start_date'] = None
            metrics['end_date'] = None
            metrics['total_days'] = 0
            metrics['years'] = 0

        # Win/Loss Statistics
        winning_trades = [t for t in all_trades if t.pl > 0]
        losing_trades = [t for t in all_trades if t.pl <= 0]

        metrics['num_wins'] = len(winning_trades)
        metrics['num_losses'] = len(losing_trades)
        metrics['win_rate'] = metrics['num_wins'] / metrics['num_trades'] if metrics['num_trades'] > 0 else 0

        # P/L Statistics
        metrics['avg_win'] = np.mean([t.pl for t in winning_trades]) if winning_trades else 0
        metrics['largest_win'] = max(t.pl for t in winning_trades) if winning_trades else 0
        metrics['avg_loss'] = np.mean([t.pl for t in losing_trades]) if losing_trades else 0
        metrics['largest_loss'] = min(t.pl for t in losing_trades) if losing_trades else 0

        # Risk metrics from equity curve
        metrics.update(self._calculate_equity_metrics(equity_df))

        # Trade quality metrics
        metrics.update(self._calculate_trade_metrics(all_trades))

        # Consistency metrics
        metrics.update(self._calculate_consistency_metrics(equity_df, all_trades))

        # Per-security contribution
        metrics['security_contribution'] = self._calculate_security_contribution(result)

        # Streak analysis
        metrics.update(self._calculate_streak_metrics(all_trades))

        # Statistical significance
        metrics.update(self._calculate_statistical_metrics(all_trades, equity_df))

        # Cost analysis
        metrics['costs_analysis'] = self._calculate_costs(all_trades, metrics['total_return'])

        # Rolling metrics
        metrics['rolling_metrics'] = self._calculate_rolling_metrics(equity_df)

        # Monthly/Yearly returns
        metrics['monthly_returns'] = self._calculate_period_returns(equity_df, 'ME')
        metrics['yearly_returns'] = self._calculate_period_returns(equity_df, 'Y')

        # Detect rolling metric anomalies for reporting and filtered Sharpe calculation
        anomalies, filtered_sharpe = PerformanceMetrics.detect_rolling_anomalies(
            equity_df,
            window=90,
            absolute_threshold=10.0,
            zscore_threshold=3.0
        )
        metrics['rolling_anomalies'] = anomalies
        metrics['filtered_sharpe_ratio'] = filtered_sharpe

        return metrics

    def _get_empty_metrics(self, base_metrics: Dict[str, Any]) -> Dict[str, Any]:
        """Return metrics with zero values for no-trade scenario."""
        base_metrics.update({
            'num_wins': 0, 'num_losses': 0, 'win_rate': 0,
            'avg_win': 0, 'avg_loss': 0, 'largest_win': 0, 'largest_loss': 0,
            'max_drawdown': 0, 'max_drawdown_pct': 0, 'avg_drawdown': 0,
            'sharpe_ratio': 0, 'sortino_ratio': 0, 'calmar_ratio': 0,
            'profit_factor': 0, 'recovery_factor': 0, 'cagr': 0,
            'volatility': 0, 'downside_deviation': 0, 'best_day': 0, 'worst_day': 0,
            'avg_trade_duration': 0, 'median_trade_duration': 0, 'risk_reward_ratio': 0,
            'max_win_streak': 0, 'max_loss_streak': 0,
            'profitable_months_pct': 0, 'monthly_consistency': 0,
            'security_contribution': {}, 'all_trades': [],
            't_statistic': 0, 'p_value': 1, 'is_significant': False,
            'costs_analysis': {}, 'rolling_metrics': pd.DataFrame(),
            'monthly_returns': pd.DataFrame(), 'yearly_returns': pd.DataFrame(),
            'rolling_anomalies': [], 'filtered_sharpe_ratio': 0,
        })
        return base_metrics

    def _calculate_equity_metrics(self, equity_df: pd.DataFrame) -> Dict[str, Any]:
        """Calculate metrics from equity curve."""
        metrics = {}

        if len(equity_df) < 2:
            return {
                'max_drawdown': 0, 'max_drawdown_pct': 0, 'avg_drawdown': 0,
                'sharpe_ratio': 0, 'sortino_ratio': 0, 'calmar_ratio': 0,
                'cagr': 0, 'volatility': 0, 'downside_deviation': 0,
                'best_day': 0, 'worst_day': 0, 'recovery_factor': 0,
            }

        equity = equity_df['equity'].values

        # Filter out NaN and invalid values
        if np.any(np.isnan(equity)) or np.any(np.isinf(equity)):
            equity = np.nan_to_num(equity, nan=0.0, posinf=0.0, neginf=0.0)

        # Ensure we have valid data
        if len(equity) == 0 or np.all(equity <= 0):
            return {
                'max_drawdown': 0, 'max_drawdown_pct': 0, 'avg_drawdown': 0,
                'sharpe_ratio': 0, 'sortino_ratio': 0, 'calmar_ratio': 0,
                'cagr': 0, 'volatility': 0, 'downside_deviation': 0,
                'best_day': 0, 'worst_day': 0, 'recovery_factor': 0,
            }

        # Drawdown calculations
        running_max = np.maximum.accumulate(equity)

        # Prevent division by zero - use safe division
        with np.errstate(divide='ignore', invalid='ignore'):
            drawdown = running_max - equity
            drawdown_pct = np.where(running_max > 0, (drawdown / running_max) * 100, 0.0)

        # Remove any NaN or inf values and cap at 100%
        drawdown = np.nan_to_num(drawdown, nan=0.0, posinf=0.0, neginf=0.0)
        drawdown_pct = np.nan_to_num(drawdown_pct, nan=0.0, posinf=0.0, neginf=0.0)
        drawdown_pct = np.clip(drawdown_pct, 0, 100)

        metrics['max_drawdown'] = np.max(drawdown)
        metrics['max_drawdown_pct'] = np.max(drawdown_pct)
        metrics['avg_drawdown'] = np.mean(drawdown[drawdown > 0]) if np.any(drawdown > 0) else 0

        # Returns
        returns = pd.Series(equity).pct_change().dropna()

        if len(returns) == 0 or returns.std() == 0:
            metrics['sharpe_ratio'] = 0
            metrics['sortino_ratio'] = 0
            metrics['volatility'] = 0
            metrics['downside_deviation'] = 0
        else:
            daily_rf = (1 + DEFAULT_RISK_FREE_RATE) ** (1/252) - 1
            excess_returns = returns - daily_rf

            # Sharpe
            metrics['sharpe_ratio'] = (excess_returns.mean() / excess_returns.std()) * np.sqrt(252)

            # Sortino
            downside_returns = excess_returns[excess_returns < 0]
            if len(downside_returns) > 0 and downside_returns.std() > 0:
                metrics['sortino_ratio'] = (excess_returns.mean() / downside_returns.std()) * np.sqrt(252)
            else:
                metrics['sortino_ratio'] = 99.99 if excess_returns.mean() > 0 else 0

            metrics['volatility'] = returns.std() * np.sqrt(252) * 100
            negative_returns = returns[returns < 0]
            metrics['downside_deviation'] = negative_returns.std() * np.sqrt(252) * 100 if len(negative_returns) > 0 else 0

        # Best/Worst day
        metrics['best_day'] = returns.max() * 100 if len(returns) > 0 else 0
        metrics['worst_day'] = returns.min() * 100 if len(returns) > 0 else 0

        # CAGR
        if len(equity_df) >= 2:
            start_date = pd.Timestamp(equity_df['date'].iloc[0])
            end_date = pd.Timestamp(equity_df['date'].iloc[-1])
            years = (end_date - start_date).days / 365.25
            initial = equity[0]
            final = equity[-1]

            if years > 0 and initial > 0:
                metrics['cagr'] = (pow(final / initial, 1 / years) - 1) * 100
            else:
                metrics['cagr'] = 0
        else:
            metrics['cagr'] = 0

        # Calmar and Recovery Factor
        metrics['calmar_ratio'] = metrics['cagr'] / metrics['max_drawdown_pct'] if metrics['max_drawdown_pct'] > 0 else 0
        metrics['recovery_factor'] = (equity[-1] - equity[0]) / metrics['max_drawdown'] if metrics['max_drawdown'] > 0 else 0

        return metrics

    def _calculate_trade_metrics(self, trades: List[Trade]) -> Dict[str, Any]:
        """Calculate trade-based metrics."""
        metrics = {}

        if not trades:
            return {'profit_factor': 0, 'avg_trade_duration': 0, 'median_trade_duration': 0,
                    'risk_reward_ratio': 0, 'expectancy': 0, 'expectancy_ratio': 0}

        winners = [t for t in trades if t.pl > 0]
        losers = [t for t in trades if t.pl <= 0]

        # Profit Factor
        gross_profit = sum(t.pl for t in winners) if winners else 0
        gross_loss = abs(sum(t.pl for t in losers)) if losers else 0

        if gross_loss > 0:
            metrics['profit_factor'] = gross_profit / gross_loss
        elif gross_profit > 0:
            metrics['profit_factor'] = 999.99
        else:
            metrics['profit_factor'] = 0

        # Duration
        durations = [t.duration_days for t in trades]
        metrics['avg_trade_duration'] = np.mean(durations)
        metrics['median_trade_duration'] = np.median(durations)

        # Risk/Reward
        avg_win = np.mean([t.pl for t in winners]) if winners else 0
        avg_loss = abs(np.mean([t.pl for t in losers])) if losers else 0
        metrics['risk_reward_ratio'] = avg_win / avg_loss if avg_loss > 0 else 0

        # Expectancy
        win_rate = len(winners) / len(trades) if trades else 0
        metrics['expectancy'] = (win_rate * avg_win) - ((1 - win_rate) * avg_loss)
        metrics['expectancy_ratio'] = metrics['expectancy'] / avg_loss if avg_loss > 0 else 0

        return metrics

    def _calculate_consistency_metrics(self, equity_df: pd.DataFrame, trades: List[Trade]) -> Dict[str, Any]:
        """Calculate consistency metrics."""
        metrics = {}

        if len(equity_df) < 2:
            return {'profitable_months_pct': 0, 'profitable_weeks_pct': 0,
                    'monthly_consistency': 0, 'consecutive_losing_months': 0}

        df = equity_df.copy()
        df['date'] = pd.to_datetime(df['date'])
        df = df.set_index('date')

        # Monthly
        monthly = df['equity'].resample('ME').last()
        monthly_returns = monthly.pct_change().dropna()

        if len(monthly_returns) > 0:
            profitable_months = (monthly_returns > 0).sum()
            metrics['profitable_months_pct'] = (profitable_months / len(monthly_returns)) * 100
            metrics['monthly_consistency'] = monthly_returns.std() * 100

            # Consecutive losing months
            max_consecutive = 0
            current = 0
            for ret in monthly_returns:
                if ret < 0:
                    current += 1
                    max_consecutive = max(max_consecutive, current)
                else:
                    current = 0
            metrics['consecutive_losing_months'] = max_consecutive
        else:
            metrics['profitable_months_pct'] = 0
            metrics['monthly_consistency'] = 0
            metrics['consecutive_losing_months'] = 0

        # Weekly
        weekly = df['equity'].resample('W').last()
        weekly_returns = weekly.pct_change().dropna()
        if len(weekly_returns) > 0:
            metrics['profitable_weeks_pct'] = ((weekly_returns > 0).sum() / len(weekly_returns)) * 100
        else:
            metrics['profitable_weeks_pct'] = 0

        return metrics

    def _calculate_security_contribution(self, result) -> Dict[str, Dict[str, float]]:
        """Calculate P/L contribution by security."""
        contributions = {}
        total_pl = result.total_return if result.total_return != 0 else 1

        for symbol, sym_result in result.symbol_results.items():
            trades = sym_result.trades
            pl = sym_result.total_return
            contributions[symbol] = {
                'pl': pl,
                'pl_pct': sym_result.total_return_pct,
                'contribution_pct': (pl / total_pl * 100) if total_pl != 0 else 0,
                'num_trades': len(trades),
                'win_rate': len([t for t in trades if t.pl > 0]) / len(trades) * 100 if trades else 0
            }

        return contributions

    def _calculate_streak_metrics(self, trades: List[Trade]) -> Dict[str, Any]:
        """Calculate win/loss streak metrics."""
        if not trades:
            return {'max_win_streak': 0, 'max_loss_streak': 0,
                    'current_streak': 0, 'current_streak_type': None,
                    'avg_win_streak': 0, 'avg_loss_streak': 0}

        sorted_trades = sorted(trades, key=lambda t: t.exit_date)

        max_win_streak = 0
        max_loss_streak = 0
        current_streak = 0
        current_type = None
        win_streaks = []
        loss_streaks = []

        for trade in sorted_trades:
            is_win = trade.pl > 0
            if current_type is None:
                current_type = is_win
                current_streak = 1
            elif is_win == current_type:
                current_streak += 1
            else:
                if current_type:
                    win_streaks.append(current_streak)
                    max_win_streak = max(max_win_streak, current_streak)
                else:
                    loss_streaks.append(current_streak)
                    max_loss_streak = max(max_loss_streak, current_streak)
                current_type = is_win
                current_streak = 1

        # Final streak
        if current_type:
            max_win_streak = max(max_win_streak, current_streak)
            win_streaks.append(current_streak)
        elif current_type is not None:
            max_loss_streak = max(max_loss_streak, current_streak)
            loss_streaks.append(current_streak)

        return {
            'max_win_streak': max_win_streak,
            'max_loss_streak': max_loss_streak,
            'current_streak': current_streak,
            'current_streak_type': 'Win' if current_type else 'Loss' if current_type is not None else 'N/A',
            'avg_win_streak': np.mean(win_streaks) if win_streaks else 0,
            'avg_loss_streak': np.mean(loss_streaks) if loss_streaks else 0,
        }

    def _calculate_statistical_metrics(self, trades: List[Trade], equity_df: pd.DataFrame) -> Dict[str, Any]:
        """Calculate statistical significance metrics."""
        metrics = {'t_statistic': 0, 'p_value': 1, 'is_significant': False,
                   'skewness': 0, 'kurtosis': 0, 'var_95': 0, 'cvar_95': 0}

        if len(trades) < 30 or not SCIPY_AVAILABLE:
            return metrics

        try:
            returns = [t.pl_pct for t in trades]

            # T-test
            t_stat, p_value = stats.ttest_1samp(returns, 0)
            metrics['t_statistic'] = t_stat
            metrics['p_value'] = p_value
            metrics['is_significant'] = p_value < 0.05

            # Distribution statistics
            metrics['skewness'] = stats.skew(returns)
            metrics['kurtosis'] = stats.kurtosis(returns)

            # Value at Risk
            metrics['var_95'] = np.percentile(returns, 5)

            # Conditional VaR
            var_threshold = metrics['var_95']
            tail_returns = [r for r in returns if r <= var_threshold]
            metrics['cvar_95'] = np.mean(tail_returns) if tail_returns else var_threshold

        except Exception:
            pass

        return metrics

    def _calculate_costs(self, trades: List[Trade], total_return: float) -> Dict[str, float]:
        """Calculate trading costs analysis."""
        if not trades:
            return {'total_commission': 0, 'avg_commission': 0, 'commission_pct_of_pl': 0, 'pl_before_costs': 0}

        total_commission = sum(t.commission_paid for t in trades)
        avg_commission = total_commission / len(trades)

        pl_before_costs = total_return + total_commission
        commission_pct = (total_commission / pl_before_costs * 100) if pl_before_costs != 0 else 0

        return {
            'total_commission': total_commission,
            'avg_commission': avg_commission,
            'commission_pct_of_pl': commission_pct,
            'pl_before_costs': pl_before_costs,
        }

    def _calculate_rolling_metrics(self, equity_df: pd.DataFrame, window: int = 90) -> pd.DataFrame:
        """Calculate rolling performance metrics."""
        if len(equity_df) < window + 10:
            return pd.DataFrame()

        df = equity_df.copy()
        df['date'] = pd.to_datetime(df['date'])
        df = df.set_index('date')
        df['returns'] = df['equity'].pct_change()

        daily_rf = (1 + DEFAULT_RISK_FREE_RATE) ** (1/252) - 1

        rolling_mean = df['returns'].rolling(window).mean()
        rolling_std = df['returns'].rolling(window).std()

        df['rolling_sharpe'] = ((rolling_mean - daily_rf) / rolling_std) * np.sqrt(252)
        df['rolling_volatility'] = rolling_std * np.sqrt(252) * 100

        return df[['returns', 'rolling_sharpe', 'rolling_volatility']].dropna()

    def _calculate_period_returns(self, equity_df: pd.DataFrame, period: str) -> pd.DataFrame:
        """Calculate returns for specified period."""
        if len(equity_df) < 2:
            return pd.DataFrame()

        df = equity_df.copy()
        df['date'] = pd.to_datetime(df['date'])
        df = df.set_index('date')

        resampled = df['equity'].resample(period).last()
        returns = resampled.pct_change().dropna() * 100

        result = pd.DataFrame({
            'period': returns.index,
            'return_pct': returns.values
        })

        return result

    # ==================== SHEET CREATION METHODS ====================

    def _create_table_of_contents(self, wb: Workbook, result):
        """Create Table of Contents sheet."""
        ws = wb.create_sheet("Table of Contents", 0)

        ws['A1'] = "PORTFOLIO BACKTEST REPORT"
        ws['A1'].font = Font(bold=True, size=20, color=self.COLORS['header_dark'])
        ws.merge_cells('A1:E1')

        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=10, color=self.COLORS['dark_gray'])

        ws['A3'] = f"Strategy: {result.strategy_name}"
        ws['A3'].font = Font(size=12, bold=True)

        ws['A4'] = f"Securities: {', '.join(result.symbol_results.keys())}"
        ws['A4'].font = Font(size=10, color=self.COLORS['dark_gray'])

        row = 6
        ws[f'A{row}'] = "TABLE OF CONTENTS"
        ws[f'A{row}'].font = self.section_font
        row += 2

        toc_items = [
            ("Executive Summary", "Key performance indicators and summary metrics"),
            ("Dashboard", "Visual overview of portfolio performance"),
            ("Performance Metrics", "Detailed performance analysis"),
            ("Per-Security", "Breakdown by individual security"),
            ("Trade Analysis", "Deep dive into trade characteristics"),
            ("Trade Log", "Complete list of all trades"),
            ("Rolling Metrics", "Time-based performance evolution"),
            ("Statistical Analysis", "Statistical significance and distribution"),
            ("Capital Allocation", "Capital utilization over time"),
            ("Costs Analysis", "Trading cost breakdown"),
            ("Monthly Heatmap", "Monthly returns visualization"),
        ]

        for i, (sheet_name, description) in enumerate(toc_items, 1):
            ws[f'A{row}'] = f"{i}."
            ws[f'B{row}'] = sheet_name
            ws[f'B{row}'].font = Font(bold=True, size=11, underline='single', color='0563C1')
            ws[f'C{row}'] = description
            ws[f'C{row}'].font = Font(size=10, color=self.COLORS['dark_gray'])
            row += 1

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 50

    def _update_toc_hyperlinks(self, wb: Workbook):
        """Add hyperlinks to Table of Contents."""
        if "Table of Contents" not in wb.sheetnames:
            return

        ws = wb["Table of Contents"]
        row = 8

        for sheet_name in wb.sheetnames:
            if sheet_name != "Table of Contents":
                cell = ws[f'B{row}']
                if cell.value:
                    try:
                        cell.hyperlink = f"#'{sheet_name}'!A1"
                    except Exception:
                        pass
                row += 1

    def _create_executive_summary(self, wb: Workbook, result, metrics: Dict[str, Any]):
        """Create Executive Summary sheet with KPIs and traffic lights."""
        ws = wb.create_sheet("Executive Summary")

        row = 1
        ws[f'A{row}'] = "EXECUTIVE SUMMARY"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:H{row}')
        row += 3

        # KPI Section
        ws[f'A{row}'] = "KEY PERFORMANCE INDICATORS"
        ws[f'A{row}'].font = self.section_font
        row += 2

        kpis = [
            ("Total Return", f"${metrics['total_return']:,.2f}", f"{metrics['total_return_pct']:.2f}%",
             metrics['total_return'] > 0),
            ("Sharpe Ratio", f"{metrics.get('sharpe_ratio', 0):.2f}", "",
             metrics.get('sharpe_ratio', 0) > 1),
            ("Max Drawdown", f"{metrics.get('max_drawdown_pct', 0):.2f}%", f"${metrics.get('max_drawdown', 0):,.0f}",
             metrics.get('max_drawdown_pct', 0) < 20),
            ("Win Rate", f"{metrics.get('win_rate', 0)*100:.1f}%", f"{metrics.get('num_wins', 0)}/{metrics.get('num_trades', 0)}",
             metrics.get('win_rate', 0) > 0.5),
            ("Profit Factor", f"{metrics.get('profit_factor', 0):.2f}", "",
             metrics.get('profit_factor', 0) > 1.5),
            ("CAGR", f"{metrics.get('cagr', 0):.2f}%", "",
             metrics.get('cagr', 0) > 10),
        ]

        col = 1
        for kpi_name, value, subvalue, is_good in kpis:
            ws.cell(row=row, column=col, value=kpi_name).font = Font(bold=True, size=10)

            value_cell = ws.cell(row=row+1, column=col, value=value)
            value_cell.font = Font(bold=True, size=14)

            indicator_cell = ws.cell(row=row+1, column=col+1)
            if is_good:
                indicator_cell.value = "●"
                indicator_cell.font = Font(color=self.COLORS['positive_dark'], size=16)
            else:
                indicator_cell.value = "●"
                indicator_cell.font = Font(color=self.COLORS['negative_dark'], size=16)

            if subvalue:
                ws.cell(row=row+2, column=col, value=subvalue).font = Font(size=9, color=self.COLORS['dark_gray'])

            col += 3
            if col > 7:
                col = 1
                row += 4

        row += 5

        # Performance Summary Table
        ws[f'A{row}'] = "PERFORMANCE SUMMARY"
        ws[f'A{row}'].font = self.section_font
        row += 2

        headers = ["Metric", "Value", "Assessment"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.thin_border
        row += 1

        performance_data = [
            ("Initial Capital", f"${metrics['initial_capital']:,.2f}", ""),
            ("Final Equity", f"${metrics['final_equity']:,.2f}",
             "+" if metrics['final_equity'] > metrics['initial_capital'] else "-"),
            ("Total Return", f"${metrics['total_return']:,.2f} ({metrics['total_return_pct']:.2f}%)",
             "Excellent" if metrics['total_return_pct'] > 20 else "Good" if metrics['total_return_pct'] > 5 else "Poor"),
            ("Number of Trades", str(metrics['num_trades']), ""),
            ("Win Rate", f"{metrics.get('win_rate', 0)*100:.1f}%",
             "Good" if metrics.get('win_rate', 0) > 0.5 else "Needs Improvement"),
            ("Sharpe Ratio", f"{metrics.get('sharpe_ratio', 0):.2f}",
             "Excellent" if metrics.get('sharpe_ratio', 0) > 2 else "Good" if metrics.get('sharpe_ratio', 0) > 1 else "Fair"),
            ("Sortino Ratio", f"{metrics.get('sortino_ratio', 0):.2f}",
             "Excellent" if metrics.get('sortino_ratio', 0) > 2 else "Good" if metrics.get('sortino_ratio', 0) > 1 else "Fair"),
            ("Max Drawdown", f"{metrics.get('max_drawdown_pct', 0):.2f}%",
             "Acceptable" if metrics.get('max_drawdown_pct', 0) < 20 else "High Risk"),
            ("Profit Factor", f"{metrics.get('profit_factor', 0):.2f}",
             "Excellent" if metrics.get('profit_factor', 0) > 2 else "Good" if metrics.get('profit_factor', 0) > 1.5 else "Fair"),
            ("Avg Trade Duration", f"{metrics.get('avg_trade_duration', 0):.1f} days", ""),
        ]

        for metric, value, assessment in performance_data:
            ws.cell(row=row, column=1, value=metric).border = self.thin_border
            ws.cell(row=row, column=2, value=value).border = self.thin_border

            assessment_cell = ws.cell(row=row, column=3, value=assessment)
            assessment_cell.border = self.thin_border

            if assessment in ["Excellent", "Good", "+", "Acceptable"]:
                assessment_cell.fill = self.positive_fill
            elif assessment in ["Poor", "-", "High Risk", "Needs Improvement"]:
                assessment_cell.fill = self.negative_fill
            elif assessment == "Fair":
                assessment_cell.fill = self.neutral_fill

            row += 1

        row += 2

        # Risk Summary
        ws[f'A{row}'] = "RISK SUMMARY"
        ws[f'A{row}'].font = self.section_font
        row += 2

        risk_headers = ["Risk Metric", "Value"]
        for col_idx, header in enumerate(risk_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = PatternFill(start_color=self.COLORS['negative_dark'],
                                     end_color=self.COLORS['negative_dark'], fill_type="solid")
            cell.font = self.header_font
            cell.border = self.thin_border
        row += 1

        risk_data = [
            ("Maximum Drawdown ($)", f"${metrics.get('max_drawdown', 0):,.2f}"),
            ("Maximum Drawdown (%)", f"{metrics.get('max_drawdown_pct', 0):.2f}%"),
            ("Average Drawdown", f"${metrics.get('avg_drawdown', 0):,.2f}"),
            ("Volatility (Annualized)", f"{metrics.get('volatility', 0):.2f}%"),
            ("Downside Deviation", f"{metrics.get('downside_deviation', 0):.2f}%"),
            ("Best Day", f"{metrics.get('best_day', 0):.2f}%"),
            ("Worst Day", f"{metrics.get('worst_day', 0):.2f}%"),
            ("Max Loss Streak", f"{metrics.get('max_loss_streak', 0)} trades"),
            ("Consecutive Losing Months", f"{metrics.get('consecutive_losing_months', 0)}"),
        ]

        for metric, value in risk_data:
            ws.cell(row=row, column=1, value=metric).border = self.thin_border
            ws.cell(row=row, column=2, value=value).border = self.thin_border
            row += 1

        # Add Rolling Metric Anomaly Note if anomalies were detected
        rolling_anomalies = metrics.get('rolling_anomalies', [])
        if rolling_anomalies:
            row += 2
            ws[f'A{row}'] = "DATA QUALITY NOTE"
            ws[f'A{row}'].font = self.section_font
            row += 2

            ws[f'A{row}'] = f"⚠ {len(rolling_anomalies)} anomalous values detected in rolling metrics"
            ws[f'A{row}'].font = Font(bold=True, color=self.COLORS['negative_dark'])
            row += 1

            # Show regular vs filtered Sharpe
            regular_sharpe = metrics.get('sharpe_ratio', 0)
            filtered_sharpe = metrics.get('filtered_sharpe_ratio', regular_sharpe)

            ws[f'A{row}'] = f"Sharpe Ratio (with anomalous periods): {regular_sharpe:.2f}"
            row += 1
            ws[f'A{row}'] = f"Sharpe Ratio (anomalous periods excluded): {filtered_sharpe:.2f}"
            ws[f'A{row}'].font = Font(bold=True)
            row += 2

            ws[f'A{row}'] = "See 'Rolling Metrics' sheet for full anomaly details."
            ws[f'A{row}'].font = Font(italic=True, color=self.COLORS['dark_gray'])
            row += 1

            # Summarize by type
            sharpe_anomalies = [a for a in rolling_anomalies if a['metric'] == 'Sharpe Ratio']
            sortino_anomalies = [a for a in rolling_anomalies if a['metric'] == 'Sortino Ratio']

            if sharpe_anomalies:
                values = [a['value'] for a in sharpe_anomalies]
                ws[f'A{row}'] = f"  • Sharpe Ratio: {len(sharpe_anomalies)} anomalies (range: {min(values):.1f} to {max(values):.1f})"
                row += 1

            if sortino_anomalies:
                values = [a['value'] for a in sortino_anomalies]
                ws[f'A{row}'] = f"  • Sortino Ratio: {len(sortino_anomalies)} anomalies (range: {min(values):.1f} to {max(values):.1f})"
                row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18
        for col in ['D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 15

    def _create_dashboard_visualizations(self, wb: Workbook, result, metrics: Dict[str, Any]):
        """Create Dashboard sheet with embedded matplotlib visualizations."""
        ws = wb.create_sheet("Dashboard")

        row = 1
        ws[f'A{row}'] = "VISUAL DASHBOARD"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:N{row}')
        row += 2

        if not self.include_matplotlib_charts or self.viz is None:
            ws[f'A{row}'] = "Note: Install matplotlib for enhanced visualizations (pip install matplotlib)"
            ws[f'A{row}'].font = Font(italic=True, color=self.COLORS['dark_gray'])
            row += 2
            self._create_native_equity_chart(ws, result, row)
            return

        equity_df = result.portfolio_equity_curve
        all_trades = metrics.get('all_trades', [])

        chart_row = row

        # 1. Equity Curve with Drawdown
        try:
            equity_img = self.viz.create_equity_curve_with_drawdown(equity_df)
            img = Image(equity_img)
            img.width = 600
            img.height = 350
            ws.add_image(img, f'A{chart_row}')
        except Exception as e:
            ws[f'A{chart_row}'] = f"Equity chart unavailable"

        chart_row += 20

        # 2. Monthly Returns Heatmap
        try:
            heatmap_img = self.viz.create_monthly_returns_heatmap(equity_df)
            img = Image(heatmap_img)
            img.width = 600
            img.height = 350
            ws.add_image(img, f'A{chart_row}')
        except Exception as e:
            ws[f'A{chart_row}'] = f"Monthly heatmap unavailable: {str(e)[:50]}"
            ws[f'A{chart_row}'].font = Font(italic=True, color=self.COLORS['dark_gray'])

        chart_row += 20

        # 3. Trade Distribution
        if all_trades:
            try:
                dist_img = self.viz.create_trade_distribution_histogram(all_trades)
                img = Image(dist_img)
                img.width = 500
                img.height = 300
                ws.add_image(img, f'A{chart_row}')
            except Exception:
                pass

        # 4. Rolling Metrics (right side) - with anomaly filtering
        try:
            rolling_img, _ = self.viz.create_rolling_metrics_chart(
                equity_df,
                window=90,
                filter_anomalies=True
            )
            img = Image(rolling_img)
            img.width = 600
            img.height = 400
            ws.add_image(img, f'I{row}')
        except Exception:
            pass

        # 5. Contribution Analysis
        if len(result.symbol_results) > 1:
            symbol_pnl = {sym: res.total_return for sym, res in result.symbol_results.items()}
            try:
                contrib_img = self.viz.create_contribution_analysis(symbol_pnl)
                img = Image(contrib_img)
                img.width = 600
                img.height = 350
                ws.add_image(img, f'I{row + 22}')
            except Exception:
                pass

    def _create_native_equity_chart(self, ws, result, start_row: int):
        """Create native Excel equity curve chart as fallback."""
        equity_df = result.portfolio_equity_curve

        ws.cell(row=start_row, column=1, value="Date").font = self.header_font
        ws.cell(row=start_row, column=2, value="Equity").font = self.header_font

        for idx, row_data in enumerate(equity_df.itertuples(), 1):
            ws.cell(row=start_row + idx, column=1,
                    value=row_data.date.strftime("%Y-%m-%d") if hasattr(row_data.date, 'strftime') else str(row_data.date))
            ws.cell(row=start_row + idx, column=2, value=row_data.equity)

        if len(equity_df) > 1:
            chart = LineChart()
            chart.title = "Portfolio Equity Curve"
            chart.x_axis.title = "Date"
            chart.y_axis.title = "Equity ($)"
            chart.width = 20
            chart.height = 10

            data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(equity_df))
            chart.add_data(data, titles_from_data=True)

            ws.add_chart(chart, "D" + str(start_row))

    def _create_performance_metrics(self, wb: Workbook, result, metrics: Dict[str, Any]):
        """Create detailed Performance Metrics sheet."""
        ws = wb.create_sheet("Performance Metrics")

        row = 1
        ws[f'A{row}'] = "COMPREHENSIVE PERFORMANCE METRICS"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:F{row}')
        row += 3

        # Section A: Return Metrics
        row = self._add_section_header(ws, row, "A. RETURN METRICS")

        return_data = [
            ("Total Return ($)", f"${metrics['total_return']:,.2f}"),
            ("Total Return (%)", f"{metrics['total_return_pct']:.2f}%"),
            ("CAGR", f"{metrics.get('cagr', 0):.2f}%"),
            ("Best Day", f"{metrics.get('best_day', 0):.2f}%"),
            ("Worst Day", f"{metrics.get('worst_day', 0):.2f}%"),
        ]
        row = self._add_metrics_table(ws, row, return_data)
        row += 2

        # Section B: Risk-Adjusted Ratios
        row = self._add_section_header(ws, row, "B. RISK-ADJUSTED RATIOS")

        ratios_data = [
            ("Sharpe Ratio", f"{metrics.get('sharpe_ratio', 0):.2f}"),
            ("Sortino Ratio", f"{metrics.get('sortino_ratio', 0):.2f}"),
            ("Calmar Ratio", f"{metrics.get('calmar_ratio', 0):.2f}"),
            ("Recovery Factor", f"{metrics.get('recovery_factor', 0):.2f}"),
            ("Profit Factor", f"{metrics.get('profit_factor', 0):.2f}"),
        ]
        row = self._add_metrics_table(ws, row, ratios_data)
        row += 2

        # Section C: Risk Metrics
        row = self._add_section_header(ws, row, "C. RISK METRICS")

        risk_data = [
            ("Max Drawdown ($)", f"${metrics.get('max_drawdown', 0):,.2f}"),
            ("Max Drawdown (%)", f"{metrics.get('max_drawdown_pct', 0):.2f}%"),
            ("Average Drawdown", f"${metrics.get('avg_drawdown', 0):,.2f}"),
            ("Volatility (Annualized)", f"{metrics.get('volatility', 0):.2f}%"),
            ("Downside Deviation", f"{metrics.get('downside_deviation', 0):.2f}%"),
            ("Value at Risk (95%)", f"{metrics.get('var_95', 0):.2f}%"),
            ("CVaR (Expected Shortfall)", f"{metrics.get('cvar_95', 0):.2f}%"),
        ]
        row = self._add_metrics_table(ws, row, risk_data)
        row += 2

        # Section D: Trade Quality
        row = self._add_section_header(ws, row, "D. TRADE QUALITY METRICS")

        trade_data = [
            ("Total Trades", str(metrics['num_trades'])),
            ("Winning Trades", str(metrics['num_wins'])),
            ("Losing Trades", str(metrics['num_losses'])),
            ("Win Rate", f"{metrics.get('win_rate', 0)*100:.1f}%"),
            ("Average Win", f"${metrics.get('avg_win', 0):,.2f}"),
            ("Average Loss", f"${metrics.get('avg_loss', 0):,.2f}"),
            ("Largest Win", f"${metrics.get('largest_win', 0):,.2f}"),
            ("Largest Loss", f"${metrics.get('largest_loss', 0):,.2f}"),
            ("Risk/Reward Ratio", f"{metrics.get('risk_reward_ratio', 0):.2f}"),
            ("Expectancy", f"${metrics.get('expectancy', 0):.2f}"),
            ("Avg Trade Duration", f"{metrics.get('avg_trade_duration', 0):.1f} days"),
        ]
        row = self._add_metrics_table(ws, row, trade_data)
        row += 2

        # Section E: Consistency
        row = self._add_section_header(ws, row, "E. CONSISTENCY METRICS")

        consistency_data = [
            ("Profitable Months", f"{metrics.get('profitable_months_pct', 0):.1f}%"),
            ("Profitable Weeks", f"{metrics.get('profitable_weeks_pct', 0):.1f}%"),
            ("Monthly Consistency (Std)", f"{metrics.get('monthly_consistency', 0):.2f}%"),
            ("Max Consecutive Losing Months", str(metrics.get('consecutive_losing_months', 0))),
            ("Max Win Streak", f"{metrics.get('max_win_streak', 0)} trades"),
            ("Max Loss Streak", f"{metrics.get('max_loss_streak', 0)} trades"),
        ]
        row = self._add_metrics_table(ws, row, consistency_data)
        row += 3

        # Section F: Calculation Methodology (Important for user understanding)
        row = self._add_section_header(ws, row, "F. CALCULATION METHODOLOGY")

        ws[f'A{row}'] = "Drawdown Calculation:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = "  Max Drawdown measures the largest peak-to-trough decline in equity."
        row += 1
        ws[f'A{row}'] = "  Formula: For each day, calculate (Peak Equity - Current Equity) / Peak Equity * 100"
        row += 1
        ws[f'A{row}'] = "  Peak Equity is the running maximum equity achieved up to that point."
        row += 1
        ws[f'A{row}'] = "  Max Drawdown (%) is the maximum of all these daily drawdown values."
        row += 2

        ws[f'A{row}'] = "Volatility Calculation:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = "  Annualized Volatility = Daily Returns Std Dev × √252 × 100"
        row += 1
        ws[f'A{row}'] = "  Where daily returns = (Today's Equity - Yesterday's Equity) / Yesterday's Equity"
        row += 2

        ws[f'A{row}'] = "Best/Worst Day:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = "  Best Day = Maximum single-day percentage return in the period"
        row += 1
        ws[f'A{row}'] = "  Worst Day = Minimum single-day percentage return in the period"
        row += 2

        ws[f'A{row}'] = "Sharpe Ratio:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = f"  Sharpe = (Mean Daily Excess Return / Std Dev of Daily Returns) × √252"
        row += 1
        ws[f'A{row}'] = f"  Risk-free rate assumed: {DEFAULT_RISK_FREE_RATE*100:.1f}% annually"
        row += 2

        ws[f'A{row}'] = "Sortino Ratio:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = "  Sortino = (Mean Daily Excess Return / Downside Deviation) × √252"
        row += 1
        ws[f'A{row}'] = "  Only considers negative returns for risk calculation (more appropriate for assymetric returns)"

        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = 20

    def _create_per_security_analysis(self, wb: Workbook, result, metrics: Dict[str, Any]):
        """Create Per-Security Analysis sheet."""
        ws = wb.create_sheet("Per-Security")

        row = 1
        ws[f'A{row}'] = "PER-SECURITY PERFORMANCE ANALYSIS"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:H{row}')
        row += 3

        headers = ["Symbol", "P/L ($)", "P/L %", "Trades", "Win Rate", "Contribution %", "Profit Factor", "Avg Duration"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.thin_border
        row += 1

        contributions = metrics.get('security_contribution', {})

        for symbol, sym_result in result.symbol_results.items():
            trades = sym_result.trades
            contrib = contributions.get(symbol, {})

            winners = [t for t in trades if t.pl > 0]
            losers = [t for t in trades if t.pl <= 0]
            gross_profit = sum(t.pl for t in winners) if winners else 0
            gross_loss = abs(sum(t.pl for t in losers)) if losers else 0
            pf = gross_profit / gross_loss if gross_loss > 0 else 999.99

            avg_dur = np.mean([t.duration_days for t in trades]) if trades else 0

            row_data = [
                symbol,
                f"${sym_result.total_return:,.2f}",
                f"{sym_result.total_return_pct:.2f}%",
                len(trades),
                f"{contrib.get('win_rate', 0):.1f}%",
                f"{contrib.get('contribution_pct', 0):.1f}%",
                f"{pf:.2f}",
                f"{avg_dur:.1f}d"
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = self.thin_border

                if col_idx == 2:
                    if sym_result.total_return > 0:
                        cell.fill = self.positive_fill
                    else:
                        cell.fill = self.negative_fill

            row += 1

        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_trade_analysis(self, wb: Workbook, result, metrics: Dict[str, Any]):
        """Create Trade Analysis sheet."""
        ws = wb.create_sheet("Trade Analysis")

        row = 1
        ws[f'A{row}'] = "TRADE ANALYSIS"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:H{row}')
        row += 3

        all_trades = metrics.get('all_trades', [])

        if not all_trades:
            ws[f'A{row}'] = "No trades to analyze"
            return

        # Trade Summary
        row = self._add_section_header(ws, row, "A. TRADE SUMMARY")

        total_trades = len(all_trades)
        winners = [t for t in all_trades if t.pl > 0]
        losers = [t for t in all_trades if t.pl <= 0]

        summary_data = [
            ("Total Trades", str(total_trades)),
            ("Winners", f"{len(winners)} ({len(winners)/total_trades*100:.1f}%)"),
            ("Losers", f"{len(losers)} ({len(losers)/total_trades*100:.1f}%)"),
            ("Avg Win", f"${np.mean([t.pl for t in winners]):,.2f}" if winners else "$0.00"),
            ("Avg Loss", f"${np.mean([t.pl for t in losers]):,.2f}" if losers else "$0.00"),
            ("Largest Win", f"${max(t.pl for t in winners):,.2f}" if winners else "$0.00"),
            ("Largest Loss", f"${min(t.pl for t in losers):,.2f}" if losers else "$0.00"),
        ]
        row = self._add_metrics_table(ws, row, summary_data)
        row += 2

        # Streak Analysis
        row = self._add_section_header(ws, row, "B. STREAK ANALYSIS")

        streak_data = [
            ("Max Win Streak", f"{metrics.get('max_win_streak', 0)} trades"),
            ("Max Loss Streak", f"{metrics.get('max_loss_streak', 0)} trades"),
            ("Current Streak", f"{metrics.get('current_streak', 0)} ({metrics.get('current_streak_type', 'N/A')})"),
            ("Avg Win Streak Length", f"{metrics.get('avg_win_streak', 0):.1f} trades"),
            ("Avg Loss Streak Length", f"{metrics.get('avg_loss_streak', 0):.1f} trades"),
        ]
        row = self._add_metrics_table(ws, row, streak_data)
        row += 2

        # Duration Analysis
        row = self._add_section_header(ws, row, "C. DURATION ANALYSIS")

        winner_durations = [t.duration_days for t in winners] if winners else [0]
        loser_durations = [t.duration_days for t in losers] if losers else [0]

        duration_data = [
            ("Avg Duration (All)", f"{np.mean([t.duration_days for t in all_trades]):.1f} days"),
            ("Avg Duration (Winners)", f"{np.mean(winner_durations):.1f} days"),
            ("Avg Duration (Losers)", f"{np.mean(loser_durations):.1f} days"),
            ("Shortest Trade", f"{min(t.duration_days for t in all_trades)} days"),
            ("Longest Trade", f"{max(t.duration_days for t in all_trades)} days"),
        ]
        row = self._add_metrics_table(ws, row, duration_data)

        # Add visualizations if available
        if self.include_matplotlib_charts and self.viz:
            # Trade distribution histogram
            try:
                dist_img = self.viz.create_trade_distribution_histogram(all_trades)
                img = Image(dist_img)
                img.width = 500
                img.height = 300
                ws.add_image(img, f'E3')
            except Exception:
                pass

            # Streak analysis visualization (new improved version)
            try:
                streak_img = self.viz.create_streak_visualization(all_trades)
                img = Image(streak_img)
                img.width = 700
                img.height = 450
                ws.add_image(img, f'A{row + 4}')
            except Exception:
                pass

            # Trade clustering analysis
            try:
                cluster_img = self.viz.create_trade_clustering_analysis(all_trades)
                img = Image(cluster_img)
                img.width = 700
                img.height = 500
                ws.add_image(img, f'A{row + 30}')
            except Exception:
                pass

    def _create_trade_log(self, wb: Workbook, result):
        """Create Enhanced Trade Log sheet with all trades including vulnerability and swap info."""
        ws = wb.create_sheet("Trade Log")

        all_trades = []
        for sym_result in result.symbol_results.values():
            all_trades.extend(sym_result.trades)
        all_trades.sort(key=lambda t: t.entry_date)

        if not all_trades:
            ws['A1'] = "No trades recorded"
            return

        # Build lookup dictionary for capital allocation events by trade_id
        capital_events_by_trade = {}
        if hasattr(result, 'capital_allocation_events') and result.capital_allocation_events:
            for event in result.capital_allocation_events:
                if event.trade_id:
                    capital_events_by_trade[event.trade_id] = event

        # Build lookup for vulnerability swaps
        swap_dates_symbols = set()
        if hasattr(result, 'vulnerability_swaps') and result.vulnerability_swaps:
            for swap in result.vulnerability_swaps:
                swap_dates_symbols.add((swap.date.strftime("%Y-%m-%d") if hasattr(swap.date, 'strftime') else str(swap.date), swap.new_symbol))

        # Enhanced headers with vulnerability and swap info
        headers = [
            "Trade ID", "Symbol", "Entry Date", "Entry Price", "Exit Date", "Exit Price",
            "Quantity", "P/L ($)", "P/L (%)", "Duration", "Entry Reason", "Exit Reason",
            "Entry Capital", "Required Capital", "Concurrent Pos", "Competing Signals",
            "Was Swap", "Vulnerability Score", "FX P/L"
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.thin_border

        for row_idx, trade in enumerate(all_trades, 2):
            entry_date_str = trade.entry_date.strftime("%Y-%m-%d") if hasattr(trade.entry_date, 'strftime') else str(trade.entry_date)

            ws.cell(row=row_idx, column=1, value=trade.trade_id)
            ws.cell(row=row_idx, column=2, value=trade.symbol)
            ws.cell(row=row_idx, column=3, value=entry_date_str)
            ws.cell(row=row_idx, column=4, value=trade.entry_price)
            ws.cell(row=row_idx, column=5, value=trade.exit_date.strftime("%Y-%m-%d") if hasattr(trade.exit_date, 'strftime') else str(trade.exit_date))
            ws.cell(row=row_idx, column=6, value=trade.exit_price)
            ws.cell(row=row_idx, column=7, value=trade.quantity)

            pl_cell = ws.cell(row=row_idx, column=8, value=trade.pl)
            pl_cell.number_format = '#,##0.00'
            pl_cell.fill = self.positive_fill if trade.pl > 0 else self.negative_fill

            pct_cell = ws.cell(row=row_idx, column=9, value=trade.pl_pct)
            pct_cell.number_format = '0.00"%"'

            ws.cell(row=row_idx, column=10, value=trade.duration_days)
            ws.cell(row=row_idx, column=11, value=trade.entry_reason or "")
            ws.cell(row=row_idx, column=12, value=trade.exit_reason or "")

            # Capital info from trade
            ws.cell(row=row_idx, column=13, value=trade.entry_capital_available).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=14, value=trade.entry_capital_required).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=15, value=trade.concurrent_positions)

            # Competing signals
            competing = ", ".join(trade.competing_signals) if trade.competing_signals else ""
            ws.cell(row=row_idx, column=16, value=competing)

            # Was this a swap entry?
            was_swap = (entry_date_str, trade.symbol) in swap_dates_symbols
            swap_cell = ws.cell(row=row_idx, column=17, value="Yes" if was_swap else "No")
            if was_swap:
                swap_cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                swap_cell.font = Font(bold=True)

            # Vulnerability score from capital allocation event
            vuln_score = ""
            if trade.trade_id in capital_events_by_trade:
                event = capital_events_by_trade[trade.trade_id]
                if event.vulnerability_scores and trade.symbol in event.vulnerability_scores:
                    vuln_score = f"{event.vulnerability_scores[trade.symbol]:.1f}"
            ws.cell(row=row_idx, column=18, value=vuln_score)

            # FX P/L
            fx_pl = getattr(trade, 'fx_pl', 0) or 0
            fx_cell = ws.cell(row=row_idx, column=19, value=fx_pl)
            fx_cell.number_format = '#,##0.00'
            if fx_pl != 0:
                fx_cell.fill = self.positive_fill if fx_pl > 0 else self.negative_fill

        # Column widths
        widths = [12, 10, 12, 10, 12, 10, 10, 12, 10, 8, 20, 20, 15, 15, 10, 25, 8, 12, 12]
        for col_idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Add note about the data at the bottom
        last_row = len(all_trades) + 3
        ws[f'A{last_row}'] = "Notes:"
        ws[f'A{last_row}'].font = Font(bold=True)
        ws[f'A{last_row + 1}'] = "- 'Was Swap' indicates if this trade was entered as a vulnerability score swap (replacing another position)"
        ws[f'A{last_row + 2}'] = "- 'Vulnerability Score' shows the score at the time of entry (if vulnerability mode was used)"
        ws[f'A{last_row + 3}'] = "- 'FX P/L' shows profit/loss from currency conversion (for non-base currency securities)"

    def _create_rolling_metrics(self, wb: Workbook, result, metrics: Dict[str, Any]):
        """Create Rolling Metrics sheet with anomaly detection and reporting."""
        ws = wb.create_sheet("Rolling Metrics")

        row = 1
        ws[f'A{row}'] = "ROLLING PERFORMANCE METRICS"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:F{row}')
        row += 2

        # Add methodology note
        ws[f'A{row}'] = "Note: Rolling window of 90 days used for stability. Anomalous values (|value| > 10 or z-score > 3) are filtered from charts."
        ws[f'A{row}'].font = Font(italic=True, size=9, color=self.COLORS['dark_gray'])
        row += 2

        rolling_df = metrics.get('rolling_metrics', pd.DataFrame())
        detected_anomalies = []

        if rolling_df.empty:
            ws[f'A{row}'] = "Insufficient data for rolling metrics calculation"
            return

        if self.include_matplotlib_charts and self.viz:
            equity_df = result.portfolio_equity_curve
            try:
                # Get rolling chart with anomaly detection (90-day window)
                rolling_img, detected_anomalies = self.viz.create_rolling_metrics_chart(
                    equity_df,
                    window=90,
                    filter_anomalies=True,
                    anomaly_absolute_threshold=10.0,
                    anomaly_zscore_threshold=3.0
                )
                img = Image(rolling_img)
                img.width = 700
                img.height = 500
                ws.add_image(img, f'A{row}')
                row += 28

                # Store anomalies in metrics for use in executive summary
                metrics['rolling_anomalies'] = detected_anomalies

            except Exception as e:
                ws[f'A{row}'] = f"Rolling metrics chart unavailable: {str(e)[:50]}"
                row += 2

            all_trades = metrics.get('all_trades', [])
            if len(all_trades) >= 25:
                try:
                    wr_img = self.viz.create_win_rate_over_time(all_trades, window=20)
                    img = Image(wr_img)
                    img.width = 700
                    img.height = 400
                    ws.add_image(img, f'A{row}')
                    row += 24
                except Exception:
                    pass

        # Add Anomaly Table if anomalies were detected
        if detected_anomalies:
            row += 2
            row = self._add_section_header(ws, row, "METRIC ANOMALIES (Excluded from Charts)")

            ws[f'A{row}'] = f"Total anomalous values detected and filtered: {len(detected_anomalies)}"
            ws[f'A{row}'].font = Font(bold=True, color=self.COLORS['negative_dark'])
            row += 2

            # Anomaly table headers
            anomaly_headers = ["Date", "Metric", "Value", "Reason for Exclusion"]
            for col_idx, header in enumerate(anomaly_headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.fill = PatternFill(start_color=self.COLORS['negative_dark'],
                                        end_color=self.COLORS['negative_dark'], fill_type="solid")
                cell.font = self.header_font
                cell.border = self.thin_border
            row += 1

            # Sort anomalies by date
            sorted_anomalies = sorted(detected_anomalies, key=lambda x: x['date'])

            for anomaly in sorted_anomalies:
                date_str = anomaly['date'].strftime('%Y-%m-%d') if hasattr(anomaly['date'], 'strftime') else str(anomaly['date'])
                ws.cell(row=row, column=1, value=date_str).border = self.thin_border
                ws.cell(row=row, column=2, value=anomaly['metric']).border = self.thin_border

                value_cell = ws.cell(row=row, column=3, value=f"{anomaly['value']:.2f}")
                value_cell.border = self.thin_border
                value_cell.fill = self.negative_fill

                ws.cell(row=row, column=4, value=anomaly['reason']).border = self.thin_border
                row += 1

            row += 2

            # Add explanation note
            ws[f'A{row}'] = "Why are these values excluded?"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = "Extreme spikes in rolling Sharpe/Sortino ratios typically occur during periods of very low volatility"
            row += 1
            ws[f'A{row}'] = "or rapid equity changes. These values distort the visual representation and are not representative"
            row += 1
            ws[f'A{row}'] = "of typical strategy performance. The underlying trade data remains intact."
            row += 2

            # Summary statistics of excluded periods
            if len(sorted_anomalies) > 0:
                sharpe_anomalies = [a for a in sorted_anomalies if a['metric'] == 'Sharpe Ratio']
                sortino_anomalies = [a for a in sorted_anomalies if a['metric'] == 'Sortino Ratio']

                ws[f'A{row}'] = "Exclusion Summary:"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1

                if sharpe_anomalies:
                    sharpe_values = [a['value'] for a in sharpe_anomalies]
                    ws[f'A{row}'] = f"  Sharpe Ratio: {len(sharpe_anomalies)} values excluded (range: {min(sharpe_values):.1f} to {max(sharpe_values):.1f})"
                    row += 1

                if sortino_anomalies:
                    sortino_values = [a['value'] for a in sortino_anomalies]
                    ws[f'A{row}'] = f"  Sortino Ratio: {len(sortino_anomalies)} values excluded (range: {min(sortino_values):.1f} to {max(sortino_values):.1f})"
                    row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 40

    def _create_statistical_analysis(self, wb: Workbook, result, metrics: Dict[str, Any]):
        """Create Statistical Analysis sheet."""
        ws = wb.create_sheet("Statistical Analysis")

        row = 1
        ws[f'A{row}'] = "STATISTICAL SIGNIFICANCE ANALYSIS"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:F{row}')
        row += 3

        row = self._add_section_header(ws, row, "A. HYPOTHESIS TESTING")

        ws[f'A{row}'] = "H0: Mean trade return = 0 (no edge)"
        ws[f'A{row}'].font = Font(italic=True)
        row += 1
        ws[f'A{row}'] = "H1: Mean trade return != 0 (strategy has edge)"
        ws[f'A{row}'].font = Font(italic=True)
        row += 2

        stat_data = [
            ("T-Statistic", f"{metrics.get('t_statistic', 0):.4f}"),
            ("P-Value", f"{metrics.get('p_value', 1):.4f}"),
            ("Statistically Significant (alpha=0.05)", "Yes" if metrics.get('is_significant', False) else "No"),
            ("Number of Trades (n)", str(metrics.get('num_trades', 0))),
        ]
        row = self._add_metrics_table(ws, row, stat_data)
        row += 2

        is_significant = metrics.get('is_significant', False)
        p_value = metrics.get('p_value', 1)

        ws[f'A{row}'] = "Interpretation:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        if is_significant:
            ws[f'A{row}'] = f"The strategy's returns are statistically significant (p={p_value:.4f} < 0.05)"
            ws[f'A{row}'].font = Font(color=self.COLORS['positive_dark'])
        else:
            ws[f'A{row}'] = f"The strategy's returns are NOT statistically significant (p={p_value:.4f} >= 0.05)"
            ws[f'A{row}'].font = Font(color=self.COLORS['negative_dark'])

        row += 3

        row = self._add_section_header(ws, row, "B. RETURN DISTRIBUTION")

        dist_data = [
            ("Skewness", f"{metrics.get('skewness', 0):.4f}"),
            ("Kurtosis", f"{metrics.get('kurtosis', 0):.4f}"),
            ("Value at Risk (95%)", f"{metrics.get('var_95', 0):.2f}%"),
            ("Conditional VaR (ES)", f"{metrics.get('cvar_95', 0):.2f}%"),
        ]
        row = self._add_metrics_table(ws, row, dist_data)

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20

    def _create_capital_allocation(self, wb: Workbook, result, metrics: Dict[str, Any]):
        """Create Capital Allocation sheet."""
        ws = wb.create_sheet("Capital Allocation")

        row = 1
        ws[f'A{row}'] = "CAPITAL ALLOCATION ANALYSIS"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:F{row}')
        row += 3

        equity_df = result.portfolio_equity_curve

        row = self._add_section_header(ws, row, "A. ALLOCATION STATISTICS")

        if 'capital' in equity_df.columns and 'position_value' in equity_df.columns:
            avg_cash_pct = (equity_df['capital'] / equity_df['equity'] * 100).mean()
            avg_position_pct = (equity_df['position_value'] / equity_df['equity'] * 100).mean()
            max_positions = equity_df.get('num_positions', pd.Series([0])).max()

            alloc_data = [
                ("Average Cash %", f"{avg_cash_pct:.1f}%"),
                ("Average Position %", f"{avg_position_pct:.1f}%"),
                ("Max Concurrent Positions", str(int(max_positions))),
                ("Signal Rejections", str(len(result.signal_rejections))),
                ("Vulnerability Swaps", str(len(result.vulnerability_swaps))),
            ]
        else:
            alloc_data = [
                ("Signal Rejections", str(len(result.signal_rejections))),
                ("Vulnerability Swaps", str(len(result.vulnerability_swaps))),
            ]

        row = self._add_metrics_table(ws, row, alloc_data)

        if self.include_matplotlib_charts and self.viz and 'capital' in equity_df.columns:
            try:
                cap_img = self.viz.create_capital_utilization_chart(equity_df)
                img = Image(cap_img)
                img.width = 700
                img.height = 300
                ws.add_image(img, f'A{row + 2}')
            except Exception:
                pass

    def _create_signal_rejections(self, wb: Workbook, result):
        """Create Signal Rejections sheet."""
        ws = wb.create_sheet("Signal Rejections")

        row = 1
        ws[f'A{row}'] = "SIGNAL REJECTION ANALYSIS"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:F{row}')
        row += 3

        ws[f'A{row}'] = "Total Rejections:"
        ws[f'B{row}'] = len(result.signal_rejections)
        row += 2

        rejection_by_symbol = {}
        for r in result.signal_rejections:
            rejection_by_symbol[r.symbol] = rejection_by_symbol.get(r.symbol, 0) + 1

        row = self._add_section_header(ws, row, "A. REJECTIONS BY SYMBOL")

        for symbol, count in sorted(rejection_by_symbol.items(), key=lambda x: -x[1]):
            ws.cell(row=row, column=1, value=symbol)
            ws.cell(row=row, column=2, value=count)
            row += 1

        row += 2

        row = self._add_section_header(ws, row, "B. DETAILED REJECTIONS")

        headers = ["Date", "Symbol", "Signal", "Reason", "Available", "Required"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
        row += 1

        for rejection in result.signal_rejections[:100]:
            ws.cell(row=row, column=1, value=rejection.date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=2, value=rejection.symbol)
            ws.cell(row=row, column=3, value=rejection.signal_type)
            ws.cell(row=row, column=4, value=rejection.reason[:50] if rejection.reason else "")
            ws.cell(row=row, column=5, value=f"${rejection.available_capital:,.0f}")
            ws.cell(row=row, column=6, value=f"${rejection.required_capital:,.0f}")
            row += 1

    def _create_vulnerability_analysis(self, wb: Workbook, result):
        """Create Vulnerability Analysis sheet."""
        ws = wb.create_sheet("Vulnerability Analysis")

        row = 1
        ws[f'A{row}'] = "VULNERABILITY SCORE ANALYSIS"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:F{row}')
        row += 3

        row = self._add_section_header(ws, row, "A. POSITION SWAPS")

        ws[f'A{row}'] = "Total Swaps:"
        ws[f'B{row}'] = len(result.vulnerability_swaps)
        row += 2

        if result.vulnerability_swaps:
            headers = ["Date", "Closed Symbol", "Score", "New Symbol"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.fill = self.header_fill
                cell.font = self.header_font
            row += 1

            for swap in result.vulnerability_swaps:
                ws.cell(row=row, column=1, value=swap.date.strftime("%Y-%m-%d"))
                ws.cell(row=row, column=2, value=swap.closed_symbol)

                score_cell = ws.cell(row=row, column=3, value=f"{swap.closed_score:.1f}")
                if swap.closed_score < 25:
                    score_cell.fill = self.negative_fill
                elif swap.closed_score < 50:
                    score_cell.fill = self.neutral_fill

                ws.cell(row=row, column=4, value=swap.new_symbol)
                row += 1

    def _create_costs_analysis(self, wb: Workbook, result, metrics: Dict[str, Any]):
        """Create Costs Analysis sheet."""
        ws = wb.create_sheet("Costs Analysis")

        row = 1
        ws[f'A{row}'] = "TRADING COSTS ANALYSIS"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 3

        costs = metrics.get('costs_analysis', {})

        row = self._add_section_header(ws, row, "A. COST BREAKDOWN")

        cost_data = [
            ("Total Commission Paid", f"${costs.get('total_commission', 0):,.2f}"),
            ("Average Commission/Trade", f"${costs.get('avg_commission', 0):,.2f}"),
            ("P/L Before Costs", f"${costs.get('pl_before_costs', 0):,.2f}"),
            ("Commission % of Gross P/L", f"{costs.get('commission_pct_of_pl', 0):.2f}%"),
        ]
        row = self._add_metrics_table(ws, row, cost_data)
        row += 2

        cost_pct = costs.get('commission_pct_of_pl', 0)
        ws[f'A{row}'] = "Cost Efficiency Assessment:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        if cost_pct < 5:
            ws[f'A{row}'] = "Excellent cost efficiency (<5% of gross P/L)"
            ws[f'A{row}'].font = Font(color=self.COLORS['positive_dark'])
        elif cost_pct < 10:
            ws[f'A{row}'] = "Acceptable cost efficiency (5-10% of gross P/L)"
            ws[f'A{row}'].font = Font(color=self.COLORS['neutral_dark'])
        else:
            ws[f'A{row}'] = "High costs (>10% of gross P/L) - consider reducing trade frequency"
            ws[f'A{row}'].font = Font(color=self.COLORS['negative_dark'])

    def _create_monthly_heatmap(self, wb: Workbook, result, metrics: Dict[str, Any]):
        """Create Monthly Returns Heatmap sheet with native Excel formatting and yearly breakdown."""
        ws = wb.create_sheet("Monthly Heatmap")

        row = 1
        ws[f'A{row}'] = "MONTHLY & YEARLY RETURNS ANALYSIS"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:N{row}')
        row += 3

        # Calculate monthly returns directly from equity curve for reliability
        equity_df = result.portfolio_equity_curve
        if len(equity_df) < 30:  # Need at least ~1 month of data
            ws[f'A{row}'] = "Insufficient data for monthly analysis (need at least 30 days)"
            return

        df = equity_df.copy()
        df['date'] = pd.to_datetime(df['date'])
        df = df.set_index('date')

        # Calculate monthly equity values and returns
        monthly_equity = df['equity'].resample('ME').last()
        monthly_returns = monthly_equity.pct_change() * 100
        monthly_returns = monthly_returns.dropna()

        if len(monthly_returns) < 1:
            ws[f'A{row}'] = "Insufficient data for monthly analysis"
            return

        # Create monthly returns DataFrame with year and month
        monthly_df = monthly_returns.reset_index()
        monthly_df.columns = ['date', 'return_pct']
        monthly_df['year'] = monthly_df['date'].dt.year
        monthly_df['month'] = monthly_df['date'].dt.month
        monthly_df['month_name'] = monthly_df['date'].dt.strftime('%b')

        years = sorted(monthly_df['year'].unique())
        month_order = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

        # ============ MONTHLY RETURNS HEATMAP ============
        ws[f'A{row}'] = "MONTHLY RETURNS HEATMAP (%)"
        ws[f'A{row}'].font = self.section_font
        row += 2

        # Header row: Month | Year1 | Year2 | ... | YearN | Avg
        ws.cell(row=row, column=1, value="Month").font = self.header_font
        ws.cell(row=row, column=1).fill = self.header_fill
        for col_idx, year in enumerate(years, 2):
            cell = ws.cell(row=row, column=col_idx, value=year)
            cell.font = self.header_font
            cell.fill = self.header_fill
        # Add Average column
        avg_col = len(years) + 2
        cell = ws.cell(row=row, column=avg_col, value="Avg")
        cell.font = self.header_font
        cell.fill = self.header_fill
        row += 1

        # Data rows for each month
        month_avgs = {}
        for month_idx, month in enumerate(month_order, 1):
            ws.cell(row=row, column=1, value=month).font = Font(bold=True)

            month_values = []
            for col_idx, year in enumerate(years, 2):
                month_data = monthly_df[(monthly_df['year'] == year) & (monthly_df['month'] == month_idx)]
                if not month_data.empty:
                    value = month_data['return_pct'].values[0]
                    month_values.append(value)
                    cell = ws.cell(row=row, column=col_idx, value=round(value, 1))
                    cell.number_format = '0.0"%"'

                    if value > 0:
                        cell.fill = self.positive_fill
                        cell.font = Font(color='006100')
                    elif value < 0:
                        cell.fill = self.negative_fill
                        cell.font = Font(color='9C0006')

            # Calculate and display average for this month
            if month_values:
                avg = np.mean(month_values)
                month_avgs[month] = avg
                cell = ws.cell(row=row, column=avg_col, value=round(avg, 1))
                cell.number_format = '0.0"%"'
                cell.font = Font(bold=True)
                if avg > 0:
                    cell.fill = self.positive_fill
                elif avg < 0:
                    cell.fill = self.negative_fill

            row += 1

        # Add Year Total row
        ws.cell(row=row, column=1, value="Year Total").font = Font(bold=True)
        ws.cell(row=row, column=1).fill = self.subheader_fill
        year_totals = []
        for col_idx, year in enumerate(years, 2):
            year_data = monthly_df[monthly_df['year'] == year]
            year_return = year_data['return_pct'].sum()
            year_totals.append(year_return)
            cell = ws.cell(row=row, column=col_idx, value=round(year_return, 1))
            cell.number_format = '0.0"%"'
            cell.font = Font(bold=True)
            cell.fill = self.subheader_fill
            if year_return > 0:
                cell.font = Font(bold=True, color='006100')
            elif year_return < 0:
                cell.font = Font(bold=True, color='9C0006')
        row += 3

        # ============ YEARLY BREAKDOWN ============
        ws[f'A{row}'] = "YEARLY PERFORMANCE BREAKDOWN"
        ws[f'A{row}'].font = self.section_font
        row += 2

        # Yearly summary headers
        yearly_headers = ["Year", "Total Return %", "Best Month", "Best %", "Worst Month", "Worst %",
                         "Positive Months", "Negative Months", "Win Rate %"]
        for col_idx, header in enumerate(yearly_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        row += 1

        for year in years:
            year_data = monthly_df[monthly_df['year'] == year].copy()
            if year_data.empty:
                continue

            year_return = year_data['return_pct'].sum()
            best_month_idx = year_data['return_pct'].idxmax()
            worst_month_idx = year_data['return_pct'].idxmin()
            best_month = year_data.loc[best_month_idx, 'month_name']
            best_pct = year_data.loc[best_month_idx, 'return_pct']
            worst_month = year_data.loc[worst_month_idx, 'month_name']
            worst_pct = year_data.loc[worst_month_idx, 'return_pct']
            positive_months = (year_data['return_pct'] > 0).sum()
            negative_months = (year_data['return_pct'] <= 0).sum()
            total_months = len(year_data)
            win_rate = (positive_months / total_months * 100) if total_months > 0 else 0

            ws.cell(row=row, column=1, value=year).font = Font(bold=True)

            cell = ws.cell(row=row, column=2, value=round(year_return, 2))
            cell.number_format = '0.00"%"'
            if year_return > 0:
                cell.fill = self.positive_fill
            elif year_return < 0:
                cell.fill = self.negative_fill

            ws.cell(row=row, column=3, value=best_month)
            cell = ws.cell(row=row, column=4, value=round(best_pct, 2))
            cell.number_format = '0.00"%"'
            cell.fill = self.positive_fill

            ws.cell(row=row, column=5, value=worst_month)
            cell = ws.cell(row=row, column=6, value=round(worst_pct, 2))
            cell.number_format = '0.00"%"'
            cell.fill = self.negative_fill

            ws.cell(row=row, column=7, value=positive_months)
            ws.cell(row=row, column=8, value=negative_months)

            cell = ws.cell(row=row, column=9, value=round(win_rate, 1))
            cell.number_format = '0.0"%"'

            row += 1

        # Overall summary row
        if years:
            row += 1
            ws.cell(row=row, column=1, value="OVERALL").font = Font(bold=True, size=12)
            total_return = monthly_df['return_pct'].sum()
            cell = ws.cell(row=row, column=2, value=round(total_return, 2))
            cell.number_format = '0.00"%"'
            cell.font = Font(bold=True)

            overall_positive = (monthly_df['return_pct'] > 0).sum()
            overall_negative = (monthly_df['return_pct'] <= 0).sum()
            overall_win_rate = (overall_positive / len(monthly_df) * 100) if len(monthly_df) > 0 else 0

            ws.cell(row=row, column=7, value=overall_positive).font = Font(bold=True)
            ws.cell(row=row, column=8, value=overall_negative).font = Font(bold=True)
            cell = ws.cell(row=row, column=9, value=round(overall_win_rate, 1))
            cell.number_format = '0.0"%"'
            cell.font = Font(bold=True)

        # Column widths
        ws.column_dimensions['A'].width = 12
        for col_idx in range(2, len(years) + 3):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
        for col_idx in range(len(years) + 3, len(years) + 10):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # ==================== HELPER METHODS ====================

    def _add_section_header(self, ws, row: int, title: str) -> int:
        """Add a section header and return next row."""
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = self.subsection_font
        ws[f'A{row}'].fill = self.light_fill
        ws.merge_cells(f'A{row}:D{row}')
        return row + 2

    def _add_metrics_table(self, ws, row: int, data: List[Tuple[str, str]]) -> int:
        """Add a metrics table and return next row."""
        for metric, value in data:
            ws.cell(row=row, column=1, value=metric).border = self.thin_border
            ws.cell(row=row, column=2, value=value).border = self.thin_border
            row += 1
        return row
