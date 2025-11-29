"""
Batch backtest summary report generation with correlation analysis.
"""
import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference
from openpyxl.formatting.rule import ColorScaleRule

from ..Engine.backtest_result import BacktestResult
from .performance_metrics import PerformanceMetrics


class BatchSummaryReportGenerator:
    """
    Generates comprehensive summary reports for batch backtests.

    Analyzes multiple securities together including:
    - Correlation analysis between securities
    - Comparative performance metrics
    - Portfolio-level statistics
    - Risk/return scatter plots
    - Relative performance visualizations
    """

    def __init__(self, output_directory: Path, initial_capital: float = 100000.0):
        """
        Initialize batch summary report generator.

        Args:
            output_directory: Directory to save summary reports
            initial_capital: Starting capital for calculations
        """
        self.output_directory = Path(output_directory)
        self.output_directory.mkdir(parents=True, exist_ok=True)
        self.initial_capital = initial_capital

        # Styling
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.section_font = Font(bold=True, size=12)
        self.metric_font = Font(size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate_batch_summary(self, results: Dict[str, BacktestResult],
                               backtest_name: str) -> Path:
        """
        Generate comprehensive batch summary report.

        Args:
            results: Dictionary mapping symbol to BacktestResult
            backtest_name: Name of the batch backtest

        Returns:
            Path to generated Excel file
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{backtest_name}_BATCH_SUMMARY_{timestamp}.xlsx"
        filepath = self.output_directory / filename

        # Calculate metrics for all securities
        all_metrics = self._calculate_batch_metrics(results)

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create all sheets
        self._create_overview_dashboard(wb, results, all_metrics, backtest_name)
        self._create_performance_comparison(wb, results, all_metrics)
        self._create_correlation_analysis(wb, results, all_metrics)
        self._create_risk_analysis(wb, results, all_metrics)
        self._create_visualizations(wb, results, all_metrics)

        # Save workbook
        wb.save(filepath)

        return filepath

    def _calculate_batch_metrics(self, results: Dict[str, BacktestResult]) -> Dict[str, Any]:
        """
        Calculate comprehensive batch-level metrics.

        Args:
            results: Dictionary of backtest results

        Returns:
            Dictionary containing batch metrics
        """
        batch_metrics = {
            'individual_metrics': {},
            'correlations': None,
            'return_correlation_matrix': None,
            'equity_correlation_matrix': None,
            'daily_returns': {},
        }

        # Calculate individual metrics for each security
        for symbol, result in results.items():
            metrics = PerformanceMetrics.calculate_metrics(result)
            batch_metrics['individual_metrics'][symbol] = metrics

        # Calculate correlation matrices
        batch_metrics['correlations'] = self._calculate_correlations(results)

        # Calculate aggregate statistics
        batch_metrics['aggregate'] = self._calculate_aggregate_stats(results)

        return batch_metrics

    def _calculate_correlations(self, results: Dict[str, BacktestResult]) -> Dict[str, Any]:
        """
        Calculate correlation matrices between securities.

        Includes:
        - Return correlation (based on equity curve changes)
        - Trade timing correlation
        - Drawdown correlation
        """
        correlations = {}

        # Align all equity curves by date
        equity_data = {}
        for symbol, result in results.items():
            if len(result.equity_curve) > 0:
                df = result.equity_curve.copy()
                df = df.set_index('date')
                equity_data[symbol] = df['equity']

        if len(equity_data) < 2:
            # Need at least 2 securities for correlation
            return {
                'equity_correlation': pd.DataFrame(),
                'return_correlation': pd.DataFrame(),
                'drawdown_correlation': pd.DataFrame(),
                'correlation_summary': "Insufficient data (need at least 2 securities)"
            }

        # Combine into single DataFrame (aligned by date)
        combined_equity = pd.DataFrame(equity_data)

        # Fill missing dates with forward fill (hold previous value)
        combined_equity = combined_equity.ffill()

        # Calculate daily returns
        combined_returns = combined_equity.pct_change().dropna()

        # Equity correlation
        equity_corr = combined_equity.corr()

        # Return correlation
        return_corr = combined_returns.corr()

        # Drawdown correlation
        drawdowns = {}
        for symbol in combined_equity.columns:
            equity = combined_equity[symbol].values
            running_max = np.maximum.accumulate(equity)
            dd_pct = (equity - running_max) / running_max * 100
            drawdowns[symbol] = dd_pct

        combined_dd = pd.DataFrame(drawdowns, index=combined_equity.index)
        dd_corr = combined_dd.corr()

        # Correlation summary statistics
        # Extract upper triangle of correlation matrix (excluding diagonal)
        mask = np.triu(np.ones_like(return_corr, dtype=bool), k=1)
        upper_triangle = return_corr.where(mask)
        correlations_list = upper_triangle.values.flatten()
        correlations_list = correlations_list[~np.isnan(correlations_list)]

        correlation_summary = {
            'avg_return_correlation': np.mean(correlations_list) if len(correlations_list) > 0 else 0,
            'max_return_correlation': np.max(correlations_list) if len(correlations_list) > 0 else 0,
            'min_return_correlation': np.min(correlations_list) if len(correlations_list) > 0 else 0,
            'num_highly_correlated': np.sum(correlations_list > 0.7) if len(correlations_list) > 0 else 0,
            'num_negatively_correlated': np.sum(correlations_list < -0.3) if len(correlations_list) > 0 else 0,
        }

        return {
            'equity_correlation': equity_corr,
            'return_correlation': return_corr,
            'drawdown_correlation': dd_corr,
            'correlation_summary': correlation_summary,
            'combined_returns': combined_returns,
        }

    def _calculate_aggregate_stats(self, results: Dict[str, BacktestResult]) -> Dict[str, Any]:
        """
        Calculate aggregate portfolio-level statistics.

        Simulates a portfolio that equally allocates across all securities.
        """
        stats = {}

        # Total P/L across all securities
        total_pl = sum(r.total_return for r in results.values())
        total_trades = sum(r.num_trades for r in results.values())

        stats['total_pl'] = total_pl
        stats['total_trades'] = total_trades
        stats['num_securities'] = len(results)
        stats['avg_pl_per_security'] = total_pl / len(results) if len(results) > 0 else 0
        stats['avg_return_pct'] = (stats['avg_pl_per_security'] / self.initial_capital) * 100

        # Count profitable securities
        profitable = sum(1 for r in results.values() if r.total_return > 0)
        stats['num_profitable'] = profitable
        stats['pct_profitable'] = (profitable / len(results) * 100) if len(results) > 0 else 0

        # Best and worst performers
        sorted_results = sorted(results.items(), key=lambda x: x[1].total_return, reverse=True)
        stats['best_symbol'] = sorted_results[0][0] if sorted_results else None
        stats['best_return'] = sorted_results[0][1].total_return if sorted_results else 0
        stats['best_return_pct'] = sorted_results[0][1].total_return_pct if sorted_results else 0
        stats['worst_symbol'] = sorted_results[-1][0] if sorted_results else None
        stats['worst_return'] = sorted_results[-1][1].total_return if sorted_results else 0
        stats['worst_return_pct'] = sorted_results[-1][1].total_return_pct if sorted_results else 0

        # Aggregate win rate
        all_trades = []
        for result in results.values():
            all_trades.extend(result.trades)

        if all_trades:
            winning = sum(1 for t in all_trades if t.is_winner)
            stats['aggregate_win_rate'] = (winning / len(all_trades)) * 100
        else:
            stats['aggregate_win_rate'] = 0

        # Portfolio diversity score (based on correlation)
        # Lower average correlation = higher diversity
        stats['diversity_note'] = "See Correlation Analysis sheet for details"

        return stats

    # ==================== SHEET CREATION ====================

    def _create_overview_dashboard(self, wb: Workbook, results: Dict[str, BacktestResult],
                                   metrics: Dict[str, Any], backtest_name: str):
        """Create Sheet 1: Overview Dashboard with aggregate statistics."""
        ws = wb.create_sheet("Overview Dashboard")

        row = 1

        # Title
        ws.merge_cells(f'A{row}:D{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = f"BATCH BACKTEST SUMMARY"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center')
        row += 1

        ws.merge_cells(f'A{row}:D{row}')
        subtitle = ws[f'A{row}']
        subtitle.value = backtest_name
        subtitle.font = Font(size=12)
        subtitle.alignment = Alignment(horizontal='center')
        row += 2

        # Batch info
        agg = metrics['aggregate']
        ws[f'A{row}'] = "Securities Tested:"
        ws[f'B{row}'] = agg['num_securities']
        row += 1
        ws[f'A{row}'] = "Total Trades:"
        ws[f'B{row}'] = agg['total_trades']
        row += 1
        ws[f'A{row}'] = "Initial Capital per Security:"
        ws[f'B{row}'] = self.initial_capital
        ws[f'B{row}'].number_format = '$#,##0.00'
        row += 2

        # Aggregate Performance
        row = self._add_section_header(ws, row, "AGGREGATE PERFORMANCE")

        perf_data = [
            ("Total P/L (All Securities)", agg['total_pl'], "currency"),
            ("Average P/L per Security", agg['avg_pl_per_security'], "currency"),
            ("Average Return %", agg['avg_return_pct'], "percentage"),
            ("Profitable Securities", agg['num_profitable'], "number"),
            ("% Profitable", agg['pct_profitable'], "percentage"),
            ("Aggregate Win Rate", agg['aggregate_win_rate'], "percentage"),
        ]

        row = self._add_metrics_table(ws, row, perf_data)
        row += 1

        # Best and Worst Performers
        row = self._add_section_header(ws, row, "BEST & WORST PERFORMERS")

        performers_data = [
            ("Best Performer", agg['best_symbol'], "text"),
            ("  Return ($)", agg['best_return'], "currency"),
            ("  Return (%)", agg['best_return_pct'], "percentage"),
            ("", "", "text"),
            ("Worst Performer", agg['worst_symbol'], "text"),
            ("  Return ($)", agg['worst_return'], "currency"),
            ("  Return (%)", agg['worst_return_pct'], "percentage"),
        ]

        row = self._add_metrics_table(ws, row, performers_data)
        row += 2

        # Individual Security Summary Table
        row = self._add_section_header(ws, row, "INDIVIDUAL SECURITY SUMMARY")

        # Create summary table
        headers = ['Symbol', 'Trades', 'Win Rate %', 'Total Return $', 'Return %',
                   'Sharpe', 'Max DD %', 'Profit Factor']

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')

        row += 1

        # Sort by total return
        sorted_results = sorted(results.items(),
                              key=lambda x: x[1].total_return,
                              reverse=True)

        for symbol, result in sorted_results:
            m = metrics['individual_metrics'][symbol]

            ws.cell(row=row, column=1, value=symbol)
            ws.cell(row=row, column=2, value=m['num_trades'])

            cell = ws.cell(row=row, column=3, value=m['win_rate'] * 100)
            cell.number_format = '0.0"%"'

            cell = ws.cell(row=row, column=4, value=m['total_return'])
            cell.number_format = '$#,##0.00'
            if m['total_return'] < 0:
                cell.font = Font(color="FF0000")
            elif m['total_return'] > 0:
                cell.font = Font(color="00B050")

            cell = ws.cell(row=row, column=5, value=m['total_return_pct'])
            cell.number_format = '0.00"%"'
            if m['total_return_pct'] < 0:
                cell.font = Font(color="FF0000")
            elif m['total_return_pct'] > 0:
                cell.font = Font(color="00B050")

            cell = ws.cell(row=row, column=6, value=m['sharpe_ratio'])
            cell.number_format = '0.00'

            cell = ws.cell(row=row, column=7, value=m['max_drawdown_pct'])
            cell.number_format = '0.00"%"'

            cell = ws.cell(row=row, column=8, value=m['profit_factor'])
            cell.number_format = '0.00'

            row += 1

        # Format columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 14

    def _create_performance_comparison(self, wb: Workbook, results: Dict[str, BacktestResult],
                                      metrics: Dict[str, Any]):
        """Create Sheet 2: Detailed Performance Comparison."""
        ws = wb.create_sheet("Performance Comparison")

        row = 1
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "DETAILED PERFORMANCE COMPARISON"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2

        # Comprehensive comparison table
        headers = ['Symbol', 'Total Return %', 'CAGR %', 'Sharpe', 'Sortino',
                   'Max DD %', 'Win Rate %', 'Avg Trade Days', 'Volatility %',
                   'Profit Factor', 'Num Trades']

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')

        row += 1
        data_start_row = row

        # Sort by Sharpe ratio for this view
        sorted_symbols = sorted(metrics['individual_metrics'].keys(),
                              key=lambda s: metrics['individual_metrics'][s]['sharpe_ratio'],
                              reverse=True)

        for symbol in sorted_symbols:
            m = metrics['individual_metrics'][symbol]

            col = 1
            ws.cell(row=row, column=col, value=symbol)
            col += 1

            # Format each metric appropriately
            metric_values = [
                (m['total_return_pct'], '0.00"%"'),
                (m.get('cagr', 0), '0.00"%"'),
                (m['sharpe_ratio'], '0.00'),
                (m.get('sortino_ratio', 0), '0.00'),
                (m['max_drawdown_pct'], '0.00"%"'),
                (m['win_rate'] * 100, '0.0"%"'),
                (m['avg_trade_duration'], '0.0'),
                (m.get('volatility', 0), '0.00"%"'),
                (m['profit_factor'], '0.00'),
                (m['num_trades'], '0'),
            ]

            for value, fmt in metric_values:
                cell = ws.cell(row=row, column=col, value=value)
                cell.number_format = fmt
                col += 1

            row += 1

        data_end_row = row - 1

        # Add summary statistics row
        row += 1
        ws.cell(row=row, column=1, value="AVERAGE").font = Font(bold=True)

        # Calculate averages for each column
        for col_idx in range(2, len(headers) + 1):
            formula = f"=AVERAGE({ws.cell(row=data_start_row, column=col_idx).coordinate}:{ws.cell(row=data_end_row, column=col_idx).coordinate})"
            cell = ws.cell(row=row, column=col_idx, value=formula)
            cell.font = Font(bold=True)
            # Copy format from above
            cell.number_format = ws.cell(row=data_start_row, column=col_idx).number_format

        # Apply color scale to Total Return % column
        return_col_letter = ws.cell(row=data_start_row, column=2).column_letter
        ws.conditional_formatting.add(
            f'{return_col_letter}{data_start_row}:{return_col_letter}{data_end_row}',
            ColorScaleRule(
                start_type='min', start_color='F8696B',
                mid_type='percentile', mid_value=50, mid_color='FFEB84',
                end_type='max', end_color='63BE7B'
            )
        )

        # Auto-fit columns
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 14

    def _create_correlation_analysis(self, wb: Workbook, results: Dict[str, BacktestResult],
                                     metrics: Dict[str, Any]):
        """Create Sheet 3: Correlation Analysis."""
        ws = wb.create_sheet("Correlation Analysis")

        row = 1
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = "CORRELATION ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2

        corr_data = metrics['correlations']

        if isinstance(corr_data.get('correlation_summary'), str):
            # Insufficient data
            ws[f'A{row}'] = corr_data['correlation_summary']
            return

        # Correlation Summary Statistics
        row = self._add_section_header(ws, row, "CORRELATION SUMMARY STATISTICS")

        corr_summary = corr_data['correlation_summary']
        summary_data = [
            ("Average Return Correlation", corr_summary['avg_return_correlation'], "decimal"),
            ("Maximum Correlation", corr_summary['max_return_correlation'], "decimal"),
            ("Minimum Correlation", corr_summary['min_return_correlation'], "decimal"),
            ("Highly Correlated Pairs (>0.7)", corr_summary['num_highly_correlated'], "number"),
            ("Negatively Correlated Pairs (<-0.3)", corr_summary['num_negatively_correlated'], "number"),
        ]

        row = self._add_metrics_table(ws, row, summary_data)
        row += 2

        # Interpretation
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "Correlation Interpretation:"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        row += 1

        interpretations = [
            "• Low correlation (<0.3): Securities move independently - good diversification",
            "• Medium correlation (0.3-0.7): Some relationship - moderate diversification",
            "• High correlation (>0.7): Securities move together - limited diversification benefit",
            "• Negative correlation (<0): Securities move opposite - potential hedging opportunities"
        ]

        for interp in interpretations:
            ws[f'A{row}'] = interp
            ws[f'A{row}'].font = Font(size=9)
            row += 1

        row += 1

        # Return Correlation Matrix
        row = self._add_section_header(ws, row, "RETURN CORRELATION MATRIX")

        return_corr = corr_data['return_correlation']

        if not return_corr.empty:
            # Write matrix
            matrix_start_row = row

            # Header row (symbols)
            ws.cell(row=row, column=1, value="").fill = self.header_fill
            for col_idx, symbol in enumerate(return_corr.columns, 2):
                cell = ws.cell(row=row, column=col_idx, value=symbol)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')

            row += 1

            # Data rows
            for row_symbol in return_corr.index:
                ws.cell(row=row, column=1, value=row_symbol).font = Font(bold=True)

                for col_idx, col_symbol in enumerate(return_corr.columns, 2):
                    value = return_corr.loc[row_symbol, col_symbol]
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.number_format = '0.00'

                    # Color code correlations
                    if row_symbol != col_symbol:  # Don't color diagonal
                        if value > 0.7:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        elif value < -0.3:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif abs(value) < 0.3:
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

                row += 1

            row += 1

        # Drawdown Correlation Matrix
        row = self._add_section_header(ws, row, "DRAWDOWN CORRELATION MATRIX")
        ws[f'A{row}'] = "(Shows if securities tend to have drawdowns at the same time)"
        ws[f'A{row}'].font = Font(italic=True, size=9)
        row += 1

        dd_corr = corr_data['drawdown_correlation']

        if not dd_corr.empty:
            # Header row
            ws.cell(row=row, column=1, value="").fill = self.header_fill
            for col_idx, symbol in enumerate(dd_corr.columns, 2):
                cell = ws.cell(row=row, column=col_idx, value=symbol)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')

            row += 1

            # Data rows
            for row_symbol in dd_corr.index:
                ws.cell(row=row, column=1, value=row_symbol).font = Font(bold=True)

                for col_idx, col_symbol in enumerate(dd_corr.columns, 2):
                    value = dd_corr.loc[row_symbol, col_symbol]
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.number_format = '0.00'

                row += 1

        # Format columns
        for col_idx in range(1, len(return_corr.columns) + 2):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 12

    def _create_risk_analysis(self, wb: Workbook, results: Dict[str, BacktestResult],
                             metrics: Dict[str, Any]):
        """Create Sheet 4: Risk Analysis."""
        ws = wb.create_sheet("Risk Analysis")

        row = 1
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "RISK ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2

        # Risk comparison table
        headers = ['Symbol', 'Max Drawdown %', 'Volatility %', 'Downside Dev %',
                   'Worst Day %', 'Sharpe', 'Sortino', 'Calmar']

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')

        row += 1

        # Sort by max drawdown (risk)
        sorted_symbols = sorted(metrics['individual_metrics'].keys(),
                              key=lambda s: metrics['individual_metrics'][s]['max_drawdown_pct'])

        for symbol in sorted_symbols:
            m = metrics['individual_metrics'][symbol]

            col = 1
            ws.cell(row=row, column=col, value=symbol)
            col += 1

            risk_values = [
                (m['max_drawdown_pct'], '0.00"%"'),
                (m.get('volatility', 0), '0.00"%"'),
                (m.get('downside_deviation', 0), '0.00"%"'),
                (m.get('worst_day', 0), '0.00"%"'),
                (m['sharpe_ratio'], '0.00'),
                (m.get('sortino_ratio', 0), '0.00'),
                (m.get('calmar_ratio', 0), '0.00'),
            ]

            for value, fmt in risk_values:
                cell = ws.cell(row=row, column=col, value=value)
                cell.number_format = fmt
                col += 1

            row += 1

        row += 2

        # Risk-Return Analysis
        row = self._add_section_header(ws, row, "RISK-RETURN PROFILE")

        ws.cell(row=row, column=1, value="Symbol").font = Font(bold=True)
        ws.cell(row=row, column=2, value="Return %").font = Font(bold=True)
        ws.cell(row=row, column=3, value="Risk (Volatility %)").font = Font(bold=True)
        ws.cell(row=row, column=4, value="Risk-Adjusted Return").font = Font(bold=True)
        row += 1

        for symbol in sorted_symbols:
            m = metrics['individual_metrics'][symbol]

            ws.cell(row=row, column=1, value=symbol)

            cell = ws.cell(row=row, column=2, value=m['total_return_pct'])
            cell.number_format = '0.00"%"'

            vol = m.get('volatility', 0.01)  # Avoid division by zero
            cell = ws.cell(row=row, column=3, value=vol)
            cell.number_format = '0.00"%"'

            risk_adj = m['total_return_pct'] / vol if vol > 0 else 0
            cell = ws.cell(row=row, column=4, value=risk_adj)
            cell.number_format = '0.00'

            row += 1

        # Format columns
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 16

    def _create_visualizations(self, wb: Workbook, results: Dict[str, BacktestResult],
                              metrics: Dict[str, Any]):
        """Create Sheet 5: Visualizations."""
        ws = wb.create_sheet("Visualizations")

        ws['A1'] = "COMPARATIVE VISUALIZATIONS"
        ws['A1'].font = Font(bold=True, size=14)

        # Prepare data for charts
        row = 4
        ws.cell(row=row, column=1, value="Symbol").font = self.header_font
        ws.cell(row=row, column=2, value="Total Return %").font = self.header_font
        ws.cell(row=row, column=3, value="Sharpe Ratio").font = self.header_font
        ws.cell(row=row, column=4, value="Max Drawdown %").font = self.header_font
        ws.cell(row=row, column=5, value="Win Rate %").font = self.header_font

        row += 1
        chart_data_start = row

        # Sort by return for chart
        sorted_symbols = sorted(metrics['individual_metrics'].keys(),
                              key=lambda s: metrics['individual_metrics'][s]['total_return_pct'],
                              reverse=True)

        for symbol in sorted_symbols:
            m = metrics['individual_metrics'][symbol]

            ws.cell(row=row, column=1, value=symbol)
            ws.cell(row=row, column=2, value=m['total_return_pct'])
            ws.cell(row=row, column=3, value=m['sharpe_ratio'])
            ws.cell(row=row, column=4, value=m['max_drawdown_pct'])
            ws.cell(row=row, column=5, value=m['win_rate'] * 100)

            row += 1

        chart_data_end = row - 1

        # Create Return Comparison Bar Chart
        chart1 = BarChart()
        chart1.title = "Total Return % Comparison"
        chart1.y_axis.title = 'Return %'
        chart1.x_axis.title = 'Security'

        data = Reference(ws, min_col=2, min_row=chart_data_start-1, max_row=chart_data_end)
        cats = Reference(ws, min_col=1, min_row=chart_data_start, max_row=chart_data_end)

        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.height = 12
        chart1.width = 20

        ws.add_chart(chart1, "G4")

        # Create Sharpe Ratio Bar Chart
        chart2 = BarChart()
        chart2.title = "Sharpe Ratio Comparison"
        chart2.y_axis.title = 'Sharpe Ratio'
        chart2.x_axis.title = 'Security'

        data2 = Reference(ws, min_col=3, min_row=chart_data_start-1, max_row=chart_data_end)

        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats)
        chart2.height = 12
        chart2.width = 20

        ws.add_chart(chart2, "G25")

        # Create Risk-Return Scatter Plot
        chart3 = ScatterChart()
        chart3.title = "Risk-Return Profile"
        chart3.x_axis.title = 'Max Drawdown %'
        chart3.y_axis.title = 'Total Return %'

        # For scatter, need X and Y data
        xvalues = Reference(ws, min_col=4, min_row=chart_data_start, max_row=chart_data_end)
        yvalues = Reference(ws, min_col=2, min_row=chart_data_start, max_row=chart_data_end)

        series = chart3.series.append(yvalues)
        chart3.series[-1].xvalues = xvalues
        chart3.height = 12
        chart3.width = 20

        ws.add_chart(chart3, "G46")

    # ==================== HELPER METHODS ====================

    def _add_section_header(self, ws, row: int, title: str) -> int:
        """Add a section header and return next row."""
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = title
        cell.font = self.section_font
        cell.fill = self.section_fill
        cell.alignment = Alignment(horizontal='left')
        return row + 1

    def _add_metrics_table(self, ws, row: int, data: List[tuple]) -> int:
        """
        Add a metrics table.

        Args:
            data: List of tuples (label, value, format_type)
        """
        for label, value, format_type in data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = self.metric_font

            value_cell = ws[f'B{row}']
            value_cell.value = value

            if format_type == 'currency':
                value_cell.number_format = '$#,##0.00'
            elif format_type == 'percentage':
                value_cell.number_format = '0.00"%"'
            elif format_type == 'decimal':
                value_cell.number_format = '0.00'
            elif format_type == 'number':
                value_cell.number_format = '#,##0'
            # 'text' - no special format

            row += 1

        return row
