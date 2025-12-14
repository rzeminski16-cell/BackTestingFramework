"""
Excel report generation for backtest results.

Supports both standard and enhanced reports with advanced visualizations.
"""
import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference, AreaChart
from openpyxl.chart.marker import Marker
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

from ..Engine.backtest_result import BacktestResult
from ..Models.trade import Trade

# Check if enhanced reports are available
try:
    from ..Analysis.enhanced_visualizations import EnhancedVisualizations, MATPLOTLIB_AVAILABLE
    ENHANCED_REPORTS_AVAILABLE = MATPLOTLIB_AVAILABLE
except ImportError:
    ENHANCED_REPORTS_AVAILABLE = False


class ExcelReportGenerator:
    """
    Generates comprehensive Excel reports for backtests.

    Creates a multi-sheet Excel workbook with:
    - Sheet 1: Summary Dashboard
    - Sheet 2: Trade Log
    - Sheet 3: Performance Metrics & Analysis
    - Sheet 4: Visualizations & Charts
    - Sheet 5: Market Condition Breakdown (optional)
    """

    def __init__(self, output_directory: Path, initial_capital: float = 100000.0,
                 risk_free_rate: float = 0.035, benchmark_name: str = "S&P 500",
                 use_enhanced: bool = False):
        """
        Initialize Excel report generator.

        Args:
            output_directory: Directory to save reports
            initial_capital: Starting capital for calculations
            risk_free_rate: Annual risk-free rate (default 3.5%)
            benchmark_name: Name of benchmark for comparison
            use_enhanced: If True, include enhanced matplotlib visualizations
        """
        self.output_directory = Path(output_directory)
        self.output_directory.mkdir(parents=True, exist_ok=True)
        self.initial_capital = initial_capital
        self.risk_free_rate = risk_free_rate
        self.benchmark_name = benchmark_name
        self.use_enhanced = use_enhanced and ENHANCED_REPORTS_AVAILABLE

        # Initialize enhanced visualization generator if available
        self._viz = None
        if self.use_enhanced:
            try:
                self._viz = EnhancedVisualizations()
            except Exception:
                self.use_enhanced = False

        # Styling
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.section_font = Font(bold=True, size=12)
        self.metric_font = Font(size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    @staticmethod
    def is_enhanced_available() -> bool:
        """Check if enhanced reports are available."""
        return ENHANCED_REPORTS_AVAILABLE

    def generate_report(self, result: BacktestResult, filename: Optional[str] = None) -> Path:
        """
        Generate comprehensive Excel report for backtest result.

        Args:
            result: Backtest result
            filename: Optional custom filename

        Returns:
            Path to generated Excel file
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{result.strategy_name}_{result.symbol}_{timestamp}.xlsx"

        filepath = self.output_directory / filename

        # Calculate comprehensive metrics
        metrics = self._calculate_all_metrics(result)

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create all sheets
        self._create_summary_dashboard(wb, result, metrics)
        self._create_random_trades(wb, result, metrics)
        self._create_trade_log(wb, result, metrics)
        self._create_period_analysis(wb, result, metrics)
        self._create_costs_analysis(wb, result, metrics)
        self._create_performance_analysis(wb, result, metrics)
        self._create_visualizations(wb, result, metrics)
        self._create_market_conditions(wb, result, metrics)

        # Add enhanced visualizations if available
        if self.use_enhanced and self._viz:
            self._create_enhanced_visualizations(wb, result, metrics)

        # Save workbook
        wb.save(filepath)

        return filepath

    def _create_enhanced_visualizations(self, wb: Workbook, result: BacktestResult, metrics: Dict[str, Any]):
        """Create enhanced visualizations sheet with matplotlib charts."""
        from openpyxl.drawing.image import Image

        ws = wb.create_sheet("Enhanced Charts")

        # Title
        ws['A1'] = "ENHANCED VISUALIZATIONS"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:G1')

        row = 3

        trades = result.trades
        equity_df = result.equity_curve

        # 1. Equity curve with drawdown overlay
        try:
            equity_img = self._viz.create_equity_curve_with_drawdown(equity_df)
            img = Image(equity_img)
            img.width = 700
            img.height = 400
            ws.add_image(img, f'A{row}')
            row += 24
        except Exception:
            pass

        # 2. Monthly returns heatmap
        try:
            heatmap_img = self._viz.create_monthly_returns_heatmap(equity_df)
            img = Image(heatmap_img)
            img.width = 600
            img.height = 350
            ws.add_image(img, f'A{row}')
            row += 20
        except Exception:
            pass

        # 3. Trade distribution histogram
        if trades:
            try:
                dist_img = self._viz.create_trade_distribution_histogram(trades)
                img = Image(dist_img)
                img.width = 500
                img.height = 300
                ws.add_image(img, f'A{row}')
                row += 18
            except Exception:
                pass

        # 4. Win/Loss streak visualization
        if trades:
            try:
                streak_img = self._viz.create_streak_visualization(trades)
                img = Image(streak_img)
                img.width = 700
                img.height = 450
                ws.add_image(img, f'A{row}')
                row += 26
            except Exception:
                pass

        # 5. Rolling metrics chart
        try:
            rolling_img = self._viz.create_rolling_metrics_chart(equity_df)
            img = Image(rolling_img)
            img.width = 700
            img.height = 400
            ws.add_image(img, f'A{row}')
            row += 24
        except Exception:
            pass

        # 6. Trade clustering analysis
        if trades:
            try:
                cluster_img = self._viz.create_trade_clustering_analysis(trades)
                img = Image(cluster_img)
                img.width = 700
                img.height = 500
                ws.add_image(img, f'A{row}')
            except Exception:
                pass

    def _calculate_all_metrics(self, result: BacktestResult) -> Dict[str, Any]:
        """
        Calculate all metrics needed for the report.

        Args:
            result: Backtest result

        Returns:
            Dictionary containing all calculated metrics
        """
        metrics = {}
        trades = result.trades
        equity_curve = result.equity_curve.copy()

        # Basic information
        metrics['symbol'] = result.symbol
        metrics['strategy_name'] = result.strategy_name
        metrics['initial_capital'] = self.initial_capital
        metrics['final_capital'] = result.final_equity
        metrics['num_trades'] = len(trades)

        # Date range
        if len(equity_curve) > 0:
            metrics['start_date'] = equity_curve['date'].iloc[0]
            metrics['end_date'] = equity_curve['date'].iloc[-1]
            metrics['total_days'] = (metrics['end_date'] - metrics['start_date']).days
            metrics['years'] = metrics['total_days'] / 365.25
        else:
            metrics['start_date'] = None
            metrics['end_date'] = None
            metrics['total_days'] = 0
            metrics['years'] = 0

        if len(trades) == 0:
            return self._get_empty_metrics(metrics)

        # A. Overall Performance Metrics
        metrics['total_return'] = result.total_return
        metrics['total_return_pct'] = result.total_return_pct
        metrics['cagr'] = self._calculate_cagr(self.initial_capital, result.final_equity, metrics['years'])
        metrics['annualized_return'] = metrics['cagr']  # Same as CAGR

        # Win/Loss statistics
        winning_trades = [t for t in trades if t.is_winner]
        losing_trades = [t for t in trades if not t.is_winner]

        metrics['num_wins'] = len(winning_trades)
        metrics['num_losses'] = len(losing_trades)
        metrics['win_rate'] = metrics['num_wins'] / metrics['num_trades'] if metrics['num_trades'] > 0 else 0
        metrics['loss_rate'] = 1 - metrics['win_rate']

        # B. Risk Metrics
        drawdown_metrics = self._calculate_drawdown_metrics(equity_curve)
        metrics.update(drawdown_metrics)

        metrics['volatility'] = self._calculate_volatility(equity_curve)
        metrics['downside_deviation'] = self._calculate_downside_deviation(equity_curve)
        metrics['best_day'] = self._calculate_best_day(equity_curve)
        metrics['worst_day'] = self._calculate_worst_day(equity_curve)

        # C. Risk-Adjusted Performance Ratios
        metrics['sharpe_ratio'] = self._calculate_sharpe_ratio(equity_curve)
        metrics['sortino_ratio'] = self._calculate_sortino_ratio(equity_curve)
        metrics['calmar_ratio'] = self._calculate_calmar_ratio(metrics['cagr'], metrics['max_drawdown_pct'])
        metrics['recovery_factor'] = self._calculate_recovery_factor(metrics['total_return'], metrics['max_drawdown'])
        metrics['profit_factor'] = self._calculate_profit_factor(winning_trades, losing_trades)

        # D. Trade Quality Metrics
        wins = [t.pl for t in winning_trades]
        losses = [t.pl for t in losing_trades]

        metrics['avg_win'] = np.mean(wins) if wins else 0
        metrics['avg_loss'] = np.mean(losses) if losses else 0
        metrics['largest_win'] = max(wins) if wins else 0
        metrics['largest_loss'] = min(losses) if losses else 0
        metrics['risk_reward_ratio'] = abs(metrics['avg_win'] / metrics['avg_loss']) if metrics['avg_loss'] != 0 else 0

        # Win/loss streaks
        streaks = self._calculate_streaks(trades)
        metrics['max_win_streak'] = streaks['max_win_streak']
        metrics['max_loss_streak'] = streaks['max_loss_streak']

        # Trade durations
        durations = [t.duration_days for t in trades]
        metrics['avg_trade_duration'] = np.mean(durations)
        metrics['median_trade_duration'] = np.median(durations)

        # E. Consistency Metrics
        consistency = self._calculate_consistency_metrics(equity_curve, trades)
        metrics.update(consistency)

        # F. FX P&L Metrics (if multi-currency trading)
        fx_metrics = self._calculate_fx_metrics(trades)
        metrics.update(fx_metrics)

        # G. Additional metrics
        metrics['strategy_exposure'] = self._calculate_exposure(equity_curve, trades)

        # Monthly/Quarterly returns
        metrics['monthly_returns'] = self._calculate_period_returns(equity_curve, 'M')
        metrics['quarterly_returns'] = self._calculate_period_returns(equity_curve, 'Q')

        # Trade distribution
        metrics['return_distribution'] = self._calculate_return_distribution(trades)

        # Period analysis (year-by-year and quarter-by-quarter)
        metrics['yearly_analysis'] = self._calculate_period_analysis(equity_curve, trades, 'Y')
        metrics['quarterly_analysis'] = self._calculate_period_analysis(equity_curve, trades, 'Q')

        # Costs analysis
        metrics['costs_analysis'] = self._calculate_costs_analysis(trades, metrics['total_return'])

        # Drawdown recovery analysis
        metrics['drawdown_recovery'] = self._calculate_drawdown_recovery(equity_curve)

        # Annual risk-adjusted ratios
        metrics['annual_ratios'] = self._calculate_annual_ratios(equity_curve, trades)

        return metrics

    def _get_empty_metrics(self, base_metrics: Dict[str, Any]) -> Dict[str, Any]:
        """Return metrics dictionary with zero values for no-trade scenario."""
        base_metrics.update({
            'total_return': 0, 'total_return_pct': 0, 'cagr': 0, 'annualized_return': 0,
            'num_wins': 0, 'num_losses': 0, 'win_rate': 0, 'loss_rate': 0,
            'max_drawdown': 0, 'max_drawdown_pct': 0, 'avg_drawdown': 0,
            'max_drawdown_duration': 0, 'volatility': 0, 'downside_deviation': 0,
            'best_day': 0, 'worst_day': 0, 'sharpe_ratio': 0, 'sortino_ratio': 0,
            'calmar_ratio': 0, 'recovery_factor': 0, 'profit_factor': 0,
            'avg_win': 0, 'avg_loss': 0, 'largest_win': 0, 'largest_loss': 0,
            'risk_reward_ratio': 0, 'max_win_streak': 0, 'max_loss_streak': 0,
            'avg_trade_duration': 0, 'median_trade_duration': 0,
            'strategy_exposure': 0, 'monthly_returns': pd.DataFrame(),
            'quarterly_returns': pd.DataFrame(), 'return_distribution': {}
        })
        return base_metrics

    # ==================== METRIC CALCULATIONS ====================

    def _calculate_cagr(self, initial: float, final: float, years: float) -> float:
        """Calculate Compound Annual Growth Rate."""
        if years <= 0 or initial <= 0:
            return 0.0
        return (pow(final / initial, 1 / years) - 1) * 100

    def _calculate_drawdown_metrics(self, equity_curve: pd.DataFrame) -> Dict[str, float]:
        """Calculate comprehensive drawdown metrics."""
        if len(equity_curve) == 0:
            return {
                'max_drawdown': 0, 'max_drawdown_pct': 0, 'avg_drawdown': 0,
                'max_drawdown_duration': 0, 'num_drawdowns_over_5pct': 0
            }

        equity = equity_curve['equity'].values
        dates = equity_curve['date'].values

        # Filter out NaN and invalid values
        if np.any(np.isnan(equity)) or np.any(np.isinf(equity)):
            equity = np.nan_to_num(equity, nan=0.0, posinf=0.0, neginf=0.0)

        # Ensure we have valid data
        if len(equity) == 0 or np.all(equity <= 0):
            return {
                'max_drawdown': 0, 'max_drawdown_pct': 0, 'avg_drawdown': 0,
                'max_drawdown_duration': 0, 'num_drawdowns_over_5pct': 0
            }

        # Running maximum
        running_max = np.maximum.accumulate(equity)

        # Drawdown in dollars and percent - with safe division
        with np.errstate(divide='ignore', invalid='ignore'):
            drawdown = running_max - equity
            drawdown_pct = np.where(running_max > 0, (drawdown / running_max) * 100, 0.0)

        # Remove any NaN or inf values and cap at 100%
        drawdown = np.nan_to_num(drawdown, nan=0.0, posinf=0.0, neginf=0.0)
        drawdown_pct = np.nan_to_num(drawdown_pct, nan=0.0, posinf=0.0, neginf=0.0)
        drawdown_pct = np.clip(drawdown_pct, 0, 100)

        max_dd = np.max(drawdown) if len(drawdown) > 0 else 0.0
        max_dd_pct = np.max(drawdown_pct) if len(drawdown_pct) > 0 else 0.0

        # Average drawdown (only count periods in drawdown)
        in_drawdown = drawdown > 0
        avg_dd = np.mean(drawdown[in_drawdown]) if np.any(in_drawdown) else 0

        # Max drawdown duration
        max_dd_duration = self._calculate_max_drawdown_duration(equity, dates)

        # Count drawdowns over 5%
        num_dd_over_5 = np.sum(drawdown_pct > 5)

        return {
            'max_drawdown': max_dd,
            'max_drawdown_pct': max_dd_pct,
            'avg_drawdown': avg_dd,
            'max_drawdown_duration': max_dd_duration,
            'num_drawdowns_over_5pct': num_dd_over_5
        }

    def _calculate_max_drawdown_duration(self, equity: np.ndarray, dates: np.ndarray) -> int:
        """Calculate maximum drawdown duration in days."""
        running_max = np.maximum.accumulate(equity)
        in_drawdown = equity < running_max

        if not np.any(in_drawdown):
            return 0

        # Find drawdown periods
        dd_start = None
        max_duration = 0

        for i in range(len(equity)):
            if in_drawdown[i] and dd_start is None:
                dd_start = dates[i]
            elif not in_drawdown[i] and dd_start is not None:
                duration = (dates[i] - dd_start).days if hasattr(dates[i], 'days') else 0
                max_duration = max(max_duration, duration)
                dd_start = None

        # Handle if still in drawdown at end
        if dd_start is not None:
            duration = (dates[-1] - dd_start).days if hasattr(dates[-1], 'days') else 0
            max_duration = max(max_duration, duration)

        return max_duration

    def _calculate_volatility(self, equity_curve: pd.DataFrame) -> float:
        """Calculate annualized volatility."""
        if len(equity_curve) < 2:
            return 0.0

        returns = equity_curve['equity'].pct_change().dropna()
        if len(returns) == 0:
            return 0.0

        return returns.std() * np.sqrt(252) * 100  # Annualized, as percentage

    def _calculate_downside_deviation(self, equity_curve: pd.DataFrame) -> float:
        """Calculate downside deviation (volatility of negative returns)."""
        if len(equity_curve) < 2:
            return 0.0

        returns = equity_curve['equity'].pct_change().dropna()
        negative_returns = returns[returns < 0]

        if len(negative_returns) == 0:
            return 0.0

        return negative_returns.std() * np.sqrt(252) * 100  # Annualized

    def _calculate_best_day(self, equity_curve: pd.DataFrame) -> float:
        """Calculate best single day return."""
        if len(equity_curve) < 2:
            return 0.0

        returns = equity_curve['equity'].pct_change().dropna()
        return returns.max() * 100 if len(returns) > 0 else 0.0

    def _calculate_worst_day(self, equity_curve: pd.DataFrame) -> float:
        """Calculate worst single day return."""
        if len(equity_curve) < 2:
            return 0.0

        returns = equity_curve['equity'].pct_change().dropna()
        return returns.min() * 100 if len(returns) > 0 else 0.0

    def _calculate_sharpe_ratio(self, equity_curve: pd.DataFrame) -> float:
        """Calculate annualized Sharpe ratio."""
        if len(equity_curve) < 2:
            return 0.0

        returns = equity_curve['equity'].pct_change().dropna()

        if len(returns) == 0 or returns.std() == 0:
            return 0.0

        daily_rf = pow(1 + self.risk_free_rate, 1/252) - 1
        excess_returns = returns - daily_rf

        sharpe = (excess_returns.mean() / excess_returns.std()) * np.sqrt(252)
        return sharpe

    def _calculate_sortino_ratio(self, equity_curve: pd.DataFrame) -> float:
        """Calculate Sortino ratio (uses downside deviation)."""
        if len(equity_curve) < 2:
            return 0.0

        returns = equity_curve['equity'].pct_change().dropna()

        if len(returns) == 0:
            return 0.0

        daily_rf = pow(1 + self.risk_free_rate, 1/252) - 1
        excess_returns = returns - daily_rf

        # Downside returns
        downside_returns = returns[returns < 0]

        if len(downside_returns) == 0 or downside_returns.std() == 0:
            return 0.0

        sortino = (excess_returns.mean() / downside_returns.std()) * np.sqrt(252)
        return sortino

    def _calculate_calmar_ratio(self, cagr: float, max_dd_pct: float) -> float:
        """Calculate Calmar ratio (CAGR / Max Drawdown)."""
        if max_dd_pct == 0:
            return 0.0
        return cagr / max_dd_pct

    def _calculate_recovery_factor(self, total_return: float, max_dd: float) -> float:
        """Calculate recovery factor (Total Profit / Max Drawdown)."""
        if max_dd == 0:
            return 0.0
        return total_return / max_dd

    def _calculate_profit_factor(self, winning_trades: List[Trade], losing_trades: List[Trade]) -> float:
        """Calculate profit factor (Gross Profit / Gross Loss)."""
        gross_profit = sum(t.pl for t in winning_trades)
        gross_loss = abs(sum(t.pl for t in losing_trades))

        if gross_loss == 0:
            return 0.0

        return gross_profit / gross_loss

    def _calculate_streaks(self, trades: List[Trade]) -> Dict[str, int]:
        """Calculate win and loss streaks."""
        if len(trades) == 0:
            return {'max_win_streak': 0, 'max_loss_streak': 0}

        max_win_streak = 0
        max_loss_streak = 0
        current_win_streak = 0
        current_loss_streak = 0

        for trade in trades:
            if trade.is_winner:
                current_win_streak += 1
                current_loss_streak = 0
                max_win_streak = max(max_win_streak, current_win_streak)
            else:
                current_loss_streak += 1
                current_win_streak = 0
                max_loss_streak = max(max_loss_streak, current_loss_streak)

        return {
            'max_win_streak': max_win_streak,
            'max_loss_streak': max_loss_streak
        }

    def _calculate_consistency_metrics(self, equity_curve: pd.DataFrame, trades: List[Trade]) -> Dict[str, Any]:
        """Calculate consistency metrics."""
        metrics = {}

        if len(equity_curve) < 2:
            return {
                'profitable_months_pct': 0, 'profitable_weeks_pct': 0,
                'monthly_consistency': 0, 'consecutive_losing_months': 0
            }

        # Calculate monthly returns
        equity_df = equity_curve.copy()
        equity_df.set_index('date', inplace=True)

        # Resample to month-end
        monthly = equity_df['equity'].resample('ME').last()
        monthly_returns = monthly.pct_change().dropna()

        # Profitable months
        profitable_months = (monthly_returns > 0).sum()
        total_months = len(monthly_returns)
        metrics['profitable_months_pct'] = (profitable_months / total_months * 100) if total_months > 0 else 0

        # Weekly returns
        weekly = equity_df['equity'].resample('W').last()
        weekly_returns = weekly.pct_change().dropna()
        profitable_weeks = (weekly_returns > 0).sum()
        total_weeks = len(weekly_returns)
        metrics['profitable_weeks_pct'] = (profitable_weeks / total_weeks * 100) if total_weeks > 0 else 0

        # Monthly consistency (std dev of monthly returns)
        metrics['monthly_consistency'] = monthly_returns.std() * 100 if len(monthly_returns) > 0 else 0

        # Consecutive losing months
        max_consecutive_losing = 0
        current_consecutive_losing = 0

        for ret in monthly_returns:
            if ret < 0:
                current_consecutive_losing += 1
                max_consecutive_losing = max(max_consecutive_losing, current_consecutive_losing)
            else:
                current_consecutive_losing = 0

        metrics['consecutive_losing_months'] = max_consecutive_losing

        return metrics

    def _calculate_fx_metrics(self, trades: List[Trade]) -> Dict[str, Any]:
        """Calculate FX P&L metrics for multi-currency trading."""
        fx_metrics = {
            'total_fx_pl': 0.0,
            'total_security_pl': 0.0,
            'fx_contribution_pct': 0.0,
            'avg_fx_pl': 0.0,
            'largest_fx_gain': 0.0,
            'largest_fx_loss': 0.0,
            'num_fx_positive': 0,
            'num_fx_negative': 0,
            'num_fx_neutral': 0,
        }

        if len(trades) == 0:
            return fx_metrics

        # Extract FX P&L data from trades
        fx_pls = []
        security_pls = []

        for trade in trades:
            fx_pl = getattr(trade, 'fx_pl', 0.0)
            security_pl = getattr(trade, 'security_pl', trade.pl)

            fx_pls.append(fx_pl)
            security_pls.append(security_pl)

            if fx_pl > 0.01:
                fx_metrics['num_fx_positive'] += 1
            elif fx_pl < -0.01:
                fx_metrics['num_fx_negative'] += 1
            else:
                fx_metrics['num_fx_neutral'] += 1

        fx_metrics['total_fx_pl'] = sum(fx_pls)
        fx_metrics['total_security_pl'] = sum(security_pls)

        total_return = sum(trade.pl for trade in trades)
        if abs(total_return) > 0.01:
            fx_metrics['fx_contribution_pct'] = (fx_metrics['total_fx_pl'] / total_return) * 100

        fx_metrics['avg_fx_pl'] = np.mean(fx_pls) if fx_pls else 0.0
        fx_metrics['largest_fx_gain'] = max(fx_pls) if fx_pls else 0.0
        fx_metrics['largest_fx_loss'] = min(fx_pls) if fx_pls else 0.0

        return fx_metrics

    def _calculate_exposure(self, equity_curve: pd.DataFrame, trades: List[Trade]) -> float:
        """Calculate percentage of time in market."""
        if len(equity_curve) == 0:
            return 0.0

        # Count days with open positions
        equity_df = equity_curve.copy()
        days_in_market = (equity_df['position_value'] > 0).sum()
        total_days = len(equity_df)

        return (days_in_market / total_days * 100) if total_days > 0 else 0.0

    def _calculate_period_returns(self, equity_curve: pd.DataFrame, period: str) -> pd.DataFrame:
        """
        Calculate returns by period (monthly 'M' or quarterly 'Q').

        Returns DataFrame with columns: period, return_pct, cumulative_return
        """
        if len(equity_curve) < 2:
            return pd.DataFrame(columns=['period', 'return_pct', 'cumulative_return'])

        equity_df = equity_curve.copy()
        equity_df.set_index('date', inplace=True)

        # Resample to period end
        period_equity = equity_df['equity'].resample(period).last()
        period_returns = period_equity.pct_change().dropna()

        # Create result DataFrame
        result = pd.DataFrame({
            'period': period_returns.index,
            'return_pct': period_returns.values * 100,
            'cumulative_return': (1 + period_returns).cumprod().values * 100 - 100
        })

        return result

    def _calculate_return_distribution(self, trades: List[Trade]) -> Dict[str, int]:
        """
        Calculate distribution of returns in bins.

        Returns dictionary with bins as keys and counts as values.
        """
        if len(trades) == 0:
            return {}

        returns = [t.pl_pct for t in trades]

        # Define bins
        bins = [-100, -20, -10, -5, 0, 5, 10, 20, 100]
        labels = ['< -20%', '-20% to -10%', '-10% to -5%', '-5% to 0%',
                  '0% to 5%', '5% to 10%', '10% to 20%', '> 20%']

        # Count returns in each bin
        hist, _ = np.histogram(returns, bins=bins)

        distribution = dict(zip(labels, hist))

        return distribution

    def _calculate_period_analysis(self, equity_curve: pd.DataFrame, trades: List[Trade], period: str) -> pd.DataFrame:
        """
        Calculate period-by-period analysis (year or quarter).

        Args:
            equity_curve: Equity curve DataFrame
            trades: List of trades
            period: 'Y' for yearly or 'Q' for quarterly

        Returns:
            DataFrame with columns: period, win_rate, num_trades, max_dd, avg_dd, expectancy, pl, sharpe, sortino, calmar
        """
        if len(equity_curve) < 2:
            return pd.DataFrame()

        equity_df = equity_curve.copy()
        equity_df.set_index('date', inplace=True)

        # Get date range
        start_date = equity_df.index.min()
        end_date = equity_df.index.max()

        # Create all periods in the date range
        if period == 'Y':
            periods = pd.period_range(start=start_date, end=end_date, freq='Y')
        else:  # Q
            periods = pd.period_range(start=start_date, end=end_date, freq='Q')

        # Group trades by period
        trades_by_period = {}
        if len(trades) > 0:
            for trade in trades:
                trade_date = pd.Timestamp(trade.exit_date)

                if period == 'Y':
                    period_key = trade_date.year
                else:  # Q
                    period_key = f"{trade_date.year}-Q{trade_date.quarter}"

                if period_key not in trades_by_period:
                    trades_by_period[period_key] = []

                trades_by_period[period_key].append({
                    'pl': trade.pl,
                    'is_winner': trade.is_winner,
                    'commission': trade.commission_paid
                })

        results = []

        for period_obj in periods:
            # Calculate period label and key
            if period == 'Q':
                period_label = f"{period_obj.year}-Q{period_obj.quarter}"
                period_key = period_label
                quarter = period_obj.quarter
                period_start = pd.Timestamp(period_obj.year, ((quarter-1)*3)+1, 1)
                # Last day of quarter
                if quarter == 4:
                    period_end = pd.Timestamp(period_obj.year, 12, 31)
                else:
                    period_end = pd.Timestamp(period_obj.year, quarter*3 + 1, 1) - pd.Timedelta(days=1)
            else:
                period_label = str(period_obj.year)
                period_key = period_obj.year
                period_start = pd.Timestamp(period_obj.year, 1, 1)
                period_end = pd.Timestamp(period_obj.year, 12, 31)

            # Get equity curve for this period
            period_equity = equity_df.loc[(equity_df.index >= period_start) & (equity_df.index <= period_end)]

            # Get trades for this period (if any)
            period_trades_list = trades_by_period.get(period_key, [])

            # Calculate trade-based metrics
            if len(period_trades_list) > 0:
                num_trades = len(period_trades_list)
                winners = [t for t in period_trades_list if t['is_winner']]
                win_rate = len(winners) / num_trades if num_trades > 0 else 0
                period_pl = sum(t['pl'] for t in period_trades_list)

                # Expectancy
                avg_win = np.mean([t['pl'] for t in winners]) if len(winners) > 0 else 0
                losers = [t for t in period_trades_list if not t['is_winner']]
                avg_loss = np.mean([t['pl'] for t in losers]) if len(losers) > 0 else 0
                expectancy = (win_rate * avg_win) + ((1 - win_rate) * avg_loss)
            else:
                # No trades in this period - set to N/A
                num_trades = 0
                win_rate = None  # Will be displayed as N/A
                expectancy = None  # Will be displayed as N/A

            # Calculate equity-based metrics (always calculate if equity data exists)
            if len(period_equity) > 0:
                period_equity_values = period_equity['equity'].values

                # Calculate P/L change from equity curve
                period_pl = period_equity_values[-1] - period_equity_values[0] if len(period_trades_list) == 0 else period_pl

                # Drawdown for period (reset at period start)
                running_max = np.maximum.accumulate(period_equity_values)
                drawdown = running_max - period_equity_values
                drawdown_pct = (drawdown / running_max) * 100
                max_dd = np.max(drawdown_pct)
                avg_dd = np.mean(drawdown_pct[drawdown_pct > 0]) if np.any(drawdown_pct > 0) else 0

                # Sharpe ratio for period
                returns = period_equity['equity'].pct_change().dropna()
                if len(returns) > 1 and returns.std() > 0:
                    daily_rf = pow(1 + self.risk_free_rate, 1/252) - 1
                    excess_returns = returns - daily_rf
                    sharpe = (excess_returns.mean() / excess_returns.std()) * np.sqrt(252)
                else:
                    sharpe = 0

                # Sortino ratio for period
                downside_returns = returns[returns < 0]
                if len(downside_returns) > 0 and downside_returns.std() > 0:
                    sortino = (excess_returns.mean() / downside_returns.std()) * np.sqrt(252)
                else:
                    sortino = 0

                # Calmar ratio for period (annualized return / max DD)
                period_days = (period_end - period_start).days
                period_return = (period_equity_values[-1] / period_equity_values[0] - 1) * 100 if period_equity_values[0] > 0 else 0
                annualized_return = period_return * (365.25 / period_days) if period_days > 0 else 0
                calmar = annualized_return / max_dd if max_dd > 0 else 0
            else:
                # No equity data for this period
                period_pl = None
                max_dd = None
                avg_dd = None
                sharpe = None
                sortino = None
                calmar = None

            results.append({
                'period': period_label,
                'num_trades': num_trades,
                'win_rate': win_rate * 100 if win_rate is not None else None,
                'pl': period_pl,
                'max_dd_pct': max_dd,
                'avg_dd_pct': avg_dd,
                'expectancy': expectancy,
                'sharpe': sharpe,
                'sortino': sortino,
                'calmar': calmar
            })

        return pd.DataFrame(results)

    def _calculate_costs_analysis(self, trades: List[Trade], total_return: float) -> Dict[str, Any]:
        """
        Calculate comprehensive costs analysis.

        Args:
            trades: List of trades
            total_return: Total return from all trades

        Returns:
            Dictionary with cost metrics
        """
        if len(trades) == 0:
            return {
                'total_commission': 0,
                'total_slippage': 0,
                'total_costs': 0,
                'avg_commission_per_trade': 0,
                'avg_slippage_per_trade': 0,
                'avg_costs_per_trade': 0,
                'costs_pct_of_pl': 0,
                'pl_before_costs': 0,
                'pl_after_costs': 0
            }

        total_commission = sum(t.commission_paid for t in trades)

        # Calculate slippage if available in metadata
        total_slippage = 0
        for t in trades:
            if hasattr(t, 'metadata') and isinstance(t.metadata, dict):
                slippage = t.metadata.get('slippage', 0)
                total_slippage += slippage

        total_costs = total_commission + abs(total_slippage)
        avg_commission = total_commission / len(trades)
        avg_slippage = total_slippage / len(trades)
        avg_costs = total_costs / len(trades)

        pl_before_costs = total_return + total_costs
        pl_after_costs = total_return

        costs_pct = (total_costs / abs(pl_before_costs)) * 100 if pl_before_costs != 0 else 0

        return {
            'total_commission': total_commission,
            'total_slippage': abs(total_slippage),
            'total_costs': total_costs,
            'avg_commission_per_trade': avg_commission,
            'avg_slippage_per_trade': abs(avg_slippage),
            'avg_costs_per_trade': avg_costs,
            'costs_pct_of_pl': costs_pct,
            'pl_before_costs': pl_before_costs,
            'pl_after_costs': pl_after_costs
        }

    def _calculate_drawdown_recovery(self, equity_curve: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Calculate drawdown periods and recovery times.

        Returns:
            List of drawdown periods with start, end, recovery dates, and durations
        """
        if len(equity_curve) < 2:
            return []

        equity = equity_curve['equity'].values
        dates = pd.to_datetime(equity_curve['date'])  # Convert to pandas datetime for proper arithmetic

        running_max = np.maximum.accumulate(equity)
        drawdown = running_max - equity
        drawdown_pct = (drawdown / running_max) * 100

        # Find drawdown periods
        in_drawdown = drawdown_pct > 0

        drawdown_periods = []
        dd_start = None
        dd_start_idx = None
        peak_value = None

        for i in range(len(equity)):
            if in_drawdown[i] and dd_start is None:
                # Start of drawdown
                dd_start = dates.iloc[i-1] if i > 0 else dates.iloc[i]
                dd_start_idx = i-1 if i > 0 else i
                peak_value = running_max[i]
            elif not in_drawdown[i] and dd_start is not None:
                # End of drawdown (recovery)
                dd_end_idx = i - 1
                dd_end = dates.iloc[dd_end_idx]
                recovery_date = dates.iloc[i]

                max_dd_in_period = np.max(drawdown_pct[dd_start_idx:i])

                # Calculate durations using pandas datetime objects
                dd_duration = (dd_end - dd_start).days
                recovery_duration = (recovery_date - dd_start).days

                drawdown_periods.append({
                    'start_date': dd_start.strftime('%Y-%m-%d'),
                    'trough_date': dd_end.strftime('%Y-%m-%d'),
                    'recovery_date': recovery_date.strftime('%Y-%m-%d'),
                    'peak_value': peak_value,
                    'trough_value': equity[dd_end_idx],
                    'max_dd_pct': max_dd_in_period,
                    'drawdown_duration_days': dd_duration,
                    'recovery_duration_days': recovery_duration
                })

                dd_start = None
                dd_start_idx = None
                peak_value = None

        # Handle if still in drawdown at end
        if dd_start is not None:
            dd_end_idx = len(equity) - 1
            dd_end = dates.iloc[dd_end_idx]
            max_dd_in_period = np.max(drawdown_pct[dd_start_idx:])
            dd_duration = (dd_end - dd_start).days

            drawdown_periods.append({
                'start_date': dd_start.strftime('%Y-%m-%d'),
                'trough_date': dd_end.strftime('%Y-%m-%d'),
                'recovery_date': None,  # Not recovered yet
                'peak_value': peak_value,
                'trough_value': equity[dd_end_idx],
                'max_dd_pct': max_dd_in_period,
                'drawdown_duration_days': dd_duration,
                'recovery_duration_days': None  # Not recovered
            })

        # Sort by max drawdown
        drawdown_periods.sort(key=lambda x: x['max_dd_pct'], reverse=True)

        return drawdown_periods

    def _calculate_annual_ratios(self, equity_curve: pd.DataFrame, trades: List[Trade]) -> pd.DataFrame:
        """
        Calculate Sharpe, Sortino, and Calmar ratios on a year-by-year basis.

        Returns:
            DataFrame with columns: year, sharpe, sortino, calmar, max_dd_pct, annual_return_pct
        """
        if len(equity_curve) < 2:
            return pd.DataFrame()

        equity_df = equity_curve.copy()
        equity_df.set_index('date', inplace=True)

        # Group by year
        years = equity_df.index.year.unique()

        results = []

        for year in years:
            year_data = equity_df[equity_df.index.year == year]

            if len(year_data) < 2:
                continue

            # Calculate returns
            returns = year_data['equity'].pct_change().dropna()

            if len(returns) == 0:
                continue

            # Sharpe
            if returns.std() > 0:
                daily_rf = pow(1 + self.risk_free_rate, 1/252) - 1
                excess_returns = returns - daily_rf
                sharpe = (excess_returns.mean() / excess_returns.std()) * np.sqrt(252)
            else:
                sharpe = 0

            # Sortino
            downside_returns = returns[returns < 0]
            if len(downside_returns) > 0 and downside_returns.std() > 0:
                sortino = (excess_returns.mean() / downside_returns.std()) * np.sqrt(252)
            else:
                sortino = 0

            # Max drawdown for year
            year_equity = year_data['equity'].values
            running_max = np.maximum.accumulate(year_equity)
            drawdown_pct = ((year_equity - running_max) / running_max) * 100
            max_dd_pct = abs(np.min(drawdown_pct))

            # Annual return
            annual_return_pct = ((year_equity[-1] / year_equity[0]) - 1) * 100 if year_equity[0] > 0 else 0

            # Calmar
            calmar = annual_return_pct / max_dd_pct if max_dd_pct > 0 else 0

            results.append({
                'year': year,
                'sharpe': sharpe,
                'sortino': sortino,
                'calmar': calmar,
                'max_dd_pct': max_dd_pct,
                'annual_return_pct': annual_return_pct
            })

        return pd.DataFrame(results)

    # ==================== SHEET CREATION ====================

    def _create_summary_dashboard(self, wb: Workbook, result: BacktestResult, metrics: Dict[str, Any]):
        """Create Sheet 1: Summary Dashboard."""
        ws = wb.create_sheet("Summary Dashboard")

        row = 1

        # Title
        ws.merge_cells(f'A{row}:D{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = f"BACKTEST SUMMARY DASHBOARD"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center')
        row += 1

        # Subtitle
        ws.merge_cells(f'A{row}:D{row}')
        subtitle = ws[f'A{row}']
        subtitle.value = f"{metrics['strategy_name']} - {metrics['symbol']}"
        subtitle.font = Font(size=12)
        subtitle.alignment = Alignment(horizontal='center')
        row += 2

        # Date range and basic info
        ws[f'A{row}'] = "Backtest Period:"
        ws[f'B{row}'] = f"{metrics['start_date'].strftime('%Y-%m-%d') if metrics['start_date'] else 'N/A'} to {metrics['end_date'].strftime('%Y-%m-%d') if metrics['end_date'] else 'N/A'}"
        row += 1
        ws[f'A{row}'] = "Total Days:"
        ws[f'B{row}'] = metrics['total_days']
        row += 2

        # Strategy Parameters Section
        if result.strategy_params:
            row = self._add_section_header(ws, row, "STRATEGY PARAMETERS")

            params_data = []
            for param_name, param_value in result.strategy_params.items():
                # Format parameter name (convert snake_case to Title Case)
                display_name = param_name.replace('_', ' ').title()
                # Keep the value as-is (will be formatted by _add_metrics_table)
                params_data.append((display_name, param_value, "auto"))

            row = self._add_metrics_table(ws, row, params_data)
            row += 2

        # A. Overall Performance Metrics
        row = self._add_section_header(ws, row, "A. OVERALL PERFORMANCE METRICS")

        performance_data = [
            ("Total Return ($)", metrics['total_return'], "currency"),
            ("Total Return (%)", metrics['total_return_pct'], "percentage"),
            ("CAGR (%)", metrics['cagr'], "percentage"),
            ("Annualized Return (%)", metrics['annualized_return'], "percentage"),
            ("Initial Capital", metrics['initial_capital'], "currency"),
            ("Final Capital", metrics['final_capital'], "currency"),
            ("Number of Trades", metrics['num_trades'], "number"),
            ("Win Rate", metrics['win_rate'] * 100, "percentage"),
            ("Loss Rate", metrics['loss_rate'] * 100, "percentage"),
        ]

        row = self._add_metrics_table(ws, row, performance_data)
        row += 1

        # B. Risk Metrics
        row = self._add_section_header(ws, row, "B. RISK METRICS")

        risk_data = [
            ("Maximum Drawdown ($)", metrics['max_drawdown'], "currency"),
            ("Maximum Drawdown (%)", metrics['max_drawdown_pct'], "percentage"),
            ("Average Drawdown ($)", metrics['avg_drawdown'], "currency"),
            ("Drawdown Duration (days)", metrics['max_drawdown_duration'], "number"),
            ("Drawdowns > 5%", metrics['num_drawdowns_over_5pct'], "number"),
            ("Volatility (annualized %)", metrics['volatility'], "percentage"),
            ("Downside Deviation (%)", metrics['downside_deviation'], "percentage"),
            ("Best Day (%)", metrics['best_day'], "percentage"),
            ("Worst Day (%)", metrics['worst_day'], "percentage"),
        ]

        row = self._add_metrics_table(ws, row, risk_data)
        row += 1

        # C. Risk-Adjusted Performance Ratios
        row = self._add_section_header(ws, row, "C. RISK-ADJUSTED PERFORMANCE RATIOS")

        # Add reference note
        ws[f'A{row}'] = f"Note: Risk-free rate used = {self.risk_free_rate*100:.1f}%"
        ws[f'A{row}'].font = Font(italic=True, size=9)
        row += 1

        ratios_data = [
            ("Sharpe Ratio", metrics['sharpe_ratio'], "decimal", 1.0, "<1: Poor, 1-2: Good, >2: Excellent"),
            ("Sortino Ratio", metrics['sortino_ratio'], "decimal", 1.0, "<1: Poor, 1-2: Good, >2: Excellent"),
            ("Calmar Ratio", metrics['calmar_ratio'], "decimal", 0.5, "<0.5: Poor, 0.5-1: Good, >1: Excellent"),
            ("Recovery Factor", metrics['recovery_factor'], "decimal", None, "Higher is better"),
            ("Profit Factor", metrics['profit_factor'], "decimal", 1.5, "<1: Losing, 1-1.5: Marginal, >1.5: Good"),
        ]

        row = self._add_metrics_with_targets(ws, row, ratios_data)
        row += 1

        # D. Trade Quality Metrics
        row = self._add_section_header(ws, row, "D. TRADE QUALITY METRICS")

        trade_quality_data = [
            ("Average Win ($)", metrics['avg_win'], "currency"),
            ("Average Loss ($)", metrics['avg_loss'], "currency"),
            ("Risk/Reward Ratio", metrics['risk_reward_ratio'], "decimal"),
            ("Largest Win ($)", metrics['largest_win'], "currency"),
            ("Largest Loss ($)", metrics['largest_loss'], "currency"),
            ("Max Win Streak", metrics['max_win_streak'], "number"),
            ("Max Loss Streak", metrics['max_loss_streak'], "number"),
            ("Average Trade Duration (days)", metrics['avg_trade_duration'], "decimal"),
            ("Median Trade Duration (days)", metrics['median_trade_duration'], "number"),
        ]

        row = self._add_metrics_table(ws, row, trade_quality_data)
        row += 1

        # E. Consistency Metrics
        row = self._add_section_header(ws, row, "E. CONSISTENCY METRICS")

        consistency_data = [
            ("Profitable Months (%)", metrics['profitable_months_pct'], "percentage"),
            ("Profitable Weeks (%)", metrics['profitable_weeks_pct'], "percentage"),
            ("Monthly Consistency (std dev %)", metrics['monthly_consistency'], "percentage"),
            ("Consecutive Losing Months (max)", metrics['consecutive_losing_months'], "number"),
            ("Strategy Exposure (%)", metrics['strategy_exposure'], "percentage"),
        ]

        row = self._add_metrics_table(ws, row, consistency_data)
        row += 1

        # F. FX P&L Breakdown (if applicable)
        if metrics.get('total_fx_pl') is not None and abs(metrics.get('total_fx_pl', 0)) > 0.01:
            row = self._add_section_header(ws, row, "F. FX P&L BREAKDOWN (Multi-Currency)")

            fx_data = [
                ("Total Security P/L ($)", metrics.get('total_security_pl', 0), "currency"),
                ("Total FX P/L ($)", metrics.get('total_fx_pl', 0), "currency"),
                ("FX Contribution (%)", metrics.get('fx_contribution_pct', 0), "percentage"),
                ("Avg FX P/L per Trade ($)", metrics.get('avg_fx_pl', 0), "currency"),
                ("Largest FX Gain ($)", metrics.get('largest_fx_gain', 0), "currency"),
                ("Largest FX Loss ($)", metrics.get('largest_fx_loss', 0), "currency"),
                ("Trades with FX Gain", metrics.get('num_fx_positive', 0), "number"),
                ("Trades with FX Loss", metrics.get('num_fx_negative', 0), "number"),
                ("Trades with No FX Impact", metrics.get('num_fx_neutral', 0), "number"),
            ]

            row = self._add_metrics_table(ws, row, fx_data)

        # Format columns
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30

    def _create_random_trades(self, wb: Workbook, result: BacktestResult, metrics: Dict[str, Any]):
        """Create Sheet 2: Random Trades Sample."""
        ws = wb.create_sheet("Random Trades Sample")

        row = 1

        # Title
        ws.merge_cells(f'A{row}:L{row}')
        ws[f'A{row}'] = "RANDOM TRADES SAMPLE (Manual Verification)"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2

        if len(result.trades) == 0:
            ws[f'A{row}'] = "No trades executed"
            return

        # Select up to 10 random trades
        import random
        num_samples = min(10, len(result.trades))
        sample_trades = random.sample(result.trades, num_samples)

        # Headers
        headers = ['Trade #', 'Entry Date', 'Entry Price', 'Exit Date', 'Exit Price',
                   'Quantity', 'Initial SL', 'Final SL', 'P/L', 'P/L %',
                   'Entry Equity', '% Equity Used']

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')

        row += 1

        # Write trade data
        for idx, trade in enumerate(sample_trades, 1):
            trade_num = result.trades.index(trade) + 1  # Get original trade number

            ws.cell(row=row, column=1, value=trade_num)

            # Entry date
            ws.cell(row=row, column=2, value=trade.entry_date.strftime('%Y-%m-%d %H:%M:%S') if hasattr(trade.entry_date, 'hour') else trade.entry_date.strftime('%Y-%m-%d'))

            # Entry price
            cell = ws.cell(row=row, column=3, value=trade.entry_price)
            cell.number_format = '$#,##0.00' if trade.entry_price > 100 else '$#,##0.0000'

            # Exit date
            ws.cell(row=row, column=4, value=trade.exit_date.strftime('%Y-%m-%d %H:%M:%S') if hasattr(trade.exit_date, 'hour') else trade.exit_date.strftime('%Y-%m-%d'))

            # Exit price
            cell = ws.cell(row=row, column=5, value=trade.exit_price)
            cell.number_format = '$#,##0.00' if trade.exit_price > 100 else '$#,##0.0000'

            # Quantity
            ws.cell(row=row, column=6, value=trade.quantity)

            # Initial SL
            initial_sl = trade.initial_stop_loss if trade.initial_stop_loss else 'N/A'
            if initial_sl != 'N/A':
                cell = ws.cell(row=row, column=7, value=initial_sl)
                cell.number_format = '$#,##0.00' if initial_sl > 100 else '$#,##0.0000'
            else:
                ws.cell(row=row, column=7, value=initial_sl)

            # Final SL
            final_sl = trade.final_stop_loss if trade.final_stop_loss else 'N/A'
            if final_sl != 'N/A':
                cell = ws.cell(row=row, column=8, value=final_sl)
                cell.number_format = '$#,##0.00' if final_sl > 100 else '$#,##0.0000'
            else:
                ws.cell(row=row, column=8, value=final_sl)

            # P/L
            cell = ws.cell(row=row, column=9, value=trade.pl)
            cell.number_format = '$#,##0.00'
            if trade.pl < 0:
                cell.font = Font(color="FF0000")
            elif trade.pl > 0:
                cell.font = Font(color="00B050")

            # P/L %
            cell = ws.cell(row=row, column=10, value=trade.pl_pct)
            cell.number_format = '0.00"%"'
            if trade.pl_pct < 0:
                cell.font = Font(color="FF0000")
            elif trade.pl_pct > 0:
                cell.font = Font(color="00B050")

            # Entry Equity
            cell = ws.cell(row=row, column=11, value=trade.entry_equity)
            cell.number_format = '$#,##0.00'

            # % Equity Used
            if trade.entry_equity > 0:
                position_value = trade.entry_price * trade.quantity
                equity_pct_used = (position_value / trade.entry_equity) * 100
            else:
                equity_pct_used = 0.0
            cell = ws.cell(row=row, column=12, value=equity_pct_used)
            cell.number_format = '0.00"%"'

            row += 1

        # Format columns
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15

        # Freeze header rows
        ws.freeze_panes = 'A4'

    def _create_trade_log(self, wb: Workbook, result: BacktestResult, metrics: Dict[str, Any]):
        """Create Sheet 2: Trade Log."""
        ws = wb.create_sheet("Trade Log")

        # Create DataFrame from trades
        if len(result.trades) == 0:
            ws['A1'] = "No trades executed"
            return

        trade_data = []
        cumulative_pl = 0

        for idx, trade in enumerate(result.trades, 1):
            cumulative_pl += trade.pl

            trade_data.append({
                'Trade #': idx,
                'Entry Date': trade.entry_date.strftime('%Y-%m-%d'),
                'Entry Time': trade.entry_date.strftime('%H:%M:%S') if hasattr(trade.entry_date, 'hour') else 'N/A',
                'Entry Price': trade.entry_price,
                'Entry Reason': trade.entry_reason,
                'Exit Date': trade.exit_date.strftime('%Y-%m-%d'),
                'Exit Price': trade.exit_price,
                'Exit Reason': trade.exit_reason,
                'Trade Type': trade.side,
                'Position Size': trade.quantity,
                'Security Currency': getattr(trade, 'security_currency', 'GBP'),
                'Entry FX Rate': getattr(trade, 'entry_fx_rate', 1.0),
                'Exit FX Rate': getattr(trade, 'exit_fx_rate', 1.0),
                'Gross P/L': trade.pl + trade.commission_paid,  # P/L before fees
                'Fees/Commission': trade.commission_paid,
                'Net P/L': trade.pl,
                'Security P/L': getattr(trade, 'security_pl', trade.pl),
                'FX P/L': getattr(trade, 'fx_pl', 0.0),
                'Return %': trade.pl_pct,
                'Trade Duration': trade.duration_days,
                'Cumulative P/L': cumulative_pl,
                'Initial Stop Loss': trade.initial_stop_loss if trade.initial_stop_loss else 'N/A',
                'Final Stop Loss': trade.final_stop_loss if trade.final_stop_loss else 'N/A',
                'Take Profit': trade.take_profit if trade.take_profit else 'N/A',
                'Partial Exits': trade.partial_exits,
            })

        df = pd.DataFrame(trade_data)

        # Write headers
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border

        # Write data
        for row_idx, row_data in enumerate(df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border

                # Format currency columns
                if headers[col_idx-1] in ['Entry Price', 'Exit Price', 'Gross P/L', 'Fees/Commission', 'Net P/L', 'Cumulative P/L']:
                    cell.number_format = '$#,##0.00'

                    # Color code P/L
                    if headers[col_idx-1] in ['Net P/L', 'Gross P/L', 'Cumulative P/L']:
                        if isinstance(value, (int, float)) and value < 0:
                            cell.font = Font(color="FF0000")  # Red
                        elif isinstance(value, (int, float)) and value > 0:
                            cell.font = Font(color="00B050")  # Green

                # Format percentage
                elif headers[col_idx-1] == 'Return %':
                    cell.number_format = '0.00"%"'
                    if isinstance(value, (int, float)) and value < 0:
                        cell.font = Font(color="FF0000")
                    elif isinstance(value, (int, float)) and value > 0:
                        cell.font = Font(color="00B050")

        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A2'

    def _create_period_analysis(self, wb: Workbook, result: BacktestResult, metrics: Dict[str, Any]):
        """Create Sheet: Period Analysis (Year-by-Year and Quarter-by-Quarter)."""
        ws = wb.create_sheet("Period Analysis")

        row = 1

        # Title
        ws.merge_cells(f'A{row}:J{row}')
        ws[f'A{row}'] = "PERIOD ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2

        # Year-by-Year Analysis
        row = self._add_section_header(ws, row, "YEAR-BY-YEAR BREAKDOWN")
        ws[f'A{row}'] = "Note: P/L resets to 0 at the start of each year for drawdown calculation"
        ws[f'A{row}'].font = Font(italic=True, size=9)
        row += 1

        yearly_df = metrics.get('yearly_analysis', pd.DataFrame())

        if not yearly_df.empty:
            # Headers
            headers = ['Year', 'Trades', 'Win Rate %', 'P/L', 'Max DD %', 'Avg DD %',
                       'Expectancy', 'Sharpe', 'Sortino', 'Calmar']

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')

            row += 1

            # Data rows
            for _, year_data in yearly_df.iterrows():
                ws.cell(row=row, column=1, value=year_data['period'])
                ws.cell(row=row, column=2, value=year_data['num_trades'])

                cell = ws.cell(row=row, column=3, value=year_data['win_rate'])
                cell.number_format = '0.0"%"'

                cell = ws.cell(row=row, column=4, value=year_data['pl'])
                cell.number_format = '$#,##0.00'
                if year_data['pl'] < 0:
                    cell.font = Font(color="FF0000")
                elif year_data['pl'] > 0:
                    cell.font = Font(color="00B050")

                cell = ws.cell(row=row, column=5, value=year_data['max_dd_pct'])
                cell.number_format = '0.00"%"'

                cell = ws.cell(row=row, column=6, value=year_data['avg_dd_pct'])
                cell.number_format = '0.00"%"'

                cell = ws.cell(row=row, column=7, value=year_data['expectancy'])
                cell.number_format = '$#,##0.00'

                cell = ws.cell(row=row, column=8, value=year_data['sharpe'])
                cell.number_format = '0.00'

                cell = ws.cell(row=row, column=9, value=year_data['sortino'])
                cell.number_format = '0.00'

                cell = ws.cell(row=row, column=10, value=year_data['calmar'])
                cell.number_format = '0.00'

                row += 1

            row += 2
        else:
            ws[f'A{row}'] = "Insufficient data for yearly analysis"
            row += 3

        # Quarter-by-Quarter Analysis
        row = self._add_section_header(ws, row, "QUARTER-BY-QUARTER BREAKDOWN")
        ws[f'A{row}'] = "Note: P/L resets to 0 at the start of each quarter for drawdown calculation"
        ws[f'A{row}'].font = Font(italic=True, size=9)
        row += 1

        quarterly_df = metrics.get('quarterly_analysis', pd.DataFrame())

        if not quarterly_df.empty:
            # Headers
            headers = ['Quarter', 'Trades', 'Win Rate %', 'P/L', 'Max DD %', 'Avg DD %',
                       'Expectancy', 'Sharpe', 'Sortino', 'Calmar']

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')

            row += 1

            # Data rows
            for _, quarter_data in quarterly_df.iterrows():
                ws.cell(row=row, column=1, value=quarter_data['period'])
                ws.cell(row=row, column=2, value=quarter_data['num_trades'])

                cell = ws.cell(row=row, column=3, value=quarter_data['win_rate'])
                cell.number_format = '0.0"%"'

                cell = ws.cell(row=row, column=4, value=quarter_data['pl'])
                cell.number_format = '$#,##0.00'
                if quarter_data['pl'] < 0:
                    cell.font = Font(color="FF0000")
                elif quarter_data['pl'] > 0:
                    cell.font = Font(color="00B050")

                cell = ws.cell(row=row, column=5, value=quarter_data['max_dd_pct'])
                cell.number_format = '0.00"%"'

                cell = ws.cell(row=row, column=6, value=quarter_data['avg_dd_pct'])
                cell.number_format = '0.00"%"'

                cell = ws.cell(row=row, column=7, value=quarter_data['expectancy'])
                cell.number_format = '$#,##0.00'

                cell = ws.cell(row=row, column=8, value=quarter_data['sharpe'])
                cell.number_format = '0.00'

                cell = ws.cell(row=row, column=9, value=quarter_data['sortino'])
                cell.number_format = '0.00'

                cell = ws.cell(row=row, column=10, value=quarter_data['calmar'])
                cell.number_format = '0.00'

                row += 1
        else:
            ws[f'A{row}'] = "Insufficient data for quarterly analysis"

        # Format columns
        for col_idx in range(1, 11):
            ws.column_dimensions[chr(64 + col_idx)].width = 14

    def _create_costs_analysis(self, wb: Workbook, result: BacktestResult, metrics: Dict[str, Any]):
        """Create Sheet: Costs Analysis."""
        ws = wb.create_sheet("Costs Analysis")

        row = 1

        # Title
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "COSTS ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2

        costs = metrics.get('costs_analysis', {})

        # Total Costs Breakdown
        row = self._add_section_header(ws, row, "TOTAL COSTS BREAKDOWN")

        costs_data = [
            ("Total Commission Paid", costs.get('total_commission', 0), "currency"),
            ("Total Slippage Cost", costs.get('total_slippage', 0), "currency"),
            ("Total Costs", costs.get('total_costs', 0), "currency"),
        ]

        row = self._add_metrics_table(ws, row, costs_data)
        row += 1

        # Per-Trade Averages
        row = self._add_section_header(ws, row, "PER-TRADE AVERAGES")

        avg_costs_data = [
            ("Avg Commission per Trade", costs.get('avg_commission_per_trade', 0), "currency"),
            ("Avg Slippage per Trade", costs.get('avg_slippage_per_trade', 0), "currency"),
            ("Avg Total Cost per Trade", costs.get('avg_costs_per_trade', 0), "currency"),
        ]

        row = self._add_metrics_table(ws, row, avg_costs_data)
        row += 1

        # Impact on P/L
        row = self._add_section_header(ws, row, "IMPACT ON P/L")

        impact_data = [
            ("P/L Before Costs", costs.get('pl_before_costs', 0), "currency"),
            ("Total Costs", costs.get('total_costs', 0), "currency"),
            ("P/L After Costs", costs.get('pl_after_costs', 0), "currency"),
            ("Costs as % of Gross P/L", costs.get('costs_pct_of_pl', 0), "percentage"),
        ]

        row = self._add_metrics_table(ws, row, impact_data)
        row += 2

        # Add interpretation
        ws[f'A{row}'] = "Interpretation:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        costs_pct = costs.get('costs_pct_of_pl', 0)
        if costs_pct < 5:
            interpretation = "✓ Costs are low (<5% of gross P/L) - Good cost efficiency"
            color = "00B050"
        elif costs_pct < 10:
            interpretation = "⚠ Costs are moderate (5-10% of gross P/L) - Acceptable"
            color = "FF8C00"
        else:
            interpretation = "✗ Costs are high (>10% of gross P/L) - Consider optimization"
            color = "FF0000"

        ws[f'A{row}'] = interpretation
        ws[f'A{row}'].font = Font(color=color, bold=True)
        row += 1

        # Format columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _create_performance_analysis(self, wb: Workbook, result: BacktestResult, metrics: Dict[str, Any]):
        """Create Sheet 3: Performance Metrics & Analysis."""
        ws = wb.create_sheet("Performance Analysis")

        row = 1

        # Title
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = "PERFORMANCE METRICS & ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2

        # A. Drawdown Analysis Table
        row = self._add_section_header(ws, row, "A. DRAWDOWN ANALYSIS")

        drawdown_data = [
            ("Max Drawdown ($)", metrics['max_drawdown'], "currency"),
            ("Max Drawdown (%)", metrics['max_drawdown_pct'], "percentage"),
            ("Average Drawdown ($)", metrics['avg_drawdown'], "currency"),
            ("Max Drawdown Duration (days)", metrics['max_drawdown_duration'], "number"),
            ("Number of Drawdowns > 5%", metrics['num_drawdowns_over_5pct'], "number"),
        ]

        row = self._add_metrics_table(ws, row, drawdown_data)
        row += 2

        # Drawdown Recovery Analysis
        row = self._add_section_header(ws, row, "MAJOR DRAWDOWN PERIODS & RECOVERY TIMES")

        drawdown_periods = metrics.get('drawdown_recovery', [])

        if drawdown_periods:
            # Show top 5 largest drawdowns
            headers = ['Peak Date', 'Trough Date', 'Recovery Date', 'Max DD %',
                       'DD Duration (days)', 'Recovery Time (days)', 'Status']

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center')

            row += 1

            for dd in drawdown_periods[:5]:  # Top 5
                ws.cell(row=row, column=1, value=dd['start_date'].strftime('%Y-%m-%d') if hasattr(dd['start_date'], 'strftime') else str(dd['start_date']))
                ws.cell(row=row, column=2, value=dd['trough_date'].strftime('%Y-%m-%d') if hasattr(dd['trough_date'], 'strftime') else str(dd['trough_date']))

                if dd['recovery_date']:
                    ws.cell(row=row, column=3, value=dd['recovery_date'].strftime('%Y-%m-%d') if hasattr(dd['recovery_date'], 'strftime') else str(dd['recovery_date']))
                else:
                    ws.cell(row=row, column=3, value='Not Recovered')

                cell = ws.cell(row=row, column=4, value=dd['max_dd_pct'])
                cell.number_format = '0.00"%"'

                ws.cell(row=row, column=5, value=dd['drawdown_duration_days'])

                if dd['recovery_duration_days']:
                    ws.cell(row=row, column=6, value=dd['recovery_duration_days'])
                else:
                    ws.cell(row=row, column=6, value='N/A')

                # Status
                if dd['recovery_date']:
                    status_cell = ws.cell(row=row, column=7, value='✓ Recovered')
                    status_cell.font = Font(color="00B050")
                else:
                    status_cell = ws.cell(row=row, column=7, value='⚠ Ongoing')
                    status_cell.font = Font(color="FF8C00")

                row += 1

            row += 1

            # Add interpretation
            ws[f'A{row}'] = "Note: Recovery time is measured from peak to full recovery (return to previous high)"
            ws[f'A{row}'].font = Font(italic=True, size=9)
            row += 2
        else:
            ws[f'A{row}'] = "No significant drawdown periods"
            row += 3

        # B. Monthly Returns Table
        row = self._add_section_header(ws, row, "B. MONTHLY RETURNS")

        monthly_df = metrics['monthly_returns']

        if len(monthly_df) > 0:
            # Add month/year columns
            monthly_df_copy = monthly_df.copy()
            monthly_df_copy['Year'] = monthly_df_copy['period'].dt.year
            monthly_df_copy['Month'] = monthly_df_copy['period'].dt.strftime('%b')

            # Pivot table: rows=months, columns=years
            pivot_data = monthly_df_copy.pivot_table(
                index='Month',
                columns='Year',
                values='return_pct',
                aggfunc='first'
            )

            # Write pivot table
            pivot_row = row
            ws.cell(row=pivot_row, column=1, value="Month")

            # Year headers
            for col_idx, year in enumerate(pivot_data.columns, 2):
                cell = ws.cell(row=pivot_row, column=col_idx, value=year)
                cell.font = self.header_font
                cell.fill = self.header_fill

            # Month rows
            month_order = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

            for month in month_order:
                if month in pivot_data.index:
                    pivot_row += 1
                    ws.cell(row=pivot_row, column=1, value=month).font = Font(bold=True)

                    for col_idx, year in enumerate(pivot_data.columns, 2):
                        value = pivot_data.loc[month, year] if not pd.isna(pivot_data.loc[month, year]) else None
                        if value is not None:
                            cell = ws.cell(row=pivot_row, column=col_idx, value=value)
                            cell.number_format = '0.00"%"'

                            # Color coding
                            if value > 0:
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                                cell.font = Font(color="006100")
                            elif value < 0:
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                                cell.font = Font(color="9C0006")

            row = pivot_row + 2
        else:
            ws[f'A{row}'] = "Insufficient data for monthly analysis"
            row += 2

        # C. Distribution Analysis
        row = self._add_section_header(ws, row, "C. TRADE RETURN DISTRIBUTION")

        distribution = metrics['return_distribution']

        if distribution:
            ws.cell(row=row, column=1, value="Return Range").font = self.header_font
            ws.cell(row=row, column=2, value="Number of Trades").font = self.header_font
            ws.cell(row=row, column=3, value="Percentage").font = self.header_font
            row += 1

            total_trades = sum(distribution.values())

            for bin_label, count in distribution.items():
                ws.cell(row=row, column=1, value=bin_label)
                ws.cell(row=row, column=2, value=count)
                pct = (count / total_trades * 100) if total_trades > 0 else 0
                cell = ws.cell(row=row, column=3, value=pct)
                cell.number_format = '0.00"%"'
                row += 1

        row += 1

        # Format columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

    def _create_visualizations(self, wb: Workbook, result: BacktestResult, metrics: Dict[str, Any]):
        """Create Sheet 4: Visualizations & Charts."""
        ws = wb.create_sheet("Visualizations")

        # Title
        ws['A1'] = "CHARTS & VISUALIZATIONS"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Prepare equity curve data
        equity_df = result.equity_curve.copy()

        if len(equity_df) == 0:
            ws['A3'] = "No data available for charts"
            return

        # Write equity curve data starting at row 4
        data_start_row = 4
        ws.cell(row=data_start_row, column=1, value="Date").font = self.header_font
        ws.cell(row=data_start_row, column=2, value="Equity").font = self.header_font
        ws.cell(row=data_start_row, column=3, value="Drawdown %").font = self.header_font

        # Calculate drawdown for chart
        equity_values = equity_df['equity'].values
        running_max = np.maximum.accumulate(equity_values)
        drawdown_pct = ((equity_values - running_max) / running_max * 100)

        for idx, (date, equity, dd) in enumerate(zip(equity_df['date'], equity_values, drawdown_pct)):
            row_num = data_start_row + 1 + idx
            ws.cell(row=row_num, column=1, value=date.strftime('%Y-%m-%d') if hasattr(date, 'strftime') else str(date))
            ws.cell(row=row_num, column=2, value=equity)
            ws.cell(row=row_num, column=3, value=dd)

        # Create Equity Curve Chart
        chart1 = LineChart()
        chart1.title = "Equity Curve"
        chart1.style = 13
        chart1.y_axis.title = 'Portfolio Value ($)'
        chart1.x_axis.title = 'Date'

        data_ref = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_start_row + len(equity_df))
        cats_ref = Reference(ws, min_col=1, min_row=data_start_row+1, max_row=data_start_row + len(equity_df))

        chart1.add_data(data_ref, titles_from_data=True)
        chart1.set_categories(cats_ref)
        chart1.height = 10
        chart1.width = 20

        ws.add_chart(chart1, "E4")

        # Create Drawdown Chart
        chart2 = AreaChart()
        chart2.title = "Drawdown %"
        chart2.style = 13
        chart2.y_axis.title = 'Drawdown %'
        chart2.x_axis.title = 'Date'

        dd_ref = Reference(ws, min_col=3, min_row=data_start_row, max_row=data_start_row + len(equity_df))

        chart2.add_data(dd_ref, titles_from_data=True)
        chart2.set_categories(cats_ref)
        chart2.height = 10
        chart2.width = 20

        ws.add_chart(chart2, "E25")

        # Create return distribution chart
        distribution = metrics['return_distribution']

        if distribution:
            # Write distribution data
            dist_start_row = data_start_row + len(equity_df) + 5
            ws.cell(row=dist_start_row, column=1, value="Return Range").font = self.header_font
            ws.cell(row=dist_start_row, column=2, value="Count").font = self.header_font

            for idx, (bin_label, count) in enumerate(distribution.items(), 1):
                ws.cell(row=dist_start_row + idx, column=1, value=bin_label)
                ws.cell(row=dist_start_row + idx, column=2, value=count)

            # Create bar chart
            chart3 = BarChart()
            chart3.title = "Trade Return Distribution"
            chart3.y_axis.title = 'Number of Trades'
            chart3.x_axis.title = 'Return Range'

            data_ref = Reference(ws, min_col=2, min_row=dist_start_row, max_row=dist_start_row + len(distribution))
            cats_ref = Reference(ws, min_col=1, min_row=dist_start_row+1, max_row=dist_start_row + len(distribution))

            chart3.add_data(data_ref, titles_from_data=True)
            chart3.set_categories(cats_ref)
            chart3.height = 10
            chart3.width = 15

            ws.add_chart(chart3, "E46")

        # Annual Sharpe/Sortino/Calmar Ratios Chart
        annual_ratios = metrics.get('annual_ratios', pd.DataFrame())

        if not annual_ratios.empty and len(annual_ratios) > 0:
            # Write annual ratios data
            annual_start_row = dist_start_row + len(distribution) + 10 if distribution else data_start_row + len(equity_df) + 15

            ws.cell(row=annual_start_row, column=1, value="Year").font = self.header_font
            ws.cell(row=annual_start_row, column=2, value="Sharpe").font = self.header_font
            ws.cell(row=annual_start_row, column=3, value="Sortino").font = self.header_font
            ws.cell(row=annual_start_row, column=4, value="Calmar").font = self.header_font

            for idx, row_data in enumerate(annual_ratios.to_dict('records')):
                row_num = annual_start_row + 1 + idx
                ws.cell(row=row_num, column=1, value=row_data['year'])
                ws.cell(row=row_num, column=2, value=row_data['sharpe'])
                ws.cell(row=row_num, column=3, value=row_data['sortino'])
                ws.cell(row=row_num, column=4, value=row_data['calmar'])

            # Only create chart if we have data
            if len(annual_ratios) > 0:
                # Create Annual Ratios Line Chart
                chart4 = LineChart()
                chart4.title = "Annual Risk-Adjusted Ratios"
                chart4.y_axis.title = 'Ratio Value'
                chart4.x_axis.title = 'Year'
                chart4.style = 13

                # Add all three series
                sharpe_ref = Reference(ws, min_col=2, min_row=annual_start_row, max_row=annual_start_row + len(annual_ratios))
                sortino_ref = Reference(ws, min_col=3, min_row=annual_start_row, max_row=annual_start_row + len(annual_ratios))
                calmar_ref = Reference(ws, min_col=4, min_row=annual_start_row, max_row=annual_start_row + len(annual_ratios))
                cats_ref = Reference(ws, min_col=1, min_row=annual_start_row+1, max_row=annual_start_row + len(annual_ratios))

                chart4.add_data(sharpe_ref, titles_from_data=True)
                chart4.add_data(sortino_ref, titles_from_data=True)
                chart4.add_data(calmar_ref, titles_from_data=True)
                chart4.set_categories(cats_ref)
                chart4.height = 12
                chart4.width = 20

                ws.add_chart(chart4, "E67")

        # Equity Usage Distribution Chart
        if len(result.trades) > 0:
            # Calculate equity usage % for each trade
            equity_usage_pcts = []
            for trade in result.trades:
                if trade.entry_equity > 0:
                    position_value = trade.entry_price * trade.quantity
                    equity_pct = (position_value / trade.entry_equity) * 100
                    equity_usage_pcts.append(equity_pct)

            if equity_usage_pcts:
                # Create bins for distribution
                bins = [0, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100]
                bin_labels = ['0-10%', '10-20%', '20-30%', '30-40%', '40-50%',
                             '50-60%', '60-70%', '70-80%', '80-90%', '90-100%']

                # Count trades in each bin
                bin_counts = [0] * len(bin_labels)
                for pct in equity_usage_pcts:
                    for i in range(len(bins) - 1):
                        if bins[i] <= pct < bins[i + 1]:
                            bin_counts[i] += 1
                            break
                    else:
                        # Handle values >= 100%
                        if pct >= bins[-1]:
                            bin_counts[-1] += 1

                # Write equity usage distribution data
                equity_usage_start_row = annual_start_row + len(annual_ratios) + 10 if not annual_ratios.empty else dist_start_row + len(distribution) + 20 if distribution else data_start_row + len(equity_df) + 25

                ws.cell(row=equity_usage_start_row, column=1, value="Equity Usage %").font = self.header_font
                ws.cell(row=equity_usage_start_row, column=2, value="Number of Trades").font = self.header_font

                for idx, (label, count) in enumerate(zip(bin_labels, bin_counts), 1):
                    ws.cell(row=equity_usage_start_row + idx, column=1, value=label)
                    ws.cell(row=equity_usage_start_row + idx, column=2, value=count)

                # Create bar chart for equity usage distribution
                chart5 = BarChart()
                chart5.title = "Equity Usage Distribution"
                chart5.y_axis.title = 'Number of Trades'
                chart5.x_axis.title = '% of Equity Used'
                chart5.style = 13

                data_ref = Reference(ws, min_col=2, min_row=equity_usage_start_row, max_row=equity_usage_start_row + len(bin_labels))
                cats_ref = Reference(ws, min_col=1, min_row=equity_usage_start_row+1, max_row=equity_usage_start_row + len(bin_labels))

                chart5.add_data(data_ref, titles_from_data=True)
                chart5.set_categories(cats_ref)
                chart5.height = 12
                chart5.width = 18

                ws.add_chart(chart5, "E88")

    def _create_market_conditions(self, wb: Workbook, result: BacktestResult, metrics: Dict[str, Any]):
        """Create Sheet 5: Market Condition Breakdown (Optional)."""
        ws = wb.create_sheet("Market Conditions")

        # Title
        ws['A1'] = "MARKET CONDITION BREAKDOWN"
        ws['A1'].font = Font(bold=True, size=14)

        # Note: This would require market regime classification
        # For now, provide a template/placeholder

        ws['A3'] = "Market Condition"
        ws['B3'] = "# Trades"
        ws['C3'] = "Win Rate %"
        ws['D3'] = "Avg Return %"
        ws['E3'] = "Sharpe Ratio"
        ws['F3'] = "Notes"

        # Apply header formatting
        for col in range(1, 7):
            cell = ws.cell(row=3, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill

        # Placeholder rows
        conditions = ["Bull Market", "Bear Market", "High Volatility", "Low Volatility", "Ranging Market"]

        for idx, condition in enumerate(conditions, 4):
            ws.cell(row=idx, column=1, value=condition)
            ws.cell(row=idx, column=2, value="N/A")
            ws.cell(row=idx, column=3, value="N/A")
            ws.cell(row=idx, column=4, value="N/A")
            ws.cell(row=idx, column=5, value="N/A")
            ws.cell(row=idx, column=6, value="Requires market regime data")

        ws['A10'] = "Note: Market condition analysis requires additional market regime classification."
        ws['A10'].font = Font(italic=True)

        # Format columns
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 20

    # ==================== HELPER METHODS ====================

    def _add_section_header(self, ws, row: int, title: str) -> int:
        """Add a section header and return next row."""
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = title
        cell.font = self.section_font
        cell.fill = self.section_fill
        cell.alignment = Alignment(horizontal='left')
        return row + 1

    def _add_metrics_table(self, ws, row: int, data: List[tuple]) -> int:
        """
        Add a metrics table.

        Args:
            ws: Worksheet
            row: Starting row
            data: List of tuples (label, value, format_type)
                  format_type: 'currency', 'percentage', 'number', 'decimal', 'auto'

        Returns:
            Next available row
        """
        for label, value, format_type in data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = self.metric_font

            value_cell = ws[f'B{row}']
            value_cell.value = value

            if format_type == 'currency':
                value_cell.number_format = '$#,##0.00'
            elif format_type == 'percentage':
                value_cell.number_format = '0.00"%"'
            elif format_type == 'decimal':
                value_cell.number_format = '0.00'
            elif format_type == 'number':
                value_cell.number_format = '#,##0'
            elif format_type == 'auto':
                # Auto-detect format based on value type
                if isinstance(value, float):
                    value_cell.number_format = '0.00'
                elif isinstance(value, int):
                    value_cell.number_format = '#,##0'
                # else: keep default (General) format for strings

            row += 1

        return row

    def _add_metrics_with_targets(self, ws, row: int, data: List[tuple]) -> int:
        """
        Add metrics with target values for comparison.

        Args:
            data: List of tuples (label, value, format_type, target, reference_info)
                  reference_info is optional
        """
        # Headers
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        ws[f'C{row}'] = "Target"
        ws[f'D{row}'] = "Status"
        ws[f'E{row}'] = "Reference"

        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        row += 1

        for item in data:
            # Handle both old and new format
            if len(item) == 5:
                label, value, format_type, target, reference_info = item
            else:
                label, value, format_type, target = item
                reference_info = None

            ws[f'A{row}'] = label
            ws[f'A{row}'].font = self.metric_font

            # Value
            value_cell = ws[f'B{row}']
            value_cell.value = value

            if format_type == 'decimal':
                value_cell.number_format = '0.00'

            # Target
            if target is not None:
                ws[f'C{row}'] = f"> {target}"

                # Status
                status_cell = ws[f'D{row}']
                if value >= target:
                    status_cell.value = "✓ Pass"
                    status_cell.font = Font(color="00B050", bold=True)
                else:
                    status_cell.value = "✗ Below"
                    status_cell.font = Font(color="FF0000", bold=True)
            else:
                ws[f'C{row}'] = "N/A"
                ws[f'D{row}'] = "-"

            # Reference info
            if reference_info:
                ws[f'E{row}'] = reference_info
                ws[f'E{row}'].font = Font(size=9, italic=True)
            else:
                ws[f'E{row}'] = ""

            row += 1

        return row
