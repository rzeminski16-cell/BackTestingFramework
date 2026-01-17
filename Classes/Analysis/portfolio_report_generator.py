"""
Portfolio Report Generator for comprehensive portfolio analysis.

Generates Excel reports with:
- Portfolio summary dashboard
- Per-security performance breakdown
- Capital allocation timeline
- Signal rejection analysis
- Vulnerability score analysis (if enabled)
- Correlation matrix between securities
- Position overlap heatmap

Enhanced features (enabled by default):
- Executive Summary Dashboard with KPIs
- Table of Contents with hyperlinks
- Deep trade analysis (MAE/MFE, clustering)
- Rolling performance metrics
- Statistical significance testing
- Comprehensive matplotlib visualizations
"""
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import LineChart, BarChart, Reference, PieChart
    from openpyxl.chart.label import DataLabelList
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from ..Analysis.performance_metrics import PerformanceMetrics

# Check if enhanced reports are available
try:
    from ..Analysis.enhanced_portfolio_report import EnhancedPortfolioReportGenerator
    ENHANCED_REPORTS_AVAILABLE = True
except ImportError:
    ENHANCED_REPORTS_AVAILABLE = False


class PortfolioReportGenerator:
    """
    Generates comprehensive Excel reports for portfolio backtests.

    For enhanced reports with advanced visualizations, matplotlib charts,
    and deeper analytics, use the generate_enhanced_report() method or
    set use_enhanced=True in generate_portfolio_report().
    """

    # Style definitions
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") if OPENPYXL_AVAILABLE else None
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11) if OPENPYXL_AVAILABLE else None
    SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid") if OPENPYXL_AVAILABLE else None
    POSITIVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") if OPENPYXL_AVAILABLE else None
    NEGATIVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") if OPENPYXL_AVAILABLE else None
    NEUTRAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") if OPENPYXL_AVAILABLE else None

    def __init__(self, output_dir: Path, use_enhanced: bool = True):
        """
        Initialize report generator.

        Args:
            output_dir: Directory to save reports
            use_enhanced: If True (default), use enhanced report generator with advanced visualizations
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for report generation. Install with: pip install openpyxl")

        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.use_enhanced = use_enhanced and ENHANCED_REPORTS_AVAILABLE

        # Initialize enhanced generator if available and requested
        self._enhanced_generator = None
        if self.use_enhanced:
            try:
                self._enhanced_generator = EnhancedPortfolioReportGenerator(self.output_dir)
            except Exception:
                self.use_enhanced = False

    def generate_portfolio_report(self, result, report_name: Optional[str] = None,
                                     use_enhanced: Optional[bool] = None) -> Path:
        """
        Generate comprehensive portfolio report.

        Args:
            result: PortfolioBacktestResult object
            report_name: Optional custom report name
            use_enhanced: Override instance setting for enhanced reports

        Returns:
            Path to generated report
        """
        # Determine whether to use enhanced reports
        should_use_enhanced = use_enhanced if use_enhanced is not None else self.use_enhanced

        # Delegate to enhanced generator if enabled
        if should_use_enhanced and self._enhanced_generator is not None:
            return self._enhanced_generator.generate_portfolio_report(result, report_name)

        # Standard report generation
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = report_name or f"portfolio_report_{timestamp}.xlsx"
        filepath = self.output_dir / filename

        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # Create sheets
        self._create_summary_sheet(wb, result)
        self._create_per_security_sheet(wb, result)
        self._create_trades_sheet(wb, result)
        self._create_equity_curve_sheet(wb, result)

        if result.signal_rejections:
            self._create_signal_rejections_sheet(wb, result)

        if result.vulnerability_swaps or result.vulnerability_history:
            self._create_vulnerability_sheet(wb, result)

        self._create_correlation_sheet(wb, result)
        self._create_capital_allocation_sheet(wb, result)
        self._create_stable_metrics_sheet(wb, result)

        if hasattr(result, 'capital_allocation_events') and result.capital_allocation_events:
            self._create_capital_events_sheet(wb, result)

        wb.save(filepath)
        print(f"Portfolio report saved to {filepath}")
        return filepath

    def generate_enhanced_report(self, result, report_name: Optional[str] = None) -> Optional[Path]:
        """
        Explicitly generate enhanced portfolio report with advanced visualizations.

        Args:
            result: PortfolioBacktestResult object
            report_name: Optional custom report name

        Returns:
            Path to generated report, or None if enhanced reports not available
        """
        if not ENHANCED_REPORTS_AVAILABLE:
            print("Warning: Enhanced reports not available. Install matplotlib and scipy.")
            return None

        if self._enhanced_generator is None:
            try:
                self._enhanced_generator = EnhancedPortfolioReportGenerator(self.output_dir)
            except Exception as e:
                print(f"Warning: Could not initialize enhanced generator: {e}")
                return None

        return self._enhanced_generator.generate_portfolio_report(result, report_name)

    @staticmethod
    def is_enhanced_available() -> bool:
        """Check if enhanced reports are available."""
        return ENHANCED_REPORTS_AVAILABLE

    def _create_summary_sheet(self, wb: Workbook, result):
        """Create portfolio summary dashboard."""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws['A1'] = "PORTFOLIO BACKTEST REPORT"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:E1')

        # Basic info
        row = 3
        ws[f'A{row}'] = "Strategy:"
        ws[f'B{row}'] = result.strategy_name
        row += 1
        ws[f'A{row}'] = "Securities:"
        ws[f'B{row}'] = ", ".join(result.symbol_results.keys())
        row += 1
        ws[f'A{row}'] = "Generated:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Portfolio metrics
        row += 2
        ws[f'A{row}'] = "PORTFOLIO METRICS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        self._apply_header_style(ws, row, 1, 2)

        row += 1
        metrics_data = [
            ("Initial Capital", f"{result.config.initial_capital:,.2f}"),
            ("Final Equity", f"{result.final_equity:,.2f}"),
            ("Total Return", f"{result.total_return:,.2f}"),
            ("Total Return %", f"{result.total_return_pct:.2f}%"),
            ("Total Trades", sum(len(r.trades) for r in result.symbol_results.values())),
            ("Signal Rejections", len(result.signal_rejections)),
            ("Vulnerability Swaps", len(result.vulnerability_swaps)),
        ]

        for metric, value in metrics_data:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            row += 1

        # Per-security summary table
        row += 2
        ws[f'A{row}'] = "PER-SECURITY PERFORMANCE"
        ws[f'A{row}'].font = Font(bold=True, size=12)

        row += 1
        headers = ["Symbol", "Trades", "Win Rate", "P/L", "P/L %", "Contribution %"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_header_style(ws, row, col, col)

        row += 1
        total_pl = result.total_return if result.total_return != 0 else 1

        for symbol, sym_result in result.symbol_results.items():
            trades = sym_result.trades
            num_trades = len(trades)
            wins = len([t for t in trades if t.pl > 0])
            win_rate = (wins / num_trades * 100) if num_trades > 0 else 0
            pl = sym_result.total_return
            pl_pct = sym_result.total_return_pct
            contribution = (pl / total_pl * 100) if total_pl != 0 else 0

            ws.cell(row=row, column=1, value=symbol)
            ws.cell(row=row, column=2, value=num_trades)
            ws.cell(row=row, column=3, value=f"{win_rate:.1f}%")
            ws.cell(row=row, column=4, value=f"{pl:,.2f}")
            ws.cell(row=row, column=5, value=f"{pl_pct:.2f}%")
            ws.cell(row=row, column=6, value=f"{contribution:.1f}%")

            # Color based on P/L
            fill = self.POSITIVE_FILL if pl > 0 else (self.NEGATIVE_FILL if pl < 0 else self.NEUTRAL_FILL)
            ws.cell(row=row, column=4).fill = fill

            row += 1

        # Capital contention mode info
        row += 2
        ws[f'A{row}'] = "CAPITAL CONTENTION"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = "Mode:"
        ws[f'B{row}'] = result.config.capital_contention.mode.value
        row += 1

        if hasattr(result.config.capital_contention, 'vulnerability_config'):
            vc = result.config.capital_contention.vulnerability_config
            ws[f'A{row}'] = "Immunity Days:"
            ws[f'B{row}'] = vc.immunity_days
            row += 1
            ws[f'A{row}'] = "Min Profit Threshold:"
            ws[f'B{row}'] = f"{vc.min_profit_threshold * 100:.1f}%"
            row += 1
            ws[f'A{row}'] = "Swap Threshold:"
            ws[f'B{row}'] = vc.swap_threshold

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15

    def _create_per_security_sheet(self, wb: Workbook, result):
        """Create detailed per-security analysis."""
        ws = wb.create_sheet("Per-Security Analysis")

        row = 1
        for symbol, sym_result in result.symbol_results.items():
            # Section header
            ws[f'A{row}'] = f"SECURITY: {symbol}"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            ws.merge_cells(f'A{row}:F{row}')
            self._apply_header_style(ws, row, 1, 6)
            row += 2

            # Calculate metrics
            trades = sym_result.trades
            if trades:
                metrics = PerformanceMetrics.calculate_from_trades(
                    trades, result.config.initial_capital
                )

                metrics_to_show = [
                    ("Total Trades", metrics.get('num_trades', 0)),
                    ("Win Rate", f"{metrics.get('win_rate', 0):.1f}%"),
                    ("Total Return", f"{sym_result.total_return:,.2f}"),
                    ("Total Return %", f"{sym_result.total_return_pct:.2f}%"),
                    ("Profit Factor", f"{metrics.get('profit_factor', 0):.2f}"),
                    ("Max Drawdown", f"{metrics.get('max_drawdown_pct', 0):.2f}%"),
                    ("Sharpe Ratio", f"{metrics.get('sharpe_ratio', 0):.2f}"),
                    ("Avg Trade Duration", f"{metrics.get('avg_trade_duration', 0):.1f} days"),
                    ("Avg Win", f"{metrics.get('avg_win', 0):,.2f}"),
                    ("Avg Loss", f"{metrics.get('avg_loss', 0):,.2f}"),
                ]

                for col in range(0, len(metrics_to_show), 2):
                    ws.cell(row=row, column=1, value=metrics_to_show[col][0])
                    ws.cell(row=row, column=2, value=metrics_to_show[col][1])
                    if col + 1 < len(metrics_to_show):
                        ws.cell(row=row, column=4, value=metrics_to_show[col + 1][0])
                        ws.cell(row=row, column=5, value=metrics_to_show[col + 1][1])
                    row += 1
            else:
                ws[f'A{row}'] = "No trades for this security"
                row += 1

            row += 2  # Space between securities

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15

    def _create_trades_sheet(self, wb: Workbook, result):
        """Create consolidated trades sheet with trade IDs and capital allocation info."""
        ws = wb.create_sheet("All Trades")

        # Collect all trades
        all_trades = []
        for sym_result in result.symbol_results.values():
            all_trades.extend(sym_result.trades)
        all_trades.sort(key=lambda t: t.entry_date)

        if not all_trades:
            ws['A1'] = "No trades recorded"
            return

        # Headers with new fields
        headers = ["Trade ID", "Symbol", "Entry Date", "Entry Price", "Exit Date", "Exit Price",
                   "Quantity", "P/L", "P/L %", "Duration", "Concurrent Positions",
                   "Capital Available", "Capital Required", "Competing Signals",
                   "Entry Reason", "Exit Reason"]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            self._apply_header_style(ws, 1, col, col)

        # Data
        for row_idx, trade in enumerate(all_trades, 2):
            ws.cell(row=row_idx, column=1, value=trade.trade_id)
            ws.cell(row=row_idx, column=2, value=trade.symbol)
            ws.cell(row=row_idx, column=3, value=trade.entry_date.strftime("%Y-%m-%d"))
            ws.cell(row=row_idx, column=4, value=f"{trade.entry_price:.4f}")
            ws.cell(row=row_idx, column=5, value=trade.exit_date.strftime("%Y-%m-%d"))
            ws.cell(row=row_idx, column=6, value=f"{trade.exit_price:.4f}")
            ws.cell(row=row_idx, column=7, value=f"{trade.quantity:.4f}")
            ws.cell(row=row_idx, column=8, value=f"{trade.pl:,.2f}")
            ws.cell(row=row_idx, column=9, value=f"{trade.pl_pct:.2f}%")
            ws.cell(row=row_idx, column=10, value=f"{trade.duration_days} days")
            ws.cell(row=row_idx, column=11, value=trade.concurrent_positions)
            ws.cell(row=row_idx, column=12, value=f"{trade.entry_capital_available:,.2f}")
            ws.cell(row=row_idx, column=13, value=f"{trade.entry_capital_required:,.2f}")
            ws.cell(row=row_idx, column=14, value=", ".join(trade.competing_signals) if trade.competing_signals else "")
            ws.cell(row=row_idx, column=15, value=trade.entry_reason or "")
            ws.cell(row=row_idx, column=16, value=trade.exit_reason or "")

            # Color P/L
            fill = self.POSITIVE_FILL if trade.pl > 0 else self.NEGATIVE_FILL
            ws.cell(row=row_idx, column=8).fill = fill

        # Adjust column widths
        widths = [10, 10, 12, 12, 12, 12, 12, 12, 10, 10, 12, 15, 15, 20, 25, 25]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

    def _create_equity_curve_sheet(self, wb: Workbook, result):
        """Create equity curve visualization."""
        ws = wb.create_sheet("Equity Curve")

        equity_df = result.portfolio_equity_curve

        # Write data
        ws['A1'] = "Date"
        ws['B1'] = "Equity"
        ws['C1'] = "Capital"
        ws['D1'] = "Position Value"
        ws['E1'] = "Num Positions"
        self._apply_header_style(ws, 1, 1, 5)

        for idx, row_data in enumerate(equity_df.itertuples(), 2):
            ws.cell(row=idx, column=1, value=row_data.date.strftime("%Y-%m-%d") if hasattr(row_data.date, 'strftime') else str(row_data.date))
            ws.cell(row=idx, column=2, value=row_data.equity)
            ws.cell(row=idx, column=3, value=row_data.capital)
            ws.cell(row=idx, column=4, value=row_data.position_value)
            ws.cell(row=idx, column=5, value=row_data.num_positions if hasattr(row_data, 'num_positions') else 0)

        # Add chart
        if len(equity_df) > 1:
            chart = LineChart()
            chart.title = "Portfolio Equity Curve"
            chart.x_axis.title = "Date"
            chart.y_axis.title = "Equity"
            chart.width = 20
            chart.height = 10

            data = Reference(ws, min_col=2, min_row=1, max_row=len(equity_df) + 1)
            chart.add_data(data, titles_from_data=True)

            ws.add_chart(chart, "G2")

    def _create_signal_rejections_sheet(self, wb: Workbook, result):
        """Create signal rejections analysis."""
        ws = wb.create_sheet("Signal Rejections")

        ws['A1'] = "SIGNAL REJECTION ANALYSIS"
        ws['A1'].font = Font(bold=True, size=14)

        # Summary
        ws['A3'] = "Total Rejections:"
        ws['B3'] = len(result.signal_rejections)

        # By symbol
        rejection_by_symbol = {}
        for r in result.signal_rejections:
            rejection_by_symbol[r.symbol] = rejection_by_symbol.get(r.symbol, 0) + 1

        row = 5
        ws[f'A{row}'] = "Rejections by Symbol"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        for symbol, count in sorted(rejection_by_symbol.items(), key=lambda x: -x[1]):
            ws[f'A{row}'] = symbol
            ws[f'B{row}'] = count
            row += 1

        # Detailed list
        row += 2
        headers = ["Date", "Symbol", "Signal Type", "Reason", "Available Capital", "Required Capital"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            self._apply_header_style(ws, row, col, col)

        row += 1
        for rejection in result.signal_rejections:
            ws.cell(row=row, column=1, value=rejection.date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=2, value=rejection.symbol)
            ws.cell(row=row, column=3, value=rejection.signal_type)
            ws.cell(row=row, column=4, value=rejection.reason)
            ws.cell(row=row, column=5, value=f"{rejection.available_capital:,.2f}")
            ws.cell(row=row, column=6, value=f"{rejection.required_capital:,.2f}")
            row += 1

        # Adjust widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18

    def _create_vulnerability_sheet(self, wb: Workbook, result):
        """Create vulnerability score analysis sheet."""
        ws = wb.create_sheet("Vulnerability Analysis")

        ws['A1'] = "VULNERABILITY SCORE ANALYSIS"
        ws['A1'].font = Font(bold=True, size=14)

        # Swaps summary
        row = 3
        ws[f'A{row}'] = "POSITION SWAPS"
        ws[f'A{row}'].font = Font(bold=True, size=12)

        row += 1
        ws[f'A{row}'] = "Total Swaps:"
        ws[f'B{row}'] = len(result.vulnerability_swaps)

        if result.vulnerability_swaps:
            row += 2
            headers = ["Date", "Closed Symbol", "Score", "Replaced With"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header)
                self._apply_header_style(ws, row, col, col)

            row += 1
            for swap in result.vulnerability_swaps:
                ws.cell(row=row, column=1, value=swap.date.strftime("%Y-%m-%d"))
                ws.cell(row=row, column=2, value=swap.closed_symbol)
                ws.cell(row=row, column=3, value=f"{swap.closed_score:.1f}")
                ws.cell(row=row, column=4, value=swap.new_symbol)

                # Color by score
                if swap.closed_score < 25:
                    ws.cell(row=row, column=3).fill = self.NEGATIVE_FILL
                elif swap.closed_score < 50:
                    ws.cell(row=row, column=3).fill = self.NEUTRAL_FILL

                row += 1

        # Vulnerability history summary
        if result.vulnerability_history:
            row += 2
            ws[f'A{row}'] = "VULNERABILITY SCORE STATISTICS"
            ws[f'A{row}'].font = Font(bold=True, size=12)

            # Calculate average scores per symbol
            symbol_scores = {}
            for day_scores in result.vulnerability_history:
                for symbol, vr in day_scores.items():
                    if symbol not in symbol_scores:
                        symbol_scores[symbol] = []
                    symbol_scores[symbol].append(vr.score)

            row += 2
            headers = ["Symbol", "Avg Score", "Min Score", "Max Score", "Days Tracked"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header)
                self._apply_header_style(ws, row, col, col)

            row += 1
            for symbol, scores in symbol_scores.items():
                ws.cell(row=row, column=1, value=symbol)
                ws.cell(row=row, column=2, value=f"{np.mean(scores):.1f}")
                ws.cell(row=row, column=3, value=f"{min(scores):.1f}")
                ws.cell(row=row, column=4, value=f"{max(scores):.1f}")
                ws.cell(row=row, column=5, value=len(scores))
                row += 1

        # Adjust widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

    def _create_correlation_sheet(self, wb: Workbook, result):
        """Create correlation analysis between securities."""
        ws = wb.create_sheet("Correlation")

        ws['A1'] = "SECURITY CORRELATION ANALYSIS"
        ws['A1'].font = Font(bold=True, size=14)

        # Calculate returns per security
        equity_df = result.portfolio_equity_curve
        symbols = list(result.symbol_results.keys())

        # We need to calculate daily returns per symbol from trades
        # This is a simplified version - in production you'd want more detailed per-symbol equity

        ws['A3'] = "Note: Correlation analysis based on trade timing overlap"

        # Trade overlap analysis
        row = 5
        ws[f'A{row}'] = "POSITION OVERLAP ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12)

        # Calculate how often securities were held together
        # This is derived from the equity curve's num_positions data

        row += 2
        ws[f'A{row}'] = "Average Concurrent Positions:"
        if 'num_positions' in equity_df.columns:
            avg_positions = equity_df['num_positions'].mean()
            ws[f'B{row}'] = f"{avg_positions:.2f}"
        else:
            ws[f'B{row}'] = "N/A"

        row += 1
        ws[f'A{row}'] = "Max Concurrent Positions:"
        if 'num_positions' in equity_df.columns:
            max_positions = equity_df['num_positions'].max()
            ws[f'B{row}'] = max_positions
        else:
            ws[f'B{row}'] = "N/A"

        # P/L contribution correlation
        row += 3
        ws[f'A{row}'] = "P/L CONTRIBUTION BY SECURITY"
        ws[f'A{row}'].font = Font(bold=True, size=12)

        row += 1
        headers = ["Symbol", "P/L", "% of Total"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            self._apply_header_style(ws, row, col, col)

        row += 1
        total_pl = max(abs(result.total_return), 0.01)  # Avoid division by zero
        for symbol, sym_result in sorted(result.symbol_results.items(),
                                          key=lambda x: -abs(x[1].total_return)):
            ws.cell(row=row, column=1, value=symbol)
            ws.cell(row=row, column=2, value=f"{sym_result.total_return:,.2f}")
            ws.cell(row=row, column=3, value=f"{sym_result.total_return / total_pl * 100:.1f}%")

            fill = self.POSITIVE_FILL if sym_result.total_return > 0 else self.NEGATIVE_FILL
            ws.cell(row=row, column=2).fill = fill
            row += 1

    def _create_capital_allocation_sheet(self, wb: Workbook, result):
        """Create capital allocation timeline."""
        ws = wb.create_sheet("Capital Allocation")

        ws['A1'] = "CAPITAL ALLOCATION OVER TIME"
        ws['A1'].font = Font(bold=True, size=14)

        equity_df = result.portfolio_equity_curve

        # Summary stats
        ws['A3'] = "Allocation Statistics"
        ws['A3'].font = Font(bold=True)

        ws['A4'] = "Average Cash %:"
        avg_cash_pct = (equity_df['capital'] / equity_df['equity'] * 100).mean()
        ws['B4'] = f"{avg_cash_pct:.1f}%"

        ws['A5'] = "Average Position %:"
        avg_pos_pct = (equity_df['position_value'] / equity_df['equity'] * 100).mean()
        ws['B5'] = f"{avg_pos_pct:.1f}%"

        ws['A6'] = "Max Drawdown:"
        equity_series = equity_df['equity']
        rolling_max = equity_series.cummax()
        drawdown = (equity_series - rolling_max) / rolling_max * 100
        max_dd = drawdown.min()
        ws['B6'] = f"{max_dd:.2f}%"

        # Allocation timeline
        row = 8
        headers = ["Date", "Equity", "Cash", "Positions", "Cash %", "Positions %"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            self._apply_header_style(ws, row, col, col)

        # Sample every Nth row if too many
        sample_rate = max(1, len(equity_df) // 100)
        sampled_df = equity_df.iloc[::sample_rate]

        row += 1
        for _, data in sampled_df.iterrows():
            equity = data['equity']
            cash = data['capital']
            positions = data['position_value']
            cash_pct = cash / equity * 100 if equity > 0 else 0
            pos_pct = positions / equity * 100 if equity > 0 else 0

            ws.cell(row=row, column=1, value=str(data['date'])[:10])
            ws.cell(row=row, column=2, value=f"{equity:,.0f}")
            ws.cell(row=row, column=3, value=f"{cash:,.0f}")
            ws.cell(row=row, column=4, value=f"{positions:,.0f}")
            ws.cell(row=row, column=5, value=f"{cash_pct:.1f}%")
            ws.cell(row=row, column=6, value=f"{pos_pct:.1f}%")
            row += 1

        # Adjust widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12

    def _create_capital_events_sheet(self, wb: Workbook, result):
        """Create detailed capital allocation events sheet."""
        ws = wb.create_sheet("Capital Events")

        ws['A1'] = "CAPITAL ALLOCATION EVENTS"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')

        # Summary statistics
        events = result.capital_allocation_events
        executed = len([e for e in events if e.signal_type == "EXECUTED"])
        rejected = len([e for e in events if e.signal_type == "REJECTED"])
        swapped_in = len([e for e in events if e.signal_type == "SWAPPED_IN"])
        swapped_out = len([e for e in events if e.signal_type == "SWAPPED_OUT"])

        row = 3
        ws[f'A{row}'] = "Event Summary"
        ws[f'A{row}'].font = Font(bold=True, size=12)

        row += 1
        summary_data = [
            ("Total Events", len(events)),
            ("Executed", executed),
            ("Rejected", rejected),
            ("Swapped In", swapped_in),
            ("Swapped Out", swapped_out),
        ]

        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        # Detailed events table
        row += 2
        ws[f'A{row}'] = "Detailed Capital Events"
        ws[f'A{row}'].font = Font(bold=True, size=12)

        row += 1
        headers = ["Date", "Symbol", "Event Type", "Available Capital", "Required Capital",
                   "Total Equity", "Open Positions", "Competing Signals", "Outcome",
                   "Vulnerability Scores"]

        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            self._apply_header_style(ws, row, col, col)

        row += 1
        for event in events:
            ws.cell(row=row, column=1, value=event.date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=2, value=event.symbol)
            ws.cell(row=row, column=3, value=event.signal_type)
            ws.cell(row=row, column=4, value=f"{event.available_capital:,.2f}")
            ws.cell(row=row, column=5, value=f"{event.required_capital:,.2f}")
            ws.cell(row=row, column=6, value=f"{event.total_equity:,.2f}")
            ws.cell(row=row, column=7, value=", ".join(event.open_position_symbols))
            ws.cell(row=row, column=8, value=", ".join(event.competing_signals) if event.competing_signals else "")
            ws.cell(row=row, column=9, value=event.outcome)

            # Format vulnerability scores
            if event.vulnerability_scores:
                scores_str = ", ".join(f"{s}: {v:.1f}" for s, v in event.vulnerability_scores.items())
                ws.cell(row=row, column=10, value=scores_str)

            # Color by event type
            if event.signal_type == "EXECUTED":
                ws.cell(row=row, column=3).fill = self.POSITIVE_FILL
            elif event.signal_type == "REJECTED":
                ws.cell(row=row, column=3).fill = self.NEGATIVE_FILL
            elif event.signal_type == "SWAPPED_IN":
                ws.cell(row=row, column=3).fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif event.signal_type == "SWAPPED_OUT":
                ws.cell(row=row, column=3).fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

            row += 1

        # Adjust column widths
        widths = [12, 10, 12, 18, 18, 18, 25, 25, 50, 40]
        for col, width in enumerate(widths, 1):
            if col <= 10:
                ws.column_dimensions[chr(64 + col)].width = width

    def _apply_header_style(self, ws, row: int, start_col: int, end_col: int):
        """Apply header styling to cells."""
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')

    def _create_stable_metrics_sheet(self, wb: Workbook, result):
        """
        Create Stable Metrics sheet with regression-based performance metrics.

        Includes:
        - Regressed Annual Return (RAR%)
        - R² of log-equity regression
        - RAR% Adjusted (RAR% × R²)
        - R-Cubed metric
        - Robust Sharpe Ratio
        - Equity curve with regression line visualization
        - Largest drawdowns table
        """
        from Classes.Core.stable_metrics import StableMetricsCalculator

        ws = wb.create_sheet("Stable Metrics")

        # Title
        ws['A1'] = "STABLE PERFORMANCE METRICS"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:E1')

        row = 3

        # Get portfolio equity curve
        equity_df = result.portfolio_equity_curve
        if equity_df is None or len(equity_df) < 2:
            ws[f'A{row}'] = "Insufficient data for stable metrics calculation"
            return

        # Calculate stable metrics
        stable_result = StableMetricsCalculator.calculate_all(equity_df)

        # Section A: Regression Analysis
        ws[f'A{row}'] = "A. REGRESSION ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        self._apply_header_style(ws, row, 1, 2)
        row += 1

        regression_metrics = [
            ("Regressed Annual Return (RAR%)", f"{stable_result.rar_pct:.2f}%"),
            ("R² (Log-Equity Regression)", f"{stable_result.r_squared:.4f}"),
            ("RAR% Adjusted (RAR% × R²)", f"{stable_result.rar_adjusted:.2f}%"),
            ("CAGR (Classical)", f"{stable_result.cagr:.2f}%"),
            ("RAR% vs CAGR Difference", f"{stable_result.rar_pct - stable_result.cagr:.2f}%"),
            ("Bars Per Year Used", str(stable_result.bars_per_year)),
        ]

        for metric, value in regression_metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            row += 1

        # R² Warning if applicable
        if stable_result.r_squared_warning:
            row += 1
            ws[f'A{row}'] = "Warning:"
            ws[f'A{row}'].font = Font(bold=True, color="9C0006")
            ws[f'B{row}'] = stable_result.r_squared_warning
            ws[f'B{row}'].font = Font(italic=True, color="9C0006")
            ws.merge_cells(f'B{row}:E{row}')
            row += 1

        row += 2

        # Section B: R-Cubed Analysis
        ws[f'A{row}'] = "B. R-CUBED ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        self._apply_header_style(ws, row, 1, 2)
        row += 1

        ws[f'A{row}'] = "R-Cubed = RAR% / (Avg Max Drawdown × Length Adjustment)"
        ws[f'A{row}'].font = Font(italic=True, size=9)
        ws.merge_cells(f'A{row}:E{row}')
        row += 1

        r_cubed_metrics = [
            ("R-Cubed", f"{stable_result.r_cubed:.2f}"),
            ("Average Max Drawdown (%)", f"{stable_result.avg_max_drawdown_pct:.2f}%"),
            ("Average Max Drawdown Length (days)", f"{stable_result.avg_max_drawdown_length_days:.1f}"),
            ("Length Adjustment Factor", f"{stable_result.length_adjustment_factor:.4f}"),
            ("Drawdowns Used in Calculation", str(stable_result.num_drawdowns_used)),
        ]

        for metric, value in r_cubed_metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            row += 1

        # R-Cubed interpretation
        row += 1
        r_cubed_val = stable_result.r_cubed
        if r_cubed_val >= 2.0:
            interpretation = "Excellent - Strong risk-adjusted returns with manageable drawdowns"
            fill = self.POSITIVE_FILL
        elif r_cubed_val >= 1.5:
            interpretation = "Good - Solid risk/reward profile"
            fill = self.POSITIVE_FILL
        elif r_cubed_val >= 1.0:
            interpretation = "Fair - Acceptable but room for improvement"
            fill = self.NEUTRAL_FILL
        else:
            interpretation = "Poor - High risk relative to returns"
            fill = self.NEGATIVE_FILL

        ws[f'A{row}'] = f"R-Cubed = {r_cubed_val:.2f}: {interpretation}"
        ws[f'A{row}'].fill = fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 1

        # Drawdown warning if applicable
        if stable_result.drawdown_warning:
            ws[f'A{row}'] = "Note: " + stable_result.drawdown_warning
            ws[f'A{row}'].font = Font(italic=True, color="DAA520")
            ws.merge_cells(f'A{row}:E{row}')
            row += 1

        row += 2

        # Section C: Robust Sharpe Ratio
        ws[f'A{row}'] = "C. ROBUST SHARPE RATIO"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        self._apply_header_style(ws, row, 1, 2)
        row += 1

        ws[f'A{row}'] = "Robust Sharpe = RAR% / Annualized Std Dev of Monthly Returns"
        ws[f'A{row}'].font = Font(italic=True, size=9)
        ws.merge_cells(f'A{row}:E{row}')
        row += 1

        robust_sharpe_metrics = [
            ("Robust Sharpe Ratio", f"{stable_result.robust_sharpe_ratio:.3f}"),
            ("Monthly Return Std (Annualized)", f"{stable_result.monthly_return_std_annualized:.2f}%"),
        ]

        for metric, value in robust_sharpe_metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            row += 1

        # Robust Sharpe interpretation
        row += 1
        robust_sharpe_val = stable_result.robust_sharpe_ratio
        if robust_sharpe_val >= 1.0:
            interpretation = "Excellent - Strong risk-adjusted performance"
            fill = self.POSITIVE_FILL
        elif robust_sharpe_val >= 0.5:
            interpretation = "Good - Acceptable risk/return trade-off"
            fill = self.POSITIVE_FILL
        else:
            interpretation = "Poor - Returns do not adequately compensate for risk"
            fill = self.NEGATIVE_FILL

        ws[f'A{row}'] = f"Robust Sharpe = {robust_sharpe_val:.3f}: {interpretation}"
        ws[f'A{row}'].fill = fill
        ws.merge_cells(f'A{row}:E{row}')
        row += 1

        if stable_result.monthly_returns_warning:
            ws[f'A{row}'] = "Note: " + stable_result.monthly_returns_warning
            ws[f'A{row}'].font = Font(italic=True, color="DAA520")
            ws.merge_cells(f'A{row}:E{row}')
            row += 1

        row += 2

        # Section D: Largest Drawdowns Table
        ws[f'A{row}'] = "D. LARGEST DRAWDOWNS (Used for R-Cubed)"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        self._apply_header_style(ws, row, 1, 6)
        row += 1

        if stable_result.largest_drawdowns:
            # Headers
            headers = ["#", "Start Date", "Trough Date", "Recovery Date", "Drawdown %", "Duration (days)"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header)
                self._apply_header_style(ws, row, col, col)
            row += 1

            # Data rows
            for i, dd in enumerate(stable_result.largest_drawdowns, 1):
                dd_dict = dd.to_dict()
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=dd_dict['start_date'])
                ws.cell(row=row, column=3, value=dd_dict['trough_date'])
                ws.cell(row=row, column=4, value=dd_dict['recovery_date'])
                ws.cell(row=row, column=5, value=dd_dict['drawdown_pct'])
                ws.cell(row=row, column=6, value=dd.duration_days)

                # Color code based on severity
                dd_pct = dd.drawdown_pct
                if dd_pct >= 20:
                    ws.cell(row=row, column=5).fill = self.NEGATIVE_FILL
                elif dd_pct >= 10:
                    ws.cell(row=row, column=5).fill = self.NEUTRAL_FILL
                else:
                    ws.cell(row=row, column=5).fill = self.POSITIVE_FILL
                row += 1
        else:
            ws[f'A{row}'] = "No significant drawdowns detected"
            row += 1

        row += 2

        # Section E: Equity Curve with Regression Line
        ws[f'A{row}'] = "E. EQUITY CURVE WITH REGRESSION LINE"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        self._apply_header_style(ws, row, 1, 3)
        row += 1

        # Get regression line data
        if 'equity' in equity_df.columns:
            equity_values = equity_df['equity'].values
        else:
            equity_values = equity_df.iloc[:, 0].values

        # Handle NaN values
        valid_mask = ~np.isnan(equity_values) & (equity_values > 0)
        if np.any(valid_mask):
            equity_clean = equity_values[valid_mask]
            regression_line = StableMetricsCalculator.get_regression_line_data(
                equity_clean,
                stable_result.regression_slope,
                stable_result.regression_intercept
            )

            # Sample data if too large
            max_points = 150
            if len(equity_clean) > max_points:
                step = len(equity_clean) // max_points
                sample_indices = range(0, len(equity_clean), step)
            else:
                sample_indices = range(len(equity_clean))

            # Headers for chart data
            ws.cell(row=row, column=1, value="Bar").font = self.HEADER_FONT
            ws.cell(row=row, column=2, value="Equity").font = self.HEADER_FONT
            ws.cell(row=row, column=3, value="Regression Line").font = self.HEADER_FONT
            row += 1

            data_start_row = row
            for idx in sample_indices:
                ws.cell(row=row, column=1, value=idx + 1)
                ws.cell(row=row, column=2, value=float(equity_clean[idx]))
                ws.cell(row=row, column=3, value=float(regression_line[idx]))
                row += 1

            data_end_row = row - 1

            # Create chart
            if data_end_row > data_start_row:
                chart = LineChart()
                chart.title = f"Equity with Regression (RAR% = {stable_result.rar_pct:.2f}%, R² = {stable_result.r_squared:.4f})"
                chart.style = 10
                chart.x_axis.title = "Bar"
                chart.y_axis.title = "Equity"
                chart.width = 16
                chart.height = 10

                # Equity curve data
                equity_data = Reference(ws, min_col=2, min_row=data_start_row - 1, max_row=data_end_row)
                chart.add_data(equity_data, titles_from_data=True)

                # Regression line data
                reg_data = Reference(ws, min_col=3, min_row=data_start_row - 1, max_row=data_end_row)
                chart.add_data(reg_data, titles_from_data=True)

                ws.add_chart(chart, f"E{data_start_row - 1}")

        # Format column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 18
