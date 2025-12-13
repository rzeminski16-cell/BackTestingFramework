"""
Enhanced Optimization Report Generator for Portfolio Optimization.

Generates comprehensive Excel reports with:
- Executive summary with overfitting assessment
- 3D parameter surface plots
- Parameter robustness zones
- Overfitting probability scores
- In-sample vs out-of-sample analysis
- Multi-security comparison
- Statistical significance testing
- Embedded matplotlib visualizations
"""
import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple
import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, AreaChart
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

try:
    from openpyxl.drawing.image import Image
    IMAGE_AVAILABLE = True
except ImportError:
    IMAGE_AVAILABLE = False

from Classes.Optimization.sensitivity_analyzer import SensitivityResults
from Classes.Optimization.walk_forward_optimizer import WalkForwardResults, MultiSecurityResults

# Try to import enhanced visualizations
try:
    from Classes.Analysis.enhanced_visualizations import EnhancedVisualizations, MATPLOTLIB_AVAILABLE
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    EnhancedVisualizations = None

logger = logging.getLogger(__name__)


class EnhancedOptimizationReportGenerator:
    """
    Enhanced optimization report generator with advanced analytics.

    Features:
    - Executive summary with traffic-light indicators
    - 3D parameter sensitivity surfaces
    - Overfitting probability assessment
    - Parameter robustness zones
    - Multi-security comparison
    - Embedded matplotlib visualizations
    """

    # Color scheme
    COLORS = {
        'header_dark': '1F4E79',
        'header_medium': '2E75B6',
        'positive': 'C6EFCE',
        'positive_dark': '28A745',
        'negative': 'FFC7CE',
        'negative_dark': 'DC3545',
        'neutral': 'FFEB9C',
        'neutral_dark': 'FFC107',
        'white': 'FFFFFF',
        'light_gray': 'F5F5F5',
        'medium_gray': 'E0E0E0',
        'dark_gray': '666666',
    }

    def __init__(self, config: Dict[str, Any]):
        """Initialize enhanced report generator."""
        self.config = config
        self.report_config = config.get('reporting', {})
        self.decimal_places = config.get('report', {}).get('excel', {}).get('decimal_places', 4)

        # Initialize visualization module
        self.include_matplotlib_charts = MATPLOTLIB_AVAILABLE and IMAGE_AVAILABLE
        if self.include_matplotlib_charts:
            try:
                self.viz = EnhancedVisualizations(dpi=150)
            except Exception:
                self.viz = None
                self.include_matplotlib_charts = False
        else:
            self.viz = None

        self._init_styles()
        logger.info("Enhanced optimization report generator initialized")

    def _init_styles(self):
        """Initialize Excel styles."""
        self.header_fill = PatternFill(start_color=self.COLORS['header_dark'],
                                        end_color=self.COLORS['header_dark'], fill_type="solid")
        self.header_font = Font(bold=True, color=self.COLORS['white'], size=11)
        self.subheader_fill = PatternFill(start_color=self.COLORS['header_medium'],
                                           end_color=self.COLORS['header_medium'], fill_type="solid")
        self.positive_fill = PatternFill(start_color=self.COLORS['positive'],
                                          end_color=self.COLORS['positive'], fill_type="solid")
        self.negative_fill = PatternFill(start_color=self.COLORS['negative'],
                                          end_color=self.COLORS['negative'], fill_type="solid")
        self.neutral_fill = PatternFill(start_color=self.COLORS['neutral'],
                                         end_color=self.COLORS['neutral'], fill_type="solid")
        self.light_fill = PatternFill(start_color=self.COLORS['light_gray'],
                                       end_color=self.COLORS['light_gray'], fill_type="solid")

        self.title_font = Font(bold=True, size=16, color=self.COLORS['header_dark'])
        self.section_font = Font(bold=True, size=14, color=self.COLORS['header_dark'])
        self.subsection_font = Font(bold=True, size=12, color=self.COLORS['header_medium'])

        self.border = Border(
            left=Side(style='thin', color=self.COLORS['medium_gray']),
            right=Side(style='thin', color=self.COLORS['medium_gray']),
            top=Side(style='thin', color=self.COLORS['medium_gray']),
            bottom=Side(style='thin', color=self.COLORS['medium_gray'])
        )

    def generate_portfolio_optimization_report(
        self,
        multi_results: MultiSecurityResults,
        sensitivity_results_dict: Optional[Dict[str, SensitivityResults]] = None,
        output_dir: str = None
    ) -> str:
        """
        Generate comprehensive portfolio optimization report.

        Args:
            multi_results: Combined results from all securities
            sensitivity_results_dict: Optional dict of symbol -> SensitivityResults
            output_dir: Output directory

        Returns:
            Path to generated report file
        """
        if output_dir is None:
            output_dir = self.config.get('report', {}).get('output_dir', 'logs/optimization_reports')

        os.makedirs(output_dir, exist_ok=True)

        # Create filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        securities_str = "_".join(multi_results.securities[:3])
        if len(multi_results.securities) > 3:
            securities_str += f"_+{len(multi_results.securities) - 3}"
        filename = f"enhanced_portfolio_optimization_{multi_results.strategy_name}_{securities_str}_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)

        # Calculate advanced metrics
        metrics = self._calculate_optimization_metrics(multi_results, sensitivity_results_dict)

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)

        # Generate sheets - Core Analysis
        self._create_toc(wb, multi_results)
        self._create_executive_summary(wb, multi_results, metrics)
        self._create_overfitting_analysis(wb, multi_results, metrics)
        self._create_security_comparison(wb, multi_results, metrics)
        self._create_parameter_robustness(wb, multi_results, metrics)

        if sensitivity_results_dict:
            self._create_sensitivity_dashboard(wb, multi_results, sensitivity_results_dict, metrics)

        # NEW: Enhanced Analysis Sheets
        self._create_walk_forward_equity_curves(wb, multi_results, metrics)
        self._create_parameter_stability(wb, multi_results, metrics)
        self._create_constraint_analysis(wb, multi_results, metrics)
        self._create_window_comparison_table(wb, multi_results, metrics)
        self._create_risk_metrics_dashboard(wb, multi_results, metrics)
        self._create_best_worst_window_analysis(wb, multi_results, metrics)
        self._create_cross_security_correlation(wb, multi_results, metrics)
        self._create_monte_carlo_analysis(wb, multi_results, metrics)
        self._create_trade_statistics_by_window(wb, multi_results, metrics)

        # Original sheets
        self._create_window_analysis(wb, multi_results, metrics)
        self._create_recommendations(wb, multi_results, sensitivity_results_dict, metrics)

        # Individual security details
        for symbol, wf_results in multi_results.individual_results.items():
            sens_results = sensitivity_results_dict.get(symbol) if sensitivity_results_dict else None
            self._create_security_detail(wb, symbol, wf_results, sens_results)

        wb.save(filepath)
        logger.info(f"Enhanced portfolio optimization report saved to {filepath}")

        return filepath

    def _calculate_optimization_metrics(
        self,
        multi_results: MultiSecurityResults,
        sensitivity_results_dict: Optional[Dict[str, SensitivityResults]]
    ) -> Dict[str, Any]:
        """Calculate advanced optimization metrics."""
        metrics = {}

        # Basic aggregates
        metrics['total_securities'] = len(multi_results.securities)
        metrics['total_windows'] = multi_results.total_windows_all_securities
        metrics['total_passed'] = multi_results.total_passed_all_securities
        metrics['success_rate'] = (metrics['total_passed'] / metrics['total_windows'] * 100
                                    if metrics['total_windows'] > 0 else 0)

        # Performance metrics - handle potential NaN/None values
        metrics['avg_is_sortino'] = self._safe_float(multi_results.combined_avg_in_sample_sortino)
        metrics['avg_oos_sortino'] = self._safe_float(multi_results.combined_avg_out_sample_sortino)
        metrics['sortino_degradation'] = self._safe_float(multi_results.combined_avg_sortino_degradation_pct)
        metrics['avg_is_sharpe'] = self._safe_float(multi_results.combined_avg_in_sample_sharpe)
        metrics['avg_oos_sharpe'] = self._safe_float(multi_results.combined_avg_out_sample_sharpe)
        metrics['sharpe_degradation'] = self._safe_float(multi_results.combined_avg_sharpe_degradation_pct)

        # Log diagnostic info if values look suspicious
        if metrics['total_windows'] == 0:
            logger.warning("No optimization windows found in results")
        if metrics['avg_is_sortino'] == 0 and metrics['avg_oos_sortino'] == 0:
            logger.warning("Both IS and OOS Sortino ratios are 0 - check optimization results")

        # Overfitting probability
        metrics['overfitting_score'] = self._calculate_overfitting_score(multi_results, sensitivity_results_dict)
        metrics['overfitting_probability'] = self._calculate_overfitting_probability(metrics)

        # Parameter stability
        metrics['parameter_stability'] = self._calculate_parameter_stability(multi_results)

        # Robustness assessment
        metrics['robustness_score'] = self._calculate_robustness_score(multi_results, sensitivity_results_dict)

        # Per-security metrics
        metrics['security_metrics'] = {}
        for symbol, wf_results in multi_results.individual_results.items():
            metrics['security_metrics'][symbol] = {
                'windows': wf_results.total_windows,
                'passed': wf_results.windows_passed_constraints,
                'is_sortino': self._safe_float(wf_results.avg_in_sample_sortino),
                'oos_sortino': self._safe_float(wf_results.avg_out_sample_sortino),
                'degradation': self._safe_float(wf_results.avg_sortino_degradation_pct),
                'is_positive_oos': wf_results.avg_out_sample_sortino > 0,
            }

        # Store diagnostic info for the report
        metrics['_diagnostic'] = {
            'has_windows': metrics['total_windows'] > 0,
            'has_passed': metrics['total_passed'] > 0,
            'has_valid_sortino': metrics['avg_is_sortino'] != 0 or metrics['avg_oos_sortino'] != 0,
            'securities_count': len(multi_results.individual_results),
        }

        return metrics

    def _safe_float(self, value) -> float:
        """Safely convert value to float, handling NaN and None."""
        if value is None:
            return 0.0
        try:
            result = float(value)
            if np.isnan(result) or np.isinf(result):
                return 0.0
            return result
        except (ValueError, TypeError):
            return 0.0

    def _calculate_overfitting_score(
        self,
        multi_results: MultiSecurityResults,
        sensitivity_results_dict: Optional[Dict[str, SensitivityResults]]
    ) -> float:
        """
        Calculate overfitting score (0-100, higher = more overfitted).

        Factors:
        - Performance degradation (IS to OOS)
        - Parameter stability across windows
        - Sensitivity to parameter variations
        """
        score = 0.0

        # Degradation factor (0-40 points)
        degradation = abs(multi_results.combined_avg_sortino_degradation_pct)
        if degradation > 50:
            score += 40
        elif degradation > 30:
            score += 30
        elif degradation > 15:
            score += 15
        else:
            score += degradation / 15 * 15

        # Parameter instability factor (0-30 points)
        param_scores = list(multi_results.param_consistency_scores.values())
        if param_scores:
            avg_consistency = np.mean(param_scores)
            instability = 100 - avg_consistency
            score += instability * 0.3

        # Sensitivity factor (0-30 points)
        if sensitivity_results_dict:
            non_robust_count = 0
            total_params = 0
            for sens in sensitivity_results_dict.values():
                if sens:
                    for param_sens in sens.parameter_sensitivities.values():
                        total_params += 1
                        if param_sens.is_unstable:
                            non_robust_count += 1

            if total_params > 0:
                instability_ratio = non_robust_count / total_params
                score += instability_ratio * 30

        return min(100, score)

    def _calculate_overfitting_probability(self, metrics: Dict[str, Any]) -> str:
        """Convert overfitting score to probability category."""
        score = metrics['overfitting_score']

        if score < 20:
            return "LOW"
        elif score < 40:
            return "MODERATE"
        elif score < 60:
            return "HIGH"
        else:
            return "VERY HIGH"

    def _calculate_parameter_stability(self, multi_results: MultiSecurityResults) -> Dict[str, Any]:
        """Calculate parameter stability metrics."""
        stability = {}

        for param_name, consistency in multi_results.param_consistency_scores.items():
            if consistency >= 80:
                status = "Very Stable"
            elif consistency >= 60:
                status = "Stable"
            elif consistency >= 40:
                status = "Moderate"
            else:
                status = "Unstable"

            stability[param_name] = {
                'consistency': consistency,
                'status': status,
                'recommended_value': multi_results.consistent_params.get(param_name),
            }

        return stability

    def _calculate_robustness_score(
        self,
        multi_results: MultiSecurityResults,
        sensitivity_results_dict: Optional[Dict[str, SensitivityResults]]
    ) -> float:
        """Calculate overall robustness score (0-100, higher = more robust)."""
        score = 0.0

        # Positive OOS performance (0-30 points)
        positive_ratio = multi_results.securities_with_positive_oos / len(multi_results.securities)
        score += positive_ratio * 30

        # Low degradation (0-30 points)
        degradation = abs(multi_results.combined_avg_sortino_degradation_pct)
        if degradation < 15:
            score += 30
        elif degradation < 30:
            score += 20
        elif degradation < 50:
            score += 10

        # High constraint pass rate (0-20 points)
        if multi_results.total_windows_all_securities > 0:
            pass_rate = multi_results.total_passed_all_securities / multi_results.total_windows_all_securities
            score += pass_rate * 20

        # Parameter consistency (0-20 points)
        param_scores = list(multi_results.param_consistency_scores.values())
        if param_scores:
            avg_consistency = np.mean(param_scores)
            score += avg_consistency * 0.2

        return min(100, score)

    # ==================== SHEET CREATION ====================

    def _create_toc(self, wb: Workbook, multi_results: MultiSecurityResults):
        """Create Table of Contents."""
        ws = wb.create_sheet("Table of Contents", 0)

        ws['A1'] = "PORTFOLIO OPTIMIZATION REPORT"
        ws['A1'].font = Font(bold=True, size=20, color=self.COLORS['header_dark'])
        ws.merge_cells('A1:E1')

        ws['A2'] = f"Strategy: {multi_results.strategy_name}"
        ws['A2'].font = Font(size=12, bold=True)

        ws['A3'] = f"Securities: {', '.join(multi_results.securities)}"
        ws['A3'].font = Font(size=10, color=self.COLORS['dark_gray'])

        ws['A4'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'].font = Font(size=10, color=self.COLORS['dark_gray'])

        row = 6
        ws[f'A{row}'] = "TABLE OF CONTENTS"
        ws[f'A{row}'].font = self.section_font
        row += 2

        toc_items = [
            ("Executive Summary", "Key findings and KPIs"),
            ("Overfitting Analysis", "Overfitting probability and assessment"),
            ("Security Comparison", "Side-by-side performance comparison"),
            ("Parameter Robustness", "Parameter stability zones"),
            ("Sensitivity Dashboard", "Parameter sensitivity analysis"),
            ("Window Analysis", "Window-by-window results"),
            ("WF Equity Curves", "Walk-forward cumulative equity visualization"),
            ("Parameter Stability", "Parameter drift and stability over time"),
            ("Constraint Analysis", "Constraint hits and violations analysis"),
            ("Window Comparison", "Detailed window-by-window comparison table"),
            ("Risk Dashboard", "Comprehensive risk metrics analysis"),
            ("Best-Worst Analysis", "Deep dive into best and worst windows"),
            ("Security Correlation", "Cross-security correlation matrix"),
            ("Monte Carlo", "Bootstrap confidence intervals"),
            ("Trade Statistics", "Trade statistics by window"),
            ("Recommendations", "Actionable insights"),
        ]

        for i, (sheet_name, description) in enumerate(toc_items, 1):
            ws[f'A{row}'] = f"{i}."
            ws[f'B{row}'] = sheet_name
            ws[f'B{row}'].font = Font(bold=True, size=11, underline='single', color='0563C1')
            ws[f'C{row}'] = description
            ws[f'C{row}'].font = Font(size=10, color=self.COLORS['dark_gray'])
            row += 1

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 40

    def _create_executive_summary(self, wb: Workbook, multi_results: MultiSecurityResults, metrics: Dict[str, Any]):
        """Create Executive Summary sheet."""
        ws = wb.create_sheet("Executive Summary")

        row = 1
        ws[f'A{row}'] = "EXECUTIVE SUMMARY"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:H{row}')
        row += 3

        # KPI Section with traffic lights
        ws[f'A{row}'] = "KEY PERFORMANCE INDICATORS"
        ws[f'A{row}'].font = self.section_font
        row += 2

        kpis = [
            ("OOS Sortino", f"{metrics['avg_oos_sortino']:.2f}", metrics['avg_oos_sortino'] > 0.5),
            ("Degradation", f"{metrics['sortino_degradation']:.1f}%", abs(metrics['sortino_degradation']) < 30),
            ("Success Rate", f"{metrics['success_rate']:.1f}%", metrics['success_rate'] > 70),
            ("Robustness", f"{metrics['robustness_score']:.0f}/100", metrics['robustness_score'] > 60),
            ("Overfitting", metrics['overfitting_probability'], metrics['overfitting_probability'] in ['LOW', 'MODERATE']),
            ("Securities +OOS", f"{multi_results.securities_with_positive_oos}/{len(multi_results.securities)}",
             multi_results.securities_with_positive_oos >= len(multi_results.securities) * 0.7),
        ]

        col = 1
        for kpi_name, value, is_good in kpis:
            ws.cell(row=row, column=col, value=kpi_name).font = Font(bold=True, size=10)

            value_cell = ws.cell(row=row+1, column=col, value=value)
            value_cell.font = Font(bold=True, size=14)

            indicator_cell = ws.cell(row=row+1, column=col+1)
            if is_good:
                indicator_cell.value = "●"
                indicator_cell.font = Font(color=self.COLORS['positive_dark'], size=16)
            else:
                indicator_cell.value = "●"
                indicator_cell.font = Font(color=self.COLORS['negative_dark'], size=16)

            col += 3
            if col > 7:
                col = 1
                row += 4

        row += 5

        # Overall Assessment
        ws[f'A{row}'] = "OVERALL ASSESSMENT"
        ws[f'A{row}'].font = self.section_font
        row += 2

        # Determine overall status
        robustness = metrics['robustness_score']
        if robustness >= 70:
            status = "ROBUST - Strategy shows strong out-of-sample performance"
            status_color = self.COLORS['positive_dark']
        elif robustness >= 50:
            status = "MODERATE - Strategy shows acceptable performance with some concerns"
            status_color = self.COLORS['neutral_dark']
        else:
            status = "WEAK - Strategy may be overfitted or unstable"
            status_color = self.COLORS['negative_dark']

        ws[f'A{row}'] = status
        ws[f'A{row}'].font = Font(bold=True, size=12, color=status_color)
        row += 2

        # Performance Summary
        row = self._add_section_header(ws, row, "PERFORMANCE SUMMARY")

        perf_data = [
            ("Total Securities Analyzed", str(metrics['total_securities'])),
            ("Total Windows", str(metrics['total_windows'])),
            ("Windows Passed Constraints", str(metrics['total_passed'])),
            ("Avg In-Sample Sortino", f"{metrics['avg_is_sortino']:.4f}"),
            ("Avg Out-of-Sample Sortino", f"{metrics['avg_oos_sortino']:.4f}"),
            ("Sortino Degradation", f"{metrics['sortino_degradation']:.2f}%"),
            ("Avg In-Sample Sharpe", f"{metrics['avg_is_sharpe']:.4f}"),
            ("Avg Out-of-Sample Sharpe", f"{metrics['avg_oos_sharpe']:.4f}"),
            ("Best Security", multi_results.best_security),
            ("Worst Security", multi_results.worst_security),
        ]

        for metric, value in perf_data:
            ws.cell(row=row, column=1, value=metric).border = self.border
            ws.cell(row=row, column=2, value=value).border = self.border
            row += 1

        row += 2

        # Add diagnostic warnings if data looks suspicious
        diag = metrics.get('_diagnostic', {})
        warnings = []

        if not diag.get('has_windows', True):
            warnings.append("WARNING: No optimization windows found. Check that optimization completed successfully.")
        if not diag.get('has_passed', True) and diag.get('has_windows', False):
            warnings.append("WARNING: No windows passed constraints. Consider loosening constraint thresholds.")
        if not diag.get('has_valid_sortino', True) and diag.get('has_windows', False):
            warnings.append("WARNING: Sortino ratios are all zero. This may indicate insufficient trades or calculation issues.")
        if metrics['total_windows'] > 0 and metrics['success_rate'] == 0:
            warnings.append("WARNING: 0% success rate. All windows failed constraints.")

        if warnings:
            row = self._add_section_header(ws, row, "⚠️ DATA QUALITY WARNINGS")
            for warning in warnings:
                ws[f'A{row}'] = warning
                ws[f'A{row}'].font = Font(color=self.COLORS['negative_dark'], italic=True)
                row += 1
            row += 1

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        for col in ['C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 15

    def _create_overfitting_analysis(self, wb: Workbook, multi_results: MultiSecurityResults, metrics: Dict[str, Any]):
        """Create Overfitting Analysis sheet."""
        ws = wb.create_sheet("Overfitting Analysis")

        row = 1
        ws[f'A{row}'] = "OVERFITTING PROBABILITY ANALYSIS"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:F{row}')
        row += 3

        # Overfitting Score
        row = self._add_section_header(ws, row, "A. OVERFITTING ASSESSMENT")

        score = metrics['overfitting_score']
        probability = metrics['overfitting_probability']

        ws[f'A{row}'] = "Overfitting Score:"
        ws[f'B{row}'] = f"{score:.1f}/100"

        if probability == "LOW":
            ws[f'C{row}'] = "LOW RISK"
            ws[f'C{row}'].fill = self.positive_fill
        elif probability == "MODERATE":
            ws[f'C{row}'] = "MODERATE RISK"
            ws[f'C{row}'].fill = self.neutral_fill
        elif probability == "HIGH":
            ws[f'C{row}'] = "HIGH RISK"
            ws[f'C{row}'].fill = self.negative_fill
        else:
            ws[f'C{row}'] = "VERY HIGH RISK"
            ws[f'C{row}'].fill = self.negative_fill

        row += 3

        # Breakdown
        row = self._add_section_header(ws, row, "B. SCORE BREAKDOWN")

        ws[f'A{row}'] = "Component"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws[f'B{row}'] = "Impact"
        ws[f'B{row}'].font = self.header_font
        ws[f'B{row}'].fill = self.header_fill
        ws[f'C{row}'] = "Assessment"
        ws[f'C{row}'].font = self.header_font
        ws[f'C{row}'].fill = self.header_fill
        row += 1

        degradation = abs(metrics['sortino_degradation'])
        components = [
            ("Performance Degradation", f"{degradation:.1f}%",
             "Good" if degradation < 15 else "Moderate" if degradation < 30 else "Poor"),
            ("Parameter Consistency", f"{np.mean(list(multi_results.param_consistency_scores.values())):.1f}%",
             "Good" if np.mean(list(multi_results.param_consistency_scores.values())) > 70 else "Moderate"),
            ("Constraint Pass Rate", f"{metrics['success_rate']:.1f}%",
             "Good" if metrics['success_rate'] > 70 else "Moderate" if metrics['success_rate'] > 50 else "Poor"),
        ]

        for component, impact, assessment in components:
            ws.cell(row=row, column=1, value=component).border = self.border
            ws.cell(row=row, column=2, value=impact).border = self.border

            assessment_cell = ws.cell(row=row, column=3, value=assessment)
            assessment_cell.border = self.border
            if assessment == "Good":
                assessment_cell.fill = self.positive_fill
            elif assessment == "Moderate":
                assessment_cell.fill = self.neutral_fill
            else:
                assessment_cell.fill = self.negative_fill
            row += 1

        row += 2

        # Interpretation
        row = self._add_section_header(ws, row, "C. INTERPRETATION")

        interpretations = []
        if probability == "LOW":
            interpretations.append("+ Strategy shows strong generalization from in-sample to out-of-sample")
            interpretations.append("+ Parameters are consistent across different time periods")
            interpretations.append("+ Safe to proceed with forward testing")
        elif probability == "MODERATE":
            interpretations.append("~ Strategy shows acceptable but not ideal generalization")
            interpretations.append("~ Consider using more regularization or constraints")
            interpretations.append("~ Proceed with caution - monitor live performance closely")
        else:
            interpretations.append("- Significant performance degradation indicates overfitting")
            interpretations.append("- Parameters vary significantly across time periods")
            interpretations.append("- Consider simplifying strategy or using larger training windows")
            interpretations.append("- NOT recommended for live trading without further refinement")

        for interp in interpretations:
            ws[f'A{row}'] = interp
            if interp.startswith("+"):
                ws[f'A{row}'].font = Font(color=self.COLORS['positive_dark'])
            elif interp.startswith("-"):
                ws[f'A{row}'].font = Font(color=self.COLORS['negative_dark'])
            row += 1

        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20

    def _create_security_comparison(self, wb: Workbook, multi_results: MultiSecurityResults, metrics: Dict[str, Any]):
        """Create Security Comparison sheet."""
        ws = wb.create_sheet("Security Comparison")

        row = 1
        ws[f'A{row}'] = "SECURITY-BY-SECURITY COMPARISON"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:I{row}')
        row += 3

        # Comparison table
        headers = ["Security", "Windows", "Passed", "IS Sortino", "OOS Sortino",
                   "Degradation", "IS Sharpe", "OOS Sharpe", "Status"]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        row += 1

        for symbol in multi_results.securities:
            result = multi_results.individual_results[symbol]
            sec_metrics = metrics['security_metrics'][symbol]

            # Determine status
            if sec_metrics['oos_sortino'] > 0 and abs(sec_metrics['degradation']) < 30:
                status = "ROBUST"
                status_color = self.positive_fill
            elif sec_metrics['oos_sortino'] > 0:
                status = "MODERATE"
                status_color = self.neutral_fill
            else:
                status = "WEAK"
                status_color = self.negative_fill

            row_data = [
                symbol,
                result.total_windows,
                result.windows_passed_constraints,
                f"{result.avg_in_sample_sortino:.4f}",
                f"{result.avg_out_sample_sortino:.4f}",
                f"{result.avg_sortino_degradation_pct:.1f}%",
                f"{result.avg_in_sample_sharpe:.4f}",
                f"{result.avg_out_sample_sharpe:.4f}",
                status
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = self.border

                # Highlight best/worst
                if col_idx == 1:
                    if symbol == multi_results.best_security:
                        cell.fill = self.positive_fill
                    elif symbol == multi_results.worst_security:
                        cell.fill = self.negative_fill

                if col_idx == 9:
                    cell.fill = status_color

            row += 1

        # Average row
        row += 1
        ws.cell(row=row, column=1, value="AVERAGE").font = Font(bold=True)
        ws.cell(row=row, column=4, value=f"{metrics['avg_is_sortino']:.4f}").font = Font(bold=True)
        ws.cell(row=row, column=5, value=f"{metrics['avg_oos_sortino']:.4f}").font = Font(bold=True)
        ws.cell(row=row, column=6, value=f"{metrics['sortino_degradation']:.1f}%").font = Font(bold=True)
        ws.cell(row=row, column=7, value=f"{metrics['avg_is_sharpe']:.4f}").font = Font(bold=True)
        ws.cell(row=row, column=8, value=f"{metrics['avg_oos_sharpe']:.4f}").font = Font(bold=True)

        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 14

    def _create_parameter_robustness(self, wb: Workbook, multi_results: MultiSecurityResults, metrics: Dict[str, Any]):
        """Create Parameter Robustness sheet."""
        ws = wb.create_sheet("Parameter Robustness")

        row = 1
        ws[f'A{row}'] = "PARAMETER ROBUSTNESS ANALYSIS"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:G{row}')
        row += 3

        # Robustness zones
        row = self._add_section_header(ws, row, "A. PARAMETER STABILITY ZONES")

        headers = ["Parameter", "Recommended", "Consistency", "Status", "Zone"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        row += 1

        for param_name, stability in metrics['parameter_stability'].items():
            consistency = stability['consistency']

            # Determine zone
            if consistency >= 80:
                zone = "GREEN - Safe to use"
                zone_fill = self.positive_fill
            elif consistency >= 60:
                zone = "YELLOW - Use with caution"
                zone_fill = self.neutral_fill
            else:
                zone = "RED - Consider fixing"
                zone_fill = self.negative_fill

            ws.cell(row=row, column=1, value=param_name).border = self.border
            ws.cell(row=row, column=2, value=f"{stability['recommended_value']:.4f}" if stability['recommended_value'] else "N/A").border = self.border
            ws.cell(row=row, column=3, value=f"{consistency:.1f}%").border = self.border
            ws.cell(row=row, column=4, value=stability['status']).border = self.border

            zone_cell = ws.cell(row=row, column=5, value=zone)
            zone_cell.border = self.border
            zone_cell.fill = zone_fill

            row += 1

        row += 2

        # Cross-security parameter values
        row = self._add_section_header(ws, row, "B. PARAMETER VALUES ACROSS SECURITIES")

        # Headers
        param_names = list(multi_results.consistent_params.keys())
        ws.cell(row=row, column=1, value="Security").font = self.header_font
        ws.cell(row=row, column=1).fill = self.header_fill
        for col_idx, param in enumerate(param_names, 2):
            cell = ws.cell(row=row, column=col_idx, value=param)
            cell.font = self.header_font
            cell.fill = self.header_fill
        row += 1

        for symbol in multi_results.securities:
            result = multi_results.individual_results[symbol]
            ws.cell(row=row, column=1, value=symbol).border = self.border

            for col_idx, param in enumerate(param_names, 2):
                value = result.most_common_params.get(param, "N/A")
                if isinstance(value, float):
                    value = f"{value:.4f}"
                ws.cell(row=row, column=col_idx, value=value).border = self.border

            row += 1

        # Recommended row
        ws.cell(row=row, column=1, value="RECOMMENDED").font = Font(bold=True)
        ws.cell(row=row, column=1).fill = self.positive_fill
        for col_idx, param in enumerate(param_names, 2):
            value = multi_results.consistent_params.get(param, "N/A")
            if isinstance(value, float):
                value = f"{value:.4f}"
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = Font(bold=True)
            cell.fill = self.positive_fill

        ws.column_dimensions['A'].width = 15
        for col in range(2, len(param_names) + 3):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_sensitivity_dashboard(
        self,
        wb: Workbook,
        multi_results: MultiSecurityResults,
        sensitivity_results_dict: Dict[str, SensitivityResults],
        metrics: Dict[str, Any]
    ):
        """Create Sensitivity Dashboard sheet."""
        ws = wb.create_sheet("Sensitivity Dashboard")

        row = 1
        ws[f'A{row}'] = "PARAMETER SENSITIVITY DASHBOARD"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:G{row}')
        row += 3

        # Overall sensitivity summary
        row = self._add_section_header(ws, row, "A. OVERALL SENSITIVITY SUMMARY")

        all_robust = []
        all_unstable = []

        for symbol, sens in sensitivity_results_dict.items():
            if sens:
                all_robust.extend(sens.most_robust_params)
                all_unstable.extend(sens.least_robust_params)

        # Count occurrences
        from collections import Counter
        robust_counts = Counter(all_robust)
        unstable_counts = Counter(all_unstable)

        ws[f'A{row}'] = "Most Robust Parameters (across all securities):"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        for param, count in robust_counts.most_common(5):
            ws[f'A{row}'] = f"  {param} ({count} securities)"
            ws[f'A{row}'].font = Font(color=self.COLORS['positive_dark'])
            row += 1

        row += 1
        ws[f'A{row}'] = "Least Robust Parameters (across all securities):"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        for param, count in unstable_counts.most_common(5):
            ws[f'A{row}'] = f"  {param} ({count} securities)"
            ws[f'A{row}'].font = Font(color=self.COLORS['negative_dark'])
            row += 1

        row += 2

        # Per-security sensitivity
        row = self._add_section_header(ws, row, "B. PER-SECURITY SENSITIVITY")

        headers = ["Security", "Overall", "Robust Params", "Unstable Params"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
        row += 1

        for symbol, sens in sensitivity_results_dict.items():
            if sens:
                ws.cell(row=row, column=1, value=symbol).border = self.border

                overall_cell = ws.cell(row=row, column=2,
                                        value="ROBUST" if sens.is_overall_robust else "SENSITIVE")
                overall_cell.border = self.border
                overall_cell.fill = self.positive_fill if sens.is_overall_robust else self.negative_fill

                ws.cell(row=row, column=3, value=", ".join(sens.most_robust_params[:3])).border = self.border
                ws.cell(row=row, column=4, value=", ".join(sens.least_robust_params[:3])).border = self.border
                row += 1

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30

    def _create_window_analysis(self, wb: Workbook, multi_results: MultiSecurityResults, metrics: Dict[str, Any]):
        """Create Window Analysis sheet."""
        ws = wb.create_sheet("Window Analysis")

        row = 1
        ws[f'A{row}'] = "WALK-FORWARD WINDOW ANALYSIS"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:H{row}')
        row += 3

        # Aggregate window statistics
        row = self._add_section_header(ws, row, "A. AGGREGATE WINDOW STATISTICS")

        all_is_sortinos = []
        all_oos_sortinos = []
        all_degradations = []

        for wf_results in multi_results.individual_results.values():
            for window in wf_results.windows:
                all_is_sortinos.append(window.in_sample_sortino)
                all_oos_sortinos.append(window.out_sample_sortino)
                all_degradations.append(window.sortino_degradation_pct)

        if all_is_sortinos:
            stats_data = [
                ("Total Windows", str(len(all_is_sortinos))),
                ("IS Sortino (Mean)", f"{np.mean(all_is_sortinos):.4f}"),
                ("IS Sortino (Std)", f"{np.std(all_is_sortinos):.4f}"),
                ("OOS Sortino (Mean)", f"{np.mean(all_oos_sortinos):.4f}"),
                ("OOS Sortino (Std)", f"{np.std(all_oos_sortinos):.4f}"),
                ("OOS Positive (%)", f"{sum(1 for s in all_oos_sortinos if s > 0) / len(all_oos_sortinos) * 100:.1f}%"),
                ("Degradation (Mean)", f"{np.mean(all_degradations):.1f}%"),
                ("Degradation (Std)", f"{np.std(all_degradations):.1f}%"),
            ]

            for metric, value in stats_data:
                ws.cell(row=row, column=1, value=metric).border = self.border
                ws.cell(row=row, column=2, value=value).border = self.border
                row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

    def _create_recommendations(
        self,
        wb: Workbook,
        multi_results: MultiSecurityResults,
        sensitivity_results_dict: Optional[Dict[str, SensitivityResults]],
        metrics: Dict[str, Any]
    ):
        """Create Recommendations sheet."""
        ws = wb.create_sheet("Recommendations")

        row = 1
        ws[f'A{row}'] = "RECOMMENDATIONS & NEXT STEPS"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 3

        # Recommended parameters
        row = self._add_section_header(ws, row, "A. RECOMMENDED PARAMETERS FOR LIVE TRADING")

        headers = ["Parameter", "Value", "Consistency", "Confidence"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
        row += 1

        for param_name, value in multi_results.consistent_params.items():
            consistency = multi_results.param_consistency_scores.get(param_name, 0)

            if consistency >= 80:
                confidence = "HIGH"
                conf_fill = self.positive_fill
            elif consistency >= 60:
                confidence = "MEDIUM"
                conf_fill = self.neutral_fill
            else:
                confidence = "LOW"
                conf_fill = self.negative_fill

            ws.cell(row=row, column=1, value=param_name).border = self.border
            ws.cell(row=row, column=2, value=f"{value:.4f}" if isinstance(value, float) else str(value)).border = self.border
            ws.cell(row=row, column=3, value=f"{consistency:.1f}%").border = self.border

            conf_cell = ws.cell(row=row, column=4, value=confidence)
            conf_cell.border = self.border
            conf_cell.fill = conf_fill

            row += 1

        row += 2

        # Action items
        row = self._add_section_header(ws, row, "B. RECOMMENDED ACTIONS")

        actions = []

        # Based on overfitting score
        if metrics['overfitting_probability'] in ['HIGH', 'VERY HIGH']:
            actions.append("CRITICAL: High overfitting risk detected. Consider:")
            actions.append("  - Increasing training window size")
            actions.append("  - Reducing number of optimized parameters")
            actions.append("  - Adding more constraints to optimization")
        elif metrics['overfitting_probability'] == 'MODERATE':
            actions.append("CAUTION: Moderate overfitting risk. Consider:")
            actions.append("  - Monitoring early live performance closely")
            actions.append("  - Using conservative position sizing initially")

        # Based on robustness
        if metrics['robustness_score'] < 50:
            actions.append("WARNING: Low robustness score. Strategy may need:")
            actions.append("  - Parameter re-optimization with different windows")
            actions.append("  - Review of strategy logic and filters")

        # Based on parameter stability
        unstable_params = [p for p, s in metrics['parameter_stability'].items() if s['status'] == 'Unstable']
        if unstable_params:
            actions.append(f"ATTENTION: Unstable parameters detected: {', '.join(unstable_params)}")
            actions.append("  - Consider fixing these to their recommended values")
            actions.append("  - Or remove them from optimization")

        # Positive actions
        if metrics['robustness_score'] >= 70:
            actions.append("GOOD: Strategy shows strong robustness")
            actions.append("  - Proceed with forward testing")
            actions.append("  - Use recommended parameters above")

        for action in actions:
            ws[f'A{row}'] = action
            if "CRITICAL" in action or "WARNING" in action:
                ws[f'A{row}'].font = Font(color=self.COLORS['negative_dark'], bold=True)
            elif "CAUTION" in action or "ATTENTION" in action:
                ws[f'A{row}'].font = Font(color=self.COLORS['neutral_dark'], bold=True)
            elif "GOOD" in action:
                ws[f'A{row}'].font = Font(color=self.COLORS['positive_dark'], bold=True)
            row += 1

        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12

    def _create_security_detail(
        self,
        wb: Workbook,
        symbol: str,
        wf_results: WalkForwardResults,
        sensitivity_results: Optional[SensitivityResults]
    ):
        """Create detail sheet for individual security."""
        sheet_name = f"Detail_{symbol[:25]}" if len(symbol) > 25 else f"Detail_{symbol}"
        ws = wb.create_sheet(sheet_name)

        row = 1
        ws[f'A{row}'] = f"DETAILED RESULTS: {symbol}"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 3

        # Performance summary
        row = self._add_section_header(ws, row, "A. PERFORMANCE SUMMARY")

        metrics_data = [
            ("Total Windows", wf_results.total_windows),
            ("Passed Constraints", wf_results.windows_passed_constraints),
            ("Avg IS Sortino", f"{wf_results.avg_in_sample_sortino:.4f}"),
            ("Avg OOS Sortino", f"{wf_results.avg_out_sample_sortino:.4f}"),
            ("Sortino Degradation", f"{wf_results.avg_sortino_degradation_pct:.2f}%"),
            ("Avg IS Sharpe", f"{wf_results.avg_in_sample_sharpe:.4f}"),
            ("Avg OOS Sharpe", f"{wf_results.avg_out_sample_sharpe:.4f}"),
        ]

        for label, value in metrics_data:
            ws.cell(row=row, column=1, value=label).border = self.border
            ws.cell(row=row, column=2, value=value).border = self.border
            row += 1

        row += 2

        # Optimal parameters
        row = self._add_section_header(ws, row, "B. OPTIMAL PARAMETERS")

        for param_name, value in wf_results.most_common_params.items():
            ws.cell(row=row, column=1, value=param_name).border = self.border
            ws.cell(row=row, column=2, value=f"{value:.4f}" if isinstance(value, float) else str(value)).border = self.border

            min_val, max_val = wf_results.parameter_ranges.get(param_name, (value, value))
            ws.cell(row=row, column=3, value=f"Range: {min_val:.2f} - {max_val:.2f}").border = self.border
            row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25

    def _add_section_header(self, ws, row: int, title: str) -> int:
        """Add a section header and return next row."""
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = self.subsection_font
        ws[f'A{row}'].fill = self.light_fill
        ws.merge_cells(f'A{row}:D{row}')
        return row + 2

    # ==================== NEW ENHANCED ANALYSIS SHEETS ====================

    def _create_walk_forward_equity_curves(self, wb: Workbook, multi_results: MultiSecurityResults, metrics: Dict[str, Any]):
        """Create Walk-Forward Equity Curves sheet showing OOS performance over time."""
        ws = wb.create_sheet("WF Equity Curves")

        row = 1
        ws[f'A{row}'] = "WALK-FORWARD EQUITY CURVES"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:H{row}')
        row += 2

        ws[f'A{row}'] = "Out-of-sample equity progression across all walk-forward windows"
        ws[f'A{row}'].font = Font(italic=True, color=self.COLORS['dark_gray'])
        row += 3

        # Create data table for each security
        for symbol, wf_results in multi_results.individual_results.items():
            row = self._add_section_header(ws, row, f"Security: {symbol}")

            # Headers
            headers = ["Window", "Test Period", "Starting Equity", "Ending Equity",
                      "Return %", "OOS Sortino", "Cumulative Return %"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
            row += 1

            cumulative_return = 0
            starting_equity = 100  # Normalized to 100

            for window in wf_results.windows:
                ending_equity = starting_equity * (1 + window.out_sample_total_return_pct / 100)
                cumulative_return += window.out_sample_total_return_pct

                ws.cell(row=row, column=1, value=f"Window {window.window_id}")
                ws.cell(row=row, column=2, value=f"{window.test_start.strftime('%Y-%m-%d')} to {window.test_end.strftime('%Y-%m-%d')}")
                ws.cell(row=row, column=3, value=f"{starting_equity:.2f}")
                ws.cell(row=row, column=4, value=f"{ending_equity:.2f}")

                ret_cell = ws.cell(row=row, column=5, value=f"{window.out_sample_total_return_pct:.2f}%")
                ret_cell.fill = self.positive_fill if window.out_sample_total_return_pct > 0 else self.negative_fill

                ws.cell(row=row, column=6, value=f"{window.out_sample_sortino:.4f}")

                cum_cell = ws.cell(row=row, column=7, value=f"{cumulative_return:.2f}%")
                cum_cell.font = Font(bold=True)
                cum_cell.fill = self.positive_fill if cumulative_return > 0 else self.negative_fill

                starting_equity = ending_equity
                row += 1

            # Summary
            row += 1
            ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True, size=11)
            ws.cell(row=row, column=7, value=f"{cumulative_return:.2f}%").font = Font(bold=True, size=11)
            row += 3

        # Visualization if matplotlib available
        if self.include_matplotlib_charts and self.viz:
            try:
                equity_img = self._create_wf_equity_chart(multi_results)
                if equity_img:
                    img = Image(equity_img)
                    img.width = 700
                    img.height = 400
                    ws.add_image(img, f'A{row}')
            except Exception as e:
                logger.warning(f"Could not create WF equity chart: {e}")

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        for col in ['C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 18

    def _create_wf_equity_chart(self, multi_results: MultiSecurityResults):
        """Create walk-forward equity chart using matplotlib."""
        if not MATPLOTLIB_AVAILABLE:
            return None

        import matplotlib.pyplot as plt
        from io import BytesIO

        fig, ax = plt.subplots(figsize=(14, 6))

        for symbol, wf_results in multi_results.individual_results.items():
            dates = []
            equity = [100]  # Start at 100

            for window in wf_results.windows:
                dates.append(window.test_end)
                new_equity = equity[-1] * (1 + window.out_sample_total_return_pct / 100)
                equity.append(new_equity)

            if dates:
                ax.plot(dates, equity[1:], marker='o', label=symbol, linewidth=2, markersize=4)

        ax.axhline(100, color='black', linestyle='--', linewidth=1, alpha=0.5, label='Starting Value')
        ax.set_xlabel('Date', fontsize=10)
        ax.set_ylabel('Equity (normalized to 100)', fontsize=10)
        ax.set_title('Walk-Forward Out-of-Sample Equity Curves', fontsize=12, fontweight='bold')
        ax.legend(loc='best', fontsize=8)
        ax.grid(True, alpha=0.3)

        plt.tight_layout()

        buf = BytesIO()
        fig.savefig(buf, format='png', dpi=150)
        plt.close(fig)
        buf.seek(0)
        return buf

    def _create_parameter_stability(self, wb: Workbook, multi_results: MultiSecurityResults, metrics: Dict[str, Any]):
        """Create Parameter Stability Over Time sheet."""
        ws = wb.create_sheet("Parameter Stability")

        row = 1
        ws[f'A{row}'] = "PARAMETER STABILITY OVER TIME"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:H{row}')
        row += 2

        ws[f'A{row}'] = "Analysis of how optimal parameters change across walk-forward windows"
        ws[f'A{row}'].font = Font(italic=True, color=self.COLORS['dark_gray'])
        row += 3

        # Check for empty data
        if not multi_results.individual_results:
            ws[f'A{row}'] = "No data available for parameter stability analysis"
            return

        # Aggregate parameter values across all securities and windows
        param_values_by_window = {}  # {param_name: {window_id: [values across securities]}}

        for symbol, wf_results in multi_results.individual_results.items():
            for window in wf_results.windows:
                for param_name, value in window.best_params.items():
                    if param_name not in param_values_by_window:
                        param_values_by_window[param_name] = {}
                    if window.window_id not in param_values_by_window[param_name]:
                        param_values_by_window[param_name][window.window_id] = []
                    try:
                        param_values_by_window[param_name][window.window_id].append(float(value))
                    except (ValueError, TypeError):
                        pass

        # Section A: Parameter Drift Analysis
        row = self._add_section_header(ws, row, "A. PARAMETER DRIFT ANALYSIS")

        headers = ["Parameter", "Min Value", "Max Value", "Range", "Std Dev", "CV %", "Drift Score", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        row += 1

        for param_name, windows_data in param_values_by_window.items():
            all_values = [v for values in windows_data.values() for v in values]
            if not all_values:
                continue

            min_val = min(all_values)
            max_val = max(all_values)
            range_val = max_val - min_val
            std_val = np.std(all_values)
            mean_val = np.mean(all_values)
            cv = (std_val / mean_val * 100) if mean_val != 0 else 0

            # Drift score: higher = more unstable (0-100)
            # Based on coefficient of variation
            drift_score = min(100, cv * 2)

            if drift_score < 20:
                status = "Stable"
                status_fill = self.positive_fill
            elif drift_score < 50:
                status = "Moderate"
                status_fill = self.neutral_fill
            else:
                status = "Unstable"
                status_fill = self.negative_fill

            ws.cell(row=row, column=1, value=param_name).font = Font(bold=True)
            ws.cell(row=row, column=2, value=f"{min_val:.4f}")
            ws.cell(row=row, column=3, value=f"{max_val:.4f}")
            ws.cell(row=row, column=4, value=f"{range_val:.4f}")
            ws.cell(row=row, column=5, value=f"{std_val:.4f}")
            ws.cell(row=row, column=6, value=f"{cv:.1f}%")
            ws.cell(row=row, column=7, value=f"{drift_score:.0f}")

            status_cell = ws.cell(row=row, column=8, value=status)
            status_cell.fill = status_fill
            status_cell.font = Font(bold=True)
            row += 1

        row += 2

        # Section B: Window-by-Window Parameter Values
        row = self._add_section_header(ws, row, "B. PARAMETER VALUES BY WINDOW")

        # Get all window IDs
        all_window_ids = set()
        for wf_results in multi_results.individual_results.values():
            for window in wf_results.windows:
                all_window_ids.add(window.window_id)
        all_window_ids = sorted(all_window_ids)

        # Headers
        headers = ["Parameter"] + [f"W{wid}" for wid in all_window_ids] + ["Trend"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        row += 1

        for param_name, windows_data in param_values_by_window.items():
            ws.cell(row=row, column=1, value=param_name).font = Font(bold=True)

            values_sequence = []
            for col_idx, wid in enumerate(all_window_ids, 2):
                if wid in windows_data:
                    avg_val = np.mean(windows_data[wid])
                    values_sequence.append(avg_val)
                    ws.cell(row=row, column=col_idx, value=f"{avg_val:.3f}")
                else:
                    values_sequence.append(None)
                    ws.cell(row=row, column=col_idx, value="-")

            # Calculate trend
            valid_values = [v for v in values_sequence if v is not None]
            if len(valid_values) >= 2:
                first_half = np.mean(valid_values[:len(valid_values)//2])
                second_half = np.mean(valid_values[len(valid_values)//2:])
                if second_half > first_half * 1.1:
                    trend = "↑ Increasing"
                elif second_half < first_half * 0.9:
                    trend = "↓ Decreasing"
                else:
                    trend = "→ Stable"
            else:
                trend = "-"

            ws.cell(row=row, column=len(headers), value=trend)
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 20
        for col_idx in range(2, len(all_window_ids) + 3):
            ws.column_dimensions[get_column_letter(col_idx)].width = 10

    def _create_constraint_analysis(self, wb: Workbook, multi_results: MultiSecurityResults, metrics: Dict[str, Any]):
        """Create Constraint Hit Analysis sheet."""
        ws = wb.create_sheet("Constraint Analysis")

        row = 1
        ws[f'A{row}'] = "CONSTRAINT HIT ANALYSIS"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:H{row}')
        row += 2

        ws[f'A{row}'] = "Analysis of which optimization constraints are being triggered most frequently"
        ws[f'A{row}'].font = Font(italic=True, color=self.COLORS['dark_gray'])
        row += 3

        # Analyze constraint violations across all windows
        constraint_failures = {
            'low_sortino': 0,
            'negative_sortino': 0,
            'high_drawdown': 0,
            'low_profit_factor': 0,
            'insufficient_trades': 0,
            'passed': 0
        }

        total_windows = 0

        for symbol, wf_results in multi_results.individual_results.items():
            for window in wf_results.windows:
                total_windows += 1

                # Analyze OOS metrics for constraint issues
                if window.out_sample_sortino < 0:
                    constraint_failures['negative_sortino'] += 1
                elif window.out_sample_sortino < 0.5:
                    constraint_failures['low_sortino'] += 1

                if window.out_sample_max_drawdown_pct > 25:
                    constraint_failures['high_drawdown'] += 1

                if window.out_sample_profit_factor < 1.0:
                    constraint_failures['low_profit_factor'] += 1

                if window.out_sample_num_trades < 5:
                    constraint_failures['insufficient_trades'] += 1

        constraint_failures['passed'] = metrics.get('total_passed', 0)

        # Section A: Summary
        row = self._add_section_header(ws, row, "A. CONSTRAINT VIOLATION SUMMARY")

        headers = ["Constraint Type", "Violations", "% of Windows", "Severity", "Recommendation"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        row += 1

        constraint_info = [
            ("Negative Sortino (OOS < 0)", constraint_failures['negative_sortino'], "HIGH",
             "Strategy may need fundamental rework"),
            ("Low Sortino (OOS < 0.5)", constraint_failures['low_sortino'], "MEDIUM",
             "Consider longer optimization periods"),
            ("High Drawdown (> 25%)", constraint_failures['high_drawdown'], "HIGH",
             "Add drawdown constraints or reduce position sizes"),
            ("Low Profit Factor (< 1.0)", constraint_failures['low_profit_factor'], "HIGH",
             "Strategy is losing money on average"),
            ("Insufficient Trades (< 5)", constraint_failures['insufficient_trades'], "LOW",
             "May need longer test periods or more active signals"),
        ]

        for name, count, severity, rec in constraint_info:
            pct = (count / total_windows * 100) if total_windows > 0 else 0

            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{pct:.1f}%")

            severity_cell = ws.cell(row=row, column=4, value=severity)
            if severity == "HIGH":
                severity_cell.fill = self.negative_fill
            elif severity == "MEDIUM":
                severity_cell.fill = self.neutral_fill
            else:
                severity_cell.fill = self.positive_fill

            ws.cell(row=row, column=5, value=rec)
            row += 1

        row += 2

        # Section B: By Security Breakdown
        row = self._add_section_header(ws, row, "B. CONSTRAINT FAILURES BY SECURITY")

        headers = ["Security", "Total Windows", "Passed", "Pass Rate", "Primary Issue"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        row += 1

        for symbol, wf_results in multi_results.individual_results.items():
            total = wf_results.total_windows
            passed = wf_results.windows_passed_constraints
            pass_rate = (passed / total * 100) if total > 0 else 0

            # Determine primary issue
            neg_sortino_count = sum(1 for w in wf_results.windows if w.out_sample_sortino < 0)
            high_dd_count = sum(1 for w in wf_results.windows if w.out_sample_max_drawdown_pct > 25)

            if neg_sortino_count > total * 0.3:
                primary_issue = "Negative OOS Returns"
            elif high_dd_count > total * 0.3:
                primary_issue = "High Drawdowns"
            elif pass_rate < 50:
                primary_issue = "Multiple Failures"
            else:
                primary_issue = "None - Performing Well"

            ws.cell(row=row, column=1, value=symbol).font = Font(bold=True)
            ws.cell(row=row, column=2, value=total)
            ws.cell(row=row, column=3, value=passed)

            rate_cell = ws.cell(row=row, column=4, value=f"{pass_rate:.1f}%")
            rate_cell.fill = self.positive_fill if pass_rate >= 70 else (
                self.neutral_fill if pass_rate >= 50 else self.negative_fill)

            ws.cell(row=row, column=5, value=primary_issue)
            row += 1

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 35

    def _create_window_comparison_table(self, wb: Workbook, multi_results: MultiSecurityResults, metrics: Dict[str, Any]):
        """Create detailed Window-by-Window Comparison Table."""
        ws = wb.create_sheet("Window Comparison")

        row = 1
        ws[f'A{row}'] = "WINDOW-BY-WINDOW COMPARISON TABLE"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:N{row}')
        row += 2

        ws[f'A{row}'] = "Comprehensive breakdown of all walk-forward windows across all securities"
        ws[f'A{row}'].font = Font(italic=True, color=self.COLORS['dark_gray'])
        row += 3

        # Headers
        headers = ["Security", "Window", "Train Period", "Test Period",
                  "IS Sortino", "OOS Sortino", "Degradation", "IS Sharpe", "OOS Sharpe",
                  "IS Return %", "OOS Return %", "IS Trades", "OOS Trades", "Max DD %"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        row += 1

        # Data rows
        all_rows_data = []
        for symbol, wf_results in multi_results.individual_results.items():
            for window in wf_results.windows:
                row_data = {
                    'symbol': symbol,
                    'window': window.window_id,
                    'train_period': f"{window.train_start.strftime('%Y-%m-%d')} to {window.train_end.strftime('%Y-%m-%d')}",
                    'test_period': f"{window.test_start.strftime('%Y-%m-%d')} to {window.test_end.strftime('%Y-%m-%d')}",
                    'is_sortino': window.in_sample_sortino,
                    'oos_sortino': window.out_sample_sortino,
                    'degradation': window.sortino_degradation_pct,
                    'is_sharpe': window.in_sample_sharpe,
                    'oos_sharpe': window.out_sample_sharpe,
                    'is_return': window.in_sample_total_return_pct,
                    'oos_return': window.out_sample_total_return_pct,
                    'is_trades': window.in_sample_num_trades,
                    'oos_trades': window.out_sample_num_trades,
                    'max_dd': window.out_sample_max_drawdown_pct
                }
                all_rows_data.append(row_data)

        # Sort by OOS Sortino descending to highlight best performers
        all_rows_data.sort(key=lambda x: x['oos_sortino'], reverse=True)

        for data in all_rows_data:
            ws.cell(row=row, column=1, value=data['symbol']).font = Font(bold=True)
            ws.cell(row=row, column=2, value=f"W{data['window']}")
            ws.cell(row=row, column=3, value=data['train_period'])
            ws.cell(row=row, column=4, value=data['test_period'])

            # IS Sortino
            ws.cell(row=row, column=5, value=f"{data['is_sortino']:.4f}")

            # OOS Sortino with color
            oos_cell = ws.cell(row=row, column=6, value=f"{data['oos_sortino']:.4f}")
            if data['oos_sortino'] > 1:
                oos_cell.fill = self.positive_fill
            elif data['oos_sortino'] < 0:
                oos_cell.fill = self.negative_fill

            # Degradation with color
            deg_cell = ws.cell(row=row, column=7, value=f"{data['degradation']:.1f}%")
            if abs(data['degradation']) < 30:
                deg_cell.fill = self.positive_fill
            elif abs(data['degradation']) > 50:
                deg_cell.fill = self.negative_fill

            ws.cell(row=row, column=8, value=f"{data['is_sharpe']:.4f}")
            ws.cell(row=row, column=9, value=f"{data['oos_sharpe']:.4f}")

            ws.cell(row=row, column=10, value=f"{data['is_return']:.2f}%")

            ret_cell = ws.cell(row=row, column=11, value=f"{data['oos_return']:.2f}%")
            ret_cell.fill = self.positive_fill if data['oos_return'] > 0 else self.negative_fill

            ws.cell(row=row, column=12, value=data['is_trades'])
            ws.cell(row=row, column=13, value=data['oos_trades'])

            dd_cell = ws.cell(row=row, column=14, value=f"{data['max_dd']:.1f}%")
            if data['max_dd'] > 25:
                dd_cell.fill = self.negative_fill

            row += 1

        # Column widths
        widths = [12, 8, 22, 22, 10, 10, 10, 10, 10, 10, 10, 9, 9, 10]
        for col_idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _create_risk_metrics_dashboard(self, wb: Workbook, multi_results: MultiSecurityResults, metrics: Dict[str, Any]):
        """Create Risk Metrics Dashboard comparing IS vs OOS risk metrics."""
        ws = wb.create_sheet("Risk Dashboard")

        row = 1
        ws[f'A{row}'] = "RISK METRICS DASHBOARD"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:H{row}')
        row += 2

        ws[f'A{row}'] = "In-sample vs out-of-sample risk metrics comparison"
        ws[f'A{row}'].font = Font(italic=True, color=self.COLORS['dark_gray'])
        row += 3

        # Check for empty data
        if not multi_results.individual_results:
            ws[f'A{row}'] = "No data available for risk metrics analysis"
            return

        # Section A: Aggregate Risk Metrics
        row = self._add_section_header(ws, row, "A. AGGREGATE RISK METRICS (AVERAGED)")

        # Calculate aggregate risk metrics
        all_is_dd = []
        all_oos_dd = []
        all_is_pf = []
        all_oos_pf = []

        for wf_results in multi_results.individual_results.values():
            for window in wf_results.windows:
                all_is_dd.append(window.in_sample_max_drawdown_pct)
                all_oos_dd.append(window.out_sample_max_drawdown_pct)
                all_is_pf.append(window.in_sample_profit_factor)
                all_oos_pf.append(window.out_sample_profit_factor)

        headers = ["Metric", "In-Sample Avg", "OOS Avg", "Change", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        row += 1

        risk_metrics = [
            ("Max Drawdown %", np.mean(all_is_dd) if all_is_dd else 0, np.mean(all_oos_dd) if all_oos_dd else 0, "lower_better"),
            ("Sortino Ratio", metrics['avg_is_sortino'], metrics['avg_oos_sortino'], "higher_better"),
            ("Sharpe Ratio", metrics['avg_is_sharpe'], metrics['avg_oos_sharpe'], "higher_better"),
            ("Profit Factor", np.mean(all_is_pf) if all_is_pf else 0, np.mean(all_oos_pf) if all_oos_pf else 0, "higher_better"),
        ]

        for name, is_val, oos_val, direction in risk_metrics:
            change = ((oos_val - is_val) / is_val * 100) if is_val != 0 else 0

            if direction == "higher_better":
                status = "GOOD" if change > -20 else ("CAUTION" if change > -50 else "POOR")
            else:
                status = "GOOD" if change < 20 else ("CAUTION" if change < 50 else "POOR")

            ws.cell(row=row, column=1, value=name).font = Font(bold=True)
            ws.cell(row=row, column=2, value=f"{is_val:.4f}")
            ws.cell(row=row, column=3, value=f"{oos_val:.4f}")

            change_cell = ws.cell(row=row, column=4, value=f"{change:+.1f}%")
            if abs(change) < 20:
                change_cell.fill = self.positive_fill
            elif abs(change) > 50:
                change_cell.fill = self.negative_fill
            else:
                change_cell.fill = self.neutral_fill

            status_cell = ws.cell(row=row, column=5, value=status)
            if status == "GOOD":
                status_cell.fill = self.positive_fill
            elif status == "POOR":
                status_cell.fill = self.negative_fill
            else:
                status_cell.fill = self.neutral_fill
            row += 1

        row += 2

        # Section B: Risk Metrics by Security
        row = self._add_section_header(ws, row, "B. RISK METRICS BY SECURITY")

        headers = ["Security", "Avg IS DD%", "Avg OOS DD%", "Max OOS DD%", "DD Stability",
                  "Avg IS PF", "Avg OOS PF", "Risk Rating"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        row += 1

        for symbol, wf_results in multi_results.individual_results.items():
            is_dds = [w.in_sample_max_drawdown_pct for w in wf_results.windows]
            oos_dds = [w.out_sample_max_drawdown_pct for w in wf_results.windows]
            is_pfs = [w.in_sample_profit_factor for w in wf_results.windows]
            oos_pfs = [w.out_sample_profit_factor for w in wf_results.windows]

            avg_is_dd = np.mean(is_dds) if is_dds else 0
            avg_oos_dd = np.mean(oos_dds) if oos_dds else 0
            max_oos_dd = max(oos_dds) if oos_dds else 0
            # Clamp stability to 0-100 range (can go negative if std > mean)
            dd_stability = max(0, min(100, 100 - (np.std(oos_dds) / avg_oos_dd * 100))) if avg_oos_dd > 0 else 0

            avg_is_pf = np.mean(is_pfs) if is_pfs else 0
            avg_oos_pf = np.mean(oos_pfs) if oos_pfs else 0

            # Calculate risk rating
            risk_score = 0
            if avg_oos_dd < 15:
                risk_score += 30
            elif avg_oos_dd < 25:
                risk_score += 15
            if avg_oos_pf > 1.5:
                risk_score += 40
            elif avg_oos_pf > 1.0:
                risk_score += 20
            if dd_stability > 70:
                risk_score += 30
            elif dd_stability > 50:
                risk_score += 15

            if risk_score >= 70:
                risk_rating = "LOW RISK"
                rating_fill = self.positive_fill
            elif risk_score >= 40:
                risk_rating = "MODERATE"
                rating_fill = self.neutral_fill
            else:
                risk_rating = "HIGH RISK"
                rating_fill = self.negative_fill

            ws.cell(row=row, column=1, value=symbol).font = Font(bold=True)
            ws.cell(row=row, column=2, value=f"{avg_is_dd:.1f}%")
            ws.cell(row=row, column=3, value=f"{avg_oos_dd:.1f}%")
            ws.cell(row=row, column=4, value=f"{max_oos_dd:.1f}%")
            ws.cell(row=row, column=5, value=f"{dd_stability:.0f}%")
            ws.cell(row=row, column=6, value=f"{avg_is_pf:.2f}")
            ws.cell(row=row, column=7, value=f"{avg_oos_pf:.2f}")

            rating_cell = ws.cell(row=row, column=8, value=risk_rating)
            rating_cell.fill = rating_fill
            rating_cell.font = Font(bold=True)
            row += 1

        ws.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 12

    def _create_best_worst_window_analysis(self, wb: Workbook, multi_results: MultiSecurityResults, metrics: Dict[str, Any]):
        """Create Best/Worst Window Deep Dive analysis."""
        ws = wb.create_sheet("Best-Worst Analysis")

        row = 1
        ws[f'A{row}'] = "BEST & WORST WINDOW ANALYSIS"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:H{row}')
        row += 2

        ws[f'A{row}'] = "Deep dive into the best and worst performing walk-forward windows"
        ws[f'A{row}'].font = Font(italic=True, color=self.COLORS['dark_gray'])
        row += 3

        # Collect all windows across all securities
        all_windows = []
        for symbol, wf_results in multi_results.individual_results.items():
            for window in wf_results.windows:
                all_windows.append((symbol, window))

        if not all_windows:
            ws[f'A{row}'] = "No windows available for analysis"
            return

        # Sort by OOS Sortino
        all_windows.sort(key=lambda x: x[1].out_sample_sortino, reverse=True)

        # Best windows (top 3)
        row = self._add_section_header(ws, row, "A. BEST PERFORMING WINDOWS (by OOS Sortino)")

        for i, (symbol, window) in enumerate(all_windows[:min(3, len(all_windows))], 1):
            row = self._write_window_detail(ws, row, f"#{i} BEST", symbol, window)
            row += 2

        row += 2

        # Worst windows (bottom 3)
        row = self._add_section_header(ws, row, "B. WORST PERFORMING WINDOWS (by OOS Sortino)")

        worst_windows = all_windows[-min(3, len(all_windows)):][::-1]  # Reverse to show worst first
        for i, (symbol, window) in enumerate(worst_windows, 1):
            row = self._write_window_detail(ws, row, f"#{i} WORST", symbol, window)
            row += 2

        row += 2

        # Section C: What separates best from worst
        row = self._add_section_header(ws, row, "C. KEY DIFFERENTIATORS")

        best_3 = all_windows[:min(3, len(all_windows))]
        worst_3 = all_windows[-min(3, len(all_windows)):]

        best_avg_trades = np.mean([w[1].out_sample_num_trades for w in best_3])
        worst_avg_trades = np.mean([w[1].out_sample_num_trades for w in worst_3])

        best_avg_dd = np.mean([w[1].out_sample_max_drawdown_pct for w in best_3])
        worst_avg_dd = np.mean([w[1].out_sample_max_drawdown_pct for w in worst_3])

        best_avg_degradation = np.mean([abs(w[1].sortino_degradation_pct) for w in best_3])
        worst_avg_degradation = np.mean([abs(w[1].sortino_degradation_pct) for w in worst_3])

        differentiators = [
            ("Avg OOS Trades", f"{best_avg_trades:.1f}", f"{worst_avg_trades:.1f}"),
            ("Avg Max Drawdown", f"{best_avg_dd:.1f}%", f"{worst_avg_dd:.1f}%"),
            ("Avg Degradation", f"{best_avg_degradation:.1f}%", f"{worst_avg_degradation:.1f}%"),
        ]

        headers = ["Metric", "Best Windows Avg", "Worst Windows Avg"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        row += 1

        for metric, best_val, worst_val in differentiators:
            ws.cell(row=row, column=1, value=metric).font = Font(bold=True)
            ws.cell(row=row, column=2, value=best_val).fill = self.positive_fill
            ws.cell(row=row, column=3, value=worst_val).fill = self.negative_fill
            row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

    def _write_window_detail(self, ws, row: int, label: str, symbol: str, window) -> int:
        """Write detailed window information."""
        ws.cell(row=row, column=1, value=f"{label}: {symbol} - Window {window.window_id}").font = Font(bold=True, size=11)
        row += 1

        details = [
            ("Test Period", f"{window.test_start.strftime('%Y-%m-%d')} to {window.test_end.strftime('%Y-%m-%d')}"),
            ("OOS Sortino", f"{window.out_sample_sortino:.4f}"),
            ("OOS Sharpe", f"{window.out_sample_sharpe:.4f}"),
            ("OOS Return", f"{window.out_sample_total_return_pct:.2f}%"),
            ("OOS Max Drawdown", f"{window.out_sample_max_drawdown_pct:.1f}%"),
            ("OOS Trades", str(window.out_sample_num_trades)),
            ("Sortino Degradation", f"{window.sortino_degradation_pct:.1f}%"),
            ("IS Sortino", f"{window.in_sample_sortino:.4f}"),
        ]

        for label_txt, value in details:
            ws.cell(row=row, column=1, value=label_txt)
            ws.cell(row=row, column=2, value=value)
            row += 1

        return row

    def _create_cross_security_correlation(self, wb: Workbook, multi_results: MultiSecurityResults, metrics: Dict[str, Any]):
        """Create Cross-Security Correlation analysis."""
        ws = wb.create_sheet("Security Correlation")

        row = 1
        ws[f'A{row}'] = "CROSS-SECURITY CORRELATION ANALYSIS"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:H{row}')
        row += 2

        ws[f'A{row}'] = "Correlation of OOS performance across securities (diversification analysis)"
        ws[f'A{row}'].font = Font(italic=True, color=self.COLORS['dark_gray'])
        row += 3

        securities = list(multi_results.individual_results.keys())

        if len(securities) < 2:
            ws[f'A{row}'] = "At least 2 securities required for correlation analysis"
            return

        # Build OOS returns matrix
        # Align by window ID
        all_window_ids = set()
        for wf_results in multi_results.individual_results.values():
            for window in wf_results.windows:
                all_window_ids.add(window.window_id)
        all_window_ids = sorted(all_window_ids)

        returns_by_security = {}
        for symbol, wf_results in multi_results.individual_results.items():
            returns_by_security[symbol] = {}
            for window in wf_results.windows:
                returns_by_security[symbol][window.window_id] = window.out_sample_total_return_pct

        # Calculate correlation matrix
        row = self._add_section_header(ws, row, "A. OOS RETURNS CORRELATION MATRIX")

        # Header row
        ws.cell(row=row, column=1, value="").font = self.header_font
        for col, symbol in enumerate(securities, 2):
            cell = ws.cell(row=row, column=col, value=symbol[:10])
            cell.font = self.header_font
            cell.fill = self.header_fill
        row += 1

        for i, symbol1 in enumerate(securities):
            ws.cell(row=row, column=1, value=symbol1[:10]).font = Font(bold=True)
            ws.cell(row=row, column=1).fill = self.header_fill

            for j, symbol2 in enumerate(securities, 2):
                # Calculate correlation
                returns1 = [returns_by_security[symbol1].get(wid, None) for wid in all_window_ids]
                returns2 = [returns_by_security[securities[j-2]].get(wid, None) for wid in all_window_ids]

                # Only use windows where both have data
                paired_returns = [(r1, r2) for r1, r2 in zip(returns1, returns2)
                                 if r1 is not None and r2 is not None]

                if len(paired_returns) >= 3:
                    r1_list = [p[0] for p in paired_returns]
                    r2_list = [p[1] for p in paired_returns]
                    corr = np.corrcoef(r1_list, r2_list)[0, 1]
                    # Handle NaN (can occur if all values are identical)
                    if np.isnan(corr):
                        corr = 1.0 if symbol1 == securities[j-2] else 0.0
                else:
                    corr = 0

                cell = ws.cell(row=row, column=j, value=f"{corr:.2f}")

                # Color based on correlation
                if i == j - 2:
                    cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                elif corr > 0.7:
                    cell.fill = self.negative_fill  # High correlation = less diversification
                elif corr < 0.3:
                    cell.fill = self.positive_fill  # Low correlation = good diversification
                else:
                    cell.fill = self.neutral_fill

            row += 1

        row += 2

        # Section B: Diversification Assessment
        row = self._add_section_header(ws, row, "B. DIVERSIFICATION ASSESSMENT")

        # Calculate average correlation (excluding diagonal)
        total_corr = 0
        count = 0
        for i, symbol1 in enumerate(securities):
            for j, symbol2 in enumerate(securities):
                if i < j:
                    returns1 = [returns_by_security[symbol1].get(wid, None) for wid in all_window_ids]
                    returns2 = [returns_by_security[symbol2].get(wid, None) for wid in all_window_ids]
                    paired = [(r1, r2) for r1, r2 in zip(returns1, returns2)
                             if r1 is not None and r2 is not None]
                    if len(paired) >= 3:
                        corr = np.corrcoef([p[0] for p in paired], [p[1] for p in paired])[0, 1]
                        # Skip NaN correlations (can occur if all values identical)
                        if not np.isnan(corr):
                            total_corr += corr
                            count += 1

        avg_corr = total_corr / count if count > 0 else 0

        if avg_corr < 0.3:
            diversification = "EXCELLENT - Securities show low correlation"
            div_fill = self.positive_fill
        elif avg_corr < 0.5:
            diversification = "GOOD - Securities have moderate correlation"
            div_fill = self.positive_fill
        elif avg_corr < 0.7:
            diversification = "FAIR - Securities show high correlation"
            div_fill = self.neutral_fill
        else:
            diversification = "POOR - Securities are highly correlated"
            div_fill = self.negative_fill

        ws.cell(row=row, column=1, value="Average Correlation:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"{avg_corr:.2f}")
        row += 1

        ws.cell(row=row, column=1, value="Diversification Assessment:").font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=diversification)
        cell.fill = div_fill
        cell.font = Font(bold=True)

        ws.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 12

    def _create_monte_carlo_analysis(self, wb: Workbook, multi_results: MultiSecurityResults, metrics: Dict[str, Any]):
        """Create Monte Carlo Confidence Intervals analysis."""
        ws = wb.create_sheet("Monte Carlo")

        row = 1
        ws[f'A{row}'] = "MONTE CARLO SIMULATION"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:H{row}')
        row += 2

        ws[f'A{row}'] = "Bootstrap analysis of OOS returns to estimate confidence intervals"
        ws[f'A{row}'].font = Font(italic=True, color=self.COLORS['dark_gray'])
        row += 3

        # Collect all OOS returns
        all_oos_returns = []
        for wf_results in multi_results.individual_results.values():
            for window in wf_results.windows:
                all_oos_returns.append(window.out_sample_total_return_pct)

        if len(all_oos_returns) < 5:
            ws[f'A{row}'] = "Insufficient data for Monte Carlo analysis (need at least 5 OOS windows)"
            return

        # Bootstrap simulation
        n_simulations = 1000
        n_windows = len(all_oos_returns)

        # Simulate cumulative returns
        simulated_cumulative_returns = []
        for _ in range(n_simulations):
            # Sample with replacement
            sampled = np.random.choice(all_oos_returns, size=n_windows, replace=True)
            cumulative = sum(sampled)
            simulated_cumulative_returns.append(cumulative)

        # Calculate statistics
        mean_return = np.mean(simulated_cumulative_returns)
        std_return = np.std(simulated_cumulative_returns)
        percentiles = {
            5: np.percentile(simulated_cumulative_returns, 5),
            25: np.percentile(simulated_cumulative_returns, 25),
            50: np.percentile(simulated_cumulative_returns, 50),
            75: np.percentile(simulated_cumulative_returns, 75),
            95: np.percentile(simulated_cumulative_returns, 95)
        }

        # Probability of positive return
        prob_positive = sum(1 for r in simulated_cumulative_returns if r > 0) / n_simulations * 100

        # Section A: Summary Statistics
        row = self._add_section_header(ws, row, "A. MONTE CARLO SUMMARY (1,000 simulations)")

        stats_data = [
            ("Number of OOS Periods", str(n_windows)),
            ("Expected Cumulative Return", f"{mean_return:.2f}%"),
            ("Standard Deviation", f"{std_return:.2f}%"),
            ("Probability of Positive Return", f"{prob_positive:.1f}%"),
        ]

        for label, value in stats_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 2

        # Section B: Confidence Intervals
        row = self._add_section_header(ws, row, "B. CONFIDENCE INTERVALS")

        headers = ["Percentile", "Cumulative Return", "Interpretation"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        row += 1

        ci_interpretations = [
            (5, "5% worst case"),
            (25, "25th percentile (pessimistic)"),
            (50, "Median expectation"),
            (75, "75th percentile (optimistic)"),
            (95, "5% best case"),
        ]

        for pct, interp in ci_interpretations:
            ws.cell(row=row, column=1, value=f"{pct}%").font = Font(bold=True)

            val_cell = ws.cell(row=row, column=2, value=f"{percentiles[pct]:.2f}%")
            if percentiles[pct] > 0:
                val_cell.fill = self.positive_fill
            else:
                val_cell.fill = self.negative_fill

            ws.cell(row=row, column=3, value=interp)
            row += 1

        row += 2

        # Section C: Risk Assessment
        row = self._add_section_header(ws, row, "C. RISK ASSESSMENT")

        # Value at Risk
        var_5 = percentiles[5]
        cvar_values = [r for r in simulated_cumulative_returns if r <= var_5]
        cvar_5 = np.mean(cvar_values) if cvar_values else var_5  # Fallback to VaR if no values

        risk_data = [
            ("Value at Risk (95%)", f"{var_5:.2f}%", "Maximum expected loss with 95% confidence"),
            ("Conditional VaR (95%)", f"{cvar_5:.2f}%", "Expected loss when worst 5% scenarios occur"),
        ]

        for label, value, desc in risk_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=3, value=desc).font = Font(italic=True, color=self.COLORS['dark_gray'])
            row += 1

        row += 2

        # Interpretation
        ws.cell(row=row, column=1, value="INTERPRETATION:").font = Font(bold=True, size=11)
        row += 1

        if prob_positive > 80:
            interpretation = f"Strategy shows strong robustness with {prob_positive:.0f}% probability of positive returns"
            interp_fill = self.positive_fill
        elif prob_positive > 60:
            interpretation = f"Strategy shows moderate robustness with {prob_positive:.0f}% probability of positive returns"
            interp_fill = self.neutral_fill
        else:
            interpretation = f"Strategy shows low robustness with only {prob_positive:.0f}% probability of positive returns"
            interp_fill = self.negative_fill

        cell = ws.cell(row=row, column=1, value=interpretation)
        cell.fill = interp_fill
        cell.font = Font(bold=True)

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 45

    def _create_trade_statistics_by_window(self, wb: Workbook, multi_results: MultiSecurityResults, metrics: Dict[str, Any]):
        """Create Trade Statistics by Window analysis."""
        ws = wb.create_sheet("Trade Statistics")

        row = 1
        ws[f'A{row}'] = "TRADE STATISTICS BY WINDOW"
        ws[f'A{row}'].font = self.title_font
        ws.merge_cells(f'A{row}:H{row}')
        row += 2

        ws[f'A{row}'] = "Trade frequency and quality metrics across walk-forward windows"
        ws[f'A{row}'].font = Font(italic=True, color=self.COLORS['dark_gray'])
        row += 3

        # Check for empty data
        if not multi_results.individual_results:
            ws[f'A{row}'] = "No data available for trade statistics analysis"
            return

        # Section A: Aggregate Trade Statistics
        row = self._add_section_header(ws, row, "A. AGGREGATE TRADE STATISTICS")

        total_is_trades = 0
        total_oos_trades = 0
        all_is_trades = []
        all_oos_trades = []

        for wf_results in multi_results.individual_results.values():
            for window in wf_results.windows:
                total_is_trades += window.in_sample_num_trades
                total_oos_trades += window.out_sample_num_trades
                all_is_trades.append(window.in_sample_num_trades)
                all_oos_trades.append(window.out_sample_num_trades)

        avg_is_trades = np.mean(all_is_trades) if all_is_trades else 0
        avg_oos_trades = np.mean(all_oos_trades) if all_oos_trades else 0
        std_oos_trades = np.std(all_oos_trades) if all_oos_trades else 0

        # Clamp stability to 0-100 range (can go negative if std > mean)
        trade_stability = max(0, min(100, 100 - (std_oos_trades/avg_oos_trades*100))) if avg_oos_trades > 0 else 0

        stats_data = [
            ("Total IS Trades (all windows)", str(total_is_trades)),
            ("Total OOS Trades (all windows)", str(total_oos_trades)),
            ("Avg IS Trades per Window", f"{avg_is_trades:.1f}"),
            ("Avg OOS Trades per Window", f"{avg_oos_trades:.1f}"),
            ("Std Dev of OOS Trades", f"{std_oos_trades:.1f}"),
            ("Trade Count Stability", f"{trade_stability:.0f}%"),
        ]

        for label, value in stats_data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 2

        # Section B: Trade Statistics by Security
        row = self._add_section_header(ws, row, "B. TRADE STATISTICS BY SECURITY")

        headers = ["Security", "Windows", "Total IS", "Total OOS", "Avg IS", "Avg OOS", "Consistency"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        row += 1

        for symbol, wf_results in multi_results.individual_results.items():
            is_trades = [w.in_sample_num_trades for w in wf_results.windows]
            oos_trades = [w.out_sample_num_trades for w in wf_results.windows]

            total_is = sum(is_trades)
            total_oos = sum(oos_trades)
            avg_is = np.mean(is_trades) if is_trades else 0
            avg_oos = np.mean(oos_trades) if oos_trades else 0
            std_oos = np.std(oos_trades) if oos_trades else 0
            # Clamp consistency to 0-100 range
            consistency = max(0, min(100, 100 - (std_oos / avg_oos * 100))) if avg_oos > 0 else 0

            ws.cell(row=row, column=1, value=symbol).font = Font(bold=True)
            ws.cell(row=row, column=2, value=len(wf_results.windows))
            ws.cell(row=row, column=3, value=total_is)
            ws.cell(row=row, column=4, value=total_oos)
            ws.cell(row=row, column=5, value=f"{avg_is:.1f}")
            ws.cell(row=row, column=6, value=f"{avg_oos:.1f}")

            cons_cell = ws.cell(row=row, column=7, value=f"{consistency:.0f}%")
            if consistency > 70:
                cons_cell.fill = self.positive_fill
            elif consistency < 50:
                cons_cell.fill = self.negative_fill
            else:
                cons_cell.fill = self.neutral_fill

            row += 1

        row += 2

        # Section C: Trade Quality Analysis
        row = self._add_section_header(ws, row, "C. TRADE QUALITY INDICATORS")

        # Calculate win rates and profit factors from trade lists if available
        quality_notes = [
            "Trade quality is indicated by profit factor and consistency across windows.",
            "Higher OOS profit factor relative to IS profit factor suggests robustness.",
            "Consistent trade count across windows indicates stable signal generation.",
        ]

        for note in quality_notes:
            ws.cell(row=row, column=1, value=f"• {note}")
            ws.merge_cells(f'A{row}:G{row}')
            row += 1

        ws.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 12
