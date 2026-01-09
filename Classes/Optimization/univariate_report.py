"""
Univariate Optimization Excel Report Generator.

Generates Excel workbooks with:
- Summary tab: control values, best values, optimization settings
- One tab per parameter: data table + line charts for each metric

Uses openpyxl for Excel generation with embedded charts.
"""

import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.series import SeriesLabel
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from Classes.Optimization.univariate_optimizer import (
    UnivariateOptimizationResult,
    METRIC_DEFINITIONS
)

logger = logging.getLogger(__name__)


class UnivariateReportGenerator:
    """
    Generate Excel reports with embedded charts for univariate optimization results.
    """

    # Styling constants
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    SUBHEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    BEST_VALUE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    CONTROL_VALUE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Chart settings
    CHART_WIDTH = 15  # in cm
    CHART_HEIGHT = 10  # in cm
    CHARTS_PER_ROW = 2

    def __init__(self, output_dir: Optional[Path] = None):
        """
        Initialize report generator.

        Args:
            output_dir: Directory for output files (default: optimization_reports/)
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel report generation")

        self.output_dir = Path(output_dir) if output_dir else Path("optimization_reports")
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_report(
        self,
        results: UnivariateOptimizationResult,
        filename: Optional[str] = None
    ) -> Path:
        """
        Generate Excel report from optimization results.

        Args:
            results: UnivariateOptimizationResult to report
            filename: Optional custom filename (default: auto-generated)

        Returns:
            Path to generated Excel file
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"univariate_optimization_{results.strategy_name}_{timestamp}.xlsx"

        filepath = self.output_dir / filename

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Create Summary sheet first
        self._create_summary_sheet(wb, results)

        # Create a sheet for each parameter
        for param_name, param_result in results.parameter_results.items():
            self._create_parameter_sheet(wb, param_name, param_result, results.metrics_calculated)

        # Save workbook
        wb.save(filepath)
        logger.info(f"Report saved to: {filepath}")

        return filepath

    def _create_summary_sheet(self, wb: Workbook, results: UnivariateOptimizationResult):
        """Create the summary sheet with control values and best values."""
        ws = wb.create_sheet("Summary", 0)

        row = 1

        # Title
        ws.cell(row=row, column=1, value="UNIVARIATE OPTIMIZATION SUMMARY")
        ws.cell(row=row, column=1).font = Font(bold=True, size=16)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 2

        # Strategy and Settings Section
        ws.cell(row=row, column=1, value="OPTIMIZATION SETTINGS")
        ws.cell(row=row, column=1).font = self.HEADER_FONT
        ws.cell(row=row, column=1).fill = self.HEADER_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

        settings = [
            ("Strategy", results.strategy_name),
            ("Securities", ", ".join(results.securities)),
            ("Run Mode", results.run_mode.capitalize()),
            ("Timestamp", results.timestamp.strftime("%Y-%m-%d %H:%M:%S")),
        ]
        settings.extend([(k, str(v)) for k, v in results.optimization_settings.items()])

        for label, value in settings:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 1

        # Control Values Section
        ws.cell(row=row, column=1, value="CONTROL VALUES (BASELINE)")
        ws.cell(row=row, column=1).font = self.HEADER_FONT
        ws.cell(row=row, column=1).fill = self.HEADER_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

        # Header row
        ws.cell(row=row, column=1, value="Parameter")
        ws.cell(row=row, column=2, value="Control Value")
        for col in [1, 2]:
            ws.cell(row=row, column=col).font = self.SUBHEADER_FONT
            ws.cell(row=row, column=col).fill = self.SUBHEADER_FILL
        row += 1

        for param_name, control_val in results.control_values.items():
            ws.cell(row=row, column=1, value=param_name)
            ws.cell(row=row, column=2, value=control_val)
            ws.cell(row=row, column=2).fill = self.CONTROL_VALUE_FILL
            row += 1

        row += 1

        # Best Values Section
        ws.cell(row=row, column=1, value="BEST VALUES BY METRIC")
        ws.cell(row=row, column=1).font = self.HEADER_FONT
        ws.cell(row=row, column=1).fill = self.HEADER_FILL
        num_metrics = len(results.metrics_calculated)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_metrics + 2)
        row += 1

        # Header row
        headers = ["Parameter", "Control"] + [
            METRIC_DEFINITIONS.get(m, {}).get("name", m)
            for m in results.metrics_calculated
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            ws.cell(row=row, column=col).font = self.SUBHEADER_FONT
            ws.cell(row=row, column=col).fill = self.SUBHEADER_FILL
        row += 1

        # Data rows
        for param_name, param_result in results.parameter_results.items():
            ws.cell(row=row, column=1, value=param_name)
            ws.cell(row=row, column=2, value=param_result.control_value)
            ws.cell(row=row, column=2).fill = self.CONTROL_VALUE_FILL

            for col, metric in enumerate(results.metrics_calculated, 3):
                metric_def = METRIC_DEFINITIONS.get(metric, {})
                higher_is_better = metric_def.get("higher_is_better", True)
                best_val = param_result.get_best_value(metric, higher_is_better)
                ws.cell(row=row, column=col, value=best_val)

                # Highlight if different from control
                if best_val != param_result.control_value:
                    ws.cell(row=row, column=col).fill = self.BEST_VALUE_FILL

            row += 1

        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

    def _create_parameter_sheet(
        self,
        wb: Workbook,
        param_name: str,
        param_result,
        metrics: List[str]
    ):
        """Create a sheet for a single parameter with data and charts."""
        # Sanitize sheet name (Excel limit: 31 chars, no special chars)
        safe_name = param_name[:31].replace("/", "_").replace("\\", "_")
        ws = wb.create_sheet(safe_name)

        row = 1

        # Title
        ws.cell(row=row, column=1, value=f"Parameter: {param_name}")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(metrics) + 1)
        row += 1

        # Control value info
        ws.cell(row=row, column=1, value=f"Control Value: {param_result.control_value}")
        ws.cell(row=row, column=1).font = Font(italic=True)
        row += 2

        # Data table header
        data_start_row = row
        headers = ["Parameter Value"] + [
            METRIC_DEFINITIONS.get(m, {}).get("name", m) for m in metrics
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            ws.cell(row=row, column=col).font = self.HEADER_FONT
            ws.cell(row=row, column=col).fill = self.HEADER_FILL
            ws.cell(row=row, column=col).border = self.BORDER
        row += 1

        # Data rows
        data_rows_start = row
        for result in param_result.results:
            ws.cell(row=row, column=1, value=result.parameter_value)

            # Highlight control value row
            if result.parameter_value == param_result.control_value:
                ws.cell(row=row, column=1).fill = self.CONTROL_VALUE_FILL

            for col, metric in enumerate(metrics, 2):
                value = result.metrics.get(metric, 0)
                ws.cell(row=row, column=col, value=value)

                # Format based on metric type
                metric_def = METRIC_DEFINITIONS.get(metric, {})
                if "%" in metric_def.get("format", ""):
                    ws.cell(row=row, column=col).number_format = '0.00"%"'
                elif "$" in metric_def.get("format", ""):
                    ws.cell(row=row, column=col).number_format = '"$"#,##0.00'
                elif ".0f" in metric_def.get("format", ""):
                    ws.cell(row=row, column=col).number_format = '0'
                else:
                    ws.cell(row=row, column=col).number_format = '0.000'

            row += 1

        data_rows_end = row - 1
        num_data_rows = data_rows_end - data_rows_start + 1

        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        row += 2  # Space before charts

        # Create line charts for each metric
        chart_row_start = row
        charts_created = 0

        for metric_idx, metric in enumerate(metrics):
            metric_def = METRIC_DEFINITIONS.get(metric, {})
            metric_name = metric_def.get("name", metric)

            # Create chart
            chart = LineChart()
            chart.title = f"{metric_name} vs {param_name}"
            chart.style = 10  # Use a clean style
            chart.x_axis.title = param_name
            chart.y_axis.title = metric_name

            # Set chart size
            chart.width = self.CHART_WIDTH
            chart.height = self.CHART_HEIGHT

            # Data reference (metric column)
            data = Reference(
                ws,
                min_col=metric_idx + 2,  # +2 because col 1 is parameter values
                min_row=data_rows_start,  # Include header for series name
                max_row=data_rows_end
            )

            # Category reference (parameter values)
            cats = Reference(
                ws,
                min_col=1,
                min_row=data_rows_start + 1,  # Skip header
                max_row=data_rows_end
            )

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            # Style the line
            if chart.series:
                series = chart.series[0]
                series.graphicalProperties.line.width = 25000  # 2.5pt
                series.marker.symbol = "circle"
                series.marker.size = 7

            # Position chart
            chart_col = (charts_created % self.CHARTS_PER_ROW) * 10 + 1
            chart_row = chart_row_start + (charts_created // self.CHARTS_PER_ROW) * 20

            cell_pos = f"{get_column_letter(chart_col)}{chart_row}"
            ws.add_chart(chart, cell_pos)

            charts_created += 1

    def generate_multi_security_report(
        self,
        results_per_security: Dict[str, UnivariateOptimizationResult],
        filename: Optional[str] = None
    ) -> Path:
        """
        Generate report with separate sections for each security.

        Args:
            results_per_security: Dict of symbol -> UnivariateOptimizationResult
            filename: Optional custom filename

        Returns:
            Path to generated Excel file
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            symbols = "_".join(list(results_per_security.keys())[:3])
            filename = f"univariate_optimization_{symbols}_{timestamp}.xlsx"

        filepath = self.output_dir / filename

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Create a combined summary first
        first_result = list(results_per_security.values())[0]
        self._create_combined_summary(wb, results_per_security)

        # Create sheets per security
        for symbol, result in results_per_security.items():
            for param_name, param_result in result.parameter_results.items():
                sheet_name = f"{symbol}_{param_name}"[:31]
                self._create_parameter_sheet(wb, sheet_name, param_result, result.metrics_calculated)

        wb.save(filepath)
        logger.info(f"Multi-security report saved to: {filepath}")

        return filepath

    def _create_combined_summary(
        self,
        wb: Workbook,
        results_per_security: Dict[str, UnivariateOptimizationResult]
    ):
        """Create a combined summary sheet for multiple securities."""
        ws = wb.create_sheet("Summary", 0)

        row = 1
        ws.cell(row=row, column=1, value="COMBINED OPTIMIZATION SUMMARY")
        ws.cell(row=row, column=1).font = Font(bold=True, size=16)
        row += 2

        for symbol, result in results_per_security.items():
            ws.cell(row=row, column=1, value=f"Security: {symbol}")
            ws.cell(row=row, column=1).font = self.HEADER_FONT
            ws.cell(row=row, column=1).fill = self.HEADER_FILL
            row += 1

            for param_name, param_result in result.parameter_results.items():
                ws.cell(row=row, column=1, value=param_name)
                ws.cell(row=row, column=2, value=f"Control: {param_result.control_value}")

                # Show best value for first metric
                if result.metrics_calculated:
                    first_metric = result.metrics_calculated[0]
                    metric_def = METRIC_DEFINITIONS.get(first_metric, {})
                    higher_is_better = metric_def.get("higher_is_better", True)
                    best = param_result.get_best_value(first_metric, higher_is_better)
                    ws.cell(row=row, column=3, value=f"Best ({first_metric}): {best}")

                row += 1

            row += 1

        # Auto-adjust columns
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20
