"""
Optimization Report Generator

Generates comprehensive Excel reports for walk-forward optimization results,
including:
- Summary statistics
- Window-by-window results
- Parameter stability analysis
- Sensitivity analysis
- In-sample vs out-of-sample comparisons
- Robustness metrics
"""

import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from Classes.Optimization.sensitivity_analyzer import SensitivityResults
from Classes.Optimization.walk_forward_optimizer import WalkForwardResults, MultiSecurityResults

logger = logging.getLogger(__name__)


class OptimizationReportGenerator:
    """
    Generates Excel reports for optimization results.

    Creates a comprehensive workbook with multiple sheets:
    1. Summary - High-level overview
    2. Window Results - Detailed window-by-window performance
    3. Parameter Stability - Parameter ranges and stability metrics
    4. Sensitivity Analysis - Robustness testing results
    5. In-Sample vs Out-Sample - Performance comparison
    6. Recommendations - Actionable insights
    """

    def __init__(self, config: Dict[str, Any]):
        """Initialize report generator."""
        self.config = config
        self.report_config = config.get('reporting', {})
        self.decimal_places = config.get('report', {}).get('excel', {}).get('decimal_places', 4)

        # Styling
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        logger.info("Optimization report generator initialized")

    def generate_report(self,
                       wf_results: WalkForwardResults,
                       sensitivity_results: Optional[SensitivityResults],
                       output_dir: str = None) -> str:
        """
        Generate Excel report for optimization results.

        Args:
            wf_results: Walk-forward optimization results
            sensitivity_results: Sensitivity analysis results (optional)
            output_dir: Output directory (defaults to config)

        Returns:
            Path to generated report file
        """
        if output_dir is None:
            output_dir = self.config.get('report', {}).get('output_dir', 'logs/optimization_reports')

        os.makedirs(output_dir, exist_ok=True)

        # Create filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"optimization_{wf_results.strategy_name}_{wf_results.symbol}_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Generate sheets
        self._create_summary_sheet(wb, wf_results, sensitivity_results)
        self._create_window_results_sheet(wb, wf_results)
        self._create_parameter_stability_sheet(wb, wf_results)

        if sensitivity_results:
            self._create_sensitivity_sheet(wb, sensitivity_results)

        self._create_in_sample_vs_out_sample_sheet(wb, wf_results)
        self._create_recommendations_sheet(wb, wf_results, sensitivity_results)

        # Save workbook
        wb.save(filepath)
        logger.info(f"Optimization report saved to {filepath}")

        return filepath

    def generate_combined_report(
        self,
        multi_results: MultiSecurityResults,
        sensitivity_results_dict: Optional[Dict[str, SensitivityResults]] = None,
        output_dir: str = None
    ) -> str:
        """
        Generate combined Excel report for multi-security optimization.

        Creates a comprehensive workbook with:
        1. Combined Summary - Aggregated metrics across all securities
        2. Per-Security Comparison - Side-by-side performance comparison
        3. Parameter Consistency - How consistent parameters are across securities
        4. Individual security sheets - Detailed results for each security
        5. Recommendations - Cross-security insights

        Args:
            multi_results: Combined results from all securities
            sensitivity_results_dict: Optional dict of symbol -> SensitivityResults
            output_dir: Output directory (defaults to config)

        Returns:
            Path to generated report file
        """
        if output_dir is None:
            output_dir = self.config.get('report', {}).get('output_dir', 'logs/optimization_reports')

        os.makedirs(output_dir, exist_ok=True)

        # Create filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        securities_str = "_".join(multi_results.securities[:3])
        if len(multi_results.securities) > 3:
            securities_str += f"_+{len(multi_results.securities) - 3}"
        filename = f"combined_optimization_{multi_results.strategy_name}_{securities_str}_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Generate combined sheets
        self._create_combined_summary_sheet(wb, multi_results)
        self._create_security_comparison_sheet(wb, multi_results)
        self._create_parameter_consistency_sheet(wb, multi_results)
        self._create_combined_recommendations_sheet(wb, multi_results)

        # Add individual security detail sheets
        for symbol, wf_results in multi_results.individual_results.items():
            sens_results = sensitivity_results_dict.get(symbol) if sensitivity_results_dict else None
            self._create_security_detail_sheet(wb, symbol, wf_results, sens_results)

        # Save workbook
        wb.save(filepath)
        logger.info(f"Combined optimization report saved to {filepath}")

        return filepath

    def _create_combined_summary_sheet(self, wb: Workbook, multi_results: MultiSecurityResults):
        """Create combined summary sheet with aggregated metrics."""
        ws = wb.create_sheet("Combined Summary", 0)

        row = 1

        # Title
        ws.merge_cells(f'A{row}:F{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = "Multi-Security Walk-Forward Optimization Report"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center')
        row += 2

        # Overview
        ws[f'A{row}'] = "Overview"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1

        overview_data = [
            ["Strategy", multi_results.strategy_name],
            ["Securities Analyzed", len(multi_results.securities)],
            ["Securities List", ", ".join(multi_results.securities)],
            ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Total Windows (All Securities)", multi_results.total_windows_all_securities],
            ["Total Passed Constraints", multi_results.total_passed_all_securities],
            ["Overall Success Rate", f"{multi_results.total_passed_all_securities / multi_results.total_windows_all_securities * 100:.1f}%" if multi_results.total_windows_all_securities > 0 else "N/A"],
            ["Securities with Positive OOS", f"{multi_results.securities_with_positive_oos}/{len(multi_results.securities)}"],
        ]

        for label, value in overview_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1

        row += 2

        # Combined Performance Metrics
        ws[f'A{row}'] = "Combined Performance Metrics (Averaged Across All Securities)"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1

        # Headers
        headers = ["Metric", "In-Sample (Avg)", "Out-of-Sample (Avg)", "Degradation (%)"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        row += 1

        metrics = [
            ["Sortino Ratio",
             f"{multi_results.combined_avg_in_sample_sortino:.4f}",
             f"{multi_results.combined_avg_out_sample_sortino:.4f}",
             f"{multi_results.combined_avg_sortino_degradation_pct:.2f}%"],
            ["Sharpe Ratio",
             f"{multi_results.combined_avg_in_sample_sharpe:.4f}",
             f"{multi_results.combined_avg_out_sample_sharpe:.4f}",
             f"{multi_results.combined_avg_sharpe_degradation_pct:.2f}%"],
        ]

        for metric_row in metrics:
            for col_idx, value in enumerate(metric_row, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = self.border
            row += 1

        row += 2

        # Best/Worst Security
        ws[f'A{row}'] = "Performance Extremes"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1

        best_result = multi_results.individual_results[multi_results.best_security]
        worst_result = multi_results.individual_results[multi_results.worst_security]

        extremes_data = [
            ["Best Performing Security", multi_results.best_security,
             f"OOS Sortino: {best_result.avg_out_sample_sortino:.4f}"],
            ["Worst Performing Security", multi_results.worst_security,
             f"OOS Sortino: {worst_result.avg_out_sample_sortino:.4f}"],
        ]

        for data in extremes_data:
            ws[f'A{row}'] = data[0]
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = data[1]
            ws[f'C{row}'] = data[2]
            row += 1

        row += 2

        # Consistent Parameters
        ws[f'A{row}'] = "Recommended Parameters (Consistent Across Securities)"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1

        param_headers = ["Parameter", "Recommended Value", "Consistency Score"]
        for col_idx, header in enumerate(param_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        row += 1

        for param_name, value in multi_results.consistent_params.items():
            consistency = multi_results.param_consistency_scores.get(param_name, 0)
            cell_a = ws.cell(row=row, column=1, value=param_name)
            cell_b = ws.cell(row=row, column=2, value=f"{value:.4f}")
            cell_c = ws.cell(row=row, column=3, value=f"{consistency:.1f}%")

            # Color code by consistency
            if consistency >= 80:
                cell_c.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif consistency >= 60:
                cell_c.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
            else:
                cell_c.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

            cell_a.border = self.border
            cell_b.border = self.border
            cell_c.border = self.border
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20

    def _create_security_comparison_sheet(self, wb: Workbook, multi_results: MultiSecurityResults):
        """Create sheet comparing performance across all securities."""
        ws = wb.create_sheet("Security Comparison")

        row = 1

        # Title
        ws.merge_cells(f'A{row}:H{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = "Security-by-Security Performance Comparison"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        row += 2

        # Headers
        headers = [
            "Security", "Windows", "Passed", "IS Sortino", "OOS Sortino",
            "Sortino Deg%", "IS Sharpe", "OOS Sharpe"
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        row += 1

        # Data rows
        for symbol in multi_results.securities:
            result = multi_results.individual_results[symbol]
            data = [
                symbol,
                result.total_windows,
                result.windows_passed_constraints,
                f"{result.avg_in_sample_sortino:.4f}",
                f"{result.avg_out_sample_sortino:.4f}",
                f"{result.avg_sortino_degradation_pct:.2f}%",
                f"{result.avg_in_sample_sharpe:.4f}",
                f"{result.avg_out_sample_sharpe:.4f}",
            ]
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = self.border

                # Highlight best/worst
                if col_idx == 1:
                    if symbol == multi_results.best_security:
                        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                    elif symbol == multi_results.worst_security:
                        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            row += 1

        # Add average row
        row += 1
        ws[f'A{row}'] = "AVERAGE"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'D{row}'] = f"{multi_results.combined_avg_in_sample_sortino:.4f}"
        ws[f'E{row}'] = f"{multi_results.combined_avg_out_sample_sortino:.4f}"
        ws[f'F{row}'] = f"{multi_results.combined_avg_sortino_degradation_pct:.2f}%"
        ws[f'G{row}'] = f"{multi_results.combined_avg_in_sample_sharpe:.4f}"
        ws[f'H{row}'] = f"{multi_results.combined_avg_out_sample_sharpe:.4f}"

        for col in range(1, 9):
            ws.cell(row=row, column=col).border = self.border
            ws.cell(row=row, column=col).font = Font(bold=True)

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        for col in 'BCDEFGH':
            ws.column_dimensions[col].width = 12

    def _create_parameter_consistency_sheet(self, wb: Workbook, multi_results: MultiSecurityResults):
        """Create sheet showing parameter values across all securities."""
        ws = wb.create_sheet("Parameter Consistency")

        row = 1

        # Title
        ws.merge_cells(f'A{row}:G{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = "Parameter Values Across Securities"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        row += 2

        # Get all parameter names
        param_names = list(multi_results.consistent_params.keys())

        # Headers: Parameter, then each security, then Consistent Value, Consistency Score
        headers = ["Parameter"] + multi_results.securities + ["Recommended", "Consistency"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        row += 1

        # Data rows for each parameter
        for param_name in param_names:
            col = 1
            ws.cell(row=row, column=col, value=param_name).border = self.border
            col += 1

            # Value for each security
            for symbol in multi_results.securities:
                result = multi_results.individual_results[symbol]
                value = result.most_common_params.get(param_name, "N/A")
                if isinstance(value, float):
                    value = f"{value:.4f}"
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                col += 1

            # Recommended value
            rec_value = multi_results.consistent_params.get(param_name, "N/A")
            if isinstance(rec_value, float):
                rec_value = f"{rec_value:.4f}"
            ws.cell(row=row, column=col, value=rec_value).border = self.border
            col += 1

            # Consistency score
            consistency = multi_results.param_consistency_scores.get(param_name, 0)
            cell = ws.cell(row=row, column=col, value=f"{consistency:.1f}%")
            cell.border = self.border

            # Color code
            if consistency >= 80:
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif consistency >= 60:
                cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        for i, _ in enumerate(multi_results.securities):
            ws.column_dimensions[chr(ord('B') + i)].width = 12

    def _create_combined_recommendations_sheet(self, wb: Workbook, multi_results: MultiSecurityResults):
        """Create recommendations sheet for multi-security analysis."""
        ws = wb.create_sheet("Recommendations")

        row = 1

        # Title
        ws.merge_cells(f'A{row}:D{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = "Multi-Security Optimization Recommendations"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center')
        row += 2

        # Overall Assessment
        ws[f'A{row}'] = "Overall Assessment"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1

        # Calculate overall robustness
        positive_ratio = multi_results.securities_with_positive_oos / len(multi_results.securities)
        avg_degradation = multi_results.combined_avg_sortino_degradation_pct

        if positive_ratio >= 0.8 and avg_degradation < 30:
            overall_status = "ROBUST - Strategy performs well across securities"
            status_color = "90EE90"
        elif positive_ratio >= 0.5 and avg_degradation < 50:
            overall_status = "MODERATE - Strategy shows mixed results"
            status_color = "FFFFE0"
        else:
            overall_status = "WEAK - Strategy may be overfitted to specific securities"
            status_color = "FFB6C1"

        cell = ws[f'A{row}']
        cell.value = overall_status
        cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
        cell.font = Font(bold=True)
        row += 2

        # Key Insights
        ws[f'A{row}'] = "Key Insights"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1

        insights = []

        # OOS performance insight
        if multi_results.combined_avg_out_sample_sortino > 0.5:
            insights.append("+ Strong average out-of-sample performance suggests strategy has edge")
        elif multi_results.combined_avg_out_sample_sortino > 0:
            insights.append("~ Positive OOS performance but room for improvement")
        else:
            insights.append("- Negative average OOS performance suggests strategy may need refinement")

        # Degradation insight
        if abs(avg_degradation) < 20:
            insights.append("+ Low performance degradation indicates robust parameters")
        elif abs(avg_degradation) < 40:
            insights.append("~ Moderate degradation - consider more regularization")
        else:
            insights.append("- High degradation indicates possible overfitting")

        # Consistency insight
        high_consistency_params = [p for p, s in multi_results.param_consistency_scores.items() if s >= 70]
        low_consistency_params = [p for p, s in multi_results.param_consistency_scores.items() if s < 50]

        if len(high_consistency_params) > len(low_consistency_params):
            insights.append(f"+ {len(high_consistency_params)} parameters show high consistency across securities")
        if low_consistency_params:
            insights.append(f"! Parameters with low consistency: {', '.join(low_consistency_params[:3])}")

        for insight in insights:
            ws[f'A{row}'] = insight
            if insight.startswith("+"):
                ws[f'A{row}'].font = Font(color="006400")
            elif insight.startswith("-") or insight.startswith("!"):
                ws[f'A{row}'].font = Font(color="8B0000")
            row += 1

        row += 2

        # Parameter Recommendations
        ws[f'A{row}'] = "Parameter Recommendations"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1

        ws[f'A{row}'] = "Use these parameters for trading across all analyzed securities:"
        row += 1

        for param_name, value in multi_results.consistent_params.items():
            consistency = multi_results.param_consistency_scores.get(param_name, 0)
            ws[f'A{row}'] = f"  {param_name}: {value:.4f}"
            ws[f'B{row}'] = f"(Consistency: {consistency:.0f}%)"
            if consistency < 50:
                ws[f'C{row}'] = "⚠ Consider security-specific tuning"
                ws[f'C{row}'].font = Font(color="FFA500")
            row += 1

        row += 2

        # Securities to Monitor
        ws[f'A{row}'] = "Securities Requiring Attention"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1

        # Find securities with poor performance
        poor_performers = [
            (s, r) for s, r in multi_results.individual_results.items()
            if r.avg_out_sample_sortino < 0 or r.avg_sortino_degradation_pct > 50
        ]

        if poor_performers:
            for symbol, result in poor_performers:
                ws[f'A{row}'] = f"  {symbol}:"
                ws[f'B{row}'] = f"OOS Sortino: {result.avg_out_sample_sortino:.4f}"
                ws[f'C{row}'] = f"Degradation: {result.avg_sortino_degradation_pct:.1f}%"
                row += 1
        else:
            ws[f'A{row}'] = "  All securities performing within acceptable range"
            ws[f'A{row}'].font = Font(color="006400")

        # Adjust column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30

    def _create_security_detail_sheet(
        self,
        wb: Workbook,
        symbol: str,
        wf_results: WalkForwardResults,
        sensitivity_results: Optional[SensitivityResults]
    ):
        """Create a detail sheet for individual security results."""
        # Truncate symbol for sheet name (max 31 chars in Excel)
        sheet_name = f"Detail_{symbol[:25]}" if len(symbol) > 25 else f"Detail_{symbol}"
        ws = wb.create_sheet(sheet_name)

        row = 1

        # Title
        ws.merge_cells(f'A{row}:D{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = f"Detailed Results: {symbol}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        row += 2

        # Summary metrics
        ws[f'A{row}'] = "Performance Summary"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        metrics = [
            ["Total Windows", wf_results.total_windows],
            ["Passed Constraints", wf_results.windows_passed_constraints],
            ["Avg IS Sortino", f"{wf_results.avg_in_sample_sortino:.4f}"],
            ["Avg OOS Sortino", f"{wf_results.avg_out_sample_sortino:.4f}"],
            ["Sortino Degradation", f"{wf_results.avg_sortino_degradation_pct:.2f}%"],
            ["Avg IS Sharpe", f"{wf_results.avg_in_sample_sharpe:.4f}"],
            ["Avg OOS Sharpe", f"{wf_results.avg_out_sample_sharpe:.4f}"],
        ]

        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1

        row += 2

        # Optimal parameters
        ws[f'A{row}'] = "Optimal Parameters"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        for param_name, value in wf_results.most_common_params.items():
            ws[f'A{row}'] = param_name
            if isinstance(value, float):
                ws[f'B{row}'] = f"{value:.4f}"
            else:
                ws[f'B{row}'] = value
            min_val, max_val = wf_results.parameter_ranges.get(param_name, (value, value))
            ws[f'C{row}'] = f"Range: {min_val:.2f} - {max_val:.2f}"
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25

    def _create_summary_sheet(self, wb: Workbook, wf_results: WalkForwardResults,
                             sensitivity_results: Optional[SensitivityResults]):
        """Create summary sheet with high-level overview."""
        ws = wb.create_sheet("Summary", 0)

        row = 1

        # Title
        ws.merge_cells(f'A{row}:D{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = "Walk-Forward Optimization Report"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center')
        row += 2

        # Basic info
        info_data = [
            ["Strategy", wf_results.strategy_name],
            ["Symbol", wf_results.symbol],
            ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Total Windows", wf_results.total_windows],
            ["Windows Passed Constraints", wf_results.windows_passed_constraints],
            ["Success Rate", f"{wf_results.windows_passed_constraints / wf_results.total_windows * 100:.2f}%"]
        ]

        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1

        row += 1

        # Performance metrics
        ws[f'A{row}'] = "Performance Metrics"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        metrics_data = [
            ["Metric", "In-Sample", "Out-of-Sample", "Degradation (%)"],
            ["Sortino Ratio", f"{wf_results.avg_in_sample_sortino:.4f}",
             f"{wf_results.avg_out_sample_sortino:.4f}",
             f"{wf_results.avg_sortino_degradation_pct:.2f}%"],
            ["Sharpe Ratio", f"{wf_results.avg_in_sample_sharpe:.4f}",
             f"{wf_results.avg_out_sample_sharpe:.4f}",
             f"{wf_results.avg_sharpe_degradation_pct:.2f}%"]
        ]

        # Add header row
        for col_idx, header in enumerate(metrics_data[0], 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')

        row += 1

        # Add data rows
        for metric_row in metrics_data[1:]:
            for col_idx, value in enumerate(metric_row, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = self.border
            row += 1

        row += 2

        # Parameter summary
        ws[f'A{row}'] = "Most Common Optimal Parameters"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        ws[f'A{row}'] = "Parameter"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws[f'B{row}'] = "Median Value"
        ws[f'B{row}'].font = self.header_font
        ws[f'B{row}'].fill = self.header_fill
        ws[f'C{row}'] = "Range"
        ws[f'C{row}'].font = self.header_font
        ws[f'C{row}'].fill = self.header_fill
        ws[f'D{row}'] = "Std Dev"
        ws[f'D{row}'].font = self.header_font
        ws[f'D{row}'].fill = self.header_fill
        row += 1

        for param_name, median_value in wf_results.most_common_params.items():
            min_val, max_val = wf_results.parameter_ranges[param_name]
            std_val = wf_results.parameter_std[param_name]

            ws[f'A{row}'] = param_name
            ws[f'B{row}'] = f"{median_value:.4f}" if isinstance(median_value, float) else str(int(median_value))
            ws[f'C{row}'] = f"{min_val:.2f} - {max_val:.2f}" if isinstance(min_val, float) else f"{int(min_val)} - {int(max_val)}"
            ws[f'D{row}'] = f"{std_val:.4f}"
            row += 1

        # Auto-adjust column widths
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 20

    def _create_window_results_sheet(self, wb: Workbook, wf_results: WalkForwardResults):
        """Create detailed window-by-window results sheet."""
        ws = wb.create_sheet("Window Results")

        # Create DataFrame
        data = []
        for window in wf_results.windows:
            data.append({
                'Window': window.window_id + 1,
                'Train Start': window.train_start.strftime("%Y-%m-%d"),
                'Train End': window.train_end.strftime("%Y-%m-%d"),
                'Test Start': window.test_start.strftime("%Y-%m-%d"),
                'Test End': window.test_end.strftime("%Y-%m-%d"),
                'In-Sample Sortino': round(window.in_sample_sortino, 4),
                'Out-Sample Sortino': round(window.out_sample_sortino, 4),
                'Sortino Degradation (%)': round(window.sortino_degradation_pct, 2),
                'In-Sample Sharpe': round(window.in_sample_sharpe, 4),
                'Out-Sample Sharpe': round(window.out_sample_sharpe, 4),
                'Sharpe Degradation (%)': round(window.sharpe_degradation_pct, 2),
                'In-Sample Return (%)': round(window.in_sample_total_return_pct, 2),
                'Out-Sample Return (%)': round(window.out_sample_total_return_pct, 2),
                'In-Sample Max DD (%)': round(window.in_sample_max_drawdown_pct, 2),
                'Out-Sample Max DD (%)': round(window.out_sample_max_drawdown_pct, 2),
                'In-Sample Trades': window.in_sample_num_trades,
                'Out-Sample Trades': window.out_sample_num_trades
            })

        df = pd.DataFrame(data)

        # Write to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Style header row
                if r_idx == 1:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = Alignment(horizontal='center')

                cell.border = self.border

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Add chart for Sortino comparison
        if len(wf_results.windows) > 0:
            chart = LineChart()
            chart.title = "In-Sample vs Out-of-Sample Sortino Ratio"
            chart.y_axis.title = "Sortino Ratio"
            chart.x_axis.title = "Window"

            # In-sample data
            in_sample_data = Reference(ws, min_col=6, min_row=1, max_row=len(wf_results.windows) + 1)
            chart.add_data(in_sample_data, titles_from_data=True)

            # Out-sample data
            out_sample_data = Reference(ws, min_col=7, min_row=1, max_row=len(wf_results.windows) + 1)
            chart.add_data(out_sample_data, titles_from_data=True)

            # Categories
            categories = Reference(ws, min_col=1, min_row=2, max_row=len(wf_results.windows) + 1)
            chart.set_categories(categories)

            ws.add_chart(chart, f"S2")

    def _create_parameter_stability_sheet(self, wb: Workbook, wf_results: WalkForwardResults):
        """Create parameter stability analysis sheet."""
        ws = wb.create_sheet("Parameter Stability")

        row = 1

        # Title
        ws[f'A{row}'] = "Parameter Stability Analysis"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2

        # Description
        ws[f'A{row}'] = "This sheet shows how parameter values varied across all optimization windows."
        row += 1
        ws[f'A{row}'] = "Stable parameters (low std dev) indicate robust values. Unstable parameters suggest overfitting."
        row += 2

        # Create summary table
        headers = ["Parameter", "Median Value", "Min Value", "Max Value", "Std Dev", "Stability"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')

        row += 1

        # Parameter data
        for param_name in wf_results.most_common_params.keys():
            median_val = wf_results.most_common_params[param_name]
            min_val, max_val = wf_results.parameter_ranges[param_name]
            std_val = wf_results.parameter_std[param_name]

            # Determine stability
            if isinstance(median_val, float):
                cv = (std_val / abs(median_val)) if median_val != 0 else 0
            else:
                cv = std_val / abs(median_val) if median_val != 0 else 0

            if cv < 0.1:
                stability = "Very Stable"
                stability_color = "90EE90"  # Light green
            elif cv < 0.3:
                stability = "Stable"
                stability_color = "FFFFE0"  # Light yellow
            else:
                stability = "Unstable"
                stability_color = "FFB6C1"  # Light red

            ws.cell(row=row, column=1, value=param_name).border = self.border
            ws.cell(row=row, column=2, value=f"{median_val:.4f}" if isinstance(median_val, float) else int(median_val)).border = self.border
            ws.cell(row=row, column=3, value=f"{min_val:.4f}" if isinstance(min_val, float) else int(min_val)).border = self.border
            ws.cell(row=row, column=4, value=f"{max_val:.4f}" if isinstance(max_val, float) else int(max_val)).border = self.border
            ws.cell(row=row, column=5, value=f"{std_val:.4f}").border = self.border

            stability_cell = ws.cell(row=row, column=6, value=stability)
            stability_cell.border = self.border
            stability_cell.fill = PatternFill(start_color=stability_color, end_color=stability_color, fill_type="solid")

            row += 1

        # Auto-adjust column widths
        for col_idx in range(1, 7):
            ws.column_dimensions[chr(64 + col_idx)].width = 18

        row += 2

        # Parameter values across windows
        ws[f'A{row}'] = "Parameter Values Across All Windows"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 2

        # Create table with all window parameters
        headers = ["Window"] + list(wf_results.most_common_params.keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border

        row += 1

        for window in wf_results.windows:
            ws.cell(row=row, column=1, value=window.window_id + 1).border = self.border
            for col_idx, param_name in enumerate(wf_results.most_common_params.keys(), 2):
                param_value = window.best_params.get(param_name, "N/A")
                if isinstance(param_value, float):
                    param_value = f"{param_value:.4f}"
                elif isinstance(param_value, int):
                    param_value = str(param_value)

                ws.cell(row=row, column=col_idx, value=param_value).border = self.border

            row += 1

    def _create_sensitivity_sheet(self, wb: Workbook, sensitivity_results: SensitivityResults):
        """Create sensitivity analysis sheet."""
        ws = wb.create_sheet("Sensitivity Analysis")

        row = 1

        # Title
        ws[f'A{row}'] = "Sensitivity Analysis Results"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2

        # Overall robustness
        ws[f'A{row}'] = "Overall Robustness Assessment"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        ws[f'A{row}'] = "Sortino Range (%)"
        ws[f'B{row}'] = f"{sensitivity_results.overall_sortino_range_pct:.2f}%"
        row += 1

        ws[f'A{row}'] = "Sharpe Range (%)"
        ws[f'B{row}'] = f"{sensitivity_results.overall_sharpe_range_pct:.2f}%"
        row += 1

        ws[f'A{row}'] = "Overall Assessment"
        assessment = "ROBUST" if sensitivity_results.is_overall_robust else "SENSITIVE"
        assessment_cell = ws[f'B{row}']
        assessment_cell.value = assessment
        assessment_cell.font = Font(bold=True)
        assessment_cell.fill = PatternFill(
            start_color="90EE90" if sensitivity_results.is_overall_robust else "FFB6C1",
            end_color="90EE90" if sensitivity_results.is_overall_robust else "FFB6C1",
            fill_type="solid"
        )
        row += 2

        # Most and least robust parameters
        ws[f'A{row}'] = "Most Robust Parameters"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'C{row}'] = "Least Robust Parameters"
        ws[f'C{row}'].font = Font(bold=True)
        row += 1

        max_len = max(len(sensitivity_results.most_robust_params), len(sensitivity_results.least_robust_params))
        for i in range(max_len):
            if i < len(sensitivity_results.most_robust_params):
                ws.cell(row=row, column=1, value=sensitivity_results.most_robust_params[i])

            if i < len(sensitivity_results.least_robust_params):
                ws.cell(row=row, column=3, value=sensitivity_results.least_robust_params[i])

            row += 1

        row += 2

        # Individual parameter sensitivities
        ws[f'A{row}'] = "Parameter-by-Parameter Sensitivity"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        headers = ["Parameter", "Base Value", "Sortino Range", "Sortino Std Dev", "Assessment"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border

        row += 1

        for param_name, sensitivity in sensitivity_results.parameter_sensitivities.items():
            ws.cell(row=row, column=1, value=param_name).border = self.border
            ws.cell(row=row, column=2, value=f"{sensitivity.base_value:.4f}" if isinstance(sensitivity.base_value, float) else int(sensitivity.base_value)).border = self.border
            ws.cell(row=row, column=3, value=f"{sensitivity.sortino_range:.4f}").border = self.border
            ws.cell(row=row, column=4, value=f"{sensitivity.sortino_std:.4f}").border = self.border

            if sensitivity.is_robust:
                assessment = "Robust"
                color = "90EE90"
            elif sensitivity.is_unstable:
                assessment = "Unstable"
                color = "FFB6C1"
            else:
                assessment = "Moderate"
                color = "FFFFE0"

            assessment_cell = ws.cell(row=row, column=5, value=assessment)
            assessment_cell.border = self.border
            assessment_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            row += 1

        # Auto-adjust column widths
        for col_idx in range(1, 6):
            ws.column_dimensions[chr(64 + col_idx)].width = 18

    def _create_in_sample_vs_out_sample_sheet(self, wb: Workbook, wf_results: WalkForwardResults):
        """Create in-sample vs out-of-sample comparison sheet."""
        ws = wb.create_sheet("In-Sample vs Out-Sample")

        row = 1

        # Title
        ws[f'A{row}'] = "In-Sample vs Out-of-Sample Performance"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2

        # Description
        ws[f'A{row}'] = "This comparison shows if the strategy's optimized parameters generalize to unseen data."
        row += 1
        ws[f'A{row}'] = "Small degradation (<15%) indicates robust parameters. Large degradation (>50%) suggests overfitting."
        row += 3

        # Summary table
        headers = ["Metric", "In-Sample Avg", "Out-Sample Avg", "Degradation (%)", "Assessment"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')

        row += 1

        # Metrics
        metrics_data = [
            ("Sortino Ratio", wf_results.avg_in_sample_sortino, wf_results.avg_out_sample_sortino,
             wf_results.avg_sortino_degradation_pct),
            ("Sharpe Ratio", wf_results.avg_in_sample_sharpe, wf_results.avg_out_sample_sharpe,
             wf_results.avg_sharpe_degradation_pct)
        ]

        for metric_name, in_val, out_val, degradation in metrics_data:
            ws.cell(row=row, column=1, value=metric_name).border = self.border
            ws.cell(row=row, column=2, value=f"{in_val:.4f}").border = self.border
            ws.cell(row=row, column=3, value=f"{out_val:.4f}").border = self.border
            ws.cell(row=row, column=4, value=f"{degradation:.2f}%").border = self.border

            # Assessment
            if abs(degradation) < 15:
                assessment = "Excellent"
                color = "90EE90"
            elif abs(degradation) < 30:
                assessment = "Good"
                color = "FFFFE0"
            else:
                assessment = "Poor (Overfitting)"
                color = "FFB6C1"

            assessment_cell = ws.cell(row=row, column=5, value=assessment)
            assessment_cell.border = self.border
            assessment_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            row += 1

        # Auto-adjust column widths
        for col_idx in range(1, 6):
            ws.column_dimensions[chr(64 + col_idx)].width = 20

    def _create_recommendations_sheet(self, wb: Workbook, wf_results: WalkForwardResults,
                                     sensitivity_results: Optional[SensitivityResults]):
        """Create recommendations sheet with actionable insights."""
        ws = wb.create_sheet("Recommendations")

        row = 1

        # Title
        ws.merge_cells(f'A{row}:C{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = "Recommendations & Actionable Insights"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        row += 2

        # Recommended parameters
        ws[f'A{row}'] = "RECOMMENDED PARAMETERS FOR LIVE TRADING"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="006400")
        ws[f'A{row}'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        row += 2

        ws[f'A{row}'] = "Parameter"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws[f'B{row}'] = "Recommended Value"
        ws[f'B{row}'].font = self.header_font
        ws[f'B{row}'].fill = self.header_fill
        ws[f'C{row}'] = "Robust Range"
        ws[f'C{row}'].font = self.header_font
        ws[f'C{row}'].fill = self.header_fill
        row += 1

        for param_name, median_val in wf_results.most_common_params.items():
            min_val, max_val = wf_results.parameter_ranges[param_name]

            ws[f'A{row}'] = param_name
            ws[f'B{row}'] = f"{median_val:.4f}" if isinstance(median_val, float) else int(median_val)
            ws[f'C{row}'] = f"{min_val:.2f} - {max_val:.2f}" if isinstance(min_val, float) else f"{int(min_val)} - {int(max_val)}"
            row += 1

        row += 2

        # Key findings
        ws[f'A{row}'] = "KEY FINDINGS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        findings = []

        # Overfitting assessment
        if abs(wf_results.avg_sortino_degradation_pct) < 15:
            findings.append("✓ Parameters show excellent generalization (degradation < 15%)")
        elif abs(wf_results.avg_sortino_degradation_pct) < 30:
            findings.append("⚠ Parameters show moderate generalization (degradation 15-30%)")
        else:
            findings.append("✗ WARNING: Significant overfitting detected (degradation > 30%)")

        # Success rate
        success_rate = wf_results.windows_passed_constraints / wf_results.total_windows * 100
        if success_rate > 80:
            findings.append(f"✓ High success rate ({success_rate:.0f}% of windows passed constraints)")
        elif success_rate > 50:
            findings.append(f"⚠ Moderate success rate ({success_rate:.0f}% of windows passed constraints)")
        else:
            findings.append(f"✗ Low success rate ({success_rate:.0f}% of windows passed constraints)")

        # Sensitivity
        if sensitivity_results and sensitivity_results.is_overall_robust:
            findings.append("✓ Parameters are robust to small variations")
        elif sensitivity_results:
            findings.append("⚠ Parameters are sensitive to variations - use with caution")

        # Write findings
        for finding in findings:
            ws[f'A{row}'] = finding
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1

        row += 2

        # Action items
        ws[f'A{row}'] = "RECOMMENDED ACTIONS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        actions = []

        if abs(wf_results.avg_sortino_degradation_pct) > 30:
            actions.append("1. Consider simplifying the strategy - it may be too complex")
            actions.append("2. Increase training window size to improve parameter stability")
            actions.append("3. Add more constraints to prevent overfitting")

        if success_rate < 70:
            actions.append("1. Review constraint settings - they may be too strict")
            actions.append("2. Consider using a broader parameter search space")

        if sensitivity_results and not sensitivity_results.is_overall_robust:
            actions.append(f"1. Focus on stabilizing these parameters: {', '.join(sensitivity_results.least_robust_params)}")
            actions.append("2. Consider fixing sensitive parameters to their median values")

        if not actions:
            actions.append("Strategy parameters appear robust. Proceed with forward testing.")

        for action in actions:
            ws[f'A{row}'] = action
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
