"""
Excel Report Generator for Factor Analysis.

Generates professional Excel reports with multiple sheets for:
- Executive summary
- Factor correlations and statistics
- Scenario analysis
- Detailed trade-level data
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Any
from pathlib import Path
from datetime import datetime
from dataclasses import dataclass

from ..logging.audit_logger import AuditLogger

# Try to import openpyxl for Excel formatting
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference, LineChart
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@dataclass
class ExcelStyle:
    """Excel styling configuration."""
    header_font: str = 'Calibri'
    header_size: int = 12
    header_bold: bool = True
    header_bg_color: str = '4472C4'
    header_font_color: str = 'FFFFFF'

    data_font: str = 'Calibri'
    data_size: int = 11

    positive_color: str = '92D050'  # Green
    negative_color: str = 'FF6B6B'  # Red
    neutral_color: str = 'FFEB9C'   # Yellow


class ExcelReportGenerator:
    """
    Generates comprehensive Excel reports for factor analysis.

    Creates multi-sheet workbooks with:
    - Summary sheet with key findings
    - Factor statistics sheets
    - Scenario analysis sheets
    - Raw data sheets
    """

    def __init__(
        self,
        style: Optional[ExcelStyle] = None,
        logger: Optional[AuditLogger] = None
    ):
        """
        Initialize ExcelReportGenerator.

        Args:
            style: Excel styling configuration
            logger: Optional audit logger
        """
        self.style = style or ExcelStyle()
        self.logger = logger

        if not OPENPYXL_AVAILABLE:
            if self.logger:
                self.logger.warning("openpyxl not available, Excel generation limited")

    def generate_report(
        self,
        results: Dict[str, Any],
        output_path: str,
        include_raw_data: bool = True
    ) -> str:
        """
        Generate complete Excel report.

        Args:
            results: Complete analysis results dictionary
            output_path: Output file path
            include_raw_data: Whether to include raw trade data

        Returns:
            Path to generated file
        """
        if self.logger:
            self.logger.start_section("EXCEL_GENERATION")

        path = Path(output_path)
        if not path.suffix:
            path = path.with_suffix('.xlsx')

        if OPENPYXL_AVAILABLE:
            wb = Workbook()
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']

            # Add sheets
            self._add_summary_sheet(wb, results)
            self._add_factor_stats_sheet(wb, results)
            self._add_correlation_sheet(wb, results)
            self._add_hypothesis_sheet(wb, results)
            self._add_ml_sheet(wb, results)
            self._add_scenario_sheet(wb, results)

            if include_raw_data and 'enriched_trades' in results:
                self._add_data_sheet(wb, results['enriched_trades'], 'Trade_Data')

            wb.save(str(path))

        else:
            # Fallback to pandas ExcelWriter
            with pd.ExcelWriter(str(path), engine='openpyxl') as writer:
                self._write_summary_pandas(writer, results)
                self._write_stats_pandas(writer, results)

        if self.logger:
            self.logger.info("Excel report generated", {'path': str(path)})
            self.logger.end_section()

        return str(path)

    def _add_summary_sheet(self, wb: 'Workbook', results: Dict) -> None:
        """Add executive summary sheet."""
        ws = wb.create_sheet("Summary")

        # Title
        ws['A1'] = "Factor Analysis Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        # Timestamp
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True)

        row = 4

        # Data summary
        ws.cell(row=row, column=1, value="Data Summary").font = Font(bold=True, size=13)
        row += 1

        if 'data_summary' in results:
            summary = results['data_summary']
            ws.cell(row=row, column=1, value="Total Trades")
            ws.cell(row=row, column=2, value=summary.get('total_trades', 'N/A'))
            row += 1

            ws.cell(row=row, column=1, value="Good Trades")
            ws.cell(row=row, column=2, value=summary.get('good_trades', 'N/A'))
            row += 1

            ws.cell(row=row, column=1, value="Bad Trades")
            ws.cell(row=row, column=2, value=summary.get('bad_trades', 'N/A'))
            row += 1

            ws.cell(row=row, column=1, value="Indeterminate Trades")
            ws.cell(row=row, column=2, value=summary.get('indeterminate_trades', 'N/A'))
            row += 1

        row += 1

        # Key findings
        ws.cell(row=row, column=1, value="Key Findings").font = Font(bold=True, size=13)
        row += 1

        if 'key_findings' in results:
            for finding in results['key_findings'][:10]:
                ws.cell(row=row, column=1, value=f"- {finding}")
                row += 1

        # Top factors
        row += 1
        ws.cell(row=row, column=1, value="Top Significant Factors").font = Font(bold=True, size=13)
        row += 1

        if 'tier2' in results and results['tier2'].get('logistic_regression'):
            reg = results['tier2']['logistic_regression']
            sig_factors = reg.get_significant_factors()[:5]
            for factor in sig_factors:
                ws.cell(row=row, column=1, value=f"- {factor.factor_name}")
                ws.cell(row=row, column=2, value=f"p={factor.p_value:.4f}")
                ws.cell(row=row, column=3, value=f"OR={factor.odds_ratio:.2f}")
                row += 1

        # Best scenarios
        row += 1
        ws.cell(row=row, column=1, value="Best Trading Scenarios").font = Font(bold=True, size=13)
        row += 1

        if 'scenarios' in results and results['scenarios'].get('best_scenarios'):
            for scenario in results['scenarios']['best_scenarios'][:3]:
                ws.cell(row=row, column=1, value=scenario.name)
                ws.cell(row=row, column=2, value=f"Lift: {scenario.lift:.2f}")
                ws.cell(row=row, column=3, value=f"N={scenario.n_trades}")
                row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20

    def _add_factor_stats_sheet(self, wb: 'Workbook', results: Dict) -> None:
        """Add factor statistics sheet."""
        ws = wb.create_sheet("Factor_Statistics")

        if 'tier1' not in results:
            ws['A1'] = "No Tier 1 results available"
            return

        tier1 = results['tier1']

        # Descriptive statistics
        if 'descriptive_stats' in tier1:
            stats_df = pd.DataFrame(tier1['descriptive_stats']).T
            self._write_dataframe(ws, stats_df, "Descriptive Statistics", start_row=1)

    def _add_correlation_sheet(self, wb: 'Workbook', results: Dict) -> None:
        """Add correlation analysis sheet."""
        ws = wb.create_sheet("Correlations")

        if 'tier1' not in results:
            ws['A1'] = "No correlation results available"
            return

        tier1 = results['tier1']
        row = 1

        # Point-biserial correlations
        if 'point_biserial' in tier1:
            ws.cell(row=row, column=1, value="Point-Biserial Correlations").font = Font(bold=True, size=13)
            row += 1

            headers = ['Factor', 'Correlation', 'P-Value', 'Significant']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=self.style.header_bg_color,
                                        end_color=self.style.header_bg_color,
                                        fill_type='solid')
            row += 1

            for corr in tier1['point_biserial']:
                ws.cell(row=row, column=1, value=corr.factor)
                ws.cell(row=row, column=2, value=round(corr.correlation, 4))
                ws.cell(row=row, column=3, value=round(corr.p_value, 4))
                ws.cell(row=row, column=4, value='Yes' if corr.significant else 'No')

                # Color significant rows
                if corr.significant:
                    for col in range(1, 5):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color=self.style.positive_color,
                            end_color=self.style.positive_color,
                            fill_type='solid'
                        )
                row += 1

        # Correlation matrix
        row += 2
        if 'correlation_matrix' in tier1:
            ws.cell(row=row, column=1, value="Factor Correlation Matrix").font = Font(bold=True, size=13)
            row += 1

            matrix = tier1['correlation_matrix']
            if isinstance(matrix, pd.DataFrame):
                self._write_dataframe(ws, matrix, None, start_row=row)

    def _add_hypothesis_sheet(self, wb: 'Workbook', results: Dict) -> None:
        """Add hypothesis testing sheet."""
        ws = wb.create_sheet("Hypothesis_Tests")

        if 'tier2' not in results:
            ws['A1'] = "No Tier 2 results available"
            return

        tier2 = results['tier2']
        row = 1

        # Logistic regression
        if tier2.get('logistic_regression'):
            reg = tier2['logistic_regression']
            ws.cell(row=row, column=1, value="Logistic Regression Results").font = Font(bold=True, size=13)
            row += 1

            # Model info
            ws.cell(row=row, column=1, value=f"N Observations: {reg.n_observations}")
            row += 1
            ws.cell(row=row, column=1, value=f"Pseudo R²: {reg.pseudo_r2:.4f}")
            row += 1
            ws.cell(row=row, column=1, value=f"AIC: {reg.aic:.2f}")
            row += 2

            # Coefficients table
            headers = ['Factor', 'Coefficient', 'Odds Ratio', 'Std Error', 'P-Value', 'CI Lower', 'CI Upper']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
            row += 1

            for factor in reg.factor_results:
                ws.cell(row=row, column=1, value=factor.factor_name)
                ws.cell(row=row, column=2, value=round(factor.coefficient, 4))
                ws.cell(row=row, column=3, value=round(factor.odds_ratio, 4))
                ws.cell(row=row, column=4, value=round(factor.std_error, 4))
                ws.cell(row=row, column=5, value=round(factor.p_value, 4))
                ws.cell(row=row, column=6, value=round(factor.ci_lower, 4))
                ws.cell(row=row, column=7, value=round(factor.ci_upper, 4))
                row += 1

        # ANOVA
        row += 2
        if tier2.get('anova'):
            ws.cell(row=row, column=1, value="ANOVA Results").font = Font(bold=True, size=13)
            row += 1

            headers = ['Factor', 'F-Statistic', 'P-Value', 'Effect Size', 'Significant']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
            row += 1

            for anova in tier2['anova']:
                ws.cell(row=row, column=1, value=anova.factor)
                ws.cell(row=row, column=2, value=round(anova.statistic, 4))
                ws.cell(row=row, column=3, value=round(anova.p_value, 4))
                ws.cell(row=row, column=4, value=round(anova.effect_size or 0, 4))
                ws.cell(row=row, column=5, value='Yes' if anova.significant else 'No')
                row += 1

    def _add_ml_sheet(self, wb: 'Workbook', results: Dict) -> None:
        """Add ML analysis sheet."""
        ws = wb.create_sheet("ML_Analysis")

        if 'tier3' not in results or not results['tier3'].get('enabled'):
            ws['A1'] = "No Tier 3 ML results available"
            return

        tier3 = results['tier3']
        row = 1

        # Model performance
        ws.cell(row=row, column=1, value="Random Forest Performance").font = Font(bold=True, size=13)
        row += 1

        if tier3.get('rf_cv_accuracy'):
            ws.cell(row=row, column=1, value="Cross-Validation Accuracy")
            ws.cell(row=row, column=2, value=f"{tier3['rf_cv_accuracy']:.1%}")
            row += 1

        row += 1

        # Feature importance
        ws.cell(row=row, column=1, value="Feature Importance").font = Font(bold=True, size=13)
        row += 1

        headers = ['Rank', 'Feature', 'Importance', 'Std Dev']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
        row += 1

        if tier3.get('rf_feature_importances'):
            for i, feat in enumerate(tier3['rf_feature_importances'][:20], 1):
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=feat.feature_name)
                ws.cell(row=row, column=3, value=round(feat.importance, 4))
                ws.cell(row=row, column=4, value=round(feat.importance_std, 4))
                row += 1

        # SHAP values
        row += 2
        if tier3.get('shap_results'):
            ws.cell(row=row, column=1, value="SHAP Analysis").font = Font(bold=True, size=13)
            row += 1

            headers = ['Rank', 'Feature', 'Mean |SHAP|', 'Direction']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
            row += 1

            for i, shap in enumerate(tier3['shap_results'][:20], 1):
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=shap.feature_name)
                ws.cell(row=row, column=3, value=round(shap.mean_abs_shap, 4))
                ws.cell(row=row, column=4, value=shap.direction)
                row += 1

    def _add_scenario_sheet(self, wb: 'Workbook', results: Dict) -> None:
        """Add scenario analysis sheet."""
        ws = wb.create_sheet("Scenarios")

        if 'scenarios' not in results:
            ws['A1'] = "No scenario analysis results available"
            return

        scenarios = results['scenarios']
        row = 1

        # Add scenario detection methodology explanation
        ws.cell(row=row, column=1, value="Scenario Analysis").font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        row += 2

        # Detection methodology section
        ws.cell(row=row, column=1, value="Detection Methodology").font = Font(bold=True, size=12)
        row += 1

        mode = scenarios.get('mode', 'binary')
        ws.cell(row=row, column=1, value="Detection Mode:")
        ws.cell(row=row, column=2, value=mode.upper())
        row += 1

        if mode == 'binary':
            ws.cell(row=row, column=1, value="Method:")
            ws.cell(row=row, column=2, value="Threshold-based conditions using percentile splits")
            row += 1
            ws.cell(row=row, column=1, value="Thresholds tested:")
            ws.cell(row=row, column=2, value="10th, 25th, 50th, 75th, 90th percentiles of each factor")
            row += 1
            ws.cell(row=row, column=1, value="Logic:")
            ws.cell(row=row, column=2, value="For each factor, tests if trades with factor >= or <= threshold have significantly different outcomes")
            row += 1
        elif mode == 'clustering':
            ws.cell(row=row, column=1, value="Method:")
            ws.cell(row=row, column=2, value="K-means clustering on standardized factors")
            row += 1
            ws.cell(row=row, column=1, value="N Clusters:")
            ws.cell(row=row, column=2, value=scenarios.get('n_clusters', 3))
            row += 1
            ws.cell(row=row, column=1, value="Logic:")
            ws.cell(row=row, column=2, value="Groups trades into natural clusters based on factor similarity, then evaluates each cluster's performance")
            row += 1

        # Baseline metrics
        baseline = scenarios.get('baseline_metrics', {})
        if baseline:
            ws.cell(row=row, column=1, value="Baseline Good Trade Rate:")
            ws.cell(row=row, column=2, value=f"{baseline.get('good_trade_rate', 0):.1%}")
            row += 1

        # Interpretation guide
        row += 1
        ws.cell(row=row, column=1, value="Interpretation Guide").font = Font(bold=True, size=12)
        row += 1
        ws.cell(row=row, column=1, value="Lift:")
        ws.cell(row=row, column=2, value="Ratio of scenario's good trade rate to baseline. Lift > 1 = better than average.")
        row += 1
        ws.cell(row=row, column=1, value="Confidence:")
        ws.cell(row=row, column=2, value="Statistical confidence that the scenario differs from baseline (binomial test).")
        row += 1
        ws.cell(row=row, column=1, value="N Trades:")
        ws.cell(row=row, column=2, value="Number of trades matching the scenario conditions (larger = more reliable).")
        row += 2

        # Best scenarios
        ws.cell(row=row, column=1, value="Best Trading Scenarios").font = Font(bold=True, size=13)
        row += 1

        if scenarios.get('best_scenarios'):
            headers = ['Scenario', 'Conditions', 'N Trades', 'Good Rate', 'Lift', 'Confidence']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=self.style.positive_color,
                                        end_color=self.style.positive_color,
                                        fill_type='solid')
            row += 1

            for scenario in scenarios['best_scenarios']:
                ws.cell(row=row, column=1, value=scenario.name)
                ws.cell(row=row, column=2, value=scenario.get_condition_string())
                ws.cell(row=row, column=3, value=scenario.n_trades)
                ws.cell(row=row, column=4, value=f"{scenario.good_trade_rate:.1%}")
                ws.cell(row=row, column=5, value=round(scenario.lift, 2))
                ws.cell(row=row, column=6, value=f"{scenario.confidence:.1%}")
                row += 1

        # Worst scenarios
        row += 2
        ws.cell(row=row, column=1, value="Worst Trading Scenarios").font = Font(bold=True, size=13)
        row += 1

        if scenarios.get('worst_scenarios'):
            headers = ['Scenario', 'Conditions', 'N Trades', 'Good Rate', 'Lift', 'Confidence']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=self.style.negative_color,
                                        end_color=self.style.negative_color,
                                        fill_type='solid')
            row += 1

            for scenario in scenarios['worst_scenarios']:
                ws.cell(row=row, column=1, value=scenario.name)
                ws.cell(row=row, column=2, value=scenario.get_condition_string())
                ws.cell(row=row, column=3, value=scenario.n_trades)
                ws.cell(row=row, column=4, value=f"{scenario.good_trade_rate:.1%}")
                ws.cell(row=row, column=5, value=round(scenario.lift, 2))
                ws.cell(row=row, column=6, value=f"{scenario.confidence:.1%}")
                row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12

    def _add_data_sheet(self, wb: 'Workbook', df: pd.DataFrame, sheet_name: str) -> None:
        """Add raw data sheet."""
        ws = wb.create_sheet(sheet_name)

        # Limit columns to avoid huge files
        max_cols = 50
        if len(df.columns) > max_cols:
            # Prioritize non-metadata columns
            priority_cols = [c for c in df.columns if not c.startswith('_')]
            other_cols = [c for c in df.columns if c.startswith('_')]
            cols = priority_cols[:max_cols - 5] + other_cols[:5]
            df = df[cols]

        # Limit rows
        max_rows = 5000
        if len(df) > max_rows:
            df = df.head(max_rows)

        self._write_dataframe(ws, df, None, start_row=1)

    def _write_dataframe(
        self,
        ws: 'Workbook',
        df: pd.DataFrame,
        title: Optional[str],
        start_row: int = 1
    ) -> int:
        """Write DataFrame to worksheet."""
        row = start_row

        if title:
            ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=13)
            row += 1

        # Headers
        for col, header in enumerate(df.columns, 1):
            cell = ws.cell(row=row, column=col, value=str(header))
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.style.header_bg_color,
                                    end_color=self.style.header_bg_color,
                                    fill_type='solid')
        row += 1

        # Data
        for _, data_row in df.iterrows():
            for col, value in enumerate(data_row, 1):
                if pd.isna(value):
                    cell_value = ''
                elif isinstance(value, float):
                    cell_value = round(value, 4)
                else:
                    cell_value = value
                ws.cell(row=row, column=col, value=cell_value)
            row += 1

        return row

    def _write_summary_pandas(self, writer: pd.ExcelWriter, results: Dict) -> None:
        """Fallback summary writing using pandas."""
        summary_data = {
            'Metric': ['Total Trades', 'Good Trades', 'Bad Trades'],
            'Value': [
                results.get('data_summary', {}).get('total_trades', 'N/A'),
                results.get('data_summary', {}).get('good_trades', 'N/A'),
                results.get('data_summary', {}).get('bad_trades', 'N/A')
            ]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)

    def _write_stats_pandas(self, writer: pd.ExcelWriter, results: Dict) -> None:
        """Fallback stats writing using pandas."""
        if 'tier1' in results and 'descriptive_stats' in results['tier1']:
            stats_df = pd.DataFrame(results['tier1']['descriptive_stats']).T
            stats_df.to_excel(writer, sheet_name='Statistics')
