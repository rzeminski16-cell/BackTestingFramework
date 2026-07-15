"""
Discretion-vs-rules comparison: pairs an interactive run with its AUTO
baseline twin and reports where the human decisions helped or hurt.

Methodology (stated honestly, also written into the workbook):

- The AGGREGATE comparison (side-by-side metrics, aligned equity curves)
  is exact: same config, same data, only the decisions differ.
- The PER-DECISION table matches each divergent decision (reject, defer,
  suppressed, modify, capital resolution) to the baseline trade opened or
  closed at the same symbol/date — "what rules-only did". Attribution is
  labeled:
    * exact           — at or before the first divergence, both runs were
                        still identical, so the counterfactual is exact.
    * approximate     — single-security runs after the first divergence
                        (sizing compounds off different capital).
    * path_dependent  — portfolio runs after the first divergence
                        (capital, contention, and sizing all differ
                        downstream; only the aggregate view is rigorous).
"""
import json
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd

from ..Analysis.performance_metrics import PerformanceMetrics
from ..Engine.backtest_result import BacktestResult
from .baseline import extract_trades_and_equity
from .models import DecisionAction, DecisionRecord, OutcomeRecord, PromptRecord, SignalEvent

DIVERGENT_ACTIONS = (DecisionAction.REJECT, DecisionAction.DEFER,
                     DecisionAction.AUTO_SUPPRESSED, DecisionAction.MODIFY)

METRIC_ROWS = [
    ('total_return', 'Total Return'),
    ('total_return_pct', 'Total Return %'),
    ('final_equity', 'Final Equity'),
    ('num_trades', 'Number of Trades'),
    ('win_rate', 'Win Rate (fraction)'),
    ('profit_factor', 'Profit Factor'),
    ('sharpe_ratio', 'Sharpe Ratio'),
    ('sortino_ratio', 'Sortino Ratio'),
    ('max_drawdown', 'Max Drawdown'),
    ('max_drawdown_pct', 'Max Drawdown %'),
    ('avg_win', 'Average Win'),
    ('avg_loss', 'Average Loss'),
    ('avg_trade_duration', 'Avg Trade Duration (days)'),
]


@dataclass
class ComparisonResult:
    interactive_metrics: Dict[str, Any]
    baseline_metrics: Dict[str, Any]
    equity_curves: pd.DataFrame          # date, interactive, baseline, delta
    decision_rows: List[Dict[str, Any]]  # per divergent decision
    headline: Dict[str, Any]
    engine_type: str = "single"

    def to_dict(self) -> Dict[str, Any]:
        return {
            'headline': self.headline,
            'interactive_metrics': self.interactive_metrics,
            'baseline_metrics': self.baseline_metrics,
            'decision_rows': self.decision_rows,
            'engine_type': self.engine_type,
        }


def _metrics_for(result) -> Dict[str, Any]:
    """PerformanceMetrics over either result type (portfolio normalized)."""
    trades, equity_curve = extract_trades_and_equity(result)
    if hasattr(result, 'portfolio_equity_curve'):
        normalized = BacktestResult(
            symbol="PORTFOLIO",
            strategy_name=result.strategy_name,
            trades=trades,
            equity_curve=equity_curve,
            final_equity=result.final_equity,
            total_return=result.total_return,
            total_return_pct=result.total_return_pct,
            strategy_params={},
        )
        return PerformanceMetrics.calculate_metrics(normalized)
    return PerformanceMetrics.calculate_metrics(result)


def _align_equity(interactive_curve: pd.DataFrame,
                  baseline_curve: pd.DataFrame) -> pd.DataFrame:
    left = interactive_curve[['date', 'equity']].rename(
        columns={'equity': 'interactive'})
    right = baseline_curve[['date', 'equity']].rename(
        columns={'equity': 'baseline'})
    merged = left.merge(right, on='date', how='outer').sort_values('date')
    merged[['interactive', 'baseline']] = merged[
        ['interactive', 'baseline']].ffill()
    merged['delta'] = merged['interactive'] - merged['baseline']
    return merged.reset_index(drop=True)


def _trade_date(value) -> str:
    return str(value)[:10]


def _match_trade(trades, symbol: str, bar_date: str, by: str = "entry"):
    """First trade for symbol whose entry/exit date matches the bar date."""
    for trade in trades:
        if trade.symbol != symbol:
            continue
        date_value = trade.entry_date if by == "entry" else trade.exit_date
        if _trade_date(date_value) == bar_date:
            return trade
    return None


def compare_runs(interactive_result, baseline_result,
                 events: List[SignalEvent],
                 decisions: List[DecisionRecord],
                 outcomes: List[OutcomeRecord]) -> ComparisonResult:
    """Build the discretion-vs-rules comparison for one paired run."""
    engine_type = ("portfolio"
                   if hasattr(interactive_result, 'portfolio_equity_curve')
                   else "single")
    interactive_trades, interactive_curve = extract_trades_and_equity(
        interactive_result)
    baseline_trades, baseline_curve = extract_trades_and_equity(
        baseline_result)

    interactive_metrics = _metrics_for(interactive_result)
    baseline_metrics = _metrics_for(baseline_result)
    equity_curves = _align_equity(interactive_curve, baseline_curve)

    events_by_id = {e.event_id: e for e in events}
    outcomes_by_id = {o.event_id: o for o in outcomes}

    divergent = [d for d in decisions
                 if d.action in DIVERGENT_ACTIONS or d.capital_resolution]
    first_divergence = min(
        (events_by_id[d.event_id].bar_date for d in divergent
         if d.event_id in events_by_id), default=None)

    decision_rows: List[Dict[str, Any]] = []
    for decision in divergent:
        event = events_by_id.get(decision.event_id)
        if event is None:
            continue
        outcome = outcomes_by_id.get(decision.event_id)
        match_by = "entry" if event.signal_type in ("BUY", "PYRAMID") else "exit"
        baseline_trade = _match_trade(baseline_trades, event.symbol,
                                      event.bar_date, by=match_by)
        interactive_trade = _match_trade(interactive_trades, event.symbol,
                                         event.bar_date, by=match_by)
        baseline_pl = float(baseline_trade.pl) if baseline_trade else None
        interactive_pl = (float(interactive_trade.pl)
                          if interactive_trade else None)
        delta = None
        if baseline_pl is not None or interactive_pl is not None:
            delta = (interactive_pl or 0.0) - (baseline_pl or 0.0)

        if first_divergence is not None and event.bar_date <= first_divergence:
            attribution = "exact"
        elif engine_type == "portfolio":
            attribution = "path_dependent"
        else:
            attribution = "approximate"

        decision_rows.append({
            'event_id': event.event_id,
            'bar_date': event.bar_date,
            'symbol': event.symbol,
            'event_kind': event.event_kind,
            'signal_type': event.signal_type,
            'action': decision.action.value,
            'source': decision.source.value,
            'rationale': decision.rationale,
            'capital_resolution': (json.dumps(decision.capital_resolution)
                                   if decision.capital_resolution else ""),
            'executed': outcome.executed if outcome else None,
            'baseline_trade_pl': baseline_pl,
            'interactive_trade_pl': interactive_pl,
            'pl_delta': delta,
            'attribution_quality': attribution,
        })

    action_counts: Dict[str, int] = {}
    for decision in decisions:
        action_counts[decision.action.value] = (
            action_counts.get(decision.action.value, 0) + 1)

    headline = {
        'interactive_total_return': float(interactive_result.total_return),
        'baseline_total_return': float(baseline_result.total_return),
        'discretion_delta': float(interactive_result.total_return
                                  - baseline_result.total_return),
        'interactive_total_return_pct': float(
            interactive_result.total_return_pct),
        'baseline_total_return_pct': float(baseline_result.total_return_pct),
        'discretion_delta_pct': float(interactive_result.total_return_pct
                                      - baseline_result.total_return_pct),
        'num_decisions': len(decisions),
        'num_divergent_decisions': len(divergent),
        'first_divergence_date': first_divergence,
        'action_counts': action_counts,
        'generated_at': datetime.now().isoformat(timespec='seconds'),
    }

    return ComparisonResult(
        interactive_metrics=interactive_metrics,
        baseline_metrics=baseline_metrics,
        equity_curves=equity_curves,
        decision_rows=decision_rows,
        headline=headline,
        engine_type=engine_type,
    )


# ---------------------------------------------------------------------------
# Workbook / JSON output
# ---------------------------------------------------------------------------

METHODOLOGY_TEXT = [
    "Discretion vs Rules comparison methodology",
    "",
    "The interactive run and the AUTO baseline share the identical",
    "configuration, data, and strategy; only the discretionary decisions",
    "differ. The baseline was run automatically after the interactive run",
    "completed (in portfolio mode it uses the configured capital-contention",
    "mode, which interactive mode supersedes).",
    "",
    "Aggregate comparison (Summary and Equity Curves sheets) is exact.",
    "",
    "Per-decision attribution labels:",
    "  exact          - at or before the first divergence, both runs were",
    "                   still identical, so the matched baseline trade is",
    "                   an exact counterfactual.",
    "  approximate    - single-security runs after the first divergence:",
    "                   position sizing compounds off different capital.",
    "  path_dependent - portfolio runs after the first divergence: capital,",
    "                   contention, and sizing all differ downstream; only",
    "                   the aggregate comparison is rigorous there.",
    "",
    "'baseline_trade_pl' is the P&L of the trade the rules-only run opened",
    "(or closed) at the same symbol and date, when one exists.",
]


def write_comparison_workbook(comparison: ComparisonResult,
                              reports_dir: Path,
                              prompts: Optional[List[PromptRecord]] = None) -> Path:
    """Write the standalone discretion_vs_rules_{timestamp}.xlsx workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from ..Analysis.excel_charts import (
        add_line_chart, sample_series, write_series_block,
    )

    reports_dir = Path(reports_dir)
    reports_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = reports_dir / f"discretion_vs_rules_{timestamp}.xlsx"

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4",
                              fill_type="solid")

    def style_header(ws, row=1):
        for cell in ws[row]:
            if cell.value is not None:
                cell.font = header_font
                cell.fill = header_fill

    # ---- Summary ----
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Discretion vs Rules Summary"])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([])
    headline = comparison.headline
    for label, key in [
        ("Interactive total return", 'interactive_total_return'),
        ("Baseline (rules-only) total return", 'baseline_total_return'),
        ("Discretion added (return)", 'discretion_delta'),
        ("Interactive total return %", 'interactive_total_return_pct'),
        ("Baseline total return %", 'baseline_total_return_pct'),
        ("Discretion added (%-points)", 'discretion_delta_pct'),
        ("Decisions logged", 'num_decisions'),
        ("Divergent decisions", 'num_divergent_decisions'),
        ("First divergence date", 'first_divergence_date'),
    ]:
        ws.append([label, headline.get(key)])
    ws.append([])
    ws.append(["Decisions by action"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    for action, count in sorted(headline.get('action_counts', {}).items()):
        ws.append([action, count])
    ws.append([])
    metrics_header_row = ws.max_row + 1
    ws.append(["Metric", "Interactive", "Baseline", "Delta"])
    style_header(ws, metrics_header_row)
    for key, label in METRIC_ROWS:
        ivalue = comparison.interactive_metrics.get(key)
        bvalue = comparison.baseline_metrics.get(key)
        delta = (ivalue - bvalue
                 if isinstance(ivalue, (int, float))
                 and isinstance(bvalue, (int, float)) else None)
        ws.append([label, ivalue, bvalue, delta])
    ws.column_dimensions['A'].width = 34
    for col in ('B', 'C', 'D'):
        ws.column_dimensions[col].width = 16

    # ---- Equity Curves ----
    ws = wb.create_sheet("Equity Curves")
    ws.append(["date", "interactive", "baseline", "delta"])
    style_header(ws)
    curves = comparison.equity_curves
    dates = [str(d)[:10] for d in curves['date'].tolist()]
    sampled_dates, interactive_vals = sample_series(
        curves['interactive'].tolist(), dates)
    _, baseline_vals = sample_series(curves['baseline'].tolist(), dates)
    _, delta_vals = sample_series(curves['delta'].tolist(), dates)
    for i, date in enumerate(sampled_dates):
        ws.append([date, interactive_vals[i], baseline_vals[i], delta_vals[i]])
    block = write_series_block(
        wb, ["date", "interactive", "baseline"],
        [[sampled_dates[i], interactive_vals[i], baseline_vals[i]]
         for i in range(len(sampled_dates))])
    if block:
        add_line_chart(ws, "F2", block,
                       title="Equity: Interactive vs Rules-Only",
                       x_title="Date", y_title="Equity")

    # ---- Decisions ----
    ws = wb.create_sheet("Decisions")
    columns = ['event_id', 'bar_date', 'symbol', 'event_kind', 'signal_type',
               'action', 'source', 'rationale', 'capital_resolution',
               'executed', 'baseline_trade_pl', 'interactive_trade_pl',
               'pl_delta', 'attribution_quality']
    ws.append(columns)
    style_header(ws)
    for row in comparison.decision_rows:
        ws.append([row.get(c) for c in columns])
    ws.column_dimensions['H'].width = 50

    # ---- Prompts ----
    if prompts:
        ws = wb.create_sheet("Prompts")
        ws.append(["prompt_id", "event_id", "generated_at", "horizon_days",
                   "prompt_text", "response_summary"])
        style_header(ws)
        for prompt in prompts:
            ws.append([prompt.prompt_id, prompt.event_id, prompt.generated_at,
                       prompt.horizon_days, prompt.prompt_text,
                       prompt.response_summary])
        ws.column_dimensions['E'].width = 80
        ws.column_dimensions['F'].width = 50

    # ---- Methodology ----
    ws = wb.create_sheet("Methodology")
    for line in METHODOLOGY_TEXT:
        ws.append([line])
    ws.column_dimensions['A'].width = 90

    wb.save(path)
    return path


def write_comparison_json(comparison: ComparisonResult, path: Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(comparison.to_dict(), f, indent=2, default=str)
    return path
