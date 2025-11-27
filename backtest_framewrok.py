"""
Trading Strategy Backtesting Framework
A comprehensive system for backtesting trading strategies across multiple securities

Key Features:
- Flexible strategy definition with base class
- Efficient data preparation with column filtering and mapping
- Comprehensive backtesting with commission and slippage
- Detailed performance analysis and reporting

Dependencies:
- pandas
- numpy
- matplotlib
- seaborn
- pathlib (standard library)
"""

import pandas as pd
import numpy as np
from abc import ABC, abstractmethod
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass, field
from datetime import datetime, timedelta
import os
import json
import time
from pathlib import Path
import matplotlib.pyplot as plt
import seaborn as sns
from collections import defaultdict
import warnings
warnings.filterwarnings('ignore')

# Excel reporting imports
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import LineChart, BarChart, Reference, PieChart
    from openpyxl.chart.label import DataLabelList
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel reporting disabled. Install with: pip install openpyxl")


# ============================================================================
# STRATEGY BASE CLASS
# ============================================================================

@dataclass
class Signal:
    """Represents a trading signal"""
    action: str  # 'BUY', 'SELL', 'CLOSE', 'PARTIAL_PROFIT'
    size: float  # Position size as percentage of capital or number of shares
    stop_loss: Optional[float] = None
    take_profit: Optional[float] = None
    reason: str = ""  # Why this signal was generated


class BaseStrategy(ABC):
    """
    Base class for all trading strategies.
    Subclass this to create custom strategies.
    """
    
    def __init__(self, params: Dict = None):
        """
        Initialize strategy with parameters
        
        Args:
            params: Dictionary of strategy-specific parameters
        """
        self.params = params or {}
        self.name = self.__class__.__name__
    
    @abstractmethod
    def required_columns(self) -> List[str]:
        """
        Return list of required column names for this strategy.
        These are the column names that the strategy expects in the data.
        Must include at least ['date', 'close']
        
        Returns:
            List of column names that will be used in generate_signal()
        
        Example:
            return ['date', 'close', 'sma_50', 'rsi_14']
        """
        pass
    
    @abstractmethod
    def generate_signal(self, data: pd.DataFrame, position: Dict) -> Optional[Signal]:
        """
        Generate trading signal based on current data and position
        
        Args:
            data: Historical data up to current point (row -1 is current)
            position: Current position info {'size': 0, 'entry_price': 0, 'stop_loss': 0}
        
        Returns:
            Signal object or None
        """
        pass
    
    def position_size(self, capital: float, price: float, signal: Signal) -> float:
        """
        Calculate position size based on available capital
        
        Args:
            capital: Available capital
            price: Current price
            signal: The generated signal
        
        Returns:
            Number of shares to trade
        """
        # Default: use signal.size as percentage of capital
        amount = capital * signal.size
        shares = amount / price
        return int(shares)  # Round down to whole shares


class MovingAverageCrossStrategy(BaseStrategy):
    """
    Example Strategy: Moving Average Crossover
    Buy when price crosses above MA, sell when crosses below
    """
    
    def __init__(self, params: Dict = None):
        default_params = {
            'ma_period': 50,
            'position_size_pct': 0.1,  # 10% of capital per trade
            'stop_loss_pct': 0.05,  # 5% stop loss
            'take_profit_pct': 0.15  # 15% take profit
        }
        if params:
            default_params.update(params)
        super().__init__(default_params)
    
    def required_columns(self) -> List[str]:
        """
        This strategy needs:
        - date: for tracking
        - close: current price
        - sma_50: the 50-period simple moving average (or whatever ma_period is set to)
        """
        return ['date', 'close', f"sma_{self.params['ma_period']}"]
    
    def generate_signal(self, data: pd.DataFrame, position: Dict) -> Optional[Signal]:
        if len(data) < 2:
            return None
        
        current = data.iloc[-1]
        previous = data.iloc[-2]
        
        sma_col = f"sma_{self.params['ma_period']}"
        
        # Check if we have the required data
        if pd.isna(current[sma_col]) or pd.isna(previous[sma_col]):
            return None
        
        # If no position
        if position['size'] == 0:
            # Buy signal: price crosses above MA
            if previous['close'] <= previous[sma_col] and current['close'] > current[sma_col]:
                stop_loss = current['close'] * (1 - self.params['stop_loss_pct'])
                return Signal(
                    action='BUY',
                    size=self.params['position_size_pct'],
                    stop_loss=stop_loss,
                    take_profit=current['close'] * (1 + self.params['take_profit_pct']),
                    reason="Price crossed above MA"
                )
        else:
            # Check stop loss
            if current['close'] <= position['stop_loss']:
                return Signal(
                    action='CLOSE',
                    size=1.0,
                    reason="Stop loss hit"
                )
            
            # Check take profit
            if position.get('take_profit') and current['close'] >= position['take_profit']:
                return Signal(
                    action='CLOSE',
                    size=1.0,
                    reason="Take profit hit"
                )
            
            # Sell signal: price crosses below MA
            if previous['close'] >= previous[sma_col] and current['close'] < current[sma_col]:
                return Signal(
                    action='CLOSE',
                    size=1.0,
                    reason="Price crossed below MA"
                )
        
        return None

class TradingViewMultiPresetStrategy(BaseStrategy):
    """
    Replication of TradingView "Multi-Preset Strategy V2"
    
    This strategy combines multiple technical indicators with three preset configurations:
    - Aggressive: More trades, higher risk/reward
    - Balanced: Moderate approach
    - Conservative: Fewer trades, more selective
    
    Base Rules (always active):
    - New Highs: Close > highest high in past N candles
    - SMA 200: Close >= 200-period SMA
    
    Position sizing based on ATR and risk percentage
    """
    
    def __init__(self, params: Dict = None):
        default_params = {
            # Preset selection
            'preset': 'Balanced',  # Options: 'Aggressive', 'Balanced', 'Conservative'
            
            # Base rules
            'new_high_n': 14,
            'new_high_m': 0,
            'new_high_k': 10,
            'sma_200_length': 200,
            
            # Position sizing & risk
            'atr_length': 14,
            'atr_multiplier': 3.0,
            'risk_percent': 0.5,
            
            # CMF parameters
            'cmf_length': 14,
            'cmf_threshold': -0.2,
            
            # ROC parameters
            'roc_sma_length': 200,
            'roc_period': 14,
            'roc_threshold': -1.5,
            
            # RSI parameters
            'rsi_length': 14,
            'rsi_threshold': 60.0,
            
            # Volume parameters
            'vol_period': 30,
            
            # Bollinger Bands parameters
            'bb_length': 20,
            'bb_std_dev': 2.0,
            
            # MA Crossover parameters
            'ma_fast_length': 7,
            'ma_slow_length': 21,
            'ma_lookback_k': 14,
            
            # SAR parameters
            'sar_start': 0.02,
            'sar_increment': 0.02,
            'sar_maximum': 0.2,
            
            # SQZMOM parameters
            'sqz_length': 20,
            'sqz_mult': 2.0,
            'sqz_length_kc': 20,
            'sqz_mult_kc': 1.5,
            
            # EMA Sell parameters
            'ema_length': 14,
            'ema_sell_threshold': 3.0
        }
        
        if params:
            default_params.update(params)
        super().__init__(default_params)
        
        # Validate preset
        if self.params['preset'] not in ['Aggressive', 'Balanced', 'Conservative']:
            raise ValueError("Preset must be 'Aggressive', 'Balanced', or 'Conservative'")
    
    def required_columns(self) -> List[str]:
        """
        All columns needed for this strategy
        """
        cols = ['date', 'open', 'high', 'low', 'close', 'volume']
        
        # Add SMA 200
        cols.append(f"sma_{self.params['sma_200_length']}")
        
        # Add EMA for fast and slow MA crossover
        cols.append(f"ema_{self.params['ma_fast_length']}")
        cols.append(f"ema_{self.params['ma_slow_length']}")
        
        # Add EMA for sell signal
        cols.append(f"ema_{self.params['ema_length']}")
        
        # Add RSI
        cols.append(f"rsi_{self.params['rsi_length']}")
        
        # Add ATR
        cols.append(f"atr_{self.params['atr_length']}")
        
        return cols
    
    def generate_signal(self, data: pd.DataFrame, position: Dict) -> Optional[Signal]:
        """Generate trading signals based on preset configuration"""
        if len(data) < max(self.params['sma_200_length'], self.params['new_high_n']) + 10:
            return None
        
        current = data.iloc[-1]
        preset = self.params['preset']
        
        # ==== BASE RULES (Always required for BUY) ====
        
        # New Highs Rule
        try:
            lookback_start = max(1, len(data) - self.params['new_high_n'] - 1)
            lookback_end = len(data) - 1
            lookback_highs = data['high'].iloc[lookback_start:lookback_end]
            
            if len(lookback_highs) == 0:
                return None
                
            new_high_condition = current['close'] > lookback_highs.max()
            
            # Count new highs in last K candles
            new_highs_count = 0
            for i in range(1, min(self.params['new_high_k'] + 1, len(data))):
                past_idx_start = max(0, len(data) - i - self.params['new_high_n'] - 1)
                past_idx_end = len(data) - i
                
                if past_idx_start < past_idx_end:
                    past_highs = data['high'].iloc[past_idx_start:past_idx_end]
                    past_close = data['close'].iloc[past_idx_end - 1]
                    
                    if len(past_highs) > 0 and past_close > past_highs.max():
                        new_highs_count += 1
            
            new_highs_rule = new_high_condition and new_highs_count >= self.params['new_high_m']
        except Exception as e:
            print(f"Error in new highs calculation: {e}")
            new_highs_rule = False
        
        # SMA 200 Rule
        sma_200_col = f"sma_{self.params['sma_200_length']}"
        if sma_200_col not in data.columns or pd.isna(current[sma_200_col]):
            return None
        sma_200_rule = current['close'] >= current[sma_200_col]
        
        base_rules = new_highs_rule and sma_200_rule
        
        # ==== CHECK SELL SIGNALS (if in position) ====
        if position['size'] > 0:
            # Check stop loss first
            if current['close'] <= position['stop_loss']:
                return Signal(action='CLOSE', size=1.0, reason="Stop loss hit")
            
            # Check sell signals based on preset
            ema_col = f"ema_{self.params['ema_length']}"
            if ema_col in data.columns and not pd.isna(current[ema_col]):
                ema_value = current[ema_col]
                ema_sell_threshold_price = current['close'] * (1 + self.params['ema_sell_threshold'] / 100)
                
                sell_signal = False
                
                if preset == 'Aggressive':
                    sqzmom = self._calculate_sqzmom(data)
                    if sqzmom < 0:
                        sell_signal = True
                        return Signal(action='CLOSE', size=1.0, reason="SQZMOM sell signal")
                    if ema_value > ema_sell_threshold_price:
                        sell_signal = True
                        return Signal(action='CLOSE', size=1.0, reason="EMA sell signal")
                
                elif preset in ['Balanced', 'Conservative']:
                    if ema_value > ema_sell_threshold_price:
                        return Signal(action='CLOSE', size=1.0, reason="EMA sell signal")
        
        # ==== BUY SIGNAL (only if not in position and base rules met) ====
        if position['size'] == 0 and base_rules:
            
            additional_buy_rules = True
            
            if preset == 'Aggressive':
                # MA Crossover Rule
                ema_fast_col = f"ema_{self.params['ma_fast_length']}"
                ema_slow_col = f"ema_{self.params['ma_slow_length']}"
                
                if ema_fast_col not in data.columns or ema_slow_col not in data.columns:
                    return None
                
                ma_rule = False
                # Check for crossover in last K candles
                lookback_range = min(self.params['ma_lookback_k'] + 1, len(data) - 1)
                for i in range(lookback_range):
                    idx = -(i + 1)
                    prev_idx = idx - 1
                    
                    if abs(prev_idx) >= len(data):
                        break
                    
                    curr_fast = data[ema_fast_col].iloc[idx]
                    curr_slow = data[ema_slow_col].iloc[idx]
                    prev_fast = data[ema_fast_col].iloc[prev_idx]
                    prev_slow = data[ema_slow_col].iloc[prev_idx]
                    
                    if not any(pd.isna([curr_fast, curr_slow, prev_fast, prev_slow])):
                        if prev_fast <= prev_slow and curr_fast > curr_slow:
                            ma_rule = True
                            break
                
                additional_buy_rules = ma_rule
            
            elif preset == 'Balanced':
                # MA Crossover Rule
                ema_fast_col = f"ema_{self.params['ma_fast_length']}"
                ema_slow_col = f"ema_{self.params['ma_slow_length']}"
                
                if ema_fast_col not in data.columns or ema_slow_col not in data.columns:
                    return None
                
                ma_rule = False
                lookback_range = min(self.params['ma_lookback_k'] + 1, len(data) - 1)
                for i in range(lookback_range):
                    idx = -(i + 1)
                    prev_idx = idx - 1
                    
                    if abs(prev_idx) >= len(data):
                        break
                    
                    curr_fast = data[ema_fast_col].iloc[idx]
                    curr_slow = data[ema_slow_col].iloc[idx]
                    prev_fast = data[ema_fast_col].iloc[prev_idx]
                    prev_slow = data[ema_slow_col].iloc[prev_idx]
                    
                    if not any(pd.isna([curr_fast, curr_slow, prev_fast, prev_slow])):
                        if prev_fast <= prev_slow and curr_fast > curr_slow:
                            ma_rule = True
                            break
                
                # SAR Buy Rule - simplified (SAR below price = uptrend)
                sar_buy_rule = True  # Simplified - in real implementation calculate actual SAR
                
                additional_buy_rules = ma_rule and sar_buy_rule
            
            elif preset == 'Conservative':
                # SAR Buy Rule - simplified
                sar_buy_rule = True  # Simplified
                additional_buy_rules = sar_buy_rule
            
            # Generate BUY signal if all conditions met
            if additional_buy_rules:
                # Calculate stop loss based on ATR
                atr_col = f"atr_{self.params['atr_length']}"
                if atr_col not in data.columns or pd.isna(current[atr_col]):
                    return None
                
                stop_distance = current[atr_col] * self.params['atr_multiplier']
                stop_loss = current['close'] - stop_distance
                
                return Signal(
                    action='BUY',
                    size=self.params['risk_percent'] / 100,
                    stop_loss=stop_loss,
                    reason=f"{preset} preset buy signal"
                )
        
        return None
    
    def _calculate_sqzmom(self, data: pd.DataFrame) -> float:
        """Calculate Squeeze Momentum LazyBear - simplified version"""
        length = self.params['sqz_length']
        
        if len(data) < length + 10:
            return 0
        
        try:
            df = data.tail(length + 10).copy()
            
            # Calculate momentum using linear regression on price deviations
            highest_high = df['high'].rolling(window=self.params['sqz_length_kc']).max()
            lowest_low = df['low'].rolling(window=self.params['sqz_length_kc']).min()
            sma_close = df['close'].rolling(window=self.params['sqz_length_kc']).mean()
            
            val = df['close'] - ((highest_high + lowest_low) / 2 + sma_close) / 2
            
            # Get last length values
            val_clean = val.dropna()
            if len(val_clean) >= length:
                y = val_clean.tail(length).values
                x = np.arange(len(y))
                
                if len(y) == len(x) and len(y) > 0:
                    # Linear regression
                    slope, _ = np.polyfit(x, y, 1)
                    return slope
            
            return 0
        except Exception as e:
            print(f"Error calculating SQZMOM: {e}")
            return 0
    
    def position_size(self, capital: float, price: float, signal: Signal) -> float:
        """
        Calculate position size based on ATR and risk percentage
        Uses the formula: position_size = (equity * risk%) / stop_distance
        """
        if signal.stop_loss is None or signal.stop_loss <= 0:
            # Fallback to default sizing
            return super().position_size(capital, price, signal)
        
        stop_distance = price - signal.stop_loss
        if stop_distance <= 0:
            return 0
        
        max_loss = capital * self.params['risk_percent'] / 100
        shares = max_loss / stop_distance
        
        return max(int(shares), 0)  # Ensure non-negative
        
        

# ============================================================================
# DATA PREPARATION CLASS
# ============================================================================

class DataPreparation:
    """
    Handles data preparation by filtering raw data to only required columns
    and optionally renaming columns based on a mapping
    """
    
    def __init__(self, raw_data_folder: str, prepared_data_folder: str):
        """
        Initialize data preparation
        
        Args:
            raw_data_folder: Path to folder with raw CSV files (with all indicators)
            prepared_data_folder: Path to save prepared data
        """
        self.raw_data_folder = Path(raw_data_folder)
        self.prepared_data_folder = Path(prepared_data_folder)
    
    def get_required_columns(self, strategy: BaseStrategy) -> List[str]:
        """
        Get list of columns required by the strategy
        
        Args:
            strategy: Strategy instance
        
        Returns:
            List of required column names
        """
        return strategy.required_columns()
    
    def prepare_data_for_strategy(self, 
                                  strategy: BaseStrategy, 
                                  security_metadata: Dict[str, Dict],
                                  column_mapping: Optional[Dict[str, str]] = None) -> Dict[str, str]:
        """
        Prepare data for all securities based on strategy requirements
        
        Args:
            strategy: Strategy instance
            security_metadata: Dict mapping symbol to metadata 
                              (e.g., {'AAPL': {'type': 'stock', 'sector': 'tech'}})
            column_mapping: Optional dict mapping strategy column names to raw data column names
                          Example: {'sma_50': 'SMA_50', 'close': 'Close', 'date': 'Date'}
                          If None, assumes column names in raw data match strategy requirements
        
        Returns:
            Dict mapping symbol to path of prepared data file
        """
        strategy_folder = self.prepared_data_folder / strategy.name
        strategy_folder.mkdir(parents=True, exist_ok=True)
        
        required_cols = strategy.required_columns()
        
        print(f"\n{'='*60}")
        print(f"Strategy: {strategy.name}")
        print(f"Required columns: {required_cols}")
        if column_mapping:
            print(f"Column mapping provided:")
            for strategy_col, raw_col in column_mapping.items():
                if strategy_col in required_cols:
                    print(f"  {strategy_col} <- {raw_col}")
        print(f"{'='*60}")
        
        prepared_files = {}
        missing_columns_report = {}
        
        for symbol, metadata in security_metadata.items():
            print(f"\nPreparing data for {symbol}...")
            
            # Load raw data
            raw_file = self.raw_data_folder / f"{symbol}.csv"
            if not raw_file.exists():
                print(f"  ⚠ Warning: Raw data file not found for {symbol}")
                continue
            
            try:
                df = pd.read_csv(raw_file)
                print(f"  Loaded {len(df)} rows")
                print(f"  Available columns: {list(df.columns)}")
                
                # Create column mapping
                if column_mapping:
                    # Use provided mapping
                    rename_dict = {}
                    for strategy_col in required_cols:
                        if strategy_col in column_mapping:
                            raw_col = column_mapping[strategy_col]
                            if raw_col in df.columns:
                                rename_dict[raw_col] = strategy_col
                            else:
                                print(f"  ⚠ Warning: Mapped column '{raw_col}' not found in {symbol}")
                                if symbol not in missing_columns_report:
                                    missing_columns_report[symbol] = []
                                missing_columns_report[symbol].append(f"{strategy_col} (mapped from {raw_col})")
                    
                    # Rename columns
                    df = df.rename(columns=rename_dict)
                
                # Check if all required columns are present
                missing_cols = [col for col in required_cols if col not in df.columns]
                if missing_cols:
                    print(f"  ⚠ Warning: Missing required columns: {missing_cols}")
                    if symbol not in missing_columns_report:
                        missing_columns_report[symbol] = []
                    missing_columns_report[symbol].extend(missing_cols)
                    continue
                
                # Keep only required columns
                df = df[required_cols].copy()
                
                # Convert date column to datetime
                df['date'] = pd.to_datetime(df['date'])
                df = df.sort_values('date').reset_index(drop=True)
                
                # Remove rows with any NaN values in required columns
                initial_rows = len(df)
                df = df.dropna()
                if len(df) < initial_rows:
                    print(f"  Removed {initial_rows - len(df)} rows with missing values")
                
                # Add metadata columns
                for key, value in metadata.items():
                    df[f'metadata_{key}'] = value
                
                # Save prepared data
                output_file = strategy_folder / f"{symbol}.csv"
                df.to_csv(output_file, index=False)
                prepared_files[symbol] = str(output_file)
                
                print(f"  ✓ Saved prepared data: {len(df)} rows, {len(df.columns)} columns")
                
            except Exception as e:
                print(f"  ✗ Error processing {symbol}: {e}")
                continue
        
        # Save metadata about this preparation run
        metadata_file = strategy_folder / "preparation_metadata.json"
        with open(metadata_file, 'w') as f:
            json.dump({
                'strategy': strategy.name,
                'strategy_params': strategy.params,
                'preparation_date': datetime.now().isoformat(),
                'required_columns': required_cols,
                'column_mapping_used': column_mapping,
                'securities_prepared': list(prepared_files.keys()),
                'securities_with_missing_columns': missing_columns_report
            }, f, indent=2)
        
        # Print summary
        print(f"\n{'='*60}")
        print(f"PREPARATION SUMMARY")
        print(f"{'='*60}")
        print(f"Successfully prepared: {len(prepared_files)} securities")
        if missing_columns_report:
            print(f"\n⚠ Securities with missing columns:")
            for symbol, missing in missing_columns_report.items():
                print(f"  {symbol}: {missing}")
        print(f"\nPrepared data saved to: {strategy_folder}")
        print(f"{'='*60}")
        
        return prepared_files


# ============================================================================
# BACKTESTING ENGINE
# ============================================================================

@dataclass
class Trade:
    """Record of a single trade"""
    symbol: str
    entry_date: datetime
    entry_price: float
    exit_date: Optional[datetime] = None
    exit_price: Optional[float] = None
    quantity: float = 0
    side: str = "LONG"  # LONG or SHORT
    initial_stop_loss: float = 0
    current_stop_loss: float = 0
    take_profit: Optional[float] = None
    pl: float = 0
    pl_pct: float = 0
    duration_days: int = 0
    exit_reason: str = ""
    commission_paid: float = 0


class BacktestEngine:
    """
    Executes backtesting of a strategy on historical data
    """
    
    def __init__(self, 
                 strategy: BaseStrategy,
                 starting_capital: float = 100000,
                 commission_pct: float = 0.001,  # 0.1%
                 slippage_pct: float = 0.0005,   # 0.05%
                 start_date: Optional[str] = None,
                 end_date: Optional[str] = None):
        """
        Initialize backtest engine
        
        Args:
            strategy: Strategy to backtest
            starting_capital: Starting capital
            commission_pct: Commission as percentage of trade value
            slippage_pct: Slippage as percentage of price
            start_date: Start date for backtest (YYYY-MM-DD)
            end_date: End date for backtest (YYYY-MM-DD)
        """
        self.strategy = strategy
        self.starting_capital = starting_capital
        self.commission_pct = commission_pct
        self.slippage_pct = slippage_pct
        self.start_date = pd.to_datetime(start_date) if start_date else None
        self.end_date = pd.to_datetime(end_date) if end_date else None
        
        self.trades_log = []
        self.equity_curve = {}  # symbol -> list of (date, equity)
    
    def _calculate_commission(self, value: float) -> float:
        """Calculate commission on trade value"""
        return value * self.commission_pct
    
    def _apply_slippage(self, price: float, side: str) -> float:
        """Apply slippage to execution price"""
        if side == 'BUY':
            return price * (1 + self.slippage_pct)
        else:  # SELL
            return price * (1 - self.slippage_pct)
    
    def backtest_security(self, symbol: str, data: pd.DataFrame) -> Tuple[List[Trade], pd.DataFrame]:
        """
        Backtest strategy on a single security
        
        Args:
            symbol: Security symbol
            data: Prepared historical data
        
        Returns:
            Tuple of (trades list, equity curve dataframe)
        """
        print(f"\nBacktesting {symbol}...")
        
        # Filter by date range
        data = data.copy()
        if self.start_date:
            data = data[data['date'] >= self.start_date]
        if self.end_date:
            data = data[data['date'] <= self.end_date]
        
        if len(data) == 0:
            print(f"No data for {symbol} in specified date range")
            return [], pd.DataFrame()
        
        # Initialize
        capital = self.starting_capital
        position = {
            'size': 0,
            'entry_price': 0,
            'stop_loss': 0,
            'take_profit': None,
            'current_trade': None
        }
        
        trades = []
        equity_history = []
        
        # Iterate through data
        for i in range(len(data)):
            current_data = data.iloc[:i+1]
            current_row = current_data.iloc[-1]
            current_price = current_row['close']
            current_date = current_row['date']
            
            # Track equity
            position_value = position['size'] * current_price if position['size'] > 0 else 0
            current_equity = capital + position_value
            equity_history.append({
                'date': current_date,
                'equity': current_equity,
                'capital': capital,
                'position_value': position_value
            })
            
            # Generate signal
            signal = self.strategy.generate_signal(current_data, position)
            
            if signal is None:
                continue
            
            # Execute signal
            if signal.action == 'BUY' and position['size'] == 0:
                # Open position
                exec_price = self._apply_slippage(current_price, 'BUY')
                shares = self.strategy.position_size(capital, exec_price, signal)
                
                if shares > 0:
                    trade_value = shares * exec_price
                    commission = self._calculate_commission(trade_value)
                    total_cost = trade_value + commission
                    
                    if total_cost <= capital:
                        capital -= total_cost
                        position['size'] = shares
                        position['entry_price'] = exec_price
                        position['stop_loss'] = signal.stop_loss or 0
                        position['take_profit'] = signal.take_profit
                        
                        trade = Trade(
                            symbol=symbol,
                            entry_date=current_date,
                            entry_price=exec_price,
                            quantity=shares,
                            initial_stop_loss=signal.stop_loss or 0,
                            current_stop_loss=signal.stop_loss or 0,
                            take_profit=signal.take_profit,
                            commission_paid=commission
                        )
                        position['current_trade'] = trade
                        
                        print(f"  BUY {shares} @ ${exec_price:.2f} on {current_date.date()} - {signal.reason}")
            
            elif signal.action in ['SELL', 'CLOSE'] and position['size'] > 0:
                # Close position
                exec_price = self._apply_slippage(current_price, 'SELL')
                shares = position['size']
                trade_value = shares * exec_price
                commission = self._calculate_commission(trade_value)
                proceeds = trade_value - commission
                
                capital += proceeds
                
                # Complete trade record
                trade = position['current_trade']
                trade.exit_date = current_date
                trade.exit_price = exec_price
                trade.pl = proceeds - (shares * trade.entry_price + trade.commission_paid)
                trade.pl_pct = (exec_price / trade.entry_price - 1) * 100
                trade.duration_days = (current_date - trade.entry_date).days
                trade.exit_reason = signal.reason
                trade.commission_paid += commission
                
                trades.append(trade)
                
                print(f"  SELL {shares} @ ${exec_price:.2f} on {current_date.date()} - P/L: ${trade.pl:.2f} ({trade.pl_pct:.2f}%) - {signal.reason}")
                
                # Reset position
                position = {
                    'size': 0,
                    'entry_price': 0,
                    'stop_loss': 0,
                    'take_profit': None,
                    'current_trade': None
                }
        
        equity_df = pd.DataFrame(equity_history)
        
        return trades, equity_df
    
    def backtest_all_securities(self, data_folder: str) -> Dict[str, Tuple[List[Trade], pd.DataFrame]]:
        """
        Backtest strategy on all securities in folder
        
        Args:
            data_folder: Path to folder with prepared data
        
        Returns:
            Dict mapping symbol to (trades, equity_curve)
        """
        results = {}
        data_path = Path(data_folder)
        
        csv_files = list(data_path.glob("*.csv"))
        print(f"\nFound {len(csv_files)} securities to backtest")
        
        for csv_file in csv_files:
            if csv_file.name == "preparation_metadata.json":
                continue
            
            symbol = csv_file.stem
            data = pd.read_csv(csv_file)
            data['date'] = pd.to_datetime(data['date'])
            
            trades, equity_curve = self.backtest_security(symbol, data)
            results[symbol] = (trades, equity_curve)
        
        self.results = results
        return results


# ============================================================================
# REPORTING AND ANALYSIS
# ============================================================================

class PerformanceAnalyzer:
    """
    Generates comprehensive performance reports (PNG charts and Excel workbooks)
    """
    
    def __init__(self, backtest_results: Dict, output_folder: str):
        """
        Initialize analyzer
        
        Args:
            backtest_results: Results from BacktestEngine
            output_folder: Path to save reports
        """
        self.results = backtest_results
        self.output_folder = Path(output_folder)
        self.output_folder.mkdir(parents=True, exist_ok=True)
        
        # Set style
        sns.set_style("whitegrid")
        plt.rcParams['figure.figsize'] = (12, 6)
    
    def _calculate_sharpe_ratio(self, returns: pd.Series, risk_free_rate: float = 0.02) -> float:
        """Calculate annualized Sharpe ratio"""
        if len(returns) == 0 or returns.std() == 0:
            return 0
        excess_returns = returns - risk_free_rate/252  # Daily risk-free rate
        return np.sqrt(252) * excess_returns.mean() / returns.std()
    
    def _calculate_max_drawdown(self, equity_curve: pd.DataFrame) -> Tuple[float, pd.Timestamp, pd.Timestamp]:
        """Calculate maximum drawdown"""
        if len(equity_curve) == 0:
            return 0, None, None
        
        equity = equity_curve['equity'].values
        running_max = np.maximum.accumulate(equity)
        drawdown = (equity - running_max) / running_max
        
        max_dd = drawdown.min()
        max_dd_idx = drawdown.argmin()
        
        # Find peak before drawdown
        peak_idx = equity[:max_dd_idx+1].argmax()
        
        peak_date = equity_curve.iloc[peak_idx]['date']
        trough_date = equity_curve.iloc[max_dd_idx]['date']
        
        return max_dd, peak_date, trough_date
    
    def analyze_security_performance(self, symbol: str, trades: List[Trade], 
                                    equity_curve: pd.DataFrame) -> Dict:
        """
        Analyze performance for a single security
        
        Args:
            symbol: Security symbol
            trades: List of trades
            equity_curve: Equity curve dataframe
        
        Returns:
            Dictionary of performance metrics
        """
        if len(trades) == 0:
            return {'symbol': symbol, 'error': 'No trades executed'}
        
        # Basic metrics
        total_pl = sum(t.pl for t in trades)
        total_pl_pct = (total_pl / equity_curve.iloc[0]['capital']) * 100
        num_trades = len(trades)
        winning_trades = [t for t in trades if t.pl > 0]
        losing_trades = [t for t in trades if t.pl <= 0]
        
        win_rate = len(winning_trades) / num_trades if num_trades > 0 else 0
        
        avg_win = np.mean([t.pl for t in winning_trades]) if winning_trades else 0
        avg_loss = np.mean([t.pl for t in losing_trades]) if losing_trades else 0
        
        profit_factor = (sum(t.pl for t in winning_trades) / 
                        abs(sum(t.pl for t in losing_trades))) if losing_trades else float('inf')
        
        avg_duration = np.mean([t.duration_days for t in trades])
        
        # Calculate returns
        equity_curve['returns'] = equity_curve['equity'].pct_change()
        sharpe = self._calculate_sharpe_ratio(equity_curve['returns'].dropna())
        
        # Max drawdown
        max_dd, peak_date, trough_date = self._calculate_max_drawdown(equity_curve)
        
        # Best and worst trades
        best_trade = max(trades, key=lambda t: t.pl)
        worst_trade = min(trades, key=lambda t: t.pl)
        
        metrics = {
            'symbol': symbol,
            'total_pl': total_pl,
            'total_pl_pct': total_pl_pct,
            'num_trades': num_trades,
            'win_rate': win_rate,
            'num_winning_trades': len(winning_trades),
            'num_losing_trades': len(losing_trades),
            'avg_win': avg_win,
            'avg_loss': avg_loss,
            'profit_factor': profit_factor,
            'avg_duration_days': avg_duration,
            'sharpe_ratio': sharpe,
            'max_drawdown': max_dd,
            'max_drawdown_peak_date': peak_date,
            'max_drawdown_trough_date': trough_date,
            'best_trade_pl': best_trade.pl,
            'best_trade_date': best_trade.entry_date,
            'worst_trade_pl': worst_trade.pl,
            'worst_trade_date': worst_trade.entry_date,
            'final_equity': equity_curve.iloc[-1]['equity'],
            'start_date': equity_curve.iloc[0]['date'],
            'end_date': equity_curve.iloc[-1]['date']
        }
        
        return metrics
    
    def generate_security_report(self, symbol: str, trades: List[Trade], 
                                equity_curve: pd.DataFrame, metadata: Dict = None):
        """
        Generate detailed report for a single security
        
        Args:
            symbol: Security symbol
            trades: List of trades
            equity_curve: Equity curve dataframe
            metadata: Optional metadata about the security
        """
        print(f"\nGenerating report for {symbol}...")
        
        metrics = self.analyze_security_performance(symbol, trades, equity_curve)
        
        if 'error' in metrics:
            print(f"  Skipping {symbol}: {metrics['error']}")
            return
        
        # Create figure with subplots
        fig = plt.figure(figsize=(16, 12))
        gs = fig.add_gridspec(4, 2, hspace=0.3, wspace=0.3)
        
        # 1. Equity Curve
        ax1 = fig.add_subplot(gs[0, :])
        ax1.plot(equity_curve['date'], equity_curve['equity'], linewidth=2)
        ax1.set_title(f'{symbol} - Equity Curve', fontsize=14, fontweight='bold')
        ax1.set_xlabel('Date')
        ax1.set_ylabel('Equity ($)')
        ax1.grid(True, alpha=0.3)
        
        # 2. P/L Distribution
        ax2 = fig.add_subplot(gs[1, 0])
        pls = [t.pl for t in trades]
        ax2.hist(pls, bins=30, edgecolor='black', alpha=0.7)
        ax2.axvline(0, color='red', linestyle='--', linewidth=2)
        ax2.set_title('P/L Distribution', fontweight='bold')
        ax2.set_xlabel('Profit/Loss ($)')
        ax2.set_ylabel('Frequency')
        
        # 3. Trade Duration Distribution
        ax3 = fig.add_subplot(gs[1, 1])
        durations = [t.duration_days for t in trades]
        ax3.hist(durations, bins=30, edgecolor='black', alpha=0.7, color='green')
        ax3.set_title('Trade Duration Distribution', fontweight='bold')
        ax3.set_xlabel('Duration (days)')
        ax3.set_ylabel('Frequency')
        
        # 4. Cumulative P/L
        ax4 = fig.add_subplot(gs[2, :])
        cumulative_pl = np.cumsum([t.pl for t in trades])
        trade_dates = [t.exit_date for t in trades]
        ax4.plot(trade_dates, cumulative_pl, linewidth=2, marker='o', markersize=4)
        ax4.set_title('Cumulative P/L Over Time', fontweight='bold')
        ax4.set_xlabel('Date')
        ax4.set_ylabel('Cumulative P/L ($)')
        ax4.grid(True, alpha=0.3)
        
        # 5. Performance Metrics Table
        ax5 = fig.add_subplot(gs[3, :])
        ax5.axis('off')
        
        metrics_text = f"""
        PERFORMANCE SUMMARY FOR {symbol}
        {'='*60}
        
        Total P/L: ${metrics['total_pl']:,.2f} ({metrics['total_pl_pct']:.2f}%)
        Number of Trades: {metrics['num_trades']}
        Win Rate: {metrics['win_rate']*100:.2f}%
        
        Winning Trades: {metrics['num_winning_trades']} (Avg: ${metrics['avg_win']:,.2f})
        Losing Trades: {metrics['num_losing_trades']} (Avg: ${metrics['avg_loss']:,.2f})
        Profit Factor: {metrics['profit_factor']:.2f}
        
        Sharpe Ratio: {metrics['sharpe_ratio']:.3f}
        Max Drawdown: {metrics['max_drawdown']*100:.2f}%
        
        Average Trade Duration: {metrics['avg_duration_days']:.1f} days
        
        Best Trade: ${metrics['best_trade_pl']:,.2f} on {metrics['best_trade_date'].date()}
        Worst Trade: ${metrics['worst_trade_pl']:,.2f} on {metrics['worst_trade_date'].date()}
        
        Period: {metrics['start_date'].date()} to {metrics['end_date'].date()}
        """
        
        if metadata:
            metrics_text += f"\nMetadata: {metadata}"
        
        ax5.text(0.1, 0.5, metrics_text, fontfamily='monospace', 
                fontsize=10, verticalalignment='center')
        
        # Save figure
        plt.tight_layout()
        plt.savefig(self.output_folder / f'{symbol}_report.png', dpi=150, bbox_inches='tight')
        plt.close()
        
        # Save detailed trade log
        trades_df = pd.DataFrame([{
            'entry_date': t.entry_date,
            'exit_date': t.exit_date,
            'entry_price': t.entry_price,
            'exit_price': t.exit_price,
            'quantity': t.quantity,
            'pl': t.pl,
            'pl_pct': t.pl_pct,
            'duration_days': t.duration_days,
            'exit_reason': t.exit_reason,
            'commission': t.commission_paid,
            'initial_stop_loss': t.initial_stop_loss
        } for t in trades])
        
        trades_df.to_csv(self.output_folder / f'{symbol}_trades.csv', index=False)
        
        # Save metrics
        with open(self.output_folder / f'{symbol}_metrics.json', 'w') as f:
            # Convert dates to strings for JSON serialization
            metrics_json = metrics.copy()
            for key, value in metrics_json.items():
                if isinstance(value, (pd.Timestamp, datetime)):
                    metrics_json[key] = value.isoformat()
            json.dump(metrics_json, f, indent=2)
        
        print(f"  Report saved for {symbol}")
    
    def generate_aggregate_report(self, security_metadata: Dict[str, Dict] = None):
        """
        Generate aggregate report across all securities
        
        Args:
            security_metadata: Optional metadata for grouping (type, sector, etc.)
        """
        print("\nGenerating aggregate report...")
        
        # Collect all metrics
        all_metrics = []
        all_trades = []
        
        for symbol, (trades, equity_curve) in self.results.items():
            if len(trades) > 0:
                metrics = self.analyze_security_performance(symbol, trades, equity_curve)
                if 'error' not in metrics:
                    if security_metadata and symbol in security_metadata:
                        metrics.update(security_metadata[symbol])
                    all_metrics.append(metrics)
                    all_trades.extend(trades)
        
        if len(all_metrics) == 0:
            print("No successful trades to analyze")
            return
        
        metrics_df = pd.DataFrame(all_metrics)
        
        # Overall summary
        total_pl = metrics_df['total_pl'].sum()
        avg_win_rate = metrics_df['win_rate'].mean()
        avg_sharpe = metrics_df['sharpe_ratio'].mean()
        total_trades = metrics_df['num_trades'].sum()
        
        # Create comprehensive report
        fig = plt.figure(figsize=(20, 16))
        gs = fig.add_gridspec(5, 3, hspace=0.4, wspace=0.3)
        
        # 1. Overall P/L by Security
        ax1 = fig.add_subplot(gs[0, :])
        metrics_df_sorted = metrics_df.sort_values('total_pl', ascending=False)
        colors = ['green' if x > 0 else 'red' for x in metrics_df_sorted['total_pl']]
        ax1.bar(range(len(metrics_df_sorted)), metrics_df_sorted['total_pl'], color=colors, alpha=0.7)
        ax1.set_xticks(range(len(metrics_df_sorted)))
        ax1.set_xticklabels(metrics_df_sorted['symbol'], rotation=45, ha='right')
        ax1.set_title('Total P/L by Security', fontsize=14, fontweight='bold')
        ax1.set_ylabel('P/L ($)')
        ax1.axhline(0, color='black', linestyle='-', linewidth=0.5)
        ax1.grid(True, alpha=0.3, axis='y')
        
        # 2. Win Rate by Security
        ax2 = fig.add_subplot(gs[1, 0])
        metrics_df_sorted_wr = metrics_df.sort_values('win_rate', ascending=False)
        ax2.barh(range(len(metrics_df_sorted_wr)), metrics_df_sorted_wr['win_rate']*100, alpha=0.7)
        ax2.set_yticks(range(len(metrics_df_sorted_wr)))
        ax2.set_yticklabels(metrics_df_sorted_wr['symbol'])
        ax2.set_title('Win Rate by Security', fontweight='bold')
        ax2.set_xlabel('Win Rate (%)')
        ax2.axvline(50, color='red', linestyle='--', linewidth=1)
        
        # 3. Sharpe Ratio by Security
        ax3 = fig.add_subplot(gs[1, 1])
        metrics_df_sorted_sharpe = metrics_df.sort_values('sharpe_ratio', ascending=False)
        ax3.barh(range(len(metrics_df_sorted_sharpe)), metrics_df_sorted_sharpe['sharpe_ratio'], 
                color='purple', alpha=0.7)
        ax3.set_yticks(range(len(metrics_df_sorted_sharpe)))
        ax3.set_yticklabels(metrics_df_sorted_sharpe['symbol'])
        ax3.set_title('Sharpe Ratio by Security', fontweight='bold')
        ax3.set_xlabel('Sharpe Ratio')
        
        # 4. Max Drawdown by Security
        ax4 = fig.add_subplot(gs[1, 2])
        metrics_df_sorted_dd = metrics_df.sort_values('max_drawdown')
        ax4.barh(range(len(metrics_df_sorted_dd)), metrics_df_sorted_dd['max_drawdown']*100, 
                color='orange', alpha=0.7)
        ax4.set_yticks(range(len(metrics_df_sorted_dd)))
        ax4.set_yticklabels(metrics_df_sorted_dd['symbol'])
        ax4.set_title('Max Drawdown by Security', fontweight='bold')
        ax4.set_xlabel('Max Drawdown (%)')
        
        # 5. Performance by Type (if metadata available)
        if security_metadata and 'metadata_type' in metrics_df.columns:
            ax5 = fig.add_subplot(gs[2, 0])
            type_performance = metrics_df.groupby('metadata_type')['total_pl'].sum().sort_values(ascending=False)
            ax5.bar(range(len(type_performance)), type_performance.values, alpha=0.7)
            ax5.set_xticks(range(len(type_performance)))
            ax5.set_xticklabels(type_performance.index, rotation=45, ha='right')
            ax5.set_title('Total P/L by Security Type', fontweight='bold')
            ax5.set_ylabel('P/L ($)')
        
        # 6. Performance by Sector (if available)
        if security_metadata and 'metadata_sector' in metrics_df.columns:
            ax6 = fig.add_subplot(gs[2, 1])
            sector_performance = metrics_df.groupby('metadata_sector')['total_pl'].sum().sort_values(ascending=False)
            ax6.bar(range(len(sector_performance)), sector_performance.values, alpha=0.7, color='teal')
            ax6.set_xticks(range(len(sector_performance)))
            ax6.set_xticklabels(sector_performance.index, rotation=45, ha='right')
            ax6.set_title('Total P/L by Sector', fontweight='bold')
            ax6.set_ylabel('P/L ($)')
        
        # 7. Trade P/L Distribution (all trades)
        ax7 = fig.add_subplot(gs[2, 2])
        all_pls = [t.pl for t in all_trades]
        ax7.hist(all_pls, bins=50, edgecolor='black', alpha=0.7)
        ax7.axvline(0, color='red', linestyle='--', linewidth=2)
        ax7.set_title('Overall P/L Distribution', fontweight='bold')
        ax7.set_xlabel('P/L per Trade ($)')
        ax7.set_ylabel('Frequency')
        
        # 8. Monthly P/L heatmap
        ax8 = fig.add_subplot(gs[3, :])
        trades_df = pd.DataFrame([{
            'exit_date': t.exit_date,
            'pl': t.pl
        } for t in all_trades if t.exit_date is not None])
        
        if len(trades_df) > 0:
            trades_df['year'] = trades_df['exit_date'].dt.year
            trades_df['month'] = trades_df['exit_date'].dt.month
            monthly_pl = trades_df.groupby(['year', 'month'])['pl'].sum().unstack(fill_value=0)
            
            sns.heatmap(monthly_pl, annot=True, fmt='.0f', cmap='RdYlGn', center=0, 
                       ax=ax8, cbar_kws={'label': 'P/L ($)'})
            ax8.set_title('Monthly P/L Heatmap', fontweight='bold')
            ax8.set_xlabel('Month')
            ax8.set_ylabel('Year')
        
        # 9. Summary Statistics Table
        ax9 = fig.add_subplot(gs[4, :])
        ax9.axis('off')
        
        summary_text = f"""
        OVERALL STRATEGY PERFORMANCE
        {'='*80}
        
        Total P/L Across All Securities: ${total_pl:,.2f}
        Number of Securities Traded: {len(metrics_df)}
        Total Number of Trades: {total_trades}
        
        Average Win Rate: {avg_win_rate*100:.2f}%
        Average Sharpe Ratio: {avg_sharpe:.3f}
        
        Best Performing Security: {metrics_df.loc[metrics_df['total_pl'].idxmax(), 'symbol']} 
            (${metrics_df['total_pl'].max():,.2f})
        Worst Performing Security: {metrics_df.loc[metrics_df['total_pl'].idxmin(), 'symbol']} 
            (${metrics_df['total_pl'].min():,.2f})
        
        Highest Win Rate: {metrics_df.loc[metrics_df['win_rate'].idxmax(), 'symbol']} 
            ({metrics_df['win_rate'].max()*100:.2f}%)
        Lowest Win Rate: {metrics_df.loc[metrics_df['win_rate'].idxmin(), 'symbol']} 
            ({metrics_df['win_rate'].min()*100:.2f}%)
        """
        
        ax9.text(0.1, 0.5, summary_text, fontfamily='monospace', 
                fontsize=11, verticalalignment='center')
        
        # Save
        plt.tight_layout()
        plt.savefig(self.output_folder / 'aggregate_report.png', dpi=150, bbox_inches='tight')
        plt.close()
        
        # Save detailed metrics
        metrics_df.to_csv(self.output_folder / 'all_securities_metrics.csv', index=False)
        
        # Save summary statistics
        summary_stats = {
            'total_pl': float(total_pl),
            'num_securities': len(metrics_df),
            'total_trades': int(total_trades),
            'avg_win_rate': float(avg_win_rate),
            'avg_sharpe': float(avg_sharpe),
            'best_security': {
                'symbol': metrics_df.loc[metrics_df['total_pl'].idxmax(), 'symbol'],
                'pl': float(metrics_df['total_pl'].max())
            },
            'worst_security': {
                'symbol': metrics_df.loc[metrics_df['total_pl'].idxmin(), 'symbol'],
                'pl': float(metrics_df['total_pl'].min())
            }
        }
        
        with open(self.output_folder / 'summary_statistics.json', 'w') as f:
            json.dump(summary_stats, f, indent=2)
        
        print("Aggregate report completed!")
        print(f"\nResults saved to: {self.output_folder}")
        print(f"Total P/L: ${total_pl:,.2f}")
        print(f"Average Win Rate: {avg_win_rate*100:.2f}%")
    
    def generate_excel_report(self, security_metadata: Dict[str, Dict] = None, 
                            strategy_name: str = "Strategy"):
        """
        Generate comprehensive Excel workbook with all performance metrics
        
        Args:
            security_metadata: Optional metadata for grouping
            strategy_name: Name of the strategy for the report
        """
        if not EXCEL_AVAILABLE:
            print("Excel reporting not available. Install openpyxl: pip install openpyxl")
            return
        
        print("\nGenerating comprehensive Excel report...")
        
        # Collect all metrics and trades
        all_metrics = []
        all_trades_by_symbol = {}
        
        for symbol, (trades, equity_curve) in self.results.items():
            if len(trades) > 0:
                metrics = self.analyze_security_performance(symbol, trades, equity_curve)
                if 'error' not in metrics:
                    if security_metadata and symbol in security_metadata:
                        metrics.update(security_metadata[symbol])
                    all_metrics.append(metrics)
                    all_trades_by_symbol[symbol] = (trades, equity_curve)
        
        if len(all_metrics) == 0:
            print("No trades to report")
            return
        
        metrics_df = pd.DataFrame(all_metrics)
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        subheader_font = Font(bold=True, size=11)
        
        positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        title_font = Font(bold=True, size=14, color="1F4E78")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ===== SUMMARY SHEET =====
        ws_summary = wb.create_sheet("Executive Summary")
        
        # Title
        ws_summary['A1'] = f"{strategy_name} - Backtest Report"
        ws_summary['A1'].font = Font(bold=True, size=16, color="1F4E78")
        ws_summary.merge_cells('A1:F1')
        
        ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_summary['A2'].font = Font(size=10, italic=True)
        ws_summary.merge_cells('A2:F2')
        
        row = 4
        
        # Overall Performance Metrics
        ws_summary[f'A{row}'] = "OVERALL PERFORMANCE"
        ws_summary[f'A{row}'].font = title_font
        ws_summary.merge_cells(f'A{row}:B{row}')
        row += 1
        
        total_pl = metrics_df['total_pl'].sum()
        total_trades = metrics_df['num_trades'].sum()
        avg_win_rate = metrics_df['win_rate'].mean()
        avg_sharpe = metrics_df['sharpe_ratio'].mean()
        
        summary_data = [
            ['Total P/L', f"${total_pl:,.2f}"],
            ['Number of Securities', len(metrics_df)],
            ['Total Trades', int(total_trades)],
            ['Average Win Rate', f"{avg_win_rate*100:.2f}%"],
            ['Average Sharpe Ratio', f"{avg_sharpe:.3f}"],
            ['Best Security', f"{metrics_df.loc[metrics_df['total_pl'].idxmax(), 'symbol']} (${metrics_df['total_pl'].max():,.2f})"],
            ['Worst Security', f"{metrics_df.loc[metrics_df['total_pl'].idxmin(), 'symbol']} (${metrics_df['total_pl'].min():,.2f})"],
        ]
        
        for label, value in summary_data:
            ws_summary[f'A{row}'] = label
            ws_summary[f'A{row}'].font = Font(bold=True)
            ws_summary[f'B{row}'] = value
            
            # Color coding for P/L
            if label == 'Total P/L':
                if total_pl > 0:
                    ws_summary[f'B{row}'].fill = positive_fill
                else:
                    ws_summary[f'B{row}'].fill = negative_fill
            
            row += 1
        
        row += 2
        
        # Top 5 Performers
        ws_summary[f'A{row}'] = "TOP 5 PERFORMERS"
        ws_summary[f'A{row}'].font = title_font
        row += 1
        
        top_5 = metrics_df.nlargest(5, 'total_pl')[['symbol', 'total_pl', 'win_rate', 'num_trades']]
        
        headers = ['Symbol', 'Total P/L', 'Win Rate', 'Trades']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_summary.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        for _, perf in top_5.iterrows():
            ws_summary.cell(row=row, column=1, value=perf['symbol'])
            ws_summary.cell(row=row, column=2, value=f"${perf['total_pl']:,.2f}")
            ws_summary.cell(row=row, column=3, value=f"{perf['win_rate']*100:.1f}%")
            ws_summary.cell(row=row, column=4, value=int(perf['num_trades']))
            
            if perf['total_pl'] > 0:
                ws_summary.cell(row=row, column=2).fill = positive_fill
            
            row += 1
        
        row += 2
        
        # Bottom 5 Performers
        ws_summary[f'A{row}'] = "BOTTOM 5 PERFORMERS"
        ws_summary[f'A{row}'].font = title_font
        row += 1
        
        bottom_5 = metrics_df.nsmallest(5, 'total_pl')[['symbol', 'total_pl', 'win_rate', 'num_trades']]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_summary.cell(row=row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        for _, perf in bottom_5.iterrows():
            ws_summary.cell(row=row, column=1, value=perf['symbol'])
            ws_summary.cell(row=row, column=2, value=f"${perf['total_pl']:,.2f}")
            ws_summary.cell(row=row, column=3, value=f"{perf['win_rate']*100:.1f}%")
            ws_summary.cell(row=row, column=4, value=int(perf['num_trades']))
            
            if perf['total_pl'] < 0:
                ws_summary.cell(row=row, column=2).fill = negative_fill
            
            row += 1
        
        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20
        ws_summary.column_dimensions['C'].width = 15
        ws_summary.column_dimensions['D'].width = 12
        
        # Add chart - Top 10 by P/L
        row += 2
        chart_row = row
        
        # Create data for chart
        top_10 = metrics_df.nlargest(10, 'total_pl')[['symbol', 'total_pl']].sort_values('total_pl')
        
        # Write chart data
        ws_summary[f'F{chart_row}'] = "Symbol"
        ws_summary[f'G{chart_row}'] = "P/L"
        ws_summary[f'F{chart_row}'].font = header_font
        ws_summary[f'G{chart_row}'].font = header_font
        
        chart_start_row = chart_row + 1
        for idx, (_, row_data) in enumerate(top_10.iterrows(), start=chart_start_row):
            ws_summary[f'F{idx}'] = row_data['symbol']
            ws_summary[f'G{idx}'] = row_data['total_pl']
        
        # Create bar chart
        chart = BarChart()
        chart.title = "Top 10 Securities by P/L"
        chart.style = 10
        chart.y_axis.title = 'Profit/Loss ($)'
        chart.x_axis.title = 'Security'
        
        data = Reference(ws_summary, min_col=7, min_row=chart_row, max_row=chart_start_row + len(top_10) - 1)
        cats = Reference(ws_summary, min_col=6, min_row=chart_start_row, max_row=chart_start_row + len(top_10) - 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 20
        
        ws_summary.add_chart(chart, f'A{chart_row + 2}')
        
        # ===== ALL SECURITIES SHEET =====
        ws_all = wb.create_sheet("All Securities")
        
        # Write headers
        metrics_display = metrics_df[[
            'symbol', 'total_pl', 'total_pl_pct', 'num_trades', 'win_rate',
            'profit_factor', 'sharpe_ratio', 'max_drawdown', 'avg_duration_days'
        ]].copy()
        
        metrics_display.columns = [
            'Symbol', 'Total P/L ($)', 'Total P/L (%)', 'Trades', 'Win Rate',
            'Profit Factor', 'Sharpe Ratio', 'Max Drawdown', 'Avg Duration (days)'
        ]
        
        for r_idx, row_data in enumerate(dataframe_to_rows(metrics_display, index=False, header=True), start=1):
            for c_idx, value in enumerate(row_data, start=1):
                cell = ws_all.cell(row=r_idx, column=c_idx, value=value)
                
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                else:
                    # Format numbers
                    if c_idx == 2:  # Total P/L
                        cell.number_format = '$#,##0.00'
                        if isinstance(value, (int, float)) and value > 0:
                            cell.fill = positive_fill
                        elif isinstance(value, (int, float)) and value < 0:
                            cell.fill = negative_fill
                    elif c_idx == 3:  # Total P/L %
                        cell.number_format = '0.00%'
                    elif c_idx == 5:  # Win Rate
                        cell.number_format = '0.00%'
                    elif c_idx in [6, 7]:  # Profit Factor, Sharpe
                        cell.number_format = '0.000'
                    elif c_idx == 8:  # Max Drawdown
                        cell.number_format = '0.00%'
                    elif c_idx == 9:  # Avg Duration
                        cell.number_format = '0.0'
                
                cell.border = border
        
        # Adjust column widths
        for col in ws_all.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws_all.column_dimensions[column].width = adjusted_width
        
        # ===== INDIVIDUAL SECURITY SHEETS =====
        for symbol, (trades, equity_curve) in all_trades_by_symbol.items():
            ws = wb.create_sheet(symbol[:30])  # Excel sheet name limit
            
            # Title
            ws['A1'] = f"{symbol} - Performance Report"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:D1')
            
            row = 3
            
            # Get metrics for this security
            symbol_metrics = metrics_df[metrics_df['symbol'] == symbol].iloc[0]
            
            # Performance Summary
            ws[f'A{row}'] = "PERFORMANCE SUMMARY"
            ws[f'A{row}'].font = title_font
            row += 1
            
            summary_items = [
                ('Total P/L', f"${symbol_metrics['total_pl']:,.2f}"),
                ('Total P/L %', f"{symbol_metrics['total_pl_pct']:.2f}%"),
                ('Number of Trades', int(symbol_metrics['num_trades'])),
                ('Win Rate', f"{symbol_metrics['win_rate']*100:.2f}%"),
                ('Profit Factor', f"{symbol_metrics['profit_factor']:.2f}"),
                ('Sharpe Ratio', f"{symbol_metrics['sharpe_ratio']:.3f}"),
                ('Max Drawdown', f"{symbol_metrics['max_drawdown']*100:.2f}%"),
                ('Avg Trade Duration', f"{symbol_metrics['avg_duration_days']:.1f} days"),
                ('Best Trade', f"${symbol_metrics['best_trade_pl']:,.2f}"),
                ('Worst Trade', f"${symbol_metrics['worst_trade_pl']:,.2f}"),
            ]
            
            for label, value in summary_items:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value
                
                if label == 'Total P/L' and symbol_metrics['total_pl'] > 0:
                    ws[f'B{row}'].fill = positive_fill
                elif label == 'Total P/L' and symbol_metrics['total_pl'] < 0:
                    ws[f'B{row}'].fill = negative_fill
                
                row += 1
            
            row += 2
            
            # Trade Log
            ws[f'A{row}'] = "TRADE LOG"
            ws[f'A{row}'].font = title_font
            row += 1
            
            trades_data = []
            for trade in trades:
                trades_data.append({
                    'Entry Date': trade.entry_date.strftime('%Y-%m-%d'),
                    'Exit Date': trade.exit_date.strftime('%Y-%m-%d') if trade.exit_date else '',
                    'Entry Price': trade.entry_price,
                    'Exit Price': trade.exit_price if trade.exit_price else 0,
                    'Quantity': trade.quantity,
                    'P/L ($)': trade.pl,
                    'P/L (%)': trade.pl_pct,
                    'Duration (days)': trade.duration_days,
                    'Exit Reason': trade.exit_reason
                })
            
            trades_df = pd.DataFrame(trades_data)
            
            # Write trade log
            for r_idx, row_data in enumerate(dataframe_to_rows(trades_df, index=False, header=True), start=row):
                for c_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    if r_idx == row:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center')
                    else:
                        # Format numbers
                        if c_idx in [3, 4]:  # Prices
                            cell.number_format = '$0.00'
                        elif c_idx == 6:  # P/L $
                            cell.number_format = '$#,##0.00'
                            if isinstance(value, (int, float)) and value > 0:
                                cell.fill = positive_fill
                            elif isinstance(value, (int, float)) and value < 0:
                                cell.fill = negative_fill
                        elif c_idx == 7:  # P/L %
                            cell.number_format = '0.00'
                    
                    cell.border = border
            
            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)
                ws.column_dimensions[column].width = adjusted_width
            
            # Add equity curve chart
            chart_row = row + len(trades_df) + 3
            
            ws[f'A{chart_row}'] = "Date"
            ws[f'B{chart_row}'] = "Equity"
            ws[f'A{chart_row}'].font = header_font
            ws[f'B{chart_row}'].font = header_font
            
            equity_start_row = chart_row + 1
            for idx, (_, eq_row) in enumerate(equity_curve.iterrows(), start=equity_start_row):
                ws[f'A{idx}'] = eq_row['date'].strftime('%Y-%m-%d')
                ws[f'B{idx}'] = eq_row['equity']
            
            # Create line chart for equity curve
            chart = LineChart()
            chart.title = f"{symbol} - Equity Curve"
            chart.style = 12
            chart.y_axis.title = 'Equity ($)'
            chart.x_axis.title = 'Date'
            
            data = Reference(ws, min_col=2, min_row=chart_row, max_row=equity_start_row + len(equity_curve) - 1)
            dates = Reference(ws, min_col=1, min_row=equity_start_row, max_row=equity_start_row + len(equity_curve) - 1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(dates)
            chart.height = 12
            chart.width = 20
            
            ws.add_chart(chart, f'E{row}')
        
        # ===== PERFORMANCE BY TYPE (if metadata available) =====
        if security_metadata and 'metadata_type' in metrics_df.columns:
            ws_type = wb.create_sheet("By Security Type")
            
            ws_type['A1'] = "Performance by Security Type"
            ws_type['A1'].font = Font(bold=True, size=14)
            ws_type.merge_cells('A1:E1')
            
            type_performance = metrics_df.groupby('metadata_type').agg({
                'total_pl': 'sum',
                'num_trades': 'sum',
                'win_rate': 'mean',
                'sharpe_ratio': 'mean'
            }).reset_index()
            
            type_performance.columns = ['Type', 'Total P/L', 'Total Trades', 'Avg Win Rate', 'Avg Sharpe']
            
            row = 3
            for r_idx, row_data in enumerate(dataframe_to_rows(type_performance, index=False, header=True), start=row):
                for c_idx, value in enumerate(row_data, start=1):
                    cell = ws_type.cell(row=r_idx, column=c_idx, value=value)
                    
                    if r_idx == row:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center')
                    else:
                        if c_idx == 2:  # P/L
                            cell.number_format = '$#,##0.00'
                            if isinstance(value, (int, float)) and value > 0:
                                cell.fill = positive_fill
                            elif isinstance(value, (int, float)) and value < 0:
                                cell.fill = negative_fill
                        elif c_idx == 4:  # Win Rate
                            cell.number_format = '0.00%'
                        elif c_idx == 5:  # Sharpe
                            cell.number_format = '0.000'
                    
                    cell.border = border
            
            # Adjust widths
            for col in ws_type.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)
                ws_type.column_dimensions[column].width = adjusted_width
            
            # Add pie chart
            chart = PieChart()
            chart.title = "P/L Distribution by Type"
            
            # Write data for pie chart
            chart_row = row + len(type_performance) + 3
            ws_type[f'A{chart_row}'] = "Type"
            ws_type[f'B{chart_row}'] = "P/L"
            
            for idx, (_, type_row) in enumerate(type_performance.iterrows(), start=chart_row+1):
                ws_type[f'A{idx}'] = type_row['Type']
                ws_type[f'B{idx}'] = max(type_row['Total P/L'], 0)  # Only positive values for pie
            
            data = Reference(ws_type, min_col=2, min_row=chart_row, max_row=chart_row + len(type_performance))
            labels = Reference(ws_type, min_col=1, min_row=chart_row+1, max_row=chart_row + len(type_performance))
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            
            ws_type.add_chart(chart, f'D{row}')
        
        # Save workbook
        excel_file = self.output_folder / f'{strategy_name}_report.xlsx'
        wb.save(excel_file)
        
        print(f"\n{'='*60}")
        print(f"Excel report generated successfully!")
        print(f"File: {excel_file}")
        print(f"Sheets: {', '.join(wb.sheetnames)}")
        print(f"{'='*60}")


# ============================================================================
# EXAMPLE USAGE
# ============================================================================

def load_config(config_path: str = 'config/config.json') -> Dict:
    """Load configuration from JSON file"""
    with open(config_path, 'r') as f:
        return json.load(f)


def load_security_metadata(metadata_path: str = 'config/security_metadata.json') -> Dict:
    """Load security metadata from JSON file"""
    with open(metadata_path, 'r') as f:
        return json.load(f)


if __name__ == "__main__":
    
    # 1. Load configuration
    config = load_config('config/config.json')
    security_metadata = load_security_metadata('config/security_metadata.json')
    
    # 2. Define strategy
    '''strategy = MovingAverageCrossStrategy(params={
        'ma_period': 50,
        'position_size_pct': 0.1,
        'stop_loss_pct': 0.05,
        'take_profit_pct': 0.15
    })'''
    
    strategy = TradingViewMultiPresetStrategy()
    
    # 3. Prepare data (run once per strategy)
    data_prep = DataPreparation(
        raw_data_folder=config['data_settings']['raw_data_folder'],
        prepared_data_folder=config['data_settings']['prepared_data_folder'],
    )
    
    prepared_files = data_prep.prepare_data_for_strategy(strategy, security_metadata, config['column_mapping']['MovingAverageCrossStrategy'])
    
    # 4. Run backtest
    engine = BacktestEngine(
        strategy=strategy,
        starting_capital=config['backtest_settings']['starting_capital'],
        commission_pct=config['backtest_settings']['commission_pct'],
        slippage_pct=config['backtest_settings']['slippage_pct'],
        start_date=config['backtest_settings']['default_start_date'],
        end_date=config['backtest_settings']['default_end_date']
    )
    
    prepared_data_folder = f"{config['data_settings']['prepared_data_folder']}/{strategy.name}"
    results = engine.backtest_all_securities(prepared_data_folder)
    
    # 5. Generate reports
    analyzer = PerformanceAnalyzer()
    
    # Individual reports
    for symbol, (trades, equity_curve) in results.items():
        analyzer.generate_security_report(
            symbol, trades, equity_curve, 
            security_metadata.get(symbol)
        )
    
    # Aggregate report
    analyzer.generate_aggregate_report(security_metadata)
    analyzer.generate_excel_report(security_metadata)
    
    print("\nBacktesting complete!")
    print(f"Reports saved to: {config['data_settings']['reports_folder']}/{strategy.name}")
    