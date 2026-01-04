"""
Excel Reporting Module for Per-Trade Analysis

Generates comprehensive multi-sheet Excel reports with formatting,
conditional styling, and detailed trade forensics.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Optional, Any
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    NamedStyle
)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

from .data_aggregation import TradeAnalysisData, MAEMFEResult, MarketRegime
from .pattern_analysis import SignalStrengthScore, PatternFlag, PatternSummary


# =============================================================================
# STYLE DEFINITIONS
# =============================================================================

# Colors
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
POSITIVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NEGATIVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
NEUTRAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ENTRY_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
EXIT_FILL = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

# Fonts
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=12)
NORMAL_FONT = Font(size=11)
POSITIVE_FONT = Font(color="006100")
NEGATIVE_FONT = Font(color="9C0006")

# Borders
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Alignment
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')
WRAP_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)


# =============================================================================
# EXCEL REPORT GENERATOR
# =============================================================================

class PerTradeExcelReport:
    """Generate comprehensive Excel reports for per-trade analysis."""

    def __init__(self, output_dir: Optional[Path] = None):
        """
        Initialize report generator.

        Args:
            output_dir: Directory for output files. Defaults to reports/per_trade_analysis/
        """
        if output_dir is None:
            self.output_dir = Path(__file__).parent.parent.parent / 'reports' / 'per_trade_analysis'
        else:
            self.output_dir = Path(output_dir)

        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_report(self,
                        trade_data: TradeAnalysisData,
                        signal_score: Optional[SignalStrengthScore] = None,
                        pattern_flags: Optional[List[PatternFlag]] = None,
                        pattern_summary: Optional[PatternSummary] = None) -> Path:
        """
        Generate complete Excel report for a single trade.

        Args:
            trade_data: TradeAnalysisData with all aggregated data
            signal_score: Signal strength score
            pattern_flags: List of pattern flags
            pattern_summary: Aggregate pattern summary (if multiple trades)

        Returns:
            Path to generated Excel file
        """
        wb = Workbook()

        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Create sheets
        self._create_summary_sheet(wb, trade_data, signal_score, pattern_flags)
        self._create_price_action_sheet(wb, trade_data)
        self._create_weekly_prices_sheet(wb, trade_data)
        self._create_fundamentals_sheet(wb, trade_data)
        self._create_insider_sheet(wb, trade_data)
        self._create_options_sheet(wb, trade_data)
        self._create_correlation_sheet(wb, trade_data)
        self._create_mae_mfe_sheet(wb, trade_data)
        self._create_signal_strength_sheet(wb, trade_data, signal_score)

        if pattern_summary:
            self._create_pattern_summary_sheet(wb, pattern_summary)

        # Generate filename
        symbol = trade_data.symbol
        trade_id = trade_data.trade_id
        export_date = datetime.now().strftime("%Y%m%d")
        filename = f"{symbol}_{trade_id}_analysis_{export_date}.xlsx"
        file_path = self.output_dir / filename

        wb.save(file_path)
        return file_path

    def generate_multi_trade_report(self,
                                     trades_data: List[TradeAnalysisData],
                                     signal_scores: Dict[str, SignalStrengthScore],
                                     pattern_flags: Dict[str, List[PatternFlag]],
                                     pattern_summary: PatternSummary) -> Path:
        """
        Generate report comparing multiple trades.

        Args:
            trades_data: List of TradeAnalysisData
            signal_scores: Dict of trade_id -> SignalStrengthScore
            pattern_flags: Dict of trade_id -> list of PatternFlags
            pattern_summary: Aggregate pattern summary

        Returns:
            Path to generated Excel file
        """
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Create overview sheet
        self._create_multi_trade_overview(wb, trades_data, signal_scores)

        # Create individual trade summary sheets
        for trade in trades_data[:10]:  # Limit to 10 trades per report
            sheet_name = f"{trade.symbol}_{trade.trade_id[-4:]}"[:31]  # Excel sheet name limit
            self._create_trade_mini_summary(wb, trade, signal_scores.get(trade.trade_id), sheet_name)

        # Create pattern summary sheet
        self._create_pattern_summary_sheet(wb, pattern_summary)

        # Generate filename
        export_date = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"multi_trade_analysis_{export_date}.xlsx"
        file_path = self.output_dir / filename

        wb.save(file_path)
        return file_path

    # =========================================================================
    # SHEET CREATION METHODS
    # =========================================================================

    def _create_summary_sheet(self,
                              wb: Workbook,
                              trade: TradeAnalysisData,
                              signal_score: Optional[SignalStrengthScore],
                              pattern_flags: Optional[List[PatternFlag]]):
        """Create trade summary sheet."""
        ws = wb.create_sheet("Trade Summary")

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = "TRADE ANALYSIS REPORT"
        ws['A1'].font = TITLE_FONT
        ws['A1'].alignment = CENTER_ALIGN

        # Basic info header
        ws.merge_cells('A3:F3')
        ws['A3'] = f"Symbol: {trade.symbol} | Trade ID: {trade.trade_id}"
        ws['A3'].font = SECTION_FONT

        info = trade.trade_info
        exit_date = info.get('exit_date', '')
        if isinstance(exit_date, datetime):
            exit_date = exit_date.strftime('%Y-%m-%d')
        duration = info.get('duration_days', 'N/A')
        side = info.get('side', 'LONG')

        ws['A4'] = f"Exit Date: {exit_date} | Duration: {duration} days | Side: {side}"

        # Position Metrics section
        row = 6
        ws[f'A{row}'] = "POSITION METRICS"
        ws[f'A{row}'].font = SECTION_FONT
        ws.merge_cells(f'A{row}:C{row}')

        ws[f'D{row}'] = "TRADE OUTCOME"
        ws[f'D{row}'].font = SECTION_FONT
        ws.merge_cells(f'D{row}:F{row}')

        # Position data
        row += 1
        metrics = [
            ('Entry Price', f"${info.get('entry_price', 0):.4f}"),
            ('Exit Price', f"${info.get('exit_price', 0):.4f}"),
            ('Quantity', f"{info.get('quantity', 0):,.2f}"),
            ('Entry Equity', f"${info.get('entry_equity', 0):,.2f}"),
        ]

        for i, (label, value) in enumerate(metrics):
            ws[f'A{row + i}'] = label
            ws[f'B{row + i}'] = value

        # Trade outcome data
        pl = info.get('pl', 0)
        pl_pct = info.get('pl_pct', 0)
        pl_display = f"${pl:,.2f} ({pl_pct:+.2f}%)"

        outcomes = [
            ('P&L', pl_display),
            ('Entry Signal Strength', f"{signal_score.total_score:.0f}/100" if signal_score else "N/A"),
            ('Exit Reason', str(info.get('exit_reason', 'N/A'))[:40]),
        ]

        for i, (label, value) in enumerate(outcomes):
            ws[f'D{row + i}'] = label
            ws[f'E{row + i}'] = value
            if 'P&L' in label:
                ws[f'E{row + i}'].font = POSITIVE_FONT if pl >= 0 else NEGATIVE_FONT

        # Risk Metrics section
        row += 6
        ws[f'A{row}'] = "RISK METRICS"
        ws[f'A{row}'].font = SECTION_FONT
        ws.merge_cells(f'A{row}:C{row}')

        ws[f'D{row}'] = "TECHNICAL CONTEXT"
        ws[f'D{row}'].font = SECTION_FONT
        ws.merge_cells(f'D{row}:F{row}')

        row += 1
        risk_metrics = [
            ('Initial Stop', f"${info.get('initial_stop_loss', 0):.4f}"),
            ('Final Stop', f"${info.get('final_stop_loss', 0):.4f}"),
            ('Take Profit', f"${info.get('take_profit', 0):.4f}" if info.get('take_profit') else "N/A"),
        ]

        if trade.mae_mfe:
            risk_metrics.append(('Max Adverse Excursion', f"{trade.mae_mfe.mae_pct:.2f}%"))

        for i, (label, value) in enumerate(risk_metrics):
            ws[f'A{row + i}'] = label
            ws[f'B{row + i}'] = value

        # Technical context
        tech_context = []
        if trade.mae_mfe:
            tech_context.append(('MAE Date', str(trade.mae_mfe.mae_date.strftime('%Y-%m-%d') if trade.mae_mfe.mae_date else 'N/A')))
            tech_context.append(('MFE', f"{trade.mae_mfe.mfe_pct:.2f}%"))

        if trade.market_regime:
            tech_context.append(('Market Regime', trade.market_regime.trend.title()))
            tech_context.append(('Volatility', trade.market_regime.volatility.title()))

        for i, (label, value) in enumerate(tech_context):
            ws[f'D{row + i}'] = label
            ws[f'E{row + i}'] = value

        # Key Insights section
        row += 6
        ws[f'A{row}'] = "KEY INSIGHTS"
        ws[f'A{row}'].font = SECTION_FONT
        ws.merge_cells(f'A{row}:F{row}')

        row += 1
        ws[f'A{row}'] = f"Entry: {info.get('entry_reason', 'N/A')}"
        ws[f'A{row}'].alignment = WRAP_ALIGN
        ws.merge_cells(f'A{row}:F{row}')

        row += 1
        ws[f'A{row}'] = f"Exit: {info.get('exit_reason', 'N/A')}"
        ws[f'A{row}'].alignment = WRAP_ALIGN
        ws.merge_cells(f'A{row}:F{row}')

        # Pattern Flags section
        if pattern_flags:
            row += 2
            ws[f'A{row}'] = "PATTERN FLAGS"
            ws[f'A{row}'].font = SECTION_FONT
            ws.merge_cells(f'A{row}:F{row}')

            for flag in pattern_flags[:5]:  # Limit to 5 flags
                row += 1
                ws[f'A{row}'] = f"  {flag.flag_id}: {flag.description}"
                ws[f'A{row}'].alignment = WRAP_ALIGN
                ws.merge_cells(f'A{row}:F{row}')

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 15

    def _create_price_action_sheet(self, wb: Workbook, trade: TradeAnalysisData):
        """Create daily price action sheet."""
        ws = wb.create_sheet("Price Action")

        if trade.price_data is None or len(trade.price_data) == 0:
            ws['A1'] = "No price data available"
            return

        # Add header
        ws['A1'] = f"Daily Price Data: {trade.symbol}"
        ws['A1'].font = TITLE_FONT
        ws.merge_cells('A1:J1')

        # Prepare data
        df = trade.price_data.copy()

        # Select key columns
        core_cols = ['date', 'open', 'high', 'low', 'close', 'volume']
        indicator_cols = [c for c in df.columns if any(x in c.lower() for x in ['atr', 'rsi', 'sma', 'macd', 'bb'])]

        available_cols = [c for c in core_cols + indicator_cols[:8] if c in df.columns]
        df = df[available_cols]

        # Convert dates to strings for Excel
        if 'date' in df.columns:
            df['date'] = pd.to_datetime(df['date']).dt.strftime('%Y-%m-%d')

        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=3, column=col_idx, value=col_name.upper())
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        # Write data
        entry_date = pd.to_datetime(trade.trade_info.get('entry_date')).strftime('%Y-%m-%d')
        exit_date = pd.to_datetime(trade.trade_info.get('exit_date')).strftime('%Y-%m-%d')

        for row_idx, row_data in enumerate(df.values, 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                if pd.isna(value):
                    cell.value = ""
                elif isinstance(value, float):
                    cell.value = round(value, 4)
                    cell.number_format = '#,##0.0000' if abs(value) < 100 else '#,##0.00'
                else:
                    cell.value = value

                cell.border = THIN_BORDER

                # Highlight entry/exit rows
                if col_idx == 1:  # Date column
                    if str(value) == entry_date:
                        for c in range(1, len(df.columns) + 1):
                            ws.cell(row=row_idx, column=c).fill = ENTRY_FILL
                    elif str(value) == exit_date:
                        for c in range(1, len(df.columns) + 1):
                            ws.cell(row=row_idx, column=c).fill = EXIT_FILL

        # Adjust column widths
        for col_idx in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14

        # Add legend
        legend_row = len(df) + 6
        ws[f'A{legend_row}'] = "Legend:"
        ws[f'A{legend_row}'].font = SECTION_FONT
        ws[f'B{legend_row}'] = "Entry Day"
        ws[f'B{legend_row}'].fill = ENTRY_FILL
        ws[f'C{legend_row}'] = "Exit Day"
        ws[f'C{legend_row}'].fill = EXIT_FILL

    def _create_weekly_prices_sheet(self, wb: Workbook, trade: TradeAnalysisData):
        """Create weekly price data sheet."""
        ws = wb.create_sheet("Weekly Prices")

        if trade.weekly_price_data is None or len(trade.weekly_price_data) == 0:
            ws['A1'] = "No weekly price data available"
            return

        ws['A1'] = f"Weekly Price Data: {trade.symbol}"
        ws['A1'].font = TITLE_FONT
        ws.merge_cells('A1:J1')

        df = trade.weekly_price_data.copy()
        core_cols = ['date', 'open', 'high', 'low', 'close', 'volume']
        available_cols = [c for c in core_cols if c in df.columns]
        df = df[available_cols]

        if 'date' in df.columns:
            df['date'] = pd.to_datetime(df['date']).dt.strftime('%Y-%m-%d')

        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=3, column=col_idx, value=col_name.upper())
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        # Write data
        for row_idx, row_data in enumerate(df.values, 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if pd.isna(value):
                    cell.value = ""
                elif isinstance(value, float):
                    cell.value = round(value, 4)
                else:
                    cell.value = value
                cell.border = THIN_BORDER

        for col_idx in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14

    def _create_fundamentals_sheet(self, wb: Workbook, trade: TradeAnalysisData):
        """Create fundamentals comparison sheet."""
        ws = wb.create_sheet("Fundamentals")

        ws['A1'] = f"Fundamental Analysis: {trade.symbol}"
        ws['A1'].font = TITLE_FONT
        ws.merge_cells('A1:F1')

        if not trade.fundamentals_entry and not trade.fundamentals_exit:
            ws['A3'] = "No fundamental data available for this symbol"
            ws['A3'].fill = WARNING_FILL
            return

        # Headers
        headers = ['Metric', 'Entry Value', 'Exit Value', 'Change', 'Status']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER

        # Metrics to display
        metrics = [
            ('pe_ratio', 'P/E Ratio'),
            ('eps', 'EPS'),
            ('earnings_growth_yoy', 'Earnings Growth YoY %'),
            ('revenue_growth_yoy', 'Revenue Growth YoY %'),
            ('profit_margin', 'Profit Margin'),
            ('return_on_assets', 'Return on Assets'),
            ('return_on_equity', 'Return on Equity'),
        ]

        entry_data = trade.fundamentals_entry or {}
        exit_data = trade.fundamentals_exit or {}
        delta_data = trade.fundamentals_delta or {}

        row = 4
        for metric_key, metric_name in metrics:
            entry_val = entry_data.get(metric_key)
            exit_val = exit_data.get(metric_key)
            delta = delta_data.get(metric_key)

            ws.cell(row=row, column=1, value=metric_name).border = THIN_BORDER

            # Entry value
            cell_entry = ws.cell(row=row, column=2)
            cell_entry.value = f"{entry_val:.2f}" if pd.notna(entry_val) else "N/A"
            cell_entry.border = THIN_BORDER

            # Exit value
            cell_exit = ws.cell(row=row, column=3)
            cell_exit.value = f"{exit_val:.2f}" if pd.notna(exit_val) else "N/A"
            cell_exit.border = THIN_BORDER

            # Change
            cell_change = ws.cell(row=row, column=4)
            if pd.notna(delta):
                cell_change.value = f"{delta:+.2f}"
                if delta > 0:
                    cell_change.fill = POSITIVE_FILL
                    cell_change.font = POSITIVE_FONT
                elif delta < 0:
                    cell_change.fill = NEGATIVE_FILL
                    cell_change.font = NEGATIVE_FONT
            else:
                cell_change.value = "N/A"
                cell_change.fill = NEUTRAL_FILL
            cell_change.border = THIN_BORDER

            # Status
            cell_status = ws.cell(row=row, column=5)
            if pd.notna(delta):
                cell_status.value = "Improved" if delta > 0 else ("Declined" if delta < 0 else "Stable")
            else:
                cell_status.value = "Unknown"
            cell_status.border = THIN_BORDER

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 15

    def _create_insider_sheet(self, wb: Workbook, trade: TradeAnalysisData):
        """Create insider activity sheet."""
        ws = wb.create_sheet("Insider Activity")

        ws['A1'] = f"Insider Activity: {trade.symbol}"
        ws['A1'].font = TITLE_FONT
        ws.merge_cells('A1:H1')

        if trade.insider_activity is None or len(trade.insider_activity) == 0:
            ws['A3'] = "No insider activity data available"
            ws['A3'].fill = WARNING_FILL
            return

        # Summary section
        ws['A3'] = "INSIDER FLAGS"
        ws['A3'].font = SECTION_FONT

        row = 4
        if trade.insider_flags:
            for flag in trade.insider_flags:
                ws[f'A{row}'] = f"  {flag}"
                ws[f'A{row}'].alignment = WRAP_ALIGN
                ws.merge_cells(f'A{row}:H{row}')
                row += 1
        else:
            ws[f'A{row}'] = "  No significant insider patterns detected"
            row += 1

        # Transaction table
        row += 2
        ws[f'A{row}'] = "TRANSACTION HISTORY"
        ws[f'A{row}'].font = SECTION_FONT

        row += 1
        headers = ['Date', 'Executive', 'Title', 'Type', 'Shares', 'Price', 'Value', 'Security Type']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER

        # Write transactions
        df = trade.insider_activity
        col_mapping = {
            'date': 0, 'executive': 1, 'insider_title': 2, 'transaction_type': 3,
            'shares': 4, 'price': 5, 'value': 6, 'security_type': 7
        }

        row += 1
        for _, tx_row in df.iterrows():
            for col_name, col_idx in col_mapping.items():
                cell = ws.cell(row=row, column=col_idx + 1)

                if col_name in tx_row.index:
                    val = tx_row[col_name]
                    if col_name == 'date' and not isinstance(val, str):
                        val = pd.to_datetime(val).strftime('%Y-%m-%d') if pd.notna(val) else ''
                    elif col_name in ['shares', 'value']:
                        val = f"{val:,.0f}" if pd.notna(val) else ''
                    elif col_name == 'price':
                        val = f"${val:.2f}" if pd.notna(val) else ''
                    else:
                        val = str(val) if pd.notna(val) else ''
                    cell.value = val
                else:
                    cell.value = ''

                cell.border = THIN_BORDER

                # Color BUY/SELL rows
                if col_name == 'transaction_type':
                    if str(val).upper() == 'BUY':
                        for c in range(1, 9):
                            ws.cell(row=row, column=c).fill = POSITIVE_FILL
                    elif str(val).upper() == 'SELL':
                        for c in range(1, 9):
                            ws.cell(row=row, column=c).fill = NEGATIVE_FILL

            row += 1

        # Adjust widths
        widths = [12, 25, 20, 8, 12, 10, 15, 15]
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    def _create_options_sheet(self, wb: Workbook, trade: TradeAnalysisData):
        """Create options market sheet."""
        ws = wb.create_sheet("Options Market")

        ws['A1'] = f"Options Analysis: {trade.symbol}"
        ws['A1'].font = TITLE_FONT
        ws.merge_cells('A1:D1')

        if not trade.options_data:
            ws['A3'] = "No options data available for this symbol"
            ws['A3'].fill = WARNING_FILL
            return

        opts = trade.options_data

        # IV Analysis section
        ws['A3'] = "IMPLIED VOLATILITY ANALYSIS"
        ws['A3'].font = SECTION_FONT

        row = 4
        data_rows = [
            ('IV at Entry', f"{opts.get('iv_at_entry', 0):.1f}%" if opts.get('iv_at_entry') else "N/A"),
            ('IV at Exit', f"{opts.get('iv_at_exit', 0):.1f}%" if opts.get('iv_at_exit') else "N/A"),
            ('IV Percentile (Entry)', f"{opts.get('iv_percentile_entry', 0):.0f}th" if opts.get('iv_percentile_entry') else "N/A"),
            ('Put/Call Ratio (Entry)', f"{opts.get('put_call_ratio_entry', 0):.2f}" if opts.get('put_call_ratio_entry') else "N/A"),
            ('Earnings During Trade', "Yes" if opts.get('earnings_during_trade') else "No"),
        ]

        for label, value in data_rows:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        # Interpretation
        row += 1
        ws[f'A{row}'] = "INTERPRETATION"
        ws[f'A{row}'].font = SECTION_FONT

        row += 1
        iv_pct = opts.get('iv_percentile_entry')
        if iv_pct:
            if iv_pct >= 75:
                ws[f'A{row}'] = "High IV at entry: Expected volatility decline may have helped/hurt depending on direction"
            elif iv_pct <= 25:
                ws[f'A{row}'] = "Low IV at entry: Quiet market priced in; moves may have been muted"
            else:
                ws[f'A{row}'] = "Normal IV at entry: Standard volatility expectations"
        else:
            ws[f'A{row}'] = "Unable to assess IV environment due to missing data"

        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].alignment = WRAP_ALIGN

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _create_correlation_sheet(self, wb: Workbook, trade: TradeAnalysisData):
        """Create sector/market correlation sheet."""
        ws = wb.create_sheet("Correlation")

        ws['A1'] = f"Sector & Market Correlation: {trade.symbol}"
        ws['A1'].font = TITLE_FONT
        ws.merge_cells('A1:D1')

        # Correlation summary
        ws['A3'] = "CORRELATION SUMMARY"
        ws['A3'].font = SECTION_FONT

        row = 4
        sector_corr = trade.sector_correlation
        index_corr = trade.index_correlation

        ws[f'A{row}'] = "Sector Correlation"
        ws[f'B{row}'] = f"{sector_corr:.3f}" if sector_corr else "N/A"
        row += 1

        ws[f'A{row}'] = "Index (SPX) Correlation"
        ws[f'B{row}'] = f"{index_corr:.3f}" if index_corr else "N/A"
        row += 2

        # Interpretation
        ws[f'A{row}'] = "ANALYSIS"
        ws[f'A{row}'].font = SECTION_FONT
        row += 1

        if sector_corr:
            if abs(sector_corr) > 0.85:
                interpretation = "Strong correlation - consider trading sector ETF for similar exposure with less single-stock risk"
            elif abs(sector_corr) > 0.70:
                interpretation = "Moderate correlation - stock generally moves with sector"
            elif abs(sector_corr) > 0.30:
                interpretation = "Weak correlation - company-specific factors dominate"
            else:
                interpretation = "Very weak correlation - idiosyncratic movement"
        else:
            interpretation = "Correlation data not available"

        ws[f'A{row}'] = interpretation
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].alignment = WRAP_ALIGN

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

    def _create_mae_mfe_sheet(self, wb: Workbook, trade: TradeAnalysisData):
        """Create MAE/MFE analysis sheet."""
        ws = wb.create_sheet("MAE-MFE Analysis")

        ws['A1'] = f"Maximum Adverse/Favorable Excursion: {trade.symbol}"
        ws['A1'].font = TITLE_FONT
        ws.merge_cells('A1:D1')

        if not trade.mae_mfe:
            ws['A3'] = "MAE/MFE data not available (insufficient price data)"
            ws['A3'].fill = WARNING_FILL
            return

        mf = trade.mae_mfe

        # MAE Section
        ws['A3'] = "MAXIMUM ADVERSE EXCURSION (MAE)"
        ws['A3'].font = SECTION_FONT

        row = 4
        mae_data = [
            ('Date Occurred', mf.mae_date.strftime('%Y-%m-%d') if mf.mae_date else 'N/A'),
            ('Worst Price', f"${mf.mae_price:.4f}"),
            ('MAE %', f"{mf.mae_pct:.2f}%"),
            ('Days into Trade', str(mf.mae_days_into_trade)),
        ]

        for label, value in mae_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].fill = NEGATIVE_FILL if 'MAE' in label else None
            row += 1

        # MFE Section
        row += 1
        ws[f'A{row}'] = "MAXIMUM FAVORABLE EXCURSION (MFE)"
        ws[f'A{row}'].font = SECTION_FONT

        row += 1
        mfe_data = [
            ('Date Occurred', mf.mfe_date.strftime('%Y-%m-%d') if mf.mfe_date else 'N/A'),
            ('Best Price', f"${mf.mfe_price:.4f}"),
            ('MFE %', f"{mf.mfe_pct:.2f}%"),
            ('Days into Trade', str(mf.mfe_days_into_trade)),
        ]

        for label, value in mfe_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].fill = POSITIVE_FILL if 'MFE' in label else None
            row += 1

        # Actual outcome
        row += 1
        ws[f'A{row}'] = "ACTUAL OUTCOME"
        ws[f'A{row}'].font = SECTION_FONT

        row += 1
        ws[f'A{row}'] = "Actual P&L %"
        ws[f'B{row}'] = f"{mf.actual_pl_pct:.2f}%"
        ws[f'B{row}'].fill = POSITIVE_FILL if mf.actual_pl_pct >= 0 else NEGATIVE_FILL

        row += 1
        ws[f'A{row}'] = "MFE Capture %"
        ws[f'B{row}'] = f"{mf.mfe_capture_pct:.1f}%"

        # Analysis
        row += 2
        ws[f'A{row}'] = "INTERPRETATION"
        ws[f'A{row}'].font = SECTION_FONT

        row += 1
        ws[f'A{row}'] = f"You weathered a {mf.mae_pct:.2f}% drawdown before exiting."
        ws.merge_cells(f'A{row}:D{row}')

        row += 1
        ws[f'A{row}'] = f"You captured {mf.mfe_capture_pct:.1f}% of the maximum opportunity ({mf.mfe_pct:.2f}%)."
        ws.merge_cells(f'A{row}:D{row}')

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18

    def _create_signal_strength_sheet(self,
                                       wb: Workbook,
                                       trade: TradeAnalysisData,
                                       signal_score: Optional[SignalStrengthScore]):
        """Create signal strength breakdown sheet."""
        ws = wb.create_sheet("Signal Strength")

        ws['A1'] = f"Entry Signal Strength: {trade.symbol}"
        ws['A1'].font = TITLE_FONT
        ws.merge_cells('A1:D1')

        if not signal_score:
            ws['A3'] = "Signal strength score not calculated"
            ws['A3'].fill = WARNING_FILL
            return

        # Overall score
        ws['A3'] = "OVERALL SCORE"
        ws['A3'].font = SECTION_FONT

        ws['A4'] = f"{signal_score.total_score:.0f} / {signal_score.max_possible:.0f}"
        ws['A4'].font = Font(size=24, bold=True)
        ws['A5'] = signal_score.description
        ws['A5'].font = Font(italic=True)

        # Tier indicator
        tier_colors = {
            'excellent': POSITIVE_FILL,
            'good': PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"),
            'moderate': WARNING_FILL,
            'weak': NEGATIVE_FILL
        }
        ws['A4'].fill = tier_colors.get(signal_score.tier, NEUTRAL_FILL)

        # Breakdown
        row = 7
        ws[f'A{row}'] = "SCORE BREAKDOWN"
        ws[f'A{row}'].font = SECTION_FONT

        row += 1
        headers = ['Factor', 'Points', 'Max', 'Met', 'Detail']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER

        row += 1
        for factor, details in signal_score.breakdown.items():
            ws.cell(row=row, column=1, value=factor.replace('_', ' ').title()).border = THIN_BORDER
            ws.cell(row=row, column=2, value=details['points']).border = THIN_BORDER
            ws.cell(row=row, column=3, value=details['max']).border = THIN_BORDER

            met_cell = ws.cell(row=row, column=4)
            met_cell.value = "" if details['met'] else ""
            met_cell.fill = POSITIVE_FILL if details['met'] else NEGATIVE_FILL
            met_cell.border = THIN_BORDER

            detail_cell = ws.cell(row=row, column=5, value=details.get('detail', '')[:50])
            detail_cell.border = THIN_BORDER

            row += 1

        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 50

    def _create_pattern_summary_sheet(self, wb: Workbook, summary: PatternSummary):
        """Create aggregate pattern summary sheet."""
        ws = wb.create_sheet("Pattern Summary")

        ws['A1'] = "AGGREGATE PATTERN ANALYSIS"
        ws['A1'].font = TITLE_FONT
        ws.merge_cells('A1:E1')

        # Overview
        ws['A3'] = f"Total Trades: {summary.total_trades} | Winners: {summary.winners} | Losers: {summary.losers} | Win Rate: {summary.win_rate:.1f}%"
        ws['A3'].font = SECTION_FONT
        ws.merge_cells('A3:E3')

        # Win Rate by Signal Strength
        row = 5
        ws[f'A{row}'] = "WIN RATE BY SIGNAL STRENGTH"
        ws[f'A{row}'].font = SECTION_FONT

        row += 1
        headers = ['Tier', 'Wins', 'Total', 'Win Rate']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER

        row += 1
        for tier, (wins, total, rate) in summary.win_rate_by_signal_strength.items():
            if total > 0:
                ws.cell(row=row, column=1, value=tier.title()).border = THIN_BORDER
                ws.cell(row=row, column=2, value=wins).border = THIN_BORDER
                ws.cell(row=row, column=3, value=total).border = THIN_BORDER
                ws.cell(row=row, column=4, value=f"{rate:.1f}%").border = THIN_BORDER
                row += 1

        # Win Rate by Market Regime
        row += 1
        ws[f'A{row}'] = "WIN RATE BY MARKET REGIME"
        ws[f'A{row}'].font = SECTION_FONT

        row += 1
        for col_idx, header in enumerate(['Regime', 'Wins', 'Total', 'Win Rate'], 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER

        row += 1
        for regime, (wins, total, rate) in summary.win_rate_by_market_regime.items():
            if total > 0:
                ws.cell(row=row, column=1, value=regime.title()).border = THIN_BORDER
                ws.cell(row=row, column=2, value=wins).border = THIN_BORDER
                ws.cell(row=row, column=3, value=total).border = THIN_BORDER
                ws.cell(row=row, column=4, value=f"{rate:.1f}%").border = THIN_BORDER
                row += 1

        # MAE/MFE Summary
        row += 1
        ws[f'A{row}'] = "MAE/MFE BY OUTCOME"
        ws[f'A{row}'].font = SECTION_FONT

        row += 1
        mae_mfe_data = [
            ('Avg MAE (Winners)', f"{summary.avg_mae_winners:.2f}%"),
            ('Avg MAE (Losers)', f"{summary.avg_mae_losers:.2f}%"),
            ('Avg MFE (Winners)', f"{summary.avg_mfe_winners:.2f}%"),
            ('Avg MFE (Losers)', f"{summary.avg_mfe_losers:.2f}%"),
            ('Avg Duration (Winners)', f"{summary.avg_duration_winners:.1f} days"),
            ('Avg Duration (Losers)', f"{summary.avg_duration_losers:.1f} days"),
        ]

        for label, value in mae_mfe_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        # Insights
        row += 1
        ws[f'A{row}'] = "KEY INSIGHTS"
        ws[f'A{row}'].font = SECTION_FONT

        row += 1
        for insight in summary.insights:
            ws[f'A{row}'] = f"  {insight}"
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'].alignment = WRAP_ALIGN
            row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12

    def _create_multi_trade_overview(self,
                                      wb: Workbook,
                                      trades: List[TradeAnalysisData],
                                      signal_scores: Dict[str, SignalStrengthScore]):
        """Create overview sheet for multi-trade report."""
        ws = wb.create_sheet("Overview")

        ws['A1'] = "MULTI-TRADE ANALYSIS OVERVIEW"
        ws['A1'].font = TITLE_FONT
        ws.merge_cells('A1:I1')

        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        # Headers
        headers = ['Trade ID', 'Symbol', 'Side', 'Entry Date', 'Exit Date', 'P&L %', 'Duration', 'Signal Score', 'Outcome']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER

        # Data rows
        row = 5
        for trade in trades:
            info = trade.trade_info
            score = signal_scores.get(trade.trade_id)

            pl_pct = info.get('pl_pct', 0)

            data = [
                trade.trade_id,
                trade.symbol,
                info.get('side', 'LONG'),
                pd.to_datetime(info.get('entry_date')).strftime('%Y-%m-%d') if info.get('entry_date') else '',
                pd.to_datetime(info.get('exit_date')).strftime('%Y-%m-%d') if info.get('exit_date') else '',
                f"{pl_pct:+.2f}%",
                f"{info.get('duration_days', 0)} days",
                f"{score.total_score:.0f}/100" if score else "N/A",
                "WIN" if pl_pct > 0 else "LOSS"
            ]

            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = THIN_BORDER

            # Color outcome
            ws.cell(row=row, column=9).fill = POSITIVE_FILL if pl_pct > 0 else NEGATIVE_FILL
            row += 1

        # Adjust widths
        widths = [12, 8, 8, 12, 12, 10, 12, 12, 8]
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    def _create_trade_mini_summary(self,
                                    wb: Workbook,
                                    trade: TradeAnalysisData,
                                    signal_score: Optional[SignalStrengthScore],
                                    sheet_name: str):
        """Create mini summary sheet for a single trade in multi-trade report."""
        ws = wb.create_sheet(sheet_name)

        info = trade.trade_info
        pl_pct = info.get('pl_pct', 0)

        ws['A1'] = f"{trade.symbol} - {trade.trade_id}"
        ws['A1'].font = TITLE_FONT

        # Key metrics
        row = 3
        metrics = [
            ('Entry', f"{info.get('entry_date', 'N/A')} @ ${info.get('entry_price', 0):.4f}"),
            ('Exit', f"{info.get('exit_date', 'N/A')} @ ${info.get('exit_price', 0):.4f}"),
            ('P&L', f"${info.get('pl', 0):,.2f} ({pl_pct:+.2f}%)"),
            ('Duration', f"{info.get('duration_days', 0)} days"),
            ('Side', info.get('side', 'LONG')),
            ('Signal Score', f"{signal_score.total_score:.0f}/100 ({signal_score.tier})" if signal_score else "N/A"),
        ]

        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        # MAE/MFE
        if trade.mae_mfe:
            row += 1
            ws[f'A{row}'] = "MAE"
            ws[f'B{row}'] = f"{trade.mae_mfe.mae_pct:.2f}%"
            row += 1
            ws[f'A{row}'] = "MFE"
            ws[f'B{row}'] = f"{trade.mae_mfe.mfe_pct:.2f}%"

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
